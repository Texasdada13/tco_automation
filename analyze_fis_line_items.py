"""
Detailed analysis of FIS line items in WORKBOOK2.xlsx
"""
import openpyxl
from openpyxl.utils import get_column_letter
import json

def analyze_fis_line_items():
    """Extract all FIS line items from WORKBOOK2.xlsx"""

    wb = openpyxl.load_workbook('WORKBOOK2.xlsx', data_only=True)  # Get calculated values
    ws = wb['Line Items']

    print(f"\n{'='*100}")
    print("FIS LINE ITEMS ANALYSIS - WORKBOOK2.xlsx")
    print(f"{'='*100}\n")

    # First, find the FIS column (looks like it's column O based on earlier search)
    # Let's scan row 1 and row 4 to find column headers
    print("FINDING FIS COLUMNS...")
    fis_col = None
    for col in range(1, ws.max_column + 1):
        val = ws.cell(1, col).value
        if val and 'FIS' in str(val).upper():
            fis_col = col
            print(f"  FIS column found: Column {get_column_letter(col)} (index {col})")
            break

    if not fis_col:
        print("  ERROR: FIS column not found!")
        return

    # Now let's understand the structure by reading headers and first few rows
    print(f"\n{'='*100}")
    print("COLUMN HEADERS (around FIS section):")
    print(f"{'='*100}\n")

    # Show columns from A to FIS column + 5
    for col in range(1, min(fis_col + 6, ws.max_column + 1)):
        header_row1 = ws.cell(1, col).value
        header_row4 = ws.cell(4, col).value
        header_row5 = ws.cell(5, col).value
        col_letter = get_column_letter(col)

        print(f"  Col {col_letter}:")
        if header_row1:
            print(f"    Row 1: {header_row1}")
        if header_row4:
            print(f"    Row 4: {header_row4}")
        if header_row5:
            print(f"    Row 5: {header_row5}")

    # Now extract FIS line items
    print(f"\n{'='*100}")
    print("FIS LINE ITEMS EXTRACTION:")
    print(f"{'='*100}\n")

    # Define sections based on what we saw
    sections = {
        'Bundle FIS Products': [],
        'Non-Bundle REQUIRED FIS Products': [],
        'Non-Bundle REQUIRED Third Parties': [],
        'Implementation Credits and One-Time Fees': [],
        'Non-Bundle OPTIONAL FIS Solutions': [],
        'Non-Bundle OPTIONAL Third-Party Solutions': []
    }

    current_section = None
    line_items = []

    # Read through rows
    for row in range(6, min(160, ws.max_row + 1)):
        # Column A might have section headers
        col_a = ws.cell(row, 1).value
        col_b = ws.cell(row, 2).value

        # Check if this is a section header
        if col_b and isinstance(col_b, str):
            if 'Bundle FIS' in col_b:
                current_section = 'Bundle FIS Products'
                print(f"\n>>> SECTION: {current_section} (Row {row})")
                continue
            elif 'Non-Bundle REQUIRED FIS' in col_b:
                current_section = 'Non-Bundle REQUIRED FIS Products'
                print(f"\n>>> SECTION: {current_section} (Row {row})")
                continue
            elif 'Non-Bundle Required Third' in col_b:
                current_section = 'Non-Bundle REQUIRED Third Parties'
                print(f"\n>>> SECTION: {current_section} (Row {row})")
                continue
            elif 'Implementation Credits' in col_b and 'FIS' in col_b:
                current_section = 'Implementation Credits and One-Time Fees'
                print(f"\n>>> SECTION: {current_section} (Row {row})")
                continue
            elif 'Non-Bundle Optional FIS' in col_b:
                current_section = 'Non-Bundle OPTIONAL FIS Solutions'
                print(f"\n>>> SECTION: {current_section} (Row {row})")
                continue
            elif 'Non-Bundle Optional Third' in col_b:
                current_section = 'Non-Bundle OPTIONAL Third-Party Solutions'
                print(f"\n>>> SECTION: {current_section} (Row {row})")
                continue

        # Extract line item data
        # Column B = Fee Type, Column O (FIS col) = Proposal value
        fee_type = ws.cell(row, 2).value  # Column B
        fis_value = ws.cell(row, fis_col).value  # FIS column

        # Try to find solution name (might be in different columns)
        # Let's check columns around FIS
        solution_name = None
        for col_offset in [-3, -2, -1, 0, 1]:
            val = ws.cell(row, fis_col + col_offset).value
            if val and isinstance(val, str) and len(val) > 3 and not val.startswith('='):
                solution_name = val
                break

        # If we have a fee type and FIS value, this is likely a line item
        if fee_type and fee_type in ['Monthly F', 'Monthly V', 'Annual', 'One-Time']:
            if fis_value and (isinstance(fis_value, (int, float)) and fis_value != 0):
                item = {
                    'row': row,
                    'section': current_section,
                    'fee_type': fee_type,
                    'fis_value': fis_value,
                    'solution_name': solution_name
                }

                line_items.append(item)
                sections[current_section if current_section else 'Unknown'].append(item)

                print(f"  Row {row:3d}: {fee_type:12s} | ${fis_value:>12,.2f} | {solution_name or 'N/A'}")

    # Summary
    print(f"\n{'='*100}")
    print("SUMMARY BY SECTION:")
    print(f"{'='*100}\n")

    for section, items in sections.items():
        if items:
            total = sum(item['fis_value'] for item in items)
            print(f"\n{section}:")
            print(f"  Count: {len(items)} items")
            print(f"  Total value: ${total:,.2f}")

    print(f"\n{'='*100}")
    print(f"TOTAL FIS LINE ITEMS: {len(line_items)}")
    print(f"{'='*100}\n")

    # Save to JSON for comparison
    with open('workbook2_fis_items.json', 'w') as f:
        json.dump(line_items, f, indent=2)
    print("Saved FIS items to: workbook2_fis_items.json\n")

    wb.close()

if __name__ == '__main__':
    analyze_fis_line_items()
