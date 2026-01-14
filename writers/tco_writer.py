"""
TCO Writer - Populates TCO Excel template with normalized vendor data
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any
import sys
import os
sys.path.append('..')
from config import TCO_COLUMNS, LINE_ITEM_START_ROWS, FEE_TYPES


class TCOWriter:
    """Writes normalized vendor data to TCO Excel template"""
    
    def __init__(self, template_path: str, output_path: str = None):
        """
        Initialize TCO Writer
        
        Args:
            template_path: Path to TCO template Excel file
            output_path: Path for output file (if None, overwrites template)
        """
        self.template_path = template_path
        self.output_path = output_path or template_path.replace('.xlsx', '_populated.xlsx')
        self.wb = load_workbook(template_path)
        self.line_items_sheet = self.wb['Line Items']
    
    def write_vendor_data(self, line_items: List[Dict[str, Any]], vendor: str):
        """
        Write vendor line items to TCO template
        
        Args:
            line_items: List of normalized line items
            vendor: 'FIS' or 'Jack Henry'
        """
        print(f"\nWriting {len(line_items)} line items for {vendor}...")
        
        # Get column mapping for this vendor
        if vendor == 'FIS':
            col_map = TCO_COLUMNS['FIS']
        elif vendor == 'Jack Henry':
            col_map = TCO_COLUMNS['JACK_HENRY']
        else:
            raise ValueError(f"Unknown vendor: {vendor}")
        
        # Group line items by category
        bundled_items = []
        non_bundle_required = []
        non_bundle_optional = []
        one_time_items = []
        
        for item in line_items:
            fee_type = item['fee_type']
            category = item['category']
            
            if fee_type == FEE_TYPES['one_time']:
                one_time_items.append(item)
            elif 'Bundle' in category:
                bundled_items.append(item)
            elif 'Optional' in category:
                non_bundle_optional.append(item)
            else:
                non_bundle_required.append(item)
        
        # Write each category starting at appropriate row
        vendor_prefix = 'FIS' if vendor == 'FIS' else 'JH'
        
        current_row = LINE_ITEM_START_ROWS[f'{vendor_prefix}_BUNDLE']
        current_row = self._write_line_items(bundled_items, col_map, current_row)
        
        current_row = LINE_ITEM_START_ROWS[f'{vendor_prefix}_NON_BUNDLE_REQUIRED']
        current_row = self._write_line_items(non_bundle_required, col_map, current_row)
        
        current_row = LINE_ITEM_START_ROWS[f'{vendor_prefix}_NON_BUNDLE_OPTIONAL']
        current_row = self._write_line_items(non_bundle_optional, col_map, current_row)
        
        current_row = LINE_ITEM_START_ROWS[f'{vendor_prefix}_ONE_TIME']
        current_row = self._write_line_items(one_time_items, col_map, current_row)
        
        print(f"Successfully wrote {vendor} data to template")
    
    def _write_line_items(self, items: List[Dict[str, Any]], col_map: Dict[str, str], start_row: int) -> int:
        """
        Write list of line items starting at specified row
        
        Returns:
            Next available row number
        """
        current_row = start_row
        
        for item in items:
            self._write_line_item(item, col_map, current_row)
            current_row += 1
        
        return current_row
    
    def _write_line_item(self, item: Dict[str, Any], col_map: Dict[str, str], row: int):
        """Write a single line item to the specified row"""
        sheet = self.line_items_sheet
        
        def safe_write(cell_ref, value):
            """Safely write to cell, handling merged cells"""
            try:
                cell = sheet[cell_ref]
                # Check if it's a merged cell
                if isinstance(cell, type(cell)) and hasattr(cell, 'value'):
                    cell.value = value
            except AttributeError:
                # Skip merged cells
                pass
            except Exception as e:
                print(f"Warning: Could not write to {cell_ref}: {e}")
        
        # Write fee type
        safe_write(f"{col_map['fee_type']}{row}", item['fee_type'])
        
        # Write proposal/per unit rate
        safe_write(f"{col_map['proposal']}{row}", item.get('per_unit_rate', 0))
        
        # Write solution name
        safe_write(f"{col_map['solution_name']}{row}", item['solution_name'])
        
        # Write category
        safe_write(f"{col_map['category']}{row}", item['category'])
        
        # Write per unit rate
        safe_write(f"{col_map['per_unit_rate']}{row}", item.get('per_unit_rate', 0))
        
        # Write quantities by year
        quantities = item.get('quantities_by_year', {})
        if quantities:
            for year_num in range(1, 8):  # Years 1-7
                year_key = f'year_{year_num}'
                qty = quantities.get(year_key, 0)
                
                # Write quantity
                qty_col = col_map.get(f'year_{year_num}_qty')
                if qty_col:
                    safe_write(f"{qty_col}{row}", qty)
        else:
            # For one-time items, set average monthly qty to 0
            safe_write(f"{col_map['avg_monthly_qty']}{row}", 0)
        
        # Write costs by year (these will be calculated by formulas in the actual template)
        # For now, we'll write the calculated values
        for year_num in range(1, 8):
            cost_col = col_map.get(f'year_{year_num}_cost')
            if not cost_col:
                continue
            
            year_key = f'year_{year_num}'
            qty = quantities.get(year_key, 0) if quantities else 0
            
            if item['fee_type'] == FEE_TYPES['one_time']:
                # One-time fees only apply in Year 1
                cost = item['one_time_fee'] if year_num == 1 else 0
            elif 'Monthly' in item['fee_type']:
                # Monthly fees: multiply by quantity (which is in months)
                cost = item.get('per_unit_rate', 0) * qty if qty > 0 else 0
            elif item['fee_type'] == FEE_TYPES['annual']:
                # Annual fees
                cost = item.get('annual_fee', 0) if qty > 0 else 0
            else:
                cost = 0
            
            safe_write(f"{cost_col}{row}", cost)
    
    def save(self):
        """Save the populated workbook"""
        self.wb.save(self.output_path)
        print(f"\nTCO template saved to: {self.output_path}")
        return self.output_path


def populate_tco_template(template_path: str, vendor_data: Dict[str, Any], 
                         vendor: str, output_path: str = None) -> str:
    """
    Convenience function to populate TCO template
    
    Args:
        template_path: Path to TCO template
        vendor_data: List of normalized line items
        vendor: 'FIS' or 'Jack Henry'
        output_path: Optional output path
    
    Returns:
        Path to populated file
    """
    writer = TCOWriter(template_path, output_path)
    writer.write_vendor_data(vendor_data, vendor)
    return writer.save()


if __name__ == "__main__":
    # Test TCO writing
    import sys
    import os
    sys.path.append('..')
    from extractors import extract_fis_proposal
    from mappers import normalize_vendor_data
    
    data_dir = os.path.join(os.path.dirname(__file__), '..', 'data')
    
    print("="*70)
    print("Testing TCO Writer with FIS Data")
    print("="*70)
    
    # Extract and normalize FIS data
    fis_file = os.path.join(data_dir, "Echelon_FIS_Proposal_10_29_25.docx")
    fis_data = extract_fis_proposal(fis_file)
    fis_normalized = normalize_vendor_data(fis_data, 'FIS', term='7_year')
    
    # Load template and populate
    template_file = os.path.join(data_dir, "Echelon_Primary_TCO_v5_11_13_25.xlsx")
    output_file = os.path.join(data_dir, "TCO_Test_Output.xlsx")
    
    result = populate_tco_template(template_file, fis_normalized, 'FIS', output_file)
    print(f"\nSuccess! Output file: {result}")
