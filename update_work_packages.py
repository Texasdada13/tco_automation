"""
Comprehensive Work Package Definition Script
Updates Project Plan with detailed activities, effort, resources, risks, and acceptance criteria
Based on codebase analysis and industry best practices (2026)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Comprehensive work package definitions
WORK_PACKAGES = {
    'WP0': {
        'name': 'Discovery Phase',
        'tasks': [
            {
                'id': '0.1',
                'name': 'Historical Proposal Inventory',
                'description': 'Catalog all historical proposals: file count, formats, vendors, date ranges',
                'activities': [
                    'Scan client file systems for proposal documents',
                    'Categorize by vendor (FIS, Jack Henry, CSI, Fiserv, etc.)',
                    'Identify file formats (Word, Excel, PDF, scanned images)',
                    'Document naming conventions and folder structures',
                    'Assess file sizes and complexity (page counts, table counts)',
                    'Create inventory spreadsheet with metadata',
                    'Identify corrupted or inaccessible files',
                    'Estimate extraction complexity scoring (1-5 scale)'
                ],
                'deliverable': 'Historical_Proposal_Inventory.xlsx (with metadata: vendor, format, date, pages, complexity score)',
                'dependencies': '-',
                'effort_days': 3,
                'resources': 'BA + Client SME',
                'risks': [
                    'Files scattered across multiple locations/systems',
                    'Missing metadata (no clear vendor/date identification)',
                    'Legacy formats no longer supported',
                    'Access permissions/security restrictions'
                ],
                'acceptance_criteria': [
                    '100% of accessible proposal files cataloged',
                    'Each file tagged with: vendor, format, date, page count',
                    'Complexity scores assigned (based on table count, layout)',
                    'Client SME validation of inventory completeness',
                    'Sample files extracted for each vendor/format combination'
                ]
            },
            {
                'id': '0.2',
                'name': 'TCO Template Variants Collection',
                'description': 'Gather all TCO output template variants and document differences',
                'activities': [
                    'Collect all Excel TCO templates currently in use',
                    'Document template versioning and evolution history',
                    'Map field differences between template versions',
                    'Identify client-specific customizations',
                    'Document formula logic in each template',
                    'Extract column mappings and cell coordinates',
                    'Analyze conditional formatting and validation rules',
                    'Create template taxonomy and decision tree'
                ],
                'deliverable': 'TCO_Template_Catalog.xlsx + Template_Mapping_Specifications.docx',
                'dependencies': '-',
                'effort_days': 4,
                'resources': 'Developer + Client SME',
                'risks': [
                    'Undocumented template modifications over time',
                    'Complex formulas with unclear business logic',
                    'Client-specific templates not shared initially',
                    'Template versions incompatible with current automation'
                ],
                'acceptance_criteria': [
                    'All template variants collected (min 3-5 expected)',
                    'Field-by-field mapping documented for each template',
                    'Formula logic reverse-engineered and documented',
                    'Decision criteria for template selection defined',
                    'Backward compatibility assessment completed'
                ]
            },
            {
                'id': '0.3',
                'name': 'Vendor Document Pattern Analysis',
                'description': 'Analyze proposal patterns for each vendor to optimize extraction strategies',
                'activities': [
                    'Select representative samples (3-5 per vendor)',
                    'Manual extraction from samples to establish baseline',
                    'Document table structures and layouts per vendor',
                    'Identify vendor-specific terminology mappings',
                    'Test current extraction methods on samples',
                    'Measure accuracy of rule-based vs AI extraction',
                    'Document extraction failure patterns',
                    'Assess need for vendor-specific extraction rules',
                    'Evaluate direct PDF vs hybrid extraction methods'
                ],
                'deliverable': 'Vendor_Pattern_Analysis.docx + Extraction_Accuracy_Baseline.xlsx',
                'dependencies': '0.1',
                'effort_days': 5,
                'resources': 'Developer + Data Analyst',
                'risks': [
                    'Vendor document formats change over time (no consistency)',
                    'Sample size too small to capture all variations',
                    'Current extraction accuracy lower than expected (<90%)',
                    'Complex layouts require hybrid/manual extraction'
                ],
                'acceptance_criteria': [
                    'Min 3 samples per vendor analyzed (FIS, JH, CSI at minimum)',
                    'Extraction accuracy baseline established (target ≥95%)',
                    'Vendor terminology dictionary created for each vendor',
                    'Recommended extraction strategy documented per vendor',
                    'Known failure scenarios cataloged with workarounds'
                ]
            },
            {
                'id': '0.4',
                'name': 'Volume & Performance Requirements Definition',
                'description': 'Define SLAs, volume projections, and performance requirements',
                'activities': [
                    'Document current manual processing times per proposal',
                    'Define monthly volume projections (current + 12 months)',
                    'Establish performance SLAs (processing time per proposal)',
                    'Define accuracy SLAs (extraction accuracy targets)',
                    'Document peak processing periods and surge capacity needs',
                    'Calculate required system capacity (concurrent jobs)',
                    'Define storage requirements (JSON, PDFs, audit logs)',
                    'Establish monitoring and alerting requirements'
                ],
                'deliverable': 'Requirements_Specification.docx (including SLAs, volume projections, NFRs)',
                'dependencies': '0.1',
                'effort_days': 3,
                'resources': 'BA + Client PM + SME',
                'risks': [
                    'Volume projections too conservative (system under-scoped)',
                    'SLAs too aggressive given technical constraints',
                    'Peak periods exceed capacity (need surge planning)',
                    'Historical baseline data not available'
                ],
                'acceptance_criteria': [
                    'Volume projections defined: current, 6mo, 12mo',
                    'Processing time SLA agreed (e.g., <5 min per proposal)',
                    'Accuracy SLA defined (e.g., ≥95% extraction accuracy)',
                    'Storage requirements calculated based on 2-year retention',
                    'Peak capacity requirements identified (surge scenarios)'
                ]
            },
            {
                'id': '0.5',
                'name': 'Technical Architecture Design',
                'description': 'Define target architecture for enterprise scale-up',
                'activities': [
                    'Review current POC architecture and identify gaps',
                    'Design database schema (proposals, extractions, jobs, audit)',
                    'Define API architecture (REST endpoints, authentication)',
                    'Design batch processing queue system (Celery/RabbitMQ)',
                    'Plan scalability approach (horizontal vs vertical)',
                    'Define deployment architecture (cloud vs on-premise)',
                    'Design security model (encryption, RBAC, secrets mgmt)',
                    'Create infrastructure diagrams (C4 model)',
                    'Define monitoring and observability stack',
                    'Document technology decisions and tradeoffs'
                ],
                'deliverable': 'Technical_Architecture_Document.docx + Architecture_Diagrams.pdf (C4 diagrams)',
                'dependencies': '0.3, 0.4',
                'effort_days': 5,
                'resources': 'Tech Lead + Architect',
                'risks': [
                    'Over-engineering solution (excessive complexity)',
                    'Technology choices conflict with client IT standards',
                    'Cloud vs on-premise decision delayed or changes',
                    'Scalability assumptions invalidated by volume reality'
                ],
                'acceptance_criteria': [
                    'C4 architecture diagrams (Context, Container, Component)',
                    'Database schema defined (ER diagrams with constraints)',
                    'API specifications documented (OpenAPI/Swagger)',
                    'Deployment model selected and approved by client IT',
                    'Security model reviewed and approved',
                    'Technology stack ratified (Python, DB, queue, API framework)'
                ]
            },
            {
                'id': '0.6',
                'name': 'Risk Assessment & Mitigation Planning',
                'description': 'Identify project risks and define mitigation strategies',
                'activities': [
                    'Conduct risk workshop with stakeholders',
                    'Identify technical risks (vendor format changes, AI accuracy)',
                    'Identify operational risks (resource availability, SME access)',
                    'Identify data risks (quality, volume, security)',
                    'Score risks (probability x impact)',
                    'Define mitigation strategies for high/medium risks',
                    'Create risk register with owners and tracking',
                    'Establish risk review cadence (weekly during dev)'
                ],
                'deliverable': 'Risk_Register.xlsx (with mitigation plans and owners)',
                'dependencies': '0.1-0.5',
                'effort_days': 2,
                'resources': 'PM + Tech Lead + Client PM',
                'risks': [
                    'New risks emerge during development (inadequate initial assessment)',
                    'Risk mitigation strategies inadequate or not executed',
                    'Risks not monitored/updated regularly',
                    'Scope creep disguised as "risk mitigation"'
                ],
                'acceptance_criteria': [
                    'Min 15-20 risks identified and documented',
                    'All high-impact risks have mitigation plans',
                    'Risk owners assigned for each risk',
                    'Risk review process established (weekly cadence)',
                    'Client sign-off on risk register'
                ]
            },
            {
                'id': '0.7',
                'name': 'Discovery Exit & Go/No-Go Decision',
                'description': 'Review discovery findings and obtain approval to proceed',
                'activities': [
                    'Compile discovery findings report',
                    'Present findings to client stakeholders',
                    'Review scope adjustments (if needed based on findings)',
                    'Confirm SLAs and acceptance criteria',
                    'Review and approve updated project plan',
                    'Obtain sign-off on discovery deliverables (DD1-DD6)',
                    'Confirm resource availability for dev phase',
                    'Execute Go/No-Go decision'
                ],
                'deliverable': 'Discovery_Exit_Report.pptx + Signed_Discovery_Acceptance.pdf',
                'dependencies': '0.1-0.6',
                'effort_days': 2,
                'resources': 'PM + Tech Lead + Client Sponsor',
                'risks': [
                    'Discovery reveals scope significantly larger than estimated',
                    'Client requests major scope changes requiring re-estimate',
                    'Resource availability conflicts delay dev start',
                    'Budget constraints halt project'
                ],
                'acceptance_criteria': [
                    'All discovery deliverables (DD1-DD6) completed',
                    'Client formal acceptance of discovery findings',
                    'Scope adjustments (if any) documented and approved',
                    'Dev phase work packages validated and approved',
                    'Project kickoff scheduled',
                    'Payment milestone M0 (10%) approved for invoicing'
                ]
            }
        ]
    },
    'WP1': {
        'name': 'Foundation Infrastructure',
        'tasks': [
            {
                'id': '1.1',
                'name': 'Database Design & Implementation',
                'description': 'Design and implement PostgreSQL database for proposals, extractions, jobs, audit',
                'activities': [
                    'Create detailed database schema (proposals, extractions, jobs, users, audit_log)',
                    'Define table relationships and foreign keys',
                    'Design indexes for performance (on vendor, date, status fields)',
                    'Implement database migration framework (Alembic)',
                    'Create seed data for development/testing',
                    'Implement database connection pooling (SQLAlchemy)',
                    'Create ORM models (SQLAlchemy or Django ORM)',
                    'Implement database backup and recovery procedures',
                    'Write SQL stored procedures for complex queries',
                    'Performance testing with simulated load (1000+ records)'
                ],
                'deliverable': 'PostgreSQL database + Migration scripts + ORM models + DB_Schema_Documentation.md',
                'dependencies': 'WP0 (Discovery)',
                'effort_days': 8,
                'resources': 'Backend Developer + DBA',
                'risks': [
                    'Schema design doesn\'t accommodate future template variations',
                    'Performance issues with large JSON fields (extracted_data JSONB)',
                    'Migration conflicts when schema changes during dev',
                    'Database choice conflicts with client IT standards (e.g., must use MSSQL)'
                ],
                'acceptance_criteria': [
                    'All tables created with proper constraints and indexes',
                    'Foreign key relationships enforced',
                    'Migration scripts tested (up and down migrations)',
                    'Query performance < 100ms for common operations',
                    'Backup/restore tested successfully',
                    'Database documented with ER diagrams'
                ]
            },
            {
                'id': '1.2',
                'name': 'Job Queue System Implementation',
                'description': 'Implement Celery + RabbitMQ for asynchronous job processing',
                'activities': [
                    'Install and configure RabbitMQ message broker',
                    'Set up Celery worker infrastructure',
                    'Design job queue architecture (task routing, priorities)',
                    'Implement extraction job task (celery task)',
                    'Implement job status tracking (pending, in_progress, completed, failed)',
                    'Create job retry logic with exponential backoff',
                    'Implement job result storage (database + file system)',
                    'Set up Flower for queue monitoring (web UI)',
                    'Configure worker concurrency and resource limits',
                    'Implement dead letter queue for failed jobs',
                    'Write queue health check and monitoring scripts'
                ],
                'deliverable': 'Celery + RabbitMQ infrastructure + Job management API + Queue_Setup_Documentation.md',
                'dependencies': '1.1',
                'effort_days': 7,
                'resources': 'Backend Developer + DevOps',
                'risks': [
                    'RabbitMQ configuration complex for team unfamiliar with it',
                    'Worker crashes under high load (memory leaks)',
                    'Job queue fills up faster than processing (backlog)',
                    'Network issues between queue and workers (distributed setup)'
                ],
                'acceptance_criteria': [
                    'RabbitMQ running and accessible',
                    'Celery workers processing jobs successfully',
                    'Job status tracked in database in real-time',
                    'Failed jobs automatically retried (max 3 attempts)',
                    'Flower dashboard shows queue metrics',
                    'Load tested: 20 concurrent jobs processed successfully'
                ]
            },
            {
                'id': '1.3',
                'name': 'Configuration Management System',
                'description': 'Implement environment-based configuration management',
                'activities': [
                    'Create config file structure (config/dev.py, config/prod.py, etc.)',
                    'Implement environment detection (.env file + env variables)',
                    'Migrate hardcoded API keys to environment variables',
                    'Implement secrets management (AWS Secrets Manager or HashiCorp Vault)',
                    'Create configuration validation on startup',
                    'Document all configuration parameters',
                    'Implement feature flags for gradual rollout',
                    'Create configuration template files (.env.example)',
                    'Implement configuration reload without restart (where possible)',
                    'Add configuration audit logging (who changed what when)'
                ],
                'deliverable': 'Config management system + Secrets vault integration + Configuration_Guide.md',
                'dependencies': '-',
                'effort_days': 4,
                'resources': 'Backend Developer',
                'risks': [
                    'Secrets accidentally committed to Git (hardcoded API keys)',
                    'Environment-specific bugs due to config differences',
                    'Configuration drift between environments',
                    'Secrets vault access issues in production'
                ],
                'acceptance_criteria': [
                    'No hardcoded secrets in codebase (Git history cleaned)',
                    'All environments (dev, staging, prod) using env-specific configs',
                    'Secrets rotated without code changes',
                    'Configuration validation prevents startup with invalid config',
                    'Config changes logged to audit trail'
                ]
            },
            {
                'id': '1.4',
                'name': 'Structured Logging Infrastructure',
                'description': 'Implement centralized structured logging with rotation and monitoring',
                'activities': [
                    'Implement structured logging (JSON format with context)',
                    'Set up log rotation policies (size-based + time-based)',
                    'Configure log levels per module (DEBUG, INFO, WARNING, ERROR)',
                    'Implement correlation IDs for request tracing',
                    'Set up centralized log aggregation (ELK stack or CloudWatch)',
                    'Create log parsing and alerting rules',
                    'Implement PII masking in logs (no sensitive data logged)',
                    'Add performance metrics logging (processing times, API calls)',
                    'Create log retention policies (30 days debug, 1 year audit)',
                    'Set up log monitoring dashboards'
                ],
                'deliverable': 'Logging framework + Log aggregation system + Logging_Standards.md',
                'dependencies': '-',
                'effort_days': 5,
                'resources': 'Backend Developer + DevOps',
                'risks': [
                    'Log volume exceeds storage capacity (need compression/archival)',
                    'PII accidentally logged (compliance violation)',
                    'Logs don\'t correlate across distributed workers',
                    'Log aggregation system performance bottleneck'
                ],
                'acceptance_criteria': [
                    'All modules using structured logging (JSON format)',
                    'Logs include: timestamp, level, correlation_id, module, message, context',
                    'Log rotation working (max 100MB per file, 30-day retention)',
                    'PII masking validated (no SSN, credit cards in logs)',
                    'Centralized dashboard shows logs across all workers',
                    'Alerting configured for ERROR and CRITICAL logs'
                ]
            },
            {
                'id': '1.5',
                'name': 'API Framework Setup',
                'description': 'Build RESTful API foundation with FastAPI',
                'activities': [
                    'Set up FastAPI application structure',
                    'Implement API authentication (JWT tokens or API keys)',
                    'Create API documentation (OpenAPI/Swagger auto-generated)',
                    'Implement request validation (Pydantic models)',
                    'Add API rate limiting and throttling',
                    'Implement CORS configuration',
                    'Create health check endpoints (/health, /ready)',
                    'Implement API versioning strategy (v1, v2 paths)',
                    'Add request/response logging middleware',
                    'Set up API testing framework (pytest + httpx)',
                    'Implement error handling and standard error responses'
                ],
                'deliverable': 'FastAPI application + API documentation + API_Standards.md',
                'dependencies': '1.1, 1.3',
                'effort_days': 6,
                'resources': 'Backend Developer',
                'risks': [
                    'API design doesn\'t meet client integration needs',
                    'Authentication mechanism too complex for clients',
                    'Rate limiting too restrictive (blocks legitimate use)',
                    'API breaking changes during development'
                ],
                'acceptance_criteria': [
                    'FastAPI app running and accessible',
                    'Swagger UI accessible at /docs endpoint',
                    'Authentication working (valid tokens accepted, invalid rejected)',
                    'All endpoints return standard error format',
                    'Health check returns 200 with system status',
                    'API tested with Postman/httpx (test collection provided)'
                ]
            }
        ]
    },
    'WP2': {
        'name': 'Batch Processing Engine',
        'tasks': [
            {
                'id': '2.1',
                'name': 'Bulk Import & Queue Management',
                'description': 'Build system to ingest multiple proposals and queue for processing',
                'activities': [
                    'Implement directory watcher for auto-ingestion (watchdog library)',
                    'Create bulk upload API endpoint (multi-file upload)',
                    'Implement file validation (format, size, virus scan)',
                    'Build proposal metadata extraction (vendor detection, date parsing)',
                    'Implement automatic vendor classification (FIS, JH, CSI detection)',
                    'Create queue prioritization logic (urgent vs normal)',
                    'Implement duplicate detection (hash-based)',
                    'Build proposal staging area (uploaded files before processing)',
                    'Create batch submission API (submit 20+ files at once)',
                    'Implement queue status API (how many jobs in queue)',
                    'Add batch validation reporting (which files accepted/rejected)'
                ],
                'deliverable': 'Bulk import system + Queue management API + Import_User_Guide.md',
                'dependencies': 'WP1 (DB, Queue)',
                'effort_days': 7,
                'resources': 'Backend Developer',
                'risks': [
                    'Large file uploads fail or timeout (>50MB PDFs)',
                    'Virus scanning delays ingestion (performance bottleneck)',
                    'Vendor auto-detection fails on ambiguous documents',
                    'Queue fills faster than workers process (backlog grows)'
                ],
                'acceptance_criteria': [
                    'Bulk upload handles 20+ files in single request',
                    'Vendor auto-detection ≥90% accuracy (tested on samples)',
                    'Duplicate files rejected before queueing',
                    'Invalid files rejected with clear error messages',
                    'Queue prioritization working (urgent jobs processed first)',
                    'Directory watcher auto-ingests files from hot folder'
                ]
            },
            {
                'id': '2.2',
                'name': 'Parallel Worker Pool Implementation',
                'description': 'Implement multi-worker system for concurrent proposal processing',
                'activities': [
                    'Configure Celery worker pool (4-8 concurrent workers)',
                    'Implement worker health monitoring and auto-restart',
                    'Set up worker resource limits (memory, CPU)',
                    'Implement graceful worker shutdown (finish current job)',
                    'Create worker scaling strategy (manual or auto-scaling)',
                    'Implement job timeout handling (max 10 min per job)',
                    'Add worker performance metrics collection',
                    'Implement worker task routing (by vendor or complexity)',
                    'Set up worker-specific logging',
                    'Create worker management commands (start, stop, scale)',
                    'Implement job cancellation (kill running job)'
                ],
                'deliverable': 'Worker pool infrastructure + Scaling playbook + Worker_Operations_Guide.md',
                'dependencies': '1.2 (Job Queue)',
                'effort_days': 6,
                'resources': 'Backend Developer + DevOps',
                'risks': [
                    'Workers crash under load (memory leaks, API timeouts)',
                    'Uneven load distribution (some workers idle, others overloaded)',
                    'Worker scaling too slow (can\'t handle surge)',
                    'Jobs timeout on complex proposals (need tuning)'
                ],
                'acceptance_criteria': [
                    'Min 4 workers processing jobs concurrently',
                    'Worker crashes auto-restart within 30 seconds',
                    'Jobs distributed evenly across workers',
                    'Worker resource usage monitored (CPU, memory)',
                    'Timeout configured (jobs fail after 10 min)',
                    'Load tested: 20 proposals processed in <30 min (avg 90 sec each)'
                ]
            },
            {
                'id': '2.3',
                'name': 'Error Handling & Retry Logic',
                'description': 'Implement comprehensive error handling with intelligent retry',
                'activities': [
                    'Define error taxonomy (transient vs permanent errors)',
                    'Implement retry logic with exponential backoff (3 attempts)',
                    'Create dead letter queue for permanent failures',
                    'Implement error notification system (email/Slack)',
                    'Build error detail capture (stack traces, context)',
                    'Create error recovery procedures documentation',
                    'Implement manual retry from UI (for dead letter queue)',
                    'Add error rate alerting (>10% failure rate)',
                    'Implement partial success handling (some items extracted)',
                    'Create error analytics dashboard (failure patterns)',
                    'Build automatic error categorization (AI timeout, PDF corrupt, etc.)'
                ],
                'deliverable': 'Error handling framework + Dead letter queue + Error_Playbook.md',
                'dependencies': '2.1, 2.2',
                'effort_days': 5,
                'resources': 'Backend Developer',
                'risks': [
                    'Retry logic creates infinite loops (retry permanently failing job)',
                    'Error notifications too noisy (alert fatigue)',
                    'Dead letter queue grows unbounded (no cleanup)',
                    'Transient errors misclassified as permanent (no retry)'
                ],
                'acceptance_criteria': [
                    'Transient errors auto-retry (max 3 attempts)',
                    'Permanent failures move to dead letter queue',
                    'Error notifications sent for all failures (email or Slack)',
                    'Error details logged with full context',
                    'Dead letter queue UI allows manual retry',
                    'Tested: API timeout retries, corrupt PDF fails permanently'
                ]
            },
            {
                'id': '2.4',
                'name': 'Progress Tracking & Reporting',
                'description': 'Real-time progress visibility for batch jobs',
                'activities': [
                    'Implement job progress tracking (0-100% for each job)',
                    'Create batch progress API (overall batch completion)',
                    'Build WebSocket for real-time progress updates',
                    'Implement stage-based progress (uploading, queued, extracting, QA, complete)',
                    'Create progress event logging (stage transitions)',
                    'Build progress dashboard UI (simple HTML + JS)',
                    'Implement ETA calculation (based on average processing time)',
                    'Add throughput metrics (proposals per hour)',
                    'Create batch summary report (X of Y complete, Z failed)',
                    'Implement progress notifications (batch 50% done, batch complete)'
                ],
                'deliverable': 'Progress tracking API + Real-time dashboard + Progress_API_Docs.md',
                'dependencies': '2.2',
                'effort_days': 6,
                'resources': 'Full Stack Developer',
                'risks': [
                    'WebSocket connection drops (progress updates lost)',
                    'ETA calculation inaccurate (wide variance in processing time)',
                    'Progress updates too frequent (performance overhead)',
                    'Progress tracking doesn\'t work across distributed workers'
                ],
                'acceptance_criteria': [
                    'Real-time progress visible in dashboard (updates every 5 sec)',
                    'Batch progress API returns: total, completed, failed, in_progress',
                    'ETA displayed (based on rolling average)',
                    'Progress notifications sent at 50% and 100%',
                    'Tested: 20-proposal batch tracked from start to finish',
                    'WebSocket reconnects automatically on disconnect'
                ]
            }
        ]
    },
    'WP3': {
        'name': 'Multi-Template TCO Support',
        'tasks': [
            {
                'id': '3.1',
                'name': 'Template Registry & Versioning',
                'description': 'Build system to manage multiple TCO template variants',
                'activities': [
                    'Design template registry database schema',
                    'Implement template upload and storage',
                    'Create template versioning system (v1.0, v1.1, v2.0)',
                    'Build template metadata extraction (column mappings, formulas)',
                    'Implement template validation (required fields present)',
                    'Create template diff tool (compare versions)',
                    'Implement template activation/deactivation (which is current)',
                    'Build template selection API (get template by ID or criteria)',
                    'Create template audit log (who uploaded, when, changes)',
                    'Implement template export/import (for backup/migration)',
                    'Build template testing framework (validate mappings work)'
                ],
                'deliverable': 'Template registry system + Versioning API + Template_Management_Guide.md',
                'dependencies': 'WP1 (Database)',
                'effort_days': 7,
                'resources': 'Backend Developer + BA',
                'risks': [
                    'Template mapping extraction fails on complex templates',
                    'Version conflicts when multiple templates updated',
                    'Template validation too strict (rejects valid templates)',
                    'Backward compatibility broken when template structure changes'
                ],
                'acceptance_criteria': [
                    'Template registry stores ≥3 template variants',
                    'Template versioning working (can rollback to previous version)',
                    'Template metadata extracted automatically on upload',
                    'Template diff shows field mapping changes',
                    'Audit log tracks all template changes',
                    'Tested: upload template, activate, generate TCO output'
                ]
            },
            {
                'id': '3.2',
                'name': 'Dynamic Field Mapping Engine',
                'description': 'Build flexible field-to-cell mapping system',
                'activities': [
                    'Design mapping configuration schema (JSON or YAML)',
                    'Implement field mapping parser (config → execution plan)',
                    'Create mapping rule engine (conditional mappings)',
                    'Build cell coordinate calculator (dynamic row/column)',
                    'Implement formula template engine (parameterized formulas)',
                    'Create data transformation layer (format conversions)',
                    'Implement mapping validation (all fields mapped)',
                    'Build mapping test framework (verify mappings correct)',
                    'Create mapping versioning (tied to template version)',
                    'Implement fallback mappings (for missing fields)',
                    'Build mapping documentation generator'
                ],
                'deliverable': 'Mapping engine + Configuration schema + Mapping_Configuration_Guide.md',
                'dependencies': '3.1',
                'effort_days': 8,
                'resources': 'Developer + Data Engineer',
                'risks': [
                    'Mapping configuration too complex for non-developers',
                    'Dynamic cell calculation errors (off-by-one)',
                    'Formula templating doesn\'t support all Excel formulas',
                    'Mapping changes break existing extractions'
                ],
                'acceptance_criteria': [
                    'Mapping config supports: fixed cells, dynamic rows, conditional logic',
                    'Formula templates support: SUM, IF, VLOOKUP, INDEX/MATCH',
                    'Mapping validation catches: missing fields, invalid cells',
                    'Mapping tested on 3+ template variants successfully',
                    'Fallback mappings prevent failures from missing fields',
                    'Mapping documentation auto-generated from config'
                ]
            },
            {
                'id': '3.3',
                'name': 'Template Auto-Detection Logic',
                'description': 'Automatically match proposal to correct TCO template',
                'activities': [
                    'Define template selection criteria (vendor, client, date, proposal type)',
                    'Implement rule-based template matching',
                    'Build template scoring algorithm (best match)',
                    'Create template override mechanism (manual selection)',
                    'Implement template recommendation system',
                    'Add template selection logging (why this template chosen)',
                    'Build template selection testing framework',
                    'Create template selection API',
                    'Implement client-specific template routing',
                    'Build default template fallback (if no match)',
                    'Add template selection audit trail'
                ],
                'deliverable': 'Template detection engine + Selection API + Template_Selection_Logic.md',
                'dependencies': '3.1, 3.2',
                'effort_days': 5,
                'resources': 'Developer',
                'risks': [
                    'Auto-detection selects wrong template (poor matching logic)',
                    'No suitable template found for proposal (need fallback)',
                    'Client-specific routing rules conflict',
                    'Template selection logic too rigid (can\'t handle edge cases)'
                ],
                'acceptance_criteria': [
                    'Template detection ≥95% accuracy (tested on sample proposals)',
                    'Selection logic considers: vendor, client, date, custom rules',
                    'Manual override available via API',
                    'Selection rationale logged (why template X chosen)',
                    'Tested: proposals from 3 vendors route to correct templates',
                    'Fallback to default template if no match'
                ]
            },
            {
                'id': '3.4',
                'name': 'Template Management UI',
                'description': 'Web interface for template configuration (non-developer)',
                'activities': [
                    'Design template management UI wireframes',
                    'Implement template upload interface',
                    'Build visual mapping configurator (point-and-click)',
                    'Create template preview (see mapping results)',
                    'Implement template testing tool (upload sample, see output)',
                    'Build template comparison view (side-by-side)',
                    'Create template activation/deactivation controls',
                    'Implement template version history view',
                    'Build template documentation viewer',
                    'Add template export/import UI',
                    'Implement user permissions (who can edit templates)'
                ],
                'deliverable': 'Template management web UI + User_Guide.pdf',
                'dependencies': '3.1, 3.2, 3.3',
                'effort_days': 10,
                'resources': 'Full Stack Developer + UX Designer',
                'risks': [
                    'UI too complex for business users (requires training)',
                    'Visual mapping configurator buggy (hard to use)',
                    'Template preview doesn\'t match actual output',
                    'UI performance slow with large templates'
                ],
                'acceptance_criteria': [
                    'Template upload via drag-and-drop',
                    'Visual mapping configurator (map fields to cells without code)',
                    'Template preview shows sample TCO output',
                    'Template testing: upload proposal + template → see result',
                    'Version history shows all template changes',
                    'Tested: non-developer user can upload and configure template'
                ]
            }
        ]
    },
    'WP4': {
        'name': 'Historical Proposal Processing',
        'tasks': [
            {
                'id': '4.1',
                'name': 'Bulk Historical Import Utility',
                'description': 'CLI tool for mass import of historical proposals',
                'activities': [
                    'Build CLI tool for directory scanning',
                    'Implement recursive folder traversal',
                    'Create file filtering (by extension, date, size)',
                    'Build batch upload (chunks of 50 files)',
                    'Implement progress tracking (X of Y uploaded)',
                    'Add dry-run mode (preview without uploading)',
                    'Create import summary report (files uploaded, skipped, failed)',
                    'Implement resume capability (continue interrupted import)',
                    'Build import validation (check files before upload)',
                    'Add import scheduling (throttle uploads to avoid overload)',
                    'Create import logging (detailed activity log)'
                ],
                'deliverable': 'Historical import CLI tool + Import_Tool_User_Guide.md',
                'dependencies': 'WP2 (Bulk Import)',
                'effort_days': 5,
                'resources': 'Developer',
                'risks': [
                    'Import overwhelms system (too many files at once)',
                    'Network failures interrupt large imports (need resume)',
                    'File filtering misses important files (wrong criteria)',
                    'Import tool doesn\'t handle edge cases (special characters in filenames)'
                ],
                'acceptance_criteria': [
                    'CLI scans directory and subdirectories',
                    'Dry-run mode shows what would be imported',
                    'Batch upload (50 files per API call) working',
                    'Resume capability tested (interrupt and restart)',
                    'Import report shows: uploaded, skipped (duplicates), failed',
                    'Tested: import 100+ historical files successfully'
                ]
            },
            {
                'id': '4.2',
                'name': 'Document Classification & Vendor Detection',
                'description': 'Automatically identify vendor and proposal type',
                'activities': [
                    'Build vendor detection algorithm (keyword matching, logo detection)',
                    'Implement format detection (Word, Excel, PDF, scanned image)',
                    'Create proposal type classification (pricing vs terms vs summary)',
                    'Build confidence scoring for classification',
                    'Implement manual classification override',
                    'Create classification training data collection',
                    'Build classification validation (manual review sample)',
                    'Implement classification reporting (accuracy metrics)',
                    'Add classification caching (store results)',
                    'Create classification API',
                    'Build classification testing framework'
                ],
                'deliverable': 'Classification engine + Vendor detection API + Classification_Guide.md',
                'dependencies': '4.1',
                'effort_days': 7,
                'resources': 'ML Engineer / Developer',
                'risks': [
                    'Vendor detection fails on older/non-standard formats',
                    'Scanned images too low quality (OCR fails)',
                    'Classification accuracy too low (<90%)',
                    'Vendor branding changes over time (detection breaks)'
                ],
                'acceptance_criteria': [
                    'Vendor detection ≥90% accuracy (tested on samples)',
                    'Format detection 100% accurate (Word, Excel, PDF, image)',
                    'Proposal type classification ≥85% accuracy',
                    'Confidence scores provided for all classifications',
                    'Manual override available for incorrect classifications',
                    'Tested: classify 50 historical proposals, validate results'
                ]
            },
            {
                'id': '4.3',
                'name': 'Legacy Format Handlers',
                'description': 'Support for older document formats (.doc, Excel 2003, scanned PDFs)',
                'activities': [
                    'Implement .doc (Word 97-2003) reader',
                    'Add Excel 2003 (.xls) support',
                    'Build scanned PDF OCR pipeline (Tesseract)',
                    'Implement image enhancement (deskew, denoise) for OCR',
                    'Create legacy format conversion (convert to modern format first)',
                    'Build format-specific extraction strategies',
                    'Implement legacy format testing',
                    'Create fallback extraction (hybrid OCR + AI)',
                    'Build legacy format validation',
                    'Add legacy format handling documentation',
                    'Implement legacy format performance optimization'
                ],
                'deliverable': 'Legacy format handlers + OCR pipeline + Legacy_Format_Guide.md',
                'dependencies': '4.2',
                'effort_days': 8,
                'resources': 'Developer + ML Engineer',
                'risks': [
                    'OCR accuracy too low on poor-quality scans (<80%)',
                    'Legacy formats too corrupted to read',
                    'OCR processing too slow (minutes per document)',
                    'Legacy format extraction accuracy significantly lower than modern'
                ],
                'acceptance_criteria': [
                    '.doc and .xls files processed successfully',
                    'OCR accuracy ≥85% on clean scans (tested with samples)',
                    'Image enhancement improves OCR accuracy by ≥10%',
                    'Legacy format processing time <2x modern formats',
                    'Fallback to hybrid extraction when OCR confidence <80%',
                    'Tested: process 20 legacy format proposals'
                ]
            },
            {
                'id': '4.4',
                'name': 'De-duplication & Historical Analysis',
                'description': 'Identify duplicate proposals and analyze historical pricing trends',
                'activities': [
                    'Implement file hash-based de-duplication',
                    'Build content similarity detection (fuzzy matching)',
                    'Create duplicate resolution workflow (keep which version)',
                    'Implement historical pricing database',
                    'Build pricing trend analysis (vendor X over time)',
                    'Create anomaly detection (unusually high/low pricing)',
                    'Implement benchmarking reports (compare to historical average)',
                    'Build historical data visualization dashboard',
                    'Create data quality metrics (completeness, accuracy over time)',
                    'Implement historical data export (for analysis)',
                    'Build historical insights API'
                ],
                'deliverable': 'De-duplication engine + Historical analytics + Historical_Analysis_Dashboard',
                'dependencies': '4.1, 4.2',
                'effort_days': 6,
                'resources': 'Data Engineer + Analyst',
                'risks': [
                    'De-duplication misses duplicates (different file names)',
                    'Similarity detection too sensitive (flags non-duplicates)',
                    'Historical data quality too poor for meaningful analysis',
                    'Pricing trends don\'t provide actionable insights'
                ],
                'acceptance_criteria': [
                    'Hash-based de-duplication catches exact duplicates',
                    'Similarity detection ≥90% accuracy (tested on sample duplicates)',
                    'Duplicate resolution workflow allows manual selection',
                    'Historical pricing database populated with all extractions',
                    'Trend analysis shows pricing changes over time (by vendor)',
                    'Anomaly detection flags outliers (>2 std dev from mean)'
                ]
            }
        ]
    },
    'WP5': {
        'name': 'Exception Management & Review Workflow',
        'tasks': [
            {
                'id': '5.1',
                'name': 'Review Queue Backend',
                'description': 'Build exception queue for manual review items',
                'activities': [
                    'Design review queue database schema (review items, comments, approvals)',
                    'Implement review queue API (get items, update, approve, reject)',
                    'Build review assignment logic (round-robin, skill-based)',
                    'Create review item detail view (show extraction + source)',
                    'Implement review item prioritization (by confidence, urgency)',
                    'Build review comment system (threaded discussions)',
                    'Implement review item locking (one user at a time)',
                    'Create review item history (all changes tracked)',
                    'Build review SLA tracking (time in queue)',
                    'Implement review queue metrics (throughput, aging)',
                    'Create review queue notifications (new items, SLA breach)'
                ],
                'deliverable': 'Review queue backend + API + Review_Queue_Schema.md',
                'dependencies': 'WP1 (Database), WP2 (Bucket Routing)',
                'effort_days': 7,
                'resources': 'Backend Developer',
                'risks': [
                    'Review queue fills faster than reviewers can process',
                    'Review assignment logic unfair (uneven distribution)',
                    'Review item locking causes conflicts (user loses work)',
                    'SLA tracking doesn\'t account for weekends/holidays'
                ],
                'acceptance_criteria': [
                    'Review queue stores items from Bucket 2 (70-89% confidence)',
                    'API supports: get queue, assign item, update, approve, reject',
                    'Review assignment distributes items evenly',
                    'Item locking prevents concurrent edits',
                    'Review history tracks all changes with timestamps',
                    'SLA alerts trigger when item >24 hours in queue'
                ]
            },
            {
                'id': '5.2',
                'name': 'Web-based Review UI',
                'description': 'Build intuitive review interface for non-technical users',
                'activities': [
                    'Design review UI wireframes and mockups',
                    'Implement review dashboard (queue overview, My Items)',
                    'Build review item detail page (side-by-side: extraction vs source)',
                    'Create inline editing (edit fields directly in UI)',
                    'Implement source document viewer (PDF, Word, Excel preview)',
                    'Build confidence indicator visualization (color-coded)',
                    'Create bulk actions (approve all, assign multiple)',
                    'Implement keyboard shortcuts (faster navigation)',
                    'Build review commenting UI (add notes, @mention)',
                    'Create review history timeline view',
                    'Implement review filters and search (by vendor, date, status)',
                    'Add review analytics (my throughput, avg review time)'
                ],
                'deliverable': 'Review web UI + User_Guide.pdf + UI_Mockups.pdf',
                'dependencies': '5.1',
                'effort_days': 12,
                'resources': 'Full Stack Developer + UX Designer',
                'risks': [
                    'UI too complex (requires extensive training)',
                    'Source document preview doesn\'t work for all formats',
                    'Inline editing buggy (data loss)',
                    'UI performance slow with large documents (>100 line items)'
                ],
                'acceptance_criteria': [
                    'Review dashboard shows: queue count, my items, completed today',
                    'Item detail shows extraction side-by-side with source',
                    'Inline editing allows field updates without page reload',
                    'Source preview works for: PDF, Word, Excel',
                    'Confidence indicators color-coded (red <70%, yellow 70-89%, green ≥90%)',
                    'Tested: non-technical user can review and approve item in <5 min'
                ]
            },
            {
                'id': '5.3',
                'name': 'Approval Workflow & State Machine',
                'description': 'Implement review state machine and approval routing',
                'activities': [
                    'Design approval state machine (draft, in_review, approved, rejected)',
                    'Implement state transitions with validation',
                    'Build multi-level approval (reviewer → manager)',
                    'Create approval delegation (assign to someone else)',
                    'Implement approval expiration (auto-escalate after timeout)',
                    'Build approval notification system (email/Slack)',
                    'Create approval audit trail (who approved when)',
                    'Implement conditional routing (high-value items → manager)',
                    'Build approval rollback (undo approval)',
                    'Create approval reporting (approval rates, bottlenecks)',
                    'Implement approval SLA tracking'
                ],
                'deliverable': 'Approval workflow engine + State machine + Approval_Workflow_Guide.md',
                'dependencies': '5.2',
                'effort_days': 6,
                'resources': 'Backend Developer + BA',
                'risks': [
                    'State machine too complex (hard to understand)',
                    'Approval bottlenecks (manager becomes blocker)',
                    'Notification fatigue (too many approval emails)',
                    'Approval delegation misused (bypass approvals)'
                ],
                'acceptance_criteria': [
                    'State machine enforces valid transitions only',
                    'Multi-level approval working (item routed to manager if >$1M)',
                    'Approval delegation tracked in audit trail',
                    'Auto-escalation after 48 hours in review',
                    'Notifications sent: on assignment, on approval, on rejection',
                    'Tested: approval workflow from submission to final approval'
                ]
            },
            {
                'id': '5.4',
                'name': 'Feedback Loop & Active Learning',
                'description': 'Capture manual corrections to improve extraction accuracy',
                'activities': [
                    'Implement correction tracking (original vs corrected values)',
                    'Build correction categorization (field type, error type)',
                    'Create correction analytics (most frequently corrected fields)',
                    'Implement prompt refinement based on corrections',
                    'Build vendor-specific learning (cache corrections per vendor)',
                    'Create correction-based re-training pipeline',
                    'Implement confidence threshold tuning (adjust based on accuracy)',
                    'Build correction suggestion system (learn patterns)',
                    'Create correction reporting dashboard (accuracy improvements)',
                    'Implement A/B testing framework (test prompt changes)',
                    'Build correction export (for external analysis)'
                ],
                'deliverable': 'Feedback loop system + Learning pipeline + Feedback_Loop_Guide.md',
                'dependencies': '5.3',
                'effort_days': 8,
                'resources': 'ML Engineer + Developer',
                'risks': [
                    'Correction data too sparse (not enough to learn from)',
                    'Learning pipeline doesn\'t improve accuracy (changes ineffective)',
                    'Prompt changes break existing accurate extractions',
                    'A/B testing framework too complex to use'
                ],
                'acceptance_criteria': [
                    'All corrections captured: field, original, corrected, reason',
                    'Correction analytics show: most common error types',
                    'Prompt refinement improves accuracy by ≥5% (tested on validation set)',
                    'Vendor-specific cache stores corrections (90-day TTL)',
                    'A/B testing framework allows comparing 2 prompt versions',
                    'Tested: 20 corrections → prompt update → re-extract → accuracy improvement'
                ]
            }
        ]
    },
    'WP6': {
        'name': 'Monitoring & Reporting',
        'tasks': [
            {
                'id': '6.1',
                'name': 'Real-time Status Dashboard',
                'description': 'Build operations dashboard for system health and job status',
                'activities': [
                    'Design dashboard wireframes (system health, jobs, queues)',
                    'Implement dashboard backend API (metrics aggregation)',
                    'Build real-time metrics display (WebSocket)',
                    'Create system health indicators (DB, queue, workers, API)',
                    'Implement job status overview (total, in progress, completed, failed)',
                    'Build queue depth visualization (jobs waiting)',
                    'Create worker status display (active, idle, crashed)',
                    'Implement recent activity feed (last 50 jobs)',
                    'Build error rate trending (last 24 hours)',
                    'Create throughput metrics (proposals per hour)',
                    'Implement dashboard refresh controls (auto-refresh, manual)',
                    'Add dashboard filters (by vendor, date range, status)'
                ],
                'deliverable': 'Operations dashboard + Metrics API + Dashboard_User_Guide.md',
                'dependencies': 'WP2 (Processing), WP1 (Logging)',
                'effort_days': 8,
                'resources': 'Full Stack Developer',
                'risks': [
                    'Dashboard performance slow with large datasets',
                    'Real-time updates don\'t scale (too many WebSocket connections)',
                    'Metrics aggregation queries slow down database',
                    'Dashboard shows stale data (caching issues)'
                ],
                'acceptance_criteria': [
                    'Dashboard shows: system health, job counts, queue depth, worker status',
                    'Real-time updates (metrics refresh every 5 seconds)',
                    'System health checks: DB (green/red), Queue (green/yellow/red), Workers (count)',
                    'Job status: total today, completed, failed, in progress',
                    'Error rate trend (line chart last 24 hours)',
                    'Tested: dashboard accurate during 20-proposal batch processing'
                ]
            },
            {
                'id': '6.2',
                'name': 'Notification & Alerting System',
                'description': 'Email and Slack notifications for critical events',
                'activities': [
                    'Implement email notification system (SMTP)',
                    'Build Slack integration (webhook or bot)',
                    'Create notification templates (job complete, job failed, SLA breach)',
                    'Implement notification preferences (user settings)',
                    'Build notification throttling (avoid alert fatigue)',
                    'Create notification channels (email, Slack, SMS)',
                    'Implement escalation rules (critical errors → manager)',
                    'Build notification scheduling (business hours only)',
                    'Create notification testing tool (send test alerts)',
                    'Implement notification audit log (all sent notifications)',
                    'Build notification digest (daily summary email)',
                    'Create notification customization (per user or team)'
                ],
                'deliverable': 'Notification system + Slack bot + Email templates + Notification_Guide.md',
                'dependencies': '6.1',
                'effort_days': 6,
                'resources': 'Backend Developer',
                'risks': [
                    'Email delivery issues (SMTP server down, spam filters)',
                    'Slack webhook rate limits (too many notifications)',
                    'Alert fatigue (too many notifications ignored)',
                    'Notification preferences too complex (users confused)'
                ],
                'acceptance_criteria': [
                    'Email notifications working (job complete, job failed)',
                    'Slack notifications posted to channel (with job details)',
                    'Notification throttling (max 1 alert per 5 min for same error)',
                    'User preferences allow: enable/disable by event type',
                    'Escalation working (critical error → manager email)',
                    'Tested: simulate job failure → email and Slack notification sent'
                ]
            },
            {
                'id': '6.3',
                'name': 'Analytics & Standard Reports',
                'description': 'Pre-built reports for common business questions',
                'activities': [
                    'Design report requirements (with client stakeholders)',
                    'Implement report data models (aggregations, summaries)',
                    'Build Extraction Accuracy Report (by vendor, over time)',
                    'Create Processing Volume Report (proposals per day/week/month)',
                    'Implement Review Queue Metrics Report (SLA compliance, throughput)',
                    'Build Cost Analysis Report (API costs, processing costs)',
                    'Create User Activity Report (who processed what)',
                    'Implement Vendor Comparison Report (accuracy by vendor)',
                    'Build Error Analysis Report (common failure patterns)',
                    'Create report scheduling (auto-generate weekly/monthly)',
                    'Implement report export (PDF, Excel, CSV)',
                    'Build report API (programmatic access)',
                    'Create custom report builder (drag-and-drop)'
                ],
                'deliverable': 'Report library + Report scheduler + Custom report builder + Reports_Catalog.pdf',
                'dependencies': '6.1, WP5 (Review)',
                'effort_days': 10,
                'resources': 'Data Engineer + Full Stack Developer',
                'risks': [
                    'Report requirements change during development',
                    'Report queries too slow (need optimization)',
                    'Custom report builder too limited (can\'t answer all questions)',
                    'Report scheduling overwhelms system (too many concurrent reports)'
                ],
                'acceptance_criteria': [
                    'Min 7 standard reports implemented (listed above)',
                    'Reports generated on-demand (< 30 seconds)',
                    'Report scheduling working (weekly email with PDF)',
                    'Report export to: PDF, Excel, CSV',
                    'Custom report builder allows: filter, group by, aggregate',
                    'Tested: generate all reports with real data, validate accuracy'
                ]
            },
            {
                'id': '6.4',
                'name': 'Audit Trail Enhancement',
                'description': 'Comprehensive audit logging for compliance and debugging',
                'activities': [
                    'Design audit trail schema (events, actors, timestamps, changes)',
                    'Implement audit logging for all critical operations',
                    'Create audit log API (query, filter, export)',
                    'Build audit log UI (searchable, filterable)',
                    'Implement data lineage tracking (proposal → extraction → TCO)',
                    'Create change tracking (who changed what when)',
                    'Build audit log retention policy (7 years)',
                    'Implement audit log export (for compliance)',
                    'Create audit log integrity verification (tamper-proof)',
                    'Build audit log analytics (user activity patterns)',
                    'Implement audit log search (full-text search)',
                    'Create audit log reporting (compliance reports)'
                ],
                'deliverable': 'Enhanced audit system + Audit UI + Audit_Log_Specification.md',
                'dependencies': 'WP1 (Logging), WP5 (Review)',
                'effort_days': 7,
                'resources': 'Backend Developer + Security Engineer',
                'risks': [
                    'Audit log volume excessive (storage costs)',
                    'Audit queries too slow (need indexing)',
                    'Audit log doesn\'t capture all required events',
                    'Audit log retention conflicts with storage limits'
                ],
                'acceptance_criteria': [
                    'All critical operations logged: upload, extract, review, approve, export',
                    'Audit log includes: timestamp, user, action, object, before/after values',
                    'Data lineage tracked (proposal ID → extraction ID → TCO file)',
                    'Audit UI allows search and filter (by user, date, action)',
                    'Audit log export to CSV (for compliance reporting)',
                    'Tested: complete audit trail from proposal upload to TCO output'
                ]
            }
        ]
    },
    'WP7': {
        'name': 'Testing & Quality Assurance',
        'tasks': [
            {
                'id': '7.1',
                'name': 'Unit Test Expansion (80% Coverage)',
                'description': 'Expand unit test coverage across all modules',
                'activities': [
                    'Audit current test coverage (generate coverage report)',
                    'Identify untested modules and functions',
                    'Write unit tests for extractors (fis, jh, pdf)',
                    'Write unit tests for mappers (schema normalization)',
                    'Write unit tests for writers (TCO Excel generation)',
                    'Write unit tests for QA validation',
                    'Write unit tests for API endpoints',
                    'Write unit tests for queue and workers',
                    'Implement test fixtures and mocks (sample proposals)',
                    'Create test data generators (synthetic proposals)',
                    'Set up test coverage enforcement (CI fails if <80%)',
                    'Implement mutation testing (test quality validation)'
                ],
                'deliverable': 'Expanded test suite + Coverage report + Test_Documentation.md',
                'dependencies': 'WP1-WP6 (All modules implemented)',
                'effort_days': 10,
                'resources': 'QA Engineer + Developers',
                'risks': [
                    'Test coverage metric achieved but tests low quality (not testing real scenarios)',
                    'Mocking too extensive (tests don\'t reflect real behavior)',
                    'Test maintenance burden (tests break on every change)',
                    'Test execution time too long (slow CI pipeline)'
                ],
                'acceptance_criteria': [
                    'Unit test coverage ≥80% (lines of code)',
                    'All critical paths covered (extraction, mapping, output)',
                    'Test execution time < 5 minutes (for unit tests)',
                    'All tests passing (0 failures)',
                    'Test documentation includes: how to run, how to add tests',
                    'Coverage report generated automatically in CI'
                ]
            },
            {
                'id': '7.2',
                'name': 'Integration Testing',
                'description': 'End-to-end integration tests across system components',
                'activities': [
                    'Design integration test scenarios (happy path, error paths)',
                    'Set up integration test environment (test DB, test queue)',
                    'Write E2E test: upload proposal → extract → output TCO',
                    'Test multi-vendor scenarios (FIS, JH, CSI in one batch)',
                    'Test review workflow (item goes to queue → reviewed → approved)',
                    'Test template selection (correct template chosen)',
                    'Test error handling (retry, dead letter queue)',
                    'Test API integrations (all endpoints)',
                    'Test database integrations (CRUD operations)',
                    'Test queue integrations (job submission → completion)',
                    'Implement integration test fixtures (sample data)',
                    'Create integration test documentation'
                ],
                'deliverable': 'Integration test suite + Test environment + Integration_Test_Plan.md',
                'dependencies': '7.1',
                'effort_days': 8,
                'resources': 'QA Engineer + Developer',
                'risks': [
                    'Integration test environment drift from production',
                    'Tests flaky (pass sometimes, fail others)',
                    'Test data setup complex and brittle',
                    'Integration tests too slow (minutes per test)'
                ],
                'acceptance_criteria': [
                    'Min 15 integration test scenarios implemented',
                    'E2E test (upload → extract → TCO) passing for 3 vendors',
                    'Review workflow tested end-to-end',
                    'Error handling scenarios tested (API timeout, corrupt file)',
                    'Integration tests run in isolated environment (no prod data)',
                    'All integration tests passing (0 failures)'
                ]
            },
            {
                'id': '7.3',
                'name': 'Volume & Performance Testing',
                'description': 'Stress test with 20+ proposals, measure performance',
                'activities': [
                    'Design volume test plan (20, 50, 100 proposals)',
                    'Create performance test dataset (varied complexity)',
                    'Set up performance test environment (production-like)',
                    'Execute 20-proposal stress test (baseline)',
                    'Measure processing times (avg, p95, p99)',
                    'Measure system resource usage (CPU, memory, disk)',
                    'Identify performance bottlenecks (profiling)',
                    'Optimize critical paths (database queries, API calls)',
                    'Re-test after optimizations',
                    'Test concurrent processing (4 workers, 8 workers)',
                    'Measure API throughput (requests per second)',
                    'Create performance baseline documentation',
                    'Test surge scenario (100 proposals submitted at once)'
                ],
                'deliverable': 'Performance test report + Optimization recommendations + Load_Test_Results.pdf',
                'dependencies': '7.2',
                'effort_days': 6,
                'resources': 'QA Engineer + Performance Engineer',
                'risks': [
                    'System performance inadequate under load',
                    'Bottlenecks not identified (poor profiling)',
                    'Optimizations don\'t improve performance significantly',
                    'Test environment doesn\'t match production (misleading results)'
                ],
                'acceptance_criteria': [
                    '20-proposal batch processed in <30 minutes (avg 90 sec/proposal)',
                    'System stable under load (no crashes, memory leaks)',
                    'Resource usage acceptable (CPU <80%, memory <16GB)',
                    'API response times <500ms (p95)',
                    'Concurrent processing scales (8 workers = 2x throughput vs 4)',
                    'Performance report documents: avg time, p95, p99, resource usage'
                ]
            },
            {
                'id': '7.4',
                'name': 'User Acceptance Testing (UAT)',
                'description': 'Client validates system meets requirements',
                'activities': [
                    'Create UAT test plan (scenarios, acceptance criteria)',
                    'Prepare UAT environment (production-like)',
                    'Create UAT test data (real proposals from client)',
                    'Conduct UAT training session (how to use system)',
                    'Execute UAT scenarios with client users',
                    'Document UAT findings (bugs, issues, enhancements)',
                    'Triage and fix UAT issues (P0, P1, P2)',
                    'Re-test fixed issues with client',
                    'Collect client feedback (usability, features)',
                    'Obtain UAT sign-off (formal acceptance)',
                    'Create UAT report (test results, issues, sign-off)',
                    'Transition to production (deployment plan)'
                ],
                'deliverable': 'UAT test plan + UAT report + UAT sign-off document + Go-live readiness checklist',
                'dependencies': '7.3',
                'effort_days': 10,
                'resources': 'PM + QA + Client Users (3-5)',
                'risks': [
                    'Client not available for UAT (delays)',
                    'UAT uncovers major issues (requires rework)',
                    'Client expectations mismatch (feature gaps)',
                    'UAT environment issues (not production-like)'
                ],
                'acceptance_criteria': [
                    'UAT test plan covers all major user scenarios (min 20 scenarios)',
                    'Client users trained (min 3 users)',
                    'All P0 and P1 UAT issues resolved',
                    'Client formal sign-off obtained (signed document)',
                    'UAT report documents: scenarios tested, pass/fail, issues, resolutions',
                    'Go-live readiness checklist 100% complete'
                ]
            }
        ]
    },
    'WP8': {
        'name': 'Documentation & Knowledge Transfer',
        'tasks': [
            {
                'id': '8.1',
                'name': 'Technical Documentation',
                'description': 'Architecture, API, and developer documentation',
                'activities': [
                    'Write architecture documentation (C4 diagrams, system overview)',
                    'Document database schema (ER diagrams, table descriptions)',
                    'Create API documentation (OpenAPI/Swagger, endpoint details)',
                    'Write developer setup guide (local dev environment)',
                    'Document code architecture (module overview, design patterns)',
                    'Create deployment guide (production setup)',
                    'Write configuration reference (all settings explained)',
                    'Document integration points (external systems)',
                    'Create troubleshooting guide (common issues, solutions)',
                    'Write security documentation (authentication, encryption, secrets)',
                    'Document testing framework (how to run, how to add tests)',
                    'Create code contribution guide (standards, PR process)'
                ],
                'deliverable': 'Technical_Documentation/ folder (Architecture, API, Developer guides) + ReadTheDocs site',
                'dependencies': 'WP7 (Testing complete)',
                'effort_days': 8,
                'resources': 'Tech Writer + Tech Lead',
                'risks': [
                    'Documentation becomes outdated quickly (code changes)',
                    'Documentation too technical (hard for new developers)',
                    'Documentation incomplete (missing critical info)',
                    'Documentation not maintained (no ownership)'
                ],
                'acceptance_criteria': [
                    'Architecture documentation includes: C4 diagrams, data flow, technology stack',
                    'API documentation auto-generated from code (Swagger UI)',
                    'Developer setup guide tested (new dev can set up in <2 hours)',
                    'Deployment guide tested (can deploy to new environment)',
                    'Troubleshooting guide covers top 10 common issues',
                    'Documentation hosted (ReadTheDocs or internal wiki)'
                ]
            },
            {
                'id': '8.2',
                'name': 'Operations Manual',
                'description': 'Daily operations procedures for support team',
                'activities': [
                    'Write operations overview (roles, responsibilities)',
                    'Create daily operations checklist (health checks)',
                    'Document batch processing procedures (how to submit, monitor)',
                    'Write review queue procedures (how to review, approve)',
                    'Create incident response procedures (errors, outages)',
                    'Document backup and recovery procedures',
                    'Write monitoring procedures (dashboard, alerts)',
                    'Create user management procedures (add, remove, permissions)',
                    'Document template management procedures (upload, configure)',
                    'Write report generation procedures (schedule, export)',
                    'Create escalation procedures (when to escalate)',
                    'Document common support tasks (FAQ for support team)'
                ],
                'deliverable': 'Operations_Manual.pdf + Runbook/ folder (step-by-step procedures)',
                'dependencies': '8.1',
                'effort_days': 6,
                'resources': 'Tech Writer + Operations Lead',
                'risks': [
                    'Operations manual too generic (not specific to this system)',
                    'Procedures not tested (don\'t work in practice)',
                    'Manual too long (support team won\'t read)',
                    'Manual not updated when system changes'
                ],
                'acceptance_criteria': [
                    'Operations manual covers: daily tasks, incident response, common issues',
                    'Runbook includes: step-by-step procedures with screenshots',
                    'Daily checklist tested (ops team can execute)',
                    'Incident response tested (simulate outage, follow procedures)',
                    'Escalation procedures define: when to escalate, who to contact',
                    'Support team trained on operations manual'
                ]
            },
            {
                'id': '8.3',
                'name': 'End-User Documentation',
                'description': 'User guides for business users',
                'activities': [
                    'Write user guide (getting started, core workflows)',
                    'Create video tutorials (screen recordings with narration)',
                    'Document batch upload workflow (with screenshots)',
                    'Write review workflow guide (how to review, approve)',
                    'Create template management guide (for template admins)',
                    'Document report generation (how to generate, schedule)',
                    'Write FAQ (common questions, troubleshooting)',
                    'Create quick reference guide (1-page cheat sheet)',
                    'Document keyboard shortcuts and tips',
                    'Write release notes (what\'s new)',
                    'Create context-sensitive help (tooltips, in-app help)',
                    'Build user documentation portal (searchable)'
                ],
                'deliverable': 'User_Guide.pdf + Video tutorials + FAQ + Help portal',
                'dependencies': '8.1',
                'effort_days': 7,
                'resources': 'Tech Writer + UX Designer',
                'risks': [
                    'User guide too technical (users confused)',
                    'Videos become outdated (UI changes)',
                    'Documentation doesn\'t match user mental models',
                    'Help portal not discoverable (users don\'t find it)'
                ],
                'acceptance_criteria': [
                    'User guide covers: upload, review, approve, reports (with screenshots)',
                    'Min 5 video tutorials (3-5 min each): upload, review, templates, reports, FAQ',
                    'FAQ answers top 20 user questions',
                    'Quick reference guide (1-page PDF)',
                    'Help portal searchable and linked from UI',
                    'User guide tested with non-technical users (feedback incorporated)'
                ]
            },
            {
                'id': '8.4',
                'name': 'Knowledge Transfer & Training',
                'description': 'Training sessions for client team',
                'activities': [
                    'Design training curriculum (admin, user, developer tracks)',
                    'Create training materials (slides, exercises, datasets)',
                    'Conduct administrator training (system config, user mgmt)',
                    'Conduct end-user training (upload, review, reports)',
                    'Conduct developer training (code overview, how to extend)',
                    'Create hands-on exercises (practice scenarios)',
                    'Conduct Q&A sessions (answer questions)',
                    'Record training sessions (for future reference)',
                    'Create training certification (knowledge check)',
                    'Provide post-training support (office hours)',
                    'Document training feedback (improvements)',
                    'Create training calendar (ongoing education)'
                ],
                'deliverable': 'Training materials + Recorded sessions + Training certification + Training_Report.pdf',
                'dependencies': '8.1, 8.2, 8.3',
                'effort_days': 6,
                'resources': 'Trainer + Tech Lead + PM',
                'risks': [
                    'Training too rushed (not enough time)',
                    'Client team not available (scheduling conflicts)',
                    'Training too technical (users overwhelmed)',
                    'Training doesn\'t stick (no reinforcement)'
                ],
                'acceptance_criteria': [
                    'Min 3 training tracks delivered: admin (4 hours), user (2 hours), developer (4 hours)',
                    'Training materials include: slides, hands-on exercises, sample data',
                    'All sessions recorded and accessible',
                    'Min 80% of trainees pass certification (knowledge check)',
                    'Post-training support provided (2 weeks office hours)',
                    'Training feedback collected and positive (avg ≥4/5)'
                ]
            }
        ]
    }
}

def format_list_for_excel(items):
    """Format list items with bullet points for Excel cell"""
    return '\n'.join([f'• {item}' for item in items])

def update_work_packages():
    """Update the Work Packages sheet with detailed information"""

    # Load workbook
    wb = openpyxl.load_workbook('Karishma Proposal Documents/Project_Plan_Phase2_1_9 SMEdits.xlsx')
    ws = wb['Work Packages']

    # Clear existing data (except header)
    ws.delete_rows(2, ws.max_row)

    # Define styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    cell_alignment = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Update headers
    headers = [
        ('A1', 'WP', 8),
        ('B1', 'Task ID', 10),
        ('C1', 'Task Name', 30),
        ('D1', 'Description', 40),
        ('E1', 'Detailed Activities', 60),
        ('F1', 'Deliverable', 40),
        ('G1', 'Dependencies', 15),
        ('H1', 'Effort (Days)', 12),
        ('I1', 'Resources', 20),
        ('J1', 'Key Risks', 50),
        ('K1', 'Acceptance Criteria', 60),
        ('L1', 'Status', 12)
    ]

    for cell_ref, header_text, width in headers:
        cell = ws[cell_ref]
        cell.value = header_text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

        # Set column width
        col_letter = cell_ref[0]
        ws.column_dimensions[col_letter].width = width

    # Populate work packages
    row = 2
    for wp_key in ['WP0', 'WP1', 'WP2', 'WP3', 'WP4', 'WP5', 'WP6', 'WP7', 'WP8']:
        wp = WORK_PACKAGES[wp_key]

        for task in wp['tasks']:
            # WP column
            ws.cell(row, 1, wp_key).alignment = cell_alignment
            ws.cell(row, 1).border = thin_border

            # Task ID
            ws.cell(row, 2, task['id']).alignment = cell_alignment
            ws.cell(row, 2).border = thin_border

            # Task Name
            ws.cell(row, 3, task['name']).alignment = cell_alignment
            ws.cell(row, 3).border = thin_border

            # Description
            ws.cell(row, 4, task['description']).alignment = cell_alignment
            ws.cell(row, 4).border = thin_border

            # Detailed Activities
            activities_text = format_list_for_excel(task['activities'])
            ws.cell(row, 5, activities_text).alignment = cell_alignment
            ws.cell(row, 5).border = thin_border

            # Deliverable
            ws.cell(row, 6, task['deliverable']).alignment = cell_alignment
            ws.cell(row, 6).border = thin_border

            # Dependencies
            ws.cell(row, 7, task['dependencies']).alignment = cell_alignment
            ws.cell(row, 7).border = thin_border

            # Effort (Days)
            ws.cell(row, 8, task['effort_days']).alignment = Alignment(horizontal='center', vertical='top')
            ws.cell(row, 8).border = thin_border

            # Resources
            ws.cell(row, 9, task['resources']).alignment = cell_alignment
            ws.cell(row, 9).border = thin_border

            # Key Risks
            risks_text = format_list_for_excel(task['risks'])
            ws.cell(row, 10, risks_text).alignment = cell_alignment
            ws.cell(row, 10).border = thin_border

            # Acceptance Criteria
            criteria_text = format_list_for_excel(task['acceptance_criteria'])
            ws.cell(row, 11, criteria_text).alignment = cell_alignment
            ws.cell(row, 11).border = thin_border

            # Status
            ws.cell(row, 12, 'Not Started').alignment = Alignment(horizontal='center', vertical='top')
            ws.cell(row, 12).border = thin_border

            # Set row height (auto-adjust based on content)
            ws.row_dimensions[row].height = max(60, len(task['activities']) * 12)

            row += 1

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Save workbook
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = f'Karishma Proposal Documents/Project_Plan_Phase2_DETAILED_{timestamp}.xlsx'
    wb.save(output_file)

    print(f'\n[SUCCESS] Work Packages updated successfully!')
    print(f'Output file: {output_file}')
    print(f'\nSummary:')
    print(f'   Total Work Packages: 9 (WP0-WP8)')
    print(f'   Total Tasks: {row - 2}')
    print(f'   New Columns Added: Detailed Activities, Effort (Days), Resources, Key Risks, Acceptance Criteria')
    print(f'\nKey Enhancements:')
    print(f'   - Granular activity breakdown for each task')
    print(f'   - Effort estimates in person-days')
    print(f'   - Resource requirements specified')
    print(f'   - Key risks and obstacles identified')
    print(f'   - Clear acceptance criteria for validation')

    # Calculate totals
    total_days = sum(task['effort_days'] for wp in WORK_PACKAGES.values() for task in wp['tasks'])
    print(f'\nTotal Effort: {total_days} person-days (~{total_days / 20:.1f} person-months)')

if __name__ == '__main__':
    update_work_packages()
