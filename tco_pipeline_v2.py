"""
TCO Pipeline V2 - Automated TCO Template Generator

Creates a new Excel workbook with the same structure as WORKBOOK2,
with hardcoded formulas for automatic calculations.
"""

import sys
import os
import json
from typing import Dict, List, Any, Optional
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from extractors import extract_fis_proposal
from dotenv import load_dotenv

load_dotenv()

# =============================================================================
# TCO TEMPLATE STRUCTURE CONFIGURATION
# =============================================================================

# Row ranges for FIS sections
FIS_ROWS = {
    'bundle_start': 7,
    'bundle_end': 13,  # Rows 7-13 for 7 years
    'non_bundle_start': 22,
    'non_bundle_end': 54,
    'third_party_start': 58,
    'third_party_end': 84,
    'impl_credits_start': 93,
    'impl_credits_end': 97,
    'optional_start': 122,
    'optional_end': 151,
}

# Column mapping for FIS
FIS_COLS = {
    'type': 'B',           # Monthly F, Monthly V, Annual, One-Time
    'proposal': 'C',       # Proposal value (quantity)
    'avg_monthly': 'D',    # Average monthly qty (formula)
    'qty_year_1': 'E',     # Year 1 quantity (formula)
    'qty_year_2': 'F',
    'qty_year_3': 'G',
    'qty_year_4': 'H',
    'qty_year_5': 'I',
    'qty_year_6': 'J',
    'qty_year_7': 'K',
    'solution': 'O',       # Solution name
    'category': 'P',       # Category
    'rate': 'Q',           # Per unit rate
    'monthly_cost': 'R',   # Year 1 monthly cost (formula)
    'cost_year_1': 'S',    # Year 1 cost (formula)
    'cost_year_2': 'T',
    'cost_year_3': 'U',
    'cost_year_4': 'V',
    'cost_year_5': 'W',
    'cost_year_6': 'X',
    'cost_year_7': 'Y',
    'cpi_total': 'AM',     # Total CPI
    'total_cost': 'AN',    # Total term cost (formula)
}


class TCOTemplateGenerator:
    """
    Generates a TCO Excel template with formulas matching WORKBOOK2 structure.
    """

    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Line Items"

        # Styles
        self.header_font = Font(bold=True, size=11)
        self.header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
        self.section_fill = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def setup_headers(self):
        """Set up column headers matching WORKBOOK2."""
        ws = self.ws

        # Row 1 - Vendor name
        ws['O1'] = 'FIS'
        ws['O1'].font = Font(bold=True, size=14)

        # Row 2 - Configuration
        ws['B2'] = 'Years->'
        ws['D2'] = 7

        # Row 3 - Growth rate
        ws['B3'] = 'Growth->'
        ws['D3'] = 0.2

        # Row 4 - Main headers
        headers_row4 = {
            'B': 'Type (Monthly/Annual)',
            'D': 'Average Monthly QTY',
            'E': 'Quantity Per Year with Growth',
            'O': 'FIS',
            'Q': 'Per Unit Rate',
        }
        for col, text in headers_row4.items():
            ws[f'{col}4'] = text
            ws[f'{col}4'].font = self.header_font
            ws[f'{col}4'].fill = self.header_fill

        # Row 5 - Sub-headers
        headers_row5 = {
            'C': f'Proposal\n({datetime.now().strftime("%m/%Y")})',
            'E': 'Year 1 Quantity',
            'F': 'Year 2 Quantity',
            'G': 'Year 3 Quantity',
            'H': 'Year 4 Quantity',
            'I': 'Year 5 Quantity',
            'J': 'Year 6 Quantity',
            'K': 'Year 7 Quantity',
            'O': 'Solution Name/Description',
            'P': 'Category',
            'R': 'Year 1 Monthly Cost',
            'S': 'Year 1 Cost',
            'T': 'Year 2 Cost',
            'U': 'Year 3 Cost',
            'V': 'Year 4 Cost',
            'W': 'Year 5 Cost',
            'X': 'Year 6 Cost',
            'Y': 'Year 7 Cost',
            'AM': 'CPI',
            'AN': 'Total Term Cost',
        }
        for col, text in headers_row5.items():
            ws[f'{col}5'] = text
            ws[f'{col}5'].font = self.header_font

        # Row 6 - Section header
        ws['B6'] = 'Recurring Monthly and Annual Fees'
        ws['B6'].font = Font(bold=True, size=12)
        ws['B6'].fill = self.section_fill

    def add_formula_row(self, row: int, fee_type: str = 'Monthly F'):
        """Add formulas for a data row."""
        ws = self.ws

        # Column D: Average Monthly QTY
        ws[f'D{row}'] = f'=IFERROR(AVERAGE(C{row}:C{row}),0)'

        # Column E: Year 1 Quantity (based on type)
        ws[f'E{row}'] = f'=IF($B{row}="Annual",1,IF($B{row}="Monthly F",12,IF($B{row}="Monthly V",D{row}*12,IF($B{row}="One-Time",1,D{row}*12))))'

        # Columns F-K: Years 2-7 Quantity (with growth for variable, same for fixed)
        for i, col in enumerate(['F', 'G', 'H', 'I', 'J', 'K'], start=2):
            prev_col = get_column_letter(ord(col) - 1 - ord('A') + 1)
            ws[f'{col}{row}'] = f'=IF($B{row}="Monthly V",{prev_col}{row}*(1+$D$3),{prev_col}{row})'

        # Column R: Year 1 Monthly Cost
        ws[f'R{row}'] = f'=S{row}/12'

        # Columns S-Y: Year 1-7 Costs
        for col, qty_col in [('S', 'E'), ('T', 'F'), ('U', 'G'), ('V', 'H'), ('W', 'I'), ('X', 'J'), ('Y', 'K')]:
            ws[f'{col}{row}'] = f'=IFERROR($Q{row}*{qty_col}{row},0)'

        # Column AM: Total CPI (placeholder - 0 for now)
        ws[f'AM{row}'] = 0

        # Column AN: Total Term Cost
        ws[f'AN{row}'] = f'=SUM(AM{row},S{row}:Y{row})'

    def add_section_header(self, row: int, title: str):
        """Add a section header row."""
        ws = self.ws
        ws[f'O{row}'] = title
        ws[f'O{row}'].font = Font(bold=True)
        ws[f'O{row}'].fill = self.section_fill

    def add_data_row(self, row: int, solution: str, category: str,
                     fee_type: str, rate: float, qty: float = 1):
        """Add a data row with values and formulas."""
        ws = self.ws

        # Set values
        ws[f'B{row}'] = fee_type
        ws[f'C{row}'] = qty
        ws[f'O{row}'] = solution
        ws[f'P{row}'] = category
        ws[f'Q{row}'] = rate

        # Add formulas
        self.add_formula_row(row, fee_type)

    def populate_fis_bundle(self, bundle_data: Dict):
        """
        Populate FIS bundle section (rows 7-13).

        Bundle data structure:
        {
            'Year 1': {'7_year': 15000, ...},
            'Year 2': {'7_year': 17500, ...},
            ...
        }
        """
        ws = self.ws
        term = '7_year'

        # Year 1-5: Use exact values from proposal
        bundle_values = {
            1: bundle_data.get('Year 1', {}).get(term, 15000),
            2: bundle_data.get('Year 2', {}).get(term, 17500),
            3: bundle_data.get('Year 3', {}).get(term, 22500),
            4: bundle_data.get('Year 4', {}).get(term, 28000),
            5: bundle_data.get('Year 5', {}).get(term, 35000),
        }

        # Years 6-7: Apply 6.5% CPI increase
        year_5_rate = bundle_values[5]
        bundle_values[6] = year_5_rate * 1.065
        bundle_values[7] = bundle_values[6] * 1.065

        for year_num in range(1, 8):
            row = 6 + year_num  # Rows 7-13
            rate = bundle_values[year_num]

            # Solution name
            if year_num <= 5:
                solution = f'Year {year_num} CORE PROCESSING (Bundle)'
            else:
                solution = f'Year {year_num} CORE PROCESSING (Bundle) - CPI Increase'

            # Set values
            ws[f'B{row}'] = 'Monthly F'
            ws[f'C{row}'] = 1
            ws[f'O{row}'] = solution
            ws[f'P{row}'] = 'HORIZON CORE ACCOUNT PROCESSING'
            ws[f'Q{row}'] = rate

            # Quantity formulas - only Year 1 active for each bundle year row
            # For bundle, each row represents one specific year
            ws[f'D{row}'] = f'=IFERROR(AVERAGE(C{row}:C{row}),0)'

            # Set quantities: only the corresponding year has 12, rest are 0
            for y, col in enumerate(['E', 'F', 'G', 'H', 'I', 'J', 'K'], start=1):
                if y == year_num:
                    ws[f'{col}{row}'] = 12
                else:
                    ws[f'{col}{row}'] = 0

            # Cost formulas
            ws[f'R{row}'] = f'=S{row}/12'
            for col, qty_col in [('S', 'E'), ('T', 'F'), ('U', 'G'), ('V', 'H'), ('W', 'I'), ('X', 'J'), ('Y', 'K')]:
                ws[f'{col}{row}'] = f'=IFERROR($Q{row}*{qty_col}{row},0)'

            ws[f'AM{row}'] = 0
            ws[f'AN{row}'] = f'=SUM(AM{row},S{row}:Y{row})'

    def populate_monthly_fees(self, monthly_fees: List[Dict], start_row: int,
                              include_third_party: bool = False):
        """
        Populate monthly fees section.

        Args:
            monthly_fees: List of fee dictionaries from extraction
            start_row: Starting row number
            include_third_party: If True, include only third-party items;
                                if False, exclude third-party items
        """
        ws = self.ws
        row = start_row

        for fee in monthly_fees:
            is_third_party = fee.get('third_party', False)

            # Filter based on third_party flag
            if include_third_party != is_third_party:
                continue

            monthly = fee.get('monthly_fee', 0)
            name = fee.get('solution_name', '').replace('\n', ' ').strip()

            # Skip items with 0 monthly fee
            if monthly <= 0:
                continue

            # Skip "Total" rows
            if name.lower() == 'total':
                continue

            # Determine category
            if is_third_party:
                category = 'Third-Party Required'
            else:
                category = 'Non-Bundle Required'

            # Add row
            self.add_data_row(row, name, category, 'Monthly F', monthly, 1)
            row += 1

        return row

    def populate_one_time_items(self, one_time_credits: Dict, start_row: int, term: str = '7_year'):
        """Populate one-time credits and fees."""
        ws = self.ws
        row = start_row

        for name, values in one_time_credits.items():
            if name.lower() == 'total':
                continue

            amount = values.get(term, 0)
            if amount == 0:
                continue

            # Determine category
            if amount < 0:
                category = 'Implementation Credit'
            else:
                category = 'Implementation Fee'

            # Set values
            ws[f'B{row}'] = 'One-Time'
            ws[f'C{row}'] = 1
            ws[f'O{row}'] = name
            ws[f'P{row}'] = category
            ws[f'Q{row}'] = amount

            # Formulas for one-time items
            ws[f'D{row}'] = f'=IFERROR(AVERAGE(C{row}:C{row}),0)'
            ws[f'E{row}'] = 1  # Only Year 1
            for col in ['F', 'G', 'H', 'I', 'J', 'K']:
                ws[f'{col}{row}'] = 0

            ws[f'R{row}'] = f'=S{row}/12'
            ws[f'S{row}'] = f'=IFERROR($Q{row}*E{row},0)'
            for col in ['T', 'U', 'V', 'W', 'X', 'Y']:
                ws[f'{col}{row}'] = 0

            ws[f'AM{row}'] = 0
            ws[f'AN{row}'] = f'=SUM(AM{row},S{row}:Y{row})'

            row += 1

        return row

    def create_summary_sheet(self):
        """Create Summary sheet with formulas referencing Line Items."""
        ws_summary = self.wb.create_sheet("Summary")

        # Headers
        ws_summary['B1'] = 'FIS'
        ws_summary['B1'].font = Font(bold=True, size=14)

        ws_summary['B2'] = 'Required Items'
        ws_summary['C2'] = 'Total 7 Year'
        ws_summary['D2'] = 'Monthly Avg.'
        for cell in ['B2', 'C2', 'D2']:
            ws_summary[cell].font = self.header_font
            ws_summary[cell].fill = self.header_fill

        # Row 3: Bundle FIS Products
        ws_summary['B3'] = 'Bundle FIS Products'
        ws_summary['C3'] = f"=SUM('Line Items'!AN7:AN13)"
        ws_summary['D3'] = '=C3/84'

        # Row 4: Non-Bundle FIS Required
        ws_summary['B4'] = 'Non-Bundle FIS Required Products'
        ws_summary['C4'] = f"=SUM('Line Items'!AN22:AN54)"
        ws_summary['D4'] = '=C4/84'

        # Row 5: Non-Bundle Third Parties
        ws_summary['B5'] = 'Non-Bundle Required Third Parties'
        ws_summary['C5'] = f"=SUM('Line Items'!AN58:AN84)"
        ws_summary['D5'] = '=C5/84'

        # Row 6: Subtotal
        ws_summary['B6'] = 'Sub-Total: Required/Included Products'
        ws_summary['C6'] = '=SUM(C3:C5)'
        ws_summary['D6'] = '=C6/84'
        ws_summary['B6'].font = Font(bold=True)

        # Row 8: Implementation Credits
        ws_summary['B8'] = 'Implementation Credits'
        ws_summary['C8'] = f"=SUM('Line Items'!AN93:AN97)"

        # Row 10: Total
        ws_summary['B10'] = 'TOTAL REQUIRED (minus credits)'
        ws_summary['C10'] = '=C6+C8'
        ws_summary['D10'] = '=C10/84'
        ws_summary['B10'].font = Font(bold=True)
        ws_summary['C10'].font = Font(bold=True)

        # Yearly breakdown
        ws_summary['B13'] = 'Yearly Breakdown'
        ws_summary['B13'].font = Font(bold=True, size=12)

        for i, year in enumerate(range(1, 8)):
            row = 14 + i
            col_letter = get_column_letter(18 + year)  # S, T, U, V, W, X, Y
            ws_summary[f'B{row}'] = f'Year {year}'
            ws_summary[f'C{row}'] = f"=SUM('Line Items'!{col_letter}7:{col_letter}97)"

    def save(self, output_path: str):
        """Save the workbook."""
        self.wb.save(output_path)
        return output_path


def run_tco_pipeline(fis_proposal_path: str, output_path: str, term: str = '7_year'):
    """
    Run the complete TCO pipeline.

    Args:
        fis_proposal_path: Path to FIS proposal document
        output_path: Path for output Excel file
        term: Contract term ('5_year', '7_year', '10_year')
    """
    print("=" * 70)
    print("TCO PIPELINE V2 - Automated Template Generator")
    print("=" * 70)

    # Step 1: Extract FIS data
    print("\n[1/4] Extracting FIS proposal data...")
    fis_data = extract_fis_proposal(fis_proposal_path)

    bundle = fis_data.get('bundle_pricing', {})
    monthly = fis_data.get('monthly_fees', [])
    one_time = fis_data.get('one_time_credits', {})

    print(f"  - Bundle years: {len(bundle)}")
    print(f"  - Monthly fees: {len(monthly)}")
    print(f"  - One-time items: {len(one_time)}")

    # Step 2: Create template
    print("\n[2/4] Creating TCO template...")
    generator = TCOTemplateGenerator()
    generator.setup_headers()

    # Step 3: Populate data
    print("\n[3/4] Populating template with extracted data...")

    # Bundle section (rows 7-13)
    print("  - Populating bundle section...")
    generator.populate_fis_bundle(bundle)

    # Non-bundle FIS items (rows 22+)
    print("  - Populating non-bundle FIS items...")
    generator.add_section_header(21, 'Non-Bundle FIS Required Products')
    next_row = generator.populate_monthly_fees(monthly, 22, include_third_party=False)

    # Third-party items (rows 58+)
    print("  - Populating third-party items...")
    generator.add_section_header(57, 'Non-Bundle Required Third Parties')
    next_row = generator.populate_monthly_fees(monthly, 58, include_third_party=True)

    # One-time items (rows 93+)
    print("  - Populating one-time items...")
    generator.add_section_header(92, 'Implementation Credits and Fees')
    generator.populate_one_time_items(one_time, 93, term)

    # Create Summary sheet
    print("  - Creating Summary sheet...")
    generator.create_summary_sheet()

    # Step 4: Save
    print("\n[4/4] Saving output...")
    result = generator.save(output_path)

    print("\n" + "=" * 70)
    print("SUCCESS!")
    print("=" * 70)
    print(f"Output file: {result}")

    # Display summary
    print("\n" + "-" * 50)
    print("EXTRACTION SUMMARY:")
    print("-" * 50)

    # Calculate totals
    bundle_total = 0
    for year_num in range(1, 8):
        if year_num <= 5:
            rate = bundle.get(f'Year {year_num}', {}).get(term, 0)
        elif year_num == 6:
            rate = bundle.get('Year 5', {}).get(term, 0) * 1.065
        else:
            rate = bundle.get('Year 5', {}).get(term, 0) * 1.065 * 1.065
        bundle_total += rate * 12

    fis_monthly = sum(f.get('monthly_fee', 0) for f in monthly if not f.get('third_party', False))
    tp_monthly = sum(f.get('monthly_fee', 0) for f in monthly if f.get('third_party', False))

    print(f"  Bundle Total (7-year):     ${bundle_total:>12,.0f}")
    print(f"  FIS Monthly Total:         ${fis_monthly:>12,.2f}/month")
    print(f"  Third-Party Monthly Total: ${tp_monthly:>12,.2f}/month")

    return result


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description='TCO Pipeline V2')
    parser.add_argument('--proposal', type=str,
                       default=os.path.join(os.path.dirname(__file__), 'proposal1.docx'),
                       help='Path to FIS proposal')
    parser.add_argument('--output', type=str,
                       default=os.path.join(os.path.dirname(__file__), 'TCO_Generated.xlsx'),
                       help='Output path')
    parser.add_argument('--term', type=str, default='7_year',
                       choices=['5_year', '7_year', '10_year'],
                       help='Contract term')

    args = parser.parse_args()

    run_tco_pipeline(args.proposal, args.output, args.term)
