"""
5-Month Compressed Project Plan - Milestones Sheet Update
Aggressive timeline with parallel work streams
Total Duration: 100 days / 20 weeks / 5 months
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# Compressed milestones with parallel execution strategy
MILESTONES_COMPRESSED = [
    {
        'id': 'M0',
        'name': 'Discovery & Requirements Complete',
        'phase': 'Discovery',
        'duration_days': 15,
        'work_packages': ['WP0'],
        'parallel_tracks': 1,
        'key_activities': [
            'Historical proposal inventory (focus on top 100 critical files)',
            'TCO template variant collection (3-5 primary templates)',
            'Vendor pattern analysis (FIS, JH, CSI - fast track)',
            'Volume & SLA requirements definition',
            'Technical architecture design (lean approach)',
            'Risk assessment (top 10 risks only)',
            'Discovery exit & go/no-go decision'
        ],
        'deliverables': [
            'DD1: Historical Proposal Inventory (top 100 files)',
            'DD2: TCO Template Catalog (3-5 templates)',
            'DD3: Vendor Pattern Analysis (lean report)',
            'DD4: Requirements Spec (SLAs, NFRs)',
            'DD5: Technical Architecture (simplified)',
            'DD6: Risk Register (top 10)',
            'DD7: Discovery Sign-off'
        ],
        'exit_criteria': [
            'Top 100 historical files cataloged',
            '3-5 TCO templates analyzed',
            'Extraction baseline ≥95% (sample tested)',
            'Architecture approved',
            'Go/no-go decision made'
        ],
        'dependencies': 'Project kickoff',
        'team_size': '5 (PM, Tech Lead, BA, Dev, Client SME)',
        'critical_path': True,
        'payment_percent': 10,
        'risk_level': 'High',
        'optimization': 'Compressed from 24 to 15 days via focus on critical items only'
    },
    {
        'id': 'M1',
        'name': 'Foundation Infrastructure + Batch Processing Operational',
        'phase': 'Development - Core',
        'duration_days': 25,
        'work_packages': ['WP1', 'WP2'],
        'parallel_tracks': 2,
        'key_activities': [
            'PARALLEL TRACK 1: Database + Queue (PostgreSQL + Celery)',
            'PARALLEL TRACK 2: API + Config + Logging (FastAPI)',
            'Integration: Queue → Workers → Processing',
            'Bulk import + parallel workers (4-8 workers)',
            'Error handling + retry + dead letter queue',
            'Progress tracking (real-time WebSocket)',
            'Integration testing + volume test (20 proposals)'
        ],
        'deliverables': [
            'D1: PostgreSQL + Celery/RabbitMQ infrastructure',
            'D2: FastAPI + Config + Logging system',
            'D3: Bulk import + worker pool',
            'D4: Error handling framework',
            'D5: Progress tracking API',
            'D6: 20-proposal batch test passed',
            'D7: Infrastructure documentation'
        ],
        'exit_criteria': [
            'Infrastructure operational (DB + Queue + API)',
            '20 proposals processed in <30 min',
            'Workers stable under load',
            'Real-time progress working',
            'Integration tests passing'
        ],
        'dependencies': 'M0',
        'team_size': '6 (2 Backend Dev, DevOps, DBA, QA, Tech Lead)',
        'critical_path': True,
        'payment_percent': 0,
        'risk_level': 'High',
        'optimization': 'Merged WP1+WP2, parallel dev tracks, compressed from 54 to 25 days'
    },
    {
        'id': 'M2',
        'name': 'Multi-Template + Historical Processing + Review Workflow Live',
        'phase': 'Development - Features',
        'duration_days': 30,
        'work_packages': ['WP3', 'WP4', 'WP5'],
        'parallel_tracks': 3,
        'key_activities': [
            'TRACK 1: Template registry + dynamic mapping + auto-detection',
            'TRACK 2: Historical import CLI + classification + OCR',
            'TRACK 3: Review queue + web UI + approval workflow',
            'Cross-track integration: Templates → Processing → Review',
            'Feedback loop implementation',
            'Multi-template testing (3+ templates)',
            'Historical batch import (100+ files)',
            'Review workflow E2E testing'
        ],
        'deliverables': [
            'D8: Template system (registry + mapping + detection)',
            'D9: Template management UI',
            'D10: Historical import CLI + OCR pipeline',
            'D11: 100+ historical files processed',
            'D12: Review queue + web UI',
            'D13: Approval workflow + feedback loop',
            'D14: Integration test suite (3 workflows)'
        ],
        'exit_criteria': [
            '3+ templates configured and working',
            'Template detection ≥95% accurate',
            '100+ historical proposals imported',
            'Review UI tested by users',
            'Feedback loop capturing corrections',
            'All 3 parallel tracks integrated'
        ],
        'dependencies': 'M1',
        'team_size': '9 (3 Full Stack Dev, ML Eng, UX, 2 Backend, QA, BA)',
        'critical_path': True,
        'payment_percent': 25,
        'risk_level': 'High',
        'optimization': 'Merged WP3+WP4+WP5 into 3 parallel tracks, compressed from 84 to 30 days'
    },
    {
        'id': 'M3',
        'name': 'Monitoring & Reporting Operational',
        'phase': 'Development - Operations',
        'duration_days': 15,
        'work_packages': ['WP6'],
        'parallel_tracks': 1,
        'key_activities': [
            'Operations dashboard (real-time metrics)',
            'Email + Slack notifications',
            'Standard reports (5 essential reports only)',
            'Enhanced audit trail',
            'Dashboard testing',
            'Stakeholder training'
        ],
        'deliverables': [
            'D15: Operations dashboard',
            'D16: Notification system (Email + Slack)',
            'D17: Report library (5 reports)',
            'D18: Audit trail system',
            'D19: Dashboard training materials'
        ],
        'exit_criteria': [
            'Dashboard showing real-time metrics',
            'Notifications working (job events)',
            '5 core reports generating',
            'Audit trail logging all operations',
            'Stakeholders trained'
        ],
        'dependencies': 'M1 (can start after infrastructure)',
        'team_size': '3 (Full Stack Dev, Data Eng, DevOps)',
        'critical_path': False,
        'payment_percent': 0,
        'risk_level': 'Low',
        'optimization': 'Runs parallel with M2, reduced reports from 7 to 5, compressed from 31 to 15 days'
    },
    {
        'id': 'M4',
        'name': 'Testing Complete & UAT Sign-off',
        'phase': 'Quality Assurance',
        'duration_days': 20,
        'work_packages': ['WP7'],
        'parallel_tracks': 3,
        'key_activities': [
            'TRACK 1: Unit tests expansion (≥80% coverage)',
            'TRACK 2: Integration + performance testing',
            'TRACK 3: UAT preparation + execution',
            'Bug fixing (parallel with testing)',
            'Performance optimization',
            'UAT scenarios (15 critical scenarios)',
            'UAT sign-off'
        ],
        'deliverables': [
            'D20: Unit test suite (≥80% coverage)',
            'D21: Integration test suite',
            'D22: Performance test report (20-proposal batch)',
            'D23: UAT test plan (15 scenarios)',
            'D24: Bug fix report (P0/P1 resolved)',
            'D25: UAT sign-off document'
        ],
        'exit_criteria': [
            'Unit test coverage ≥80%',
            '20 proposals in <30 min',
            'UAT scenarios passed (100%)',
            'All P0/P1 bugs resolved',
            'Client UAT sign-off obtained'
        ],
        'dependencies': 'M2, M3',
        'team_size': '6 (2 QA, Performance Eng, 2 Dev, Client Users)',
        'critical_path': True,
        'payment_percent': 25,
        'risk_level': 'High',
        'optimization': 'Parallel testing tracks, reduced UAT scenarios from 20 to 15, compressed from 34 to 20 days'
    },
    {
        'id': 'M5',
        'name': 'Documentation, Training & Go-Live',
        'phase': 'Deployment',
        'duration_days': 15,
        'work_packages': ['WP8'],
        'parallel_tracks': 2,
        'key_activities': [
            'TRACK 1: Technical docs + Operations manual',
            'TRACK 2: User guides + video tutorials (3 videos)',
            'Training sessions (2 tracks: Admin + User)',
            'Production deployment',
            'Go-live execution',
            'Post-go-live stabilization (1 week)'
        ],
        'deliverables': [
            'D26: Technical Documentation',
            'D27: Operations Manual',
            'D28: User Guide + 3 Video Tutorials',
            'D29: Training materials (2 tracks)',
            'D30: Training certificates',
            'D31: Go-live sign-off'
        ],
        'exit_criteria': [
            'All documentation complete',
            '≥70% users trained',
            'Production system live',
            'Go-live successful (no rollback)',
            'Client final acceptance signed'
        ],
        'dependencies': 'M4',
        'team_size': '5 (Tech Writer, Trainer, Tech Lead, PM, DevOps)',
        'critical_path': True,
        'payment_percent': 25,
        'risk_level': 'Medium',
        'optimization': 'Parallel doc tracks, reduced training from 3 to 2 tracks, compressed from 27 to 15 days'
    }
]

# Compressed phase summary
PHASES_COMPRESSED = {
    'Discovery': {
        'color': '4472C4',
        'description': 'Fast-track requirements and planning (3 weeks)',
        'milestones': ['M0']
    },
    'Development - Core': {
        'color': '70AD47',
        'description': 'Foundation + batch processing (5 weeks)',
        'milestones': ['M1']
    },
    'Development - Features': {
        'color': 'FFC000',
        'description': '3 parallel dev tracks (6 weeks)',
        'milestones': ['M2']
    },
    'Development - Operations': {
        'color': 'ED7D31',
        'description': 'Monitoring (parallel, 3 weeks)',
        'milestones': ['M3']
    },
    'Quality Assurance': {
        'color': '5B9BD5',
        'description': 'Parallel testing + UAT (4 weeks)',
        'milestones': ['M4']
    },
    'Deployment': {
        'color': 'A5A5A5',
        'description': 'Documentation + go-live (3 weeks)',
        'milestones': ['M5']
    }
}

def format_list_for_excel(items):
    """Format list items with bullet points for Excel cell"""
    return '\n'.join([f'• {item}' for item in items])

def calculate_compressed_timeline():
    """Calculate compressed timeline with parallel execution"""
    timeline = []
    current_week = 1

    # M0: Discovery (sequential)
    m0 = MILESTONES_COMPRESSED[0]
    m0_weeks = (m0['duration_days'] + 4) // 5
    timeline.append({
        'milestone_id': 'M0',
        'start_week': current_week,
        'end_week': current_week + m0_weeks - 1,
        'duration_weeks': m0_weeks,
        'duration_days': m0['duration_days']
    })
    current_week += m0_weeks

    # M1: Core (sequential after M0)
    m1 = MILESTONES_COMPRESSED[1]
    m1_weeks = (m1['duration_days'] + 4) // 5
    timeline.append({
        'milestone_id': 'M1',
        'start_week': current_week,
        'end_week': current_week + m1_weeks - 1,
        'duration_weeks': m1_weeks,
        'duration_days': m1['duration_days']
    })
    m1_end_week = current_week + m1_weeks - 1
    current_week += m1_weeks

    # M2 and M3 run in parallel (M3 starts after M1, runs during M2)
    m2 = MILESTONES_COMPRESSED[2]
    m3 = MILESTONES_COMPRESSED[3]

    m2_weeks = (m2['duration_days'] + 4) // 5
    m3_weeks = (m3['duration_days'] + 4) // 5

    # M2 starts after M1
    timeline.append({
        'milestone_id': 'M2',
        'start_week': current_week,
        'end_week': current_week + m2_weeks - 1,
        'duration_weeks': m2_weeks,
        'duration_days': m2['duration_days']
    })

    # M3 starts after M1 (parallel with M2)
    m3_start = m1_end_week + 1
    timeline.append({
        'milestone_id': 'M3',
        'start_week': m3_start,
        'end_week': m3_start + m3_weeks - 1,
        'duration_weeks': m3_weeks,
        'duration_days': m3['duration_days']
    })

    # Next milestone starts after M2 completes
    current_week = current_week + m2_weeks

    # M4: Testing (after M2 and M3)
    m4 = MILESTONES_COMPRESSED[4]
    m4_weeks = (m4['duration_days'] + 4) // 5
    timeline.append({
        'milestone_id': 'M4',
        'start_week': current_week,
        'end_week': current_week + m4_weeks - 1,
        'duration_weeks': m4_weeks,
        'duration_days': m4['duration_days']
    })
    current_week += m4_weeks

    # M5: Deployment (after M4)
    m5 = MILESTONES_COMPRESSED[5]
    m5_weeks = (m5['duration_days'] + 4) // 5
    timeline.append({
        'milestone_id': 'M5',
        'start_week': current_week,
        'end_week': current_week + m5_weeks - 1,
        'duration_weeks': m5_weeks,
        'duration_days': m5['duration_days']
    })

    return timeline

def update_milestones_compressed():
    """Update milestones sheet with 5-month compressed timeline"""

    # Load the specific workbook
    wb = openpyxl.load_workbook('Karishma Proposal Documents/Project_Plan_Phase2_DETAILED_20260109_130815.xlsx')
    ws = wb['Milestones']

    # Clear existing data
    ws.delete_rows(1, ws.max_row)

    # Define styles
    title_font = Font(bold=True, size=16, color='FFFFFF')
    title_fill = PatternFill(start_color='203864', end_color='203864', fill_type='solid')

    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    phase_font = Font(bold=True, size=12, color='FFFFFF')
    phase_alignment = Alignment(horizontal='left', vertical='center')

    cell_alignment = Alignment(vertical='top', wrap_text=True)
    center_alignment = Alignment(horizontal='center', vertical='center')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title row
    ws.merge_cells('A1:O1')
    title_cell = ws['A1']
    title_cell.value = 'TCO AUTOMATION SYSTEM - PHASE 2 PROJECT PLAN (5-MONTH AGGRESSIVE TIMELINE)'
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Project summary row
    ws.merge_cells('A2:O2')
    summary_cell = ws['A2']
    timeline = calculate_compressed_timeline()
    total_weeks = max(t['end_week'] for t in timeline)
    total_days = sum(m['duration_days'] for m in MILESTONES_COMPRESSED)

    summary_cell.value = f'COMPRESSED TIMELINE: {total_weeks} weeks (~{total_days} days / 5 months) | 6 Milestones | 40 Tasks | 9 Parallel Work Streams | Estimated: 259 person-days'
    summary_cell.font = Font(bold=True, size=10, color='C00000')
    summary_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20

    # Headers
    headers = [
        ('A3', 'Milestone', 12),
        ('B3', 'Name', 35),
        ('C3', 'Phase', 20),
        ('D3', 'Timeline\n(Weeks)', 12),
        ('E3', 'Duration\n(Days)', 10),
        ('F3', 'Parallel\nTracks', 10),
        ('G3', 'Key Activities', 55),
        ('H3', 'Deliverables', 45),
        ('I3', 'Exit Criteria', 45),
        ('J3', 'Dependencies', 20),
        ('K3', 'Team Size', 15),
        ('L3', 'Critical\nPath', 10),
        ('M3', 'Risk\nLevel', 10),
        ('N3', 'Payment\n%', 10),
        ('O3', 'Status', 12)
    ]

    for cell_ref, header_text, width in headers:
        cell = ws[cell_ref]
        cell.value = header_text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

        col_letter = cell_ref[0]
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[3].height = 35

    # Populate milestones
    row = 4
    current_phase = None

    for idx, milestone in enumerate(MILESTONES_COMPRESSED):
        # Add phase header if phase changes
        if milestone['phase'] != current_phase:
            current_phase = milestone['phase']
            phase_info = PHASES_COMPRESSED[current_phase]

            ws.merge_cells(f'A{row}:O{row}')
            phase_cell = ws[f'A{row}']
            phase_cell.value = f'{current_phase.upper()}: {phase_info["description"]}'
            phase_cell.font = phase_font
            phase_cell.fill = PatternFill(start_color=phase_info['color'], end_color=phase_info['color'], fill_type='solid')
            phase_cell.alignment = phase_alignment
            ws.row_dimensions[row].height = 25
            row += 1

        # Get timeline info
        time_info = next(t for t in timeline if t['milestone_id'] == milestone['id'])

        # Milestone ID
        ws.cell(row, 1, milestone['id']).alignment = center_alignment
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 1).border = thin_border

        # Name
        ws.cell(row, 2, milestone['name']).alignment = cell_alignment
        ws.cell(row, 2).font = Font(bold=True)
        ws.cell(row, 2).border = thin_border

        # Phase
        ws.cell(row, 3, milestone['phase']).alignment = cell_alignment
        ws.cell(row, 3).border = thin_border

        # Timeline
        timeline_text = f"W{time_info['start_week']}-W{time_info['end_week']}\n({time_info['duration_weeks']} weeks)"
        ws.cell(row, 4, timeline_text).alignment = center_alignment
        ws.cell(row, 4).border = thin_border

        # Duration (Days)
        ws.cell(row, 5, milestone['duration_days']).alignment = center_alignment
        ws.cell(row, 5).border = thin_border

        # Parallel Tracks
        ws.cell(row, 6, milestone['parallel_tracks']).alignment = center_alignment
        ws.cell(row, 6).font = Font(bold=True, color='C00000')
        ws.cell(row, 6).border = thin_border

        # Key Activities
        activities_text = format_list_for_excel(milestone['key_activities'])
        ws.cell(row, 7, activities_text).alignment = cell_alignment
        ws.cell(row, 7).border = thin_border

        # Deliverables
        deliverables_text = format_list_for_excel(milestone['deliverables'])
        ws.cell(row, 8, deliverables_text).alignment = cell_alignment
        ws.cell(row, 8).border = thin_border

        # Exit Criteria
        criteria_text = format_list_for_excel(milestone['exit_criteria'])
        ws.cell(row, 9, criteria_text).alignment = cell_alignment
        ws.cell(row, 9).border = thin_border

        # Dependencies
        ws.cell(row, 10, milestone['dependencies']).alignment = cell_alignment
        ws.cell(row, 10).border = thin_border

        # Team Size
        ws.cell(row, 11, milestone['team_size']).alignment = cell_alignment
        ws.cell(row, 11).border = thin_border

        # Critical Path
        critical_text = 'YES' if milestone['critical_path'] else 'No'
        ws.cell(row, 12, critical_text).alignment = center_alignment
        if milestone['critical_path']:
            ws.cell(row, 12).font = Font(bold=True, color='C00000')
        ws.cell(row, 12).border = thin_border

        # Risk Level
        risk_colors = {'Low': '00B050', 'Medium': 'FFC000', 'High': 'C00000'}
        ws.cell(row, 13, milestone['risk_level']).alignment = center_alignment
        ws.cell(row, 13).font = Font(bold=True, color=risk_colors.get(milestone['risk_level'], '000000'))
        ws.cell(row, 13).border = thin_border

        # Payment %
        payment_text = f"{milestone['payment_percent']}%" if milestone['payment_percent'] > 0 else '-'
        ws.cell(row, 14, payment_text).alignment = center_alignment
        if milestone['payment_percent'] > 0:
            ws.cell(row, 14).font = Font(bold=True)
        ws.cell(row, 14).border = thin_border

        # Status
        ws.cell(row, 15, 'Not Started').alignment = center_alignment
        ws.cell(row, 15).border = thin_border

        # Set row height
        ws.row_dimensions[row].height = max(80, len(milestone['key_activities']) * 12)

        row += 1

    # Add summary section
    row += 1
    ws.merge_cells(f'A{row}:O{row}')
    summary_header = ws[f'A{row}']
    summary_header.value = 'COMPRESSED TIMELINE SUMMARY & OPTIMIZATION STRATEGY'
    summary_header.font = Font(bold=True, size=12, color='FFFFFF')
    summary_header.fill = PatternFill(start_color='203864', end_color='203864', fill_type='solid')
    summary_header.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 25

    row += 1

    # Summary metrics
    critical_path_milestones = [m['id'] for m in MILESTONES_COMPRESSED if m['critical_path']]
    total_payment = sum(m['payment_percent'] for m in MILESTONES_COMPRESSED)
    total_parallel_tracks = sum(m['parallel_tracks'] for m in MILESTONES_COMPRESSED)

    summary_data = [
        ['Original Timeline:', '54 weeks (254 days)', 'Compressed Timeline:', f'{total_weeks} weeks ({total_days} days)'],
        ['Time Reduction:', f'{254 - total_days} days saved ({int((254-total_days)/254*100)}% reduction)', 'Parallel Work Streams:', f'{total_parallel_tracks} concurrent tracks'],
        ['Total Milestones:', '6 (merged from 9)', 'Critical Path Milestones:', ', '.join(critical_path_milestones)],
        ['Payment Milestones:', '4', 'Total Payment Points:', f'{total_payment}%'],
        ['Peak Team Size:', '9 resources (M2 phase)', 'Avg Team Size:', '5-6 resources'],
        ['Optimization Strategy:', 'Parallel execution + lean discovery + merged milestones + aggressive timelines', '', '']
    ]

    for data_row in summary_data:
        ws.cell(row, 1, data_row[0]).font = Font(bold=True)
        ws.cell(row, 2, data_row[1])
        ws.cell(row, 3, data_row[2]).font = Font(bold=True)
        ws.merge_cells(f'D{row}:F{row}')
        ws.cell(row, 4, data_row[3])
        row += 1

    row += 1
    ws.merge_cells(f'A{row}:O{row}')
    risk_header = ws[f'A{row}']
    risk_header.value = 'KEY RISKS & MITIGATION FOR COMPRESSED TIMELINE'
    risk_header.font = Font(bold=True, size=11, color='FFFFFF')
    risk_header.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    risk_header.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 20

    row += 1

    risks_data = [
        ['Risk:', 'Aggressive timeline may compromise quality', 'Mitigation:', 'Automated testing (80% coverage), continuous integration, daily standups'],
        ['Risk:', 'Parallel work requires excellent coordination', 'Mitigation:', 'Daily integration meetings, clear interface contracts, dedicated integration lead'],
        ['Risk:', 'Reduced discovery may miss requirements', 'Mitigation:', 'Focus on top 100 critical files, weekly client validation, iterative refinement'],
        ['Risk:', 'Team burnout from compressed schedule', 'Mitigation:', 'Realistic daily targets, no weekend work, rotate on-call duties'],
        ['Risk:', 'Testing compressed from 34 to 20 days', 'Mitigation:', 'Parallel test tracks, automated regression, UAT focus on 15 critical scenarios']
    ]

    for risk_row in risks_data:
        ws.cell(row, 1, risk_row[0]).font = Font(bold=True, color='C00000')
        ws.merge_cells(f'B{row}:C{row}')
        ws.cell(row, 2, risk_row[1])
        ws.cell(row, 4, risk_row[2]).font = Font(bold=True, color='00B050')
        ws.merge_cells(f'E{row}:F{row}')
        ws.cell(row, 5, risk_row[3])
        row += 1

    # Freeze panes
    ws.freeze_panes = 'A4'

    # Save workbook
    wb.save('Karishma Proposal Documents/Project_Plan_Phase2_DETAILED_20260109_130815.xlsx')

    print('\n[SUCCESS] Milestones sheet updated with 5-MONTH COMPRESSED timeline!')
    print(f'\nCompressed Timeline Summary:')
    print(f'  Original: 54 weeks (254 days)')
    print(f'  Compressed: {total_weeks} weeks ({total_days} days)')
    print(f'  Time Saved: {254 - total_days} days ({int((254-total_days)/254*100)}% reduction)')
    print(f'  Target: 5 months / 20 weeks [ACHIEVED]')
    print(f'\nMilestones:')
    for m in MILESTONES_COMPRESSED:
        t = next(t for t in timeline if t['milestone_id'] == m['id'])
        print(f'  {m["id"]}: W{t["start_week"]}-W{t["end_week"]} ({m["duration_days"]} days) - {m["name"]}')
    print(f'\nParallel Execution:')
    print(f'  Total Parallel Tracks: {total_parallel_tracks}')
    print(f'  M1: {MILESTONES_COMPRESSED[1]["parallel_tracks"]} tracks (Infrastructure)')
    print(f'  M2: {MILESTONES_COMPRESSED[2]["parallel_tracks"]} tracks (Features)')
    print(f'  M3: Runs parallel with M2 (Monitoring)')
    print(f'  M4: {MILESTONES_COMPRESSED[4]["parallel_tracks"]} tracks (Testing)')

if __name__ == '__main__':
    update_milestones_compressed()
