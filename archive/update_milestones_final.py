"""
Industry-Grade 5-Month Project Plan - Milestones Sheet
Professional project plan with 9 milestones over 21 weeks
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# Define 9 comprehensive milestones for 5-month timeline
MILESTONES = [
    {
        'id': 'M0',
        'name': 'Discovery & Requirements Complete',
        'phase': 'Discovery',
        'duration_days': 15,
        'work_packages': ['WP0'],
        'key_activities': [
            'Historical proposal inventory and cataloging',
            'TCO template variant collection and analysis',
            'Vendor document pattern analysis (FIS, JH, CSI)',
            'Volume and performance requirements definition',
            'Technical architecture design and approval',
            'Risk assessment and mitigation planning',
            'Discovery exit review and stakeholder sign-off'
        ],
        'deliverables': [
            'DD1: Historical Proposal Inventory with metadata',
            'DD2: TCO Template Catalog and mapping specifications',
            'DD3: Vendor Pattern Analysis Report',
            'DD4: Requirements Specification (SLAs, NFRs)',
            'DD5: Technical Architecture Document with C4 diagrams',
            'DD6: Risk Register with mitigation plans',
            'DD7: Discovery Exit Report and client sign-off'
        ],
        'exit_criteria': [
            'All discovery questions answered',
            'Historical inventory complete with categorization',
            'TCO template variants cataloged (min 3-5)',
            'Extraction accuracy baseline established (≥95%)',
            'Technical architecture approved by client',
            'Risk register reviewed and accepted',
            'Client stakeholder sign-off obtained'
        ],
        'dependencies': 'Project kickoff',
        'team_size': '5 (PM, Tech Lead, BA, Developer, Client SME)',
        'critical_path': True,
        'payment_percent': 10,
        'risk_level': 'Medium'
    },
    {
        'id': 'M1',
        'name': 'Foundation Infrastructure Complete',
        'phase': 'Development - Core',
        'duration_days': 12,
        'work_packages': ['WP1'],
        'key_activities': [
            'PostgreSQL database design and implementation',
            'Celery + RabbitMQ job queue system setup',
            'Configuration management and secrets vault',
            'Structured logging infrastructure implementation',
            'FastAPI REST framework with authentication',
            'Infrastructure integration testing',
            'Documentation and setup guides'
        ],
        'deliverables': [
            'D1: PostgreSQL database with ORM models and migrations',
            'D2: Celery + RabbitMQ queue system',
            'D3: Configuration management system',
            'D4: Structured logging framework',
            'D5: FastAPI application with Swagger docs',
            'D6: Infrastructure test suite',
            'D7: Infrastructure documentation'
        ],
        'exit_criteria': [
            'Database operational with all tables and relationships',
            'Queue system processing test jobs',
            'API endpoints accessible with authentication',
            'Logging aggregating in centralized system',
            'Infrastructure tests passing (≥80% coverage)',
            'Security audit completed (no critical issues)',
            'Team trained on infrastructure components'
        ],
        'dependencies': 'M0',
        'team_size': '4 (Backend Dev, DevOps, DBA, Tech Lead)',
        'critical_path': True,
        'payment_percent': 0,
        'risk_level': 'Medium'
    },
    {
        'id': 'M2',
        'name': 'Batch Processing Operational',
        'phase': 'Development - Core',
        'duration_days': 13,
        'work_packages': ['WP2'],
        'key_activities': [
            'Bulk import and queue management implementation',
            'Parallel worker pool configuration (4-8 workers)',
            'Error handling and retry logic with dead letter queue',
            'Real-time progress tracking (WebSocket)',
            'Integration with extraction pipeline',
            'Volume testing (20+ proposals)',
            'Performance optimization and tuning'
        ],
        'deliverables': [
            'D8: Bulk import system with directory watcher',
            'D9: Worker pool infrastructure',
            'D10: Error handling framework',
            'D11: Progress tracking API and dashboard',
            'D12: Batch processing integration tests',
            'D13: Volume test report (20-proposal baseline)',
            'D14: Batch processing user guide'
        ],
        'exit_criteria': [
            '20 proposals processed successfully in batch',
            'Average processing time ≤90 seconds per proposal',
            'Workers stable under load (no crashes)',
            'Error handling validated (retry + dead letter queue)',
            'Progress tracking updating in real-time',
            'Integration tests passing (E2E workflows)',
            'Performance SLAs met'
        ],
        'dependencies': 'M1',
        'team_size': '3 (Backend Dev, QA, DevOps)',
        'critical_path': True,
        'payment_percent': 0,
        'risk_level': 'Medium'
    },
    {
        'id': 'M3',
        'name': 'Multi-Template Support Enabled',
        'phase': 'Development - Features',
        'duration_days': 15,
        'work_packages': ['WP3'],
        'key_activities': [
            'Template registry and versioning system',
            'Dynamic field mapping engine development',
            'Template auto-detection algorithm (ML-based)',
            'Template management web UI implementation',
            'Multi-template testing with variants',
            'Template configuration documentation',
            'User training on template management'
        ],
        'deliverables': [
            'D15: Template registry database with versioning',
            'D16: Dynamic mapping engine and configuration schema',
            'D17: Template detection algorithm',
            'D18: Template management web UI',
            'D19: Multi-template test suite',
            'D20: Template configuration guide',
            'D21: Template admin training materials'
        ],
        'exit_criteria': [
            'Minimum 3 template variants configured',
            'Template auto-detection ≥95% accurate',
            'Dynamic mapping working for all templates',
            'Non-developer can configure template via UI',
            'Template switching tested successfully',
            'Template versioning and rollback working',
            'Template administrators trained'
        ],
        'dependencies': 'M2',
        'team_size': '4 (Full Stack Dev, UX Designer, BA, QA)',
        'critical_path': False,
        'payment_percent': 0,
        'risk_level': 'Medium'
    },
    {
        'id': 'M4',
        'name': 'Historical Processing Complete',
        'phase': 'Development - Features',
        'duration_days': 15,
        'work_packages': ['WP4'],
        'key_activities': [
            'Historical import CLI tool development',
            'Document classification (vendor and format detection)',
            'Legacy format handlers and OCR pipeline',
            'De-duplication logic implementation',
            'Mass import of historical proposals (100+)',
            'Data quality validation and cleanup',
            'Historical analytics dashboard'
        ],
        'deliverables': [
            'D22: Historical import CLI tool',
            'D23: Classification engine',
            'D24: Legacy format handlers with OCR',
            'D25: De-duplication engine',
            'D26: Historical proposals database (100+ files)',
            'D27: Data quality report',
            'D28: Historical insights dashboard'
        ],
        'exit_criteria': [
            'Minimum 100 historical proposals imported',
            'Vendor classification ≥90% accurate',
            'OCR accuracy ≥85% on scanned documents',
            'Duplicates identified and resolved',
            'Data quality validated (completeness)',
            'Historical pricing trends analyzed',
            'Insights dashboard operational'
        ],
        'dependencies': 'M2, M3',
        'team_size': '4 (Developer, ML Engineer, Data Analyst, QA)',
        'critical_path': False,
        'payment_percent': 25,
        'risk_level': 'Medium'
    },
    {
        'id': 'M5',
        'name': 'Exception Management & Review Workflow Live',
        'phase': 'Development - Features',
        'duration_days': 14,
        'work_packages': ['WP5'],
        'key_activities': [
            'Review queue backend implementation',
            'Web-based review UI development',
            'Approval workflow and state machine',
            'Feedback loop and active learning integration',
            'Review workflow E2E testing',
            'Reviewer training and certification',
            'SLA tracking and notification setup'
        ],
        'deliverables': [
            'D29: Review queue backend and API',
            'D30: Review web UI',
            'D31: Approval workflow engine',
            'D32: Feedback loop system',
            'D33: Review workflow integration tests',
            'D34: Reviewer training materials',
            'D35: Review SLA dashboard'
        ],
        'exit_criteria': [
            'Review queue operational with auto-routing',
            'Review UI validated by non-technical users',
            'Approval workflow tested (multi-level)',
            'Feedback loop capturing corrections',
            'Review SLA tracking functional',
            'Minimum 3 reviewers trained',
            'Review throughput ≥10 items/hour per reviewer'
        ],
        'dependencies': 'M2',
        'team_size': '5 (Full Stack Dev, UX, Backend Dev, BA, QA)',
        'critical_path': True,
        'payment_percent': 0,
        'risk_level': 'Medium'
    },
    {
        'id': 'M6',
        'name': 'Monitoring & Reporting Operational',
        'phase': 'Development - Operations',
        'duration_days': 15,
        'work_packages': ['WP6'],
        'key_activities': [
            'Real-time operations dashboard development',
            'Email and Slack notification system',
            'Standard reports library implementation',
            'Enhanced audit trail system',
            'Monitoring and alerting configuration',
            'Report scheduling and automation',
            'Stakeholder training on dashboards'
        ],
        'deliverables': [
            'D36: Operations dashboard',
            'D37: Notification system (Email + Slack)',
            'D38: Report library (7 standard reports)',
            'D39: Enhanced audit trail',
            'D40: Monitoring integration tests',
            'D41: Report catalog and user guide',
            'D42: Dashboard training materials'
        ],
        'exit_criteria': [
            'Dashboard displaying real-time metrics',
            'Notifications working for all job events',
            'All 7 standard reports generating',
            'Audit trail logging all operations',
            'Monitoring alerts tested and configured',
            'Reports scheduled for automated delivery',
            'Stakeholders trained on dashboards'
        ],
        'dependencies': 'M2, M5',
        'team_size': '4 (Full Stack Dev, Data Engineer, DevOps, BA)',
        'critical_path': False,
        'payment_percent': 0,
        'risk_level': 'Low'
    },
    {
        'id': 'M7',
        'name': 'Testing Complete & UAT Sign-off',
        'phase': 'Quality Assurance',
        'duration_days': 20,
        'work_packages': ['WP7'],
        'key_activities': [
            'Unit test expansion (≥80% coverage)',
            'Integration testing (E2E scenarios)',
            'Volume and performance testing',
            'User Acceptance Testing with client',
            'Bug fixing and retesting (P0/P1)',
            'Performance optimization',
            'UAT sign-off and approval'
        ],
        'deliverables': [
            'D43: Unit test suite (≥80% coverage)',
            'D44: Integration test suite',
            'D45: Performance test report',
            'D46: UAT test plan',
            'D47: Bug fix report (P0/P1 resolved)',
            'D48: UAT test results and sign-off',
            'D49: Go-live readiness checklist'
        ],
        'exit_criteria': [
            'Unit test coverage ≥80%',
            '20 proposals processed in <30 minutes',
            'UAT scenarios executed (100%)',
            'All P0 and P1 bugs resolved',
            'Client UAT sign-off obtained',
            'Performance benchmarks met',
            'Go-live checklist complete'
        ],
        'dependencies': 'M3, M4, M5, M6',
        'team_size': '6 (2 QA, Performance Eng, 2 Dev, Client Users)',
        'critical_path': True,
        'payment_percent': 25,
        'risk_level': 'Medium'
    },
    {
        'id': 'M8',
        'name': 'Documentation, Training & Go-Live',
        'phase': 'Deployment',
        'duration_days': 15,
        'work_packages': ['WP8'],
        'key_activities': [
            'Technical documentation completion',
            'Operations manual and runbooks',
            'End-user documentation and video tutorials',
            'Knowledge transfer training sessions',
            'Production deployment preparation',
            'Go-live execution and cutover',
            'Post-go-live stabilization and support'
        ],
        'deliverables': [
            'D50: Technical Documentation',
            'D51: Operations Manual and Runbooks',
            'D52: User Guide and Video Tutorials',
            'D53: Training materials (Admin, User, Developer)',
            'D54: Training completion certificates',
            'D55: Deployment guide and checklist',
            'D56: Go-live sign-off and warranty start'
        ],
        'exit_criteria': [
            'All documentation complete and reviewed',
            'Minimum 80% of users trained',
            'Production environment validated',
            'Go-live checklist 100% complete',
            'System operational in production',
            'Client final acceptance signed',
            'Post-go-live stabilization complete'
        ],
        'dependencies': 'M7',
        'team_size': '6 (Tech Writer, Trainer, Tech Lead, PM, DevOps, Support)',
        'critical_path': True,
        'payment_percent': 25,
        'risk_level': 'Low'
    }
]

# Phase summary
PHASES = {
    'Discovery': {
        'color': '4472C4',
        'description': 'Requirements gathering, analysis, and planning',
        'milestones': ['M0']
    },
    'Development - Core': {
        'color': '70AD47',
        'description': 'Foundation infrastructure and core processing engine',
        'milestones': ['M1', 'M2']
    },
    'Development - Features': {
        'color': 'FFC000',
        'description': 'Feature development and enhancements',
        'milestones': ['M3', 'M4', 'M5']
    },
    'Development - Operations': {
        'color': 'ED7D31',
        'description': 'Monitoring, reporting, and operational readiness',
        'milestones': ['M6']
    },
    'Quality Assurance': {
        'color': '5B9BD5',
        'description': 'Testing, validation, and client acceptance',
        'milestones': ['M7']
    },
    'Deployment': {
        'color': 'A5A5A5',
        'description': 'Documentation, training, and production deployment',
        'milestones': ['M8']
    }
}

def format_list_for_excel(items):
    """Format list items with bullet points for Excel cell"""
    return '\n'.join([f'• {item}' for item in items])

def calculate_timeline():
    """Calculate project timeline"""
    timeline = []
    current_week = 1

    for milestone in MILESTONES:
        duration_weeks = (milestone['duration_days'] + 4) // 5

        timeline.append({
            'milestone_id': milestone['id'],
            'start_week': current_week,
            'end_week': current_week + duration_weeks - 1,
            'duration_weeks': duration_weeks,
            'duration_days': milestone['duration_days']
        })

        current_week += duration_weeks

    return timeline

def update_milestones_final():
    """Update milestones sheet with clean 5-month plan"""

    # Load the workbook
    wb = openpyxl.load_workbook('Karishma Proposal Documents/Project_Plan_Phase2_DETAILED_20260109_132826.xlsx')
    ws = wb['Milestones']

    # Clear existing data
    ws.delete_rows(1, ws.max_row)

    # Define styles
    title_font = Font(bold=True, size=16, color='FFFFFF')
    title_fill = PatternFill(start_color='203864', end_color='203864', fill_type='solid')

    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    phase_font = Font(bold=True, size=12, color='FFFFFF')
    phase_alignment = Alignment(horizontal='left', vertical='center')

    cell_alignment = Alignment(vertical='top', wrap_text=True)
    center_alignment = Alignment(horizontal='center', vertical='center')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title row
    ws.merge_cells('A1:N1')
    title_cell = ws['A1']
    title_cell.value = 'TCO AUTOMATION SYSTEM - PHASE 2 PROJECT PLAN'
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Project summary row
    ws.merge_cells('A2:N2')
    summary_cell = ws['A2']
    timeline = calculate_timeline()
    total_weeks = timeline[-1]['end_week']
    total_days = sum(m['duration_days'] for m in MILESTONES)

    summary_cell.value = f'Duration: {total_weeks} weeks ({total_days} days / ~5 months) | 9 Milestones | 40 Tasks | 259 person-days effort'
    summary_cell.font = Font(bold=True, size=10)
    summary_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20

    # Headers
    headers = [
        ('A3', 'Milestone', 12),
        ('B3', 'Name', 35),
        ('C3', 'Phase', 20),
        ('D3', 'Timeline\n(Weeks)', 12),
        ('E3', 'Duration\n(Days)', 10),
        ('F3', 'Key Activities', 50),
        ('G3', 'Deliverables', 50),
        ('H3', 'Exit Criteria', 50),
        ('I3', 'Dependencies', 20),
        ('J3', 'Team Size', 15),
        ('K3', 'Critical\nPath', 10),
        ('L3', 'Risk\nLevel', 10),
        ('M3', 'Payment\n%', 10),
        ('N3', 'Status', 12)
    ]

    for cell_ref, header_text, width in headers:
        cell = ws[cell_ref]
        cell.value = header_text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

        col_letter = cell_ref[0]
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[3].height = 35

    # Populate milestones
    row = 4
    current_phase = None

    for idx, milestone in enumerate(MILESTONES):
        # Add phase header if phase changes
        if milestone['phase'] != current_phase:
            current_phase = milestone['phase']
            phase_info = PHASES[current_phase]

            ws.merge_cells(f'A{row}:N{row}')
            phase_cell = ws[f'A{row}']
            phase_cell.value = f'{current_phase.upper()}: {phase_info["description"]}'
            phase_cell.font = phase_font
            phase_cell.fill = PatternFill(start_color=phase_info['color'], end_color=phase_info['color'], fill_type='solid')
            phase_cell.alignment = phase_alignment
            ws.row_dimensions[row].height = 25
            row += 1

        # Get timeline info
        time_info = timeline[idx]

        # Milestone ID
        ws.cell(row, 1, milestone['id']).alignment = center_alignment
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 1).border = thin_border

        # Name
        ws.cell(row, 2, milestone['name']).alignment = cell_alignment
        ws.cell(row, 2).font = Font(bold=True)
        ws.cell(row, 2).border = thin_border

        # Phase
        ws.cell(row, 3, milestone['phase']).alignment = cell_alignment
        ws.cell(row, 3).border = thin_border

        # Timeline
        timeline_text = f"W{time_info['start_week']}-W{time_info['end_week']}\n({time_info['duration_weeks']} weeks)"
        ws.cell(row, 4, timeline_text).alignment = center_alignment
        ws.cell(row, 4).border = thin_border

        # Duration (Days)
        ws.cell(row, 5, milestone['duration_days']).alignment = center_alignment
        ws.cell(row, 5).border = thin_border

        # Key Activities
        activities_text = format_list_for_excel(milestone['key_activities'])
        ws.cell(row, 6, activities_text).alignment = cell_alignment
        ws.cell(row, 6).border = thin_border

        # Deliverables
        deliverables_text = format_list_for_excel(milestone['deliverables'])
        ws.cell(row, 7, deliverables_text).alignment = cell_alignment
        ws.cell(row, 7).border = thin_border

        # Exit Criteria
        criteria_text = format_list_for_excel(milestone['exit_criteria'])
        ws.cell(row, 8, criteria_text).alignment = cell_alignment
        ws.cell(row, 8).border = thin_border

        # Dependencies
        ws.cell(row, 9, milestone['dependencies']).alignment = cell_alignment
        ws.cell(row, 9).border = thin_border

        # Team Size
        ws.cell(row, 10, milestone['team_size']).alignment = cell_alignment
        ws.cell(row, 10).border = thin_border

        # Critical Path
        critical_text = 'YES' if milestone['critical_path'] else 'No'
        ws.cell(row, 11, critical_text).alignment = center_alignment
        if milestone['critical_path']:
            ws.cell(row, 11).font = Font(bold=True, color='C00000')
        ws.cell(row, 11).border = thin_border

        # Risk Level
        risk_colors = {'Low': '00B050', 'Medium': 'FFC000', 'High': 'C00000'}
        ws.cell(row, 12, milestone['risk_level']).alignment = center_alignment
        ws.cell(row, 12).font = Font(bold=True, color=risk_colors.get(milestone['risk_level'], '000000'))
        ws.cell(row, 12).border = thin_border

        # Payment %
        payment_text = f"{milestone['payment_percent']}%" if milestone['payment_percent'] > 0 else '-'
        ws.cell(row, 13, payment_text).alignment = center_alignment
        if milestone['payment_percent'] > 0:
            ws.cell(row, 13).font = Font(bold=True)
        ws.cell(row, 13).border = thin_border

        # Status
        ws.cell(row, 14, 'Not Started').alignment = center_alignment
        ws.cell(row, 14).border = thin_border

        # Set row height
        ws.row_dimensions[row].height = max(80, len(milestone['key_activities']) * 12)

        row += 1

    # Add summary section
    row += 1
    ws.merge_cells(f'A{row}:N{row}')
    summary_header = ws[f'A{row}']
    summary_header.value = 'PROJECT SUMMARY & KEY METRICS'
    summary_header.font = Font(bold=True, size=12, color='FFFFFF')
    summary_header.fill = PatternFill(start_color='203864', end_color='203864', fill_type='solid')
    summary_header.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 25

    row += 1

    # Summary metrics
    critical_path_milestones = [m['id'] for m in MILESTONES if m['critical_path']]
    total_payment = sum(m['payment_percent'] for m in MILESTONES)

    summary_data = [
        ['Total Milestones:', '9', 'Critical Path Milestones:', ', '.join(critical_path_milestones)],
        ['Total Duration:', f'{total_weeks} weeks ({total_days} days / ~5 months)', 'Total Payment Points:', f'{total_payment}%'],
        ['Phase Count:', '6 phases', 'Estimated Team:', '6-8 resources (peak)'],
        ['Total Deliverables:', f'{sum(len(m["deliverables"]) for m in MILESTONES)} items', 'High Risk Milestones:', f'{sum(1 for m in MILESTONES if m["risk_level"] == "High")}']
    ]

    for data_row in summary_data:
        ws.cell(row, 1, data_row[0]).font = Font(bold=True)
        ws.cell(row, 2, data_row[1])
        ws.cell(row, 3, data_row[2]).font = Font(bold=True)
        ws.merge_cells(f'D{row}:F{row}')
        ws.cell(row, 4, data_row[3])
        row += 1

    # Freeze panes
    ws.freeze_panes = 'A4'

    # Save workbook
    wb.save('Karishma Proposal Documents/Project_Plan_Phase2_DETAILED_20260109_132826.xlsx')

    print('\n[SUCCESS] Project Plan updated with 5-month timeline!')
    print(f'\nProject Timeline:')
    print(f'  Total Duration: {total_weeks} weeks ({total_days} days / ~5 months)')
    print(f'  Total Milestones: 9')
    print(f'  Total Deliverables: {sum(len(m["deliverables"]) for m in MILESTONES)}')
    print(f'\nMilestone Breakdown:')
    for m in MILESTONES:
        t = timeline[MILESTONES.index(m)]
        print(f'  {m["id"]}: W{t["start_week"]}-W{t["end_week"]} ({m["duration_days"]} days) - {m["name"]}')
    print(f'\nPhase Summary:')
    for phase_name, phase_info in PHASES.items():
        phase_milestones = [m for m in MILESTONES if m['phase'] == phase_name]
        phase_days = sum(m['duration_days'] for m in phase_milestones)
        phase_weeks = (phase_days + 4) // 5
        print(f'  {phase_name}: {phase_weeks} weeks ({phase_days} days) - {len(phase_milestones)} milestone(s)')
    print(f'\nPayment Schedule:')
    for m in MILESTONES:
        if m['payment_percent'] > 0:
            print(f'  {m["id"]}: {m["payment_percent"]}% - {m["name"]}')

if __name__ == '__main__':
    update_milestones_final()
