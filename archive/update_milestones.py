"""
Industry-Grade Project Plan - Milestones Sheet Update
Based on work package analysis and 2026 project management best practices
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# Define comprehensive milestones with industry-standard details
MILESTONES = [
    {
        'id': 'M0',
        'name': 'Discovery & Requirements Complete',
        'phase': 'Discovery',
        'duration_days': 24,
        'work_packages': ['WP0'],
        'key_activities': [
            'Historical proposal inventory (500+ files)',
            'TCO template variant collection & analysis',
            'Vendor document pattern analysis (FIS, JH, CSI)',
            'Volume & performance requirements definition',
            'Technical architecture design (C4 diagrams)',
            'Risk assessment & mitigation planning',
            'Discovery exit review & go/no-go decision'
        ],
        'deliverables': [
            'DD1: Historical Proposal Inventory (Excel + metadata)',
            'DD2: TCO Template Catalog & Mapping Specs',
            'DD3: Vendor Pattern Analysis Report',
            'DD4: Requirements Specification (SLAs, NFRs)',
            'DD5: Technical Architecture Document + C4 Diagrams',
            'DD6: Risk Register with mitigation plans',
            'DD7: Discovery Exit Report + Signed Acceptance'
        ],
        'exit_criteria': [
            'All discovery questions answered (100%)',
            'Historical inventory complete (file count, formats, vendors)',
            'All TCO template variants cataloged (min 3-5)',
            'Vendor extraction accuracy baseline established (≥95%)',
            'Technical architecture approved by client IT',
            'Risk register reviewed (min 15-20 risks)',
            'Client stakeholder sign-off on findings',
            'Dev phase work packages validated'
        ],
        'dependencies': 'Project kickoff',
        'team_size': '4-5 (PM, Tech Lead, BA, Developer, Client SME)',
        'critical_path': True,
        'payment_percent': 10,
        'risk_level': 'High',
        'success_metrics': [
            'Discovery completed within 24 days',
            'All discovery deliverables accepted',
            'No major scope changes required',
            'Client approval to proceed to development'
        ]
    },
    {
        'id': 'M1',
        'name': 'Foundation Infrastructure Operational',
        'phase': 'Development - Core',
        'duration_days': 30,
        'work_packages': ['WP1'],
        'key_activities': [
            'PostgreSQL database implementation (schema, migrations)',
            'Celery + RabbitMQ queue system setup',
            'Configuration management & secrets vault',
            'Structured logging infrastructure (ELK/CloudWatch)',
            'FastAPI REST framework with authentication',
            'Infrastructure testing & validation',
            'Performance baseline testing'
        ],
        'deliverables': [
            'D1: PostgreSQL database + ORM models + migrations',
            'D2: Celery + RabbitMQ job queue (4 workers)',
            'D3: Configuration management system + secrets vault',
            'D4: Structured logging framework + aggregation',
            'D5: FastAPI application + Swagger documentation',
            'D6: Infrastructure test suite (unit + integration)',
            'D7: Infrastructure setup guide + runbook'
        ],
        'exit_criteria': [
            'Database operational with all tables created',
            'Queue system processing test jobs successfully',
            'API endpoints returning 200 OK (health checks)',
            'Logs aggregating in centralized system',
            'No hardcoded secrets in codebase (Git audit clean)',
            'Infrastructure tests passing (≥80% coverage)',
            'Performance baseline documented'
        ],
        'dependencies': 'M0 (Discovery complete)',
        'team_size': '4 (Backend Dev, DevOps, DBA, Tech Lead)',
        'critical_path': True,
        'payment_percent': 0,
        'risk_level': 'Medium',
        'success_metrics': [
            'Infrastructure deployed and stable',
            'API response time <500ms',
            'Queue processing 4+ concurrent jobs',
            'Zero critical security vulnerabilities'
        ]
    },
    {
        'id': 'M2',
        'name': 'Batch Processing Engine Live',
        'phase': 'Development - Core',
        'duration_days': 24,
        'work_packages': ['WP2'],
        'key_activities': [
            'Bulk import & queue management implementation',
            'Parallel worker pool (4-8 workers)',
            'Error handling & retry logic with dead letter queue',
            'Real-time progress tracking (WebSocket)',
            'Volume testing (20+ proposals)',
            'Integration with extraction pipeline',
            'Performance optimization'
        ],
        'deliverables': [
            'D8: Bulk import system + directory watcher',
            'D9: Worker pool infrastructure (scalable)',
            'D10: Error handling framework + dead letter queue',
            'D11: Progress tracking API + dashboard',
            'D12: Batch processing integration tests',
            'D13: Volume test report (20-proposal baseline)',
            'D14: Batch processing user guide'
        ],
        'exit_criteria': [
            '20 proposals processed successfully in batch',
            'Avg processing time ≤90 seconds per proposal',
            'Workers stable (no crashes under load)',
            'Error handling tested (retry + dead letter queue)',
            'Progress tracking real-time (5-sec updates)',
            'Integration tests passing (E2E workflows)',
            'Performance meets SLA (defined in M0)'
        ],
        'dependencies': 'M1 (Infrastructure ready)',
        'team_size': '3 (Backend Dev, QA, DevOps)',
        'critical_path': True,
        'payment_percent': 0,
        'risk_level': 'High',
        'success_metrics': [
            '20 proposals processed in <30 minutes',
            'System throughput: 40-50 proposals/hour',
            'Error rate <5%',
            'Worker utilization ≥75%'
        ]
    },
    {
        'id': 'M3',
        'name': 'Multi-Template Support Enabled',
        'phase': 'Development - Features',
        'duration_days': 30,
        'work_packages': ['WP3'],
        'key_activities': [
            'Template registry & versioning system',
            'Dynamic field mapping engine',
            'Template auto-detection (≥95% accuracy)',
            'Template management web UI (no-code)',
            'Multi-template testing (3+ variants)',
            'Template configuration documentation',
            'User training materials'
        ],
        'deliverables': [
            'D15: Template registry database + versioning',
            'D16: Dynamic mapping engine + config schema',
            'D17: Template detection algorithm (ML-based)',
            'D18: Template management web UI',
            'D19: Multi-template test suite (3+ templates)',
            'D20: Template configuration guide',
            'D21: Template admin training materials'
        ],
        'exit_criteria': [
            'Min 3 template variants configured in registry',
            'Template auto-detection ≥95% accurate',
            'Dynamic mapping working for all templates',
            'Non-developer can upload & configure template via UI',
            'Template switching tested (same proposal → different templates)',
            'Template versioning working (rollback tested)',
            'Template admin trained and certified'
        ],
        'dependencies': 'M2 (Batch processing working)',
        'team_size': '4 (Full Stack Dev, UX Designer, BA, QA)',
        'critical_path': False,
        'payment_percent': 0,
        'risk_level': 'Medium',
        'success_metrics': [
            '3+ templates configured and working',
            'Template detection accuracy ≥95%',
            'Template configuration time <2 hours (per template)',
            'Zero code changes needed for new template'
        ]
    },
    {
        'id': 'M4',
        'name': 'Historical Processing Complete',
        'phase': 'Development - Features',
        'duration_days': 26,
        'work_packages': ['WP4'],
        'key_activities': [
            'Historical import CLI tool development',
            'Document classification (vendor + format detection)',
            'Legacy format handlers (OCR for scanned PDFs)',
            'De-duplication & historical analysis',
            'Mass import of historical proposals (100+)',
            'Data quality validation & cleanup',
            'Historical insights dashboard'
        ],
        'deliverables': [
            'D22: Historical import CLI tool',
            'D23: Classification engine (vendor + format)',
            'D24: Legacy format handlers (OCR pipeline)',
            'D25: De-duplication engine + analytics',
            'D26: Historical proposals database (100+ files)',
            'D27: Data quality report',
            'D28: Historical insights dashboard'
        ],
        'exit_criteria': [
            'Min 100 historical proposals imported successfully',
            'Vendor classification ≥90% accurate',
            'OCR accuracy ≥85% on scanned documents',
            'Duplicates identified and resolved',
            'Historical data quality validated (completeness check)',
            'Historical pricing trends analyzed',
            'Insights dashboard showing historical patterns'
        ],
        'dependencies': 'M2 (Batch processing), M3 (Multi-template)',
        'team_size': '4 (Developer, ML Engineer, Data Analyst, QA)',
        'critical_path': False,
        'payment_percent': 25,
        'risk_level': 'High',
        'success_metrics': [
            '100+ historical files processed',
            'Classification accuracy ≥90%',
            'OCR success rate ≥80% (readable scans)',
            'Historical database provides actionable insights'
        ]
    },
    {
        'id': 'M5',
        'name': 'Exception Management & Review Workflow Live',
        'phase': 'Development - Features',
        'duration_days': 28,
        'work_packages': ['WP5'],
        'key_activities': [
            'Review queue backend implementation',
            'Web-based review UI (side-by-side comparison)',
            'Approval workflow & state machine',
            'Feedback loop & active learning',
            'Review workflow testing (E2E)',
            'Reviewer training & certification',
            'SLA tracking & notifications'
        ],
        'deliverables': [
            'D29: Review queue backend + API',
            'D30: Review web UI (intuitive, non-technical)',
            'D31: Approval workflow engine + state machine',
            'D32: Feedback loop system + learning pipeline',
            'D33: Review workflow integration tests',
            'D34: Reviewer training materials',
            'D35: Review SLA dashboard'
        ],
        'exit_criteria': [
            'Review queue operational (items auto-routed)',
            'Review UI tested by non-technical users',
            'Approval workflow tested (multi-level)',
            'Feedback loop capturing corrections',
            'Review SLA tracking working (24-hour alerts)',
            'Min 3 reviewers trained and certified',
            'Review throughput ≥10 items/hour (per reviewer)'
        ],
        'dependencies': 'M2 (Batch processing)',
        'team_size': '5 (Full Stack Dev, UX, Backend Dev, BA, QA)',
        'critical_path': True,
        'payment_percent': 0,
        'risk_level': 'Medium',
        'success_metrics': [
            'Review queue processing ≥20 items/day',
            'Review SLA compliance ≥90%',
            'User satisfaction ≥4/5 (UI usability)',
            'Feedback loop improves accuracy by ≥5%'
        ]
    },
    {
        'id': 'M6',
        'name': 'Monitoring, Reporting & Analytics Operational',
        'phase': 'Development - Operations',
        'duration_days': 31,
        'work_packages': ['WP6'],
        'key_activities': [
            'Real-time operations dashboard',
            'Email & Slack notification system',
            'Standard reports library (7+ reports)',
            'Enhanced audit trail (compliance)',
            'Monitoring & alerting testing',
            'Report scheduling & automation',
            'Stakeholder training on dashboards'
        ],
        'deliverables': [
            'D36: Operations dashboard (real-time)',
            'D37: Notification system (Email + Slack)',
            'D38: Report library (7+ standard reports)',
            'D39: Enhanced audit trail system',
            'D40: Monitoring integration tests',
            'D41: Report catalog & user guide',
            'D42: Dashboard training materials'
        ],
        'exit_criteria': [
            'Dashboard shows real-time system health',
            'Notifications working (job complete, failures, SLA)',
            'All 7 standard reports generating successfully',
            'Audit trail tracks all critical operations',
            'Monitoring alerts tested (simulate failures)',
            'Reports scheduled (weekly automated delivery)',
            'Stakeholders trained on dashboards'
        ],
        'dependencies': 'M2 (Batch), M5 (Review)',
        'team_size': '4 (Full Stack Dev, Data Engineer, DevOps, BA)',
        'critical_path': False,
        'payment_percent': 0,
        'risk_level': 'Low',
        'success_metrics': [
            'Dashboard uptime ≥99%',
            'Notification delivery rate ≥98%',
            'Report generation time <30 seconds',
            'Audit trail 100% complete (all ops logged)'
        ]
    },
    {
        'id': 'M7',
        'name': 'Testing Complete & UAT Sign-off',
        'phase': 'Quality Assurance',
        'duration_days': 34,
        'work_packages': ['WP7'],
        'key_activities': [
            'Unit test expansion (≥80% coverage)',
            'Integration testing (E2E scenarios)',
            'Volume & performance testing (20+ proposals)',
            'User Acceptance Testing with client',
            'Bug fixing & retesting (P0, P1 issues)',
            'Performance optimization',
            'UAT sign-off & approval'
        ],
        'deliverables': [
            'D43: Unit test suite (≥80% coverage)',
            'D44: Integration test suite (15+ scenarios)',
            'D45: Performance test report (load testing)',
            'D46: UAT test plan (20+ scenarios)',
            'D47: Bug fix report (all P0/P1 resolved)',
            'D48: UAT test results & sign-off',
            'D49: Go-live readiness checklist'
        ],
        'exit_criteria': [
            'Unit test coverage ≥80%',
            'All integration tests passing',
            '20-proposal batch processed in <30 min',
            'Client UAT scenarios executed (100%)',
            'All P0 and P1 bugs resolved',
            'Client formal UAT sign-off obtained',
            'Go-live checklist 100% complete'
        ],
        'dependencies': 'M1-M6 (All dev complete)',
        'team_size': '5 (QA, Performance Eng, Developers, Client Users)',
        'critical_path': True,
        'payment_percent': 25,
        'risk_level': 'High',
        'success_metrics': [
            'Zero critical bugs in production',
            'Performance SLAs met (90-sec avg)',
            'UAT pass rate ≥95%',
            'Client satisfaction ≥4.5/5'
        ]
    },
    {
        'id': 'M8',
        'name': 'Documentation, Training & Go-Live',
        'phase': 'Deployment',
        'duration_days': 27,
        'work_packages': ['WP8'],
        'key_activities': [
            'Technical documentation (Architecture, API, Dev)',
            'Operations manual & runbooks',
            'End-user documentation & videos',
            'Knowledge transfer training (3 tracks)',
            'Production deployment preparation',
            'Go-live execution & cutover',
            'Post-go-live support (2 weeks)'
        ],
        'deliverables': [
            'D50: Technical Documentation (ReadTheDocs)',
            'D51: Operations Manual + Runbooks',
            'D52: User Guide + Video Tutorials (5+)',
            'D53: Training materials (Admin, User, Developer)',
            'D54: Training completion certificates',
            'D55: Deployment guide + checklist',
            'D56: Go-live sign-off & warranty start'
        ],
        'exit_criteria': [
            'All documentation complete and reviewed',
            'Min 80% of users trained and certified',
            'Production environment validated',
            'Go-live checklist 100% complete',
            'System operational in production',
            'Client final acceptance signed',
            '2-week stabilization period complete'
        ],
        'dependencies': 'M7 (UAT sign-off)',
        'team_size': '6 (Tech Writer, Trainer, Tech Lead, PM, DevOps, Support)',
        'critical_path': True,
        'payment_percent': 25,
        'risk_level': 'Medium',
        'success_metrics': [
            'Documentation complete and accessible',
            'Training completion rate ≥80%',
            'Go-live successful (no rollback)',
            'Production stability ≥99% uptime (week 1)'
        ]
    }
]

# Phase summary for visual organization
PHASES = {
    'Discovery': {
        'color': '4472C4',
        'description': 'Requirements gathering, analysis, and planning',
        'milestones': ['M0']
    },
    'Development - Core': {
        'color': '70AD47',
        'description': 'Foundation infrastructure and core processing engine',
        'milestones': ['M1', 'M2']
    },
    'Development - Features': {
        'color': 'FFC000',
        'description': 'Feature development and enhancements',
        'milestones': ['M3', 'M4', 'M5']
    },
    'Development - Operations': {
        'color': 'ED7D31',
        'description': 'Monitoring, reporting, and operational readiness',
        'milestones': ['M6']
    },
    'Quality Assurance': {
        'color': '5B9BD5',
        'description': 'Testing, validation, and client acceptance',
        'milestones': ['M7']
    },
    'Deployment': {
        'color': 'A5A5A5',
        'description': 'Documentation, training, and production deployment',
        'milestones': ['M8']
    }
}

def format_list_for_excel(items):
    """Format list items with bullet points for Excel cell"""
    return '\n'.join([f'• {item}' for item in items])

def calculate_timeline():
    """Calculate project timeline with start/end weeks"""
    timeline = []
    current_week = 1

    for milestone in MILESTONES:
        duration_weeks = (milestone['duration_days'] + 4) // 5  # Convert days to weeks, round up

        timeline.append({
            'milestone_id': milestone['id'],
            'start_week': current_week,
            'end_week': current_week + duration_weeks - 1,
            'duration_weeks': duration_weeks,
            'duration_days': milestone['duration_days']
        })

        current_week += duration_weeks

    return timeline

def update_milestones_sheet():
    """Update the Milestones sheet with comprehensive project plan"""

    # Load workbook
    wb = openpyxl.load_workbook('Karishma Proposal Documents/Project_Plan_Phase2_1_9 SMEdits.xlsx')
    ws = wb['Milestones']

    # Clear existing data
    ws.delete_rows(1, ws.max_row)

    # Define styles
    title_font = Font(bold=True, size=16, color='FFFFFF')
    title_fill = PatternFill(start_color='203864', end_color='203864', fill_type='solid')

    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    phase_font = Font(bold=True, size=12, color='FFFFFF')
    phase_alignment = Alignment(horizontal='left', vertical='center')

    cell_alignment = Alignment(vertical='top', wrap_text=True)
    center_alignment = Alignment(horizontal='center', vertical='center')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title row
    ws.merge_cells('A1:N1')
    title_cell = ws['A1']
    title_cell.value = 'TCO AUTOMATION SYSTEM - PHASE 2 PROJECT PLAN'
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Project summary row
    ws.merge_cells('A2:N2')
    summary_cell = ws['A2']
    timeline = calculate_timeline()
    total_weeks = timeline[-1]['end_week']
    total_days = sum(m['duration_days'] for m in MILESTONES)

    summary_cell.value = f'Total Duration: {total_weeks} weeks ({total_days} days) | 9 Milestones | 40 Tasks | Estimated Effort: 259 person-days'
    summary_cell.font = Font(bold=True, size=10)
    summary_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20

    # Headers
    headers = [
        ('A3', 'Milestone', 12),
        ('B3', 'Name', 35),
        ('C3', 'Phase', 20),
        ('D3', 'Timeline\n(Weeks)', 12),
        ('E3', 'Duration\n(Days)', 10),
        ('F3', 'Key Activities', 50),
        ('G3', 'Deliverables', 50),
        ('H3', 'Exit Criteria', 50),
        ('I3', 'Dependencies', 20),
        ('J3', 'Team Size', 15),
        ('K3', 'Critical\nPath', 10),
        ('L3', 'Risk\nLevel', 10),
        ('M3', 'Payment\n%', 10),
        ('N3', 'Status', 12)
    ]

    for cell_ref, header_text, width in headers:
        cell = ws[cell_ref]
        cell.value = header_text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

        col_letter = cell_ref[0]
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[3].height = 35

    # Populate milestones grouped by phase
    row = 4
    current_phase = None

    for idx, milestone in enumerate(MILESTONES):
        # Add phase header row if phase changes
        if milestone['phase'] != current_phase:
            current_phase = milestone['phase']
            phase_info = PHASES[current_phase]

            ws.merge_cells(f'A{row}:N{row}')
            phase_cell = ws[f'A{row}']
            phase_cell.value = f'{current_phase.upper()}: {phase_info["description"]}'
            phase_cell.font = phase_font
            phase_cell.fill = PatternFill(start_color=phase_info['color'], end_color=phase_info['color'], fill_type='solid')
            phase_cell.alignment = phase_alignment
            ws.row_dimensions[row].height = 25
            row += 1

        # Milestone row
        time_info = timeline[idx]

        # Milestone ID
        ws.cell(row, 1, milestone['id']).alignment = center_alignment
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 1).border = thin_border

        # Name
        ws.cell(row, 2, milestone['name']).alignment = cell_alignment
        ws.cell(row, 2).font = Font(bold=True)
        ws.cell(row, 2).border = thin_border

        # Phase
        ws.cell(row, 3, milestone['phase']).alignment = cell_alignment
        ws.cell(row, 3).border = thin_border

        # Timeline
        timeline_text = f"W{time_info['start_week']}-W{time_info['end_week']}\n({time_info['duration_weeks']} weeks)"
        ws.cell(row, 4, timeline_text).alignment = center_alignment
        ws.cell(row, 4).border = thin_border

        # Duration (Days)
        ws.cell(row, 5, milestone['duration_days']).alignment = center_alignment
        ws.cell(row, 5).border = thin_border

        # Key Activities
        activities_text = format_list_for_excel(milestone['key_activities'])
        ws.cell(row, 6, activities_text).alignment = cell_alignment
        ws.cell(row, 6).border = thin_border

        # Deliverables
        deliverables_text = format_list_for_excel(milestone['deliverables'])
        ws.cell(row, 7, deliverables_text).alignment = cell_alignment
        ws.cell(row, 7).border = thin_border

        # Exit Criteria
        criteria_text = format_list_for_excel(milestone['exit_criteria'])
        ws.cell(row, 8, criteria_text).alignment = cell_alignment
        ws.cell(row, 8).border = thin_border

        # Dependencies
        ws.cell(row, 9, milestone['dependencies']).alignment = cell_alignment
        ws.cell(row, 9).border = thin_border

        # Team Size
        ws.cell(row, 10, milestone['team_size']).alignment = cell_alignment
        ws.cell(row, 10).border = thin_border

        # Critical Path
        critical_text = 'YES' if milestone['critical_path'] else 'No'
        ws.cell(row, 11, critical_text).alignment = center_alignment
        if milestone['critical_path']:
            ws.cell(row, 11).font = Font(bold=True, color='C00000')
        ws.cell(row, 11).border = thin_border

        # Risk Level
        risk_colors = {'Low': '00B050', 'Medium': 'FFC000', 'High': 'C00000'}
        ws.cell(row, 12, milestone['risk_level']).alignment = center_alignment
        ws.cell(row, 12).font = Font(bold=True, color=risk_colors.get(milestone['risk_level'], '000000'))
        ws.cell(row, 12).border = thin_border

        # Payment %
        payment_text = f"{milestone['payment_percent']}%" if milestone['payment_percent'] > 0 else '-'
        ws.cell(row, 13, payment_text).alignment = center_alignment
        if milestone['payment_percent'] > 0:
            ws.cell(row, 13).font = Font(bold=True)
        ws.cell(row, 13).border = thin_border

        # Status
        ws.cell(row, 14, 'Not Started').alignment = center_alignment
        ws.cell(row, 14).border = thin_border

        # Set row height
        ws.row_dimensions[row].height = max(80, len(milestone['key_activities']) * 12)

        row += 1

    # Add summary section
    row += 1
    ws.merge_cells(f'A{row}:N{row}')
    summary_header = ws[f'A{row}']
    summary_header.value = 'PROJECT SUMMARY & KEY METRICS'
    summary_header.font = Font(bold=True, size=12, color='FFFFFF')
    summary_header.fill = PatternFill(start_color='203864', end_color='203864', fill_type='solid')
    summary_header.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 25

    row += 1

    # Summary metrics
    critical_path_milestones = [m['id'] for m in MILESTONES if m['critical_path']]
    total_payment = sum(m['payment_percent'] for m in MILESTONES)

    summary_data = [
        ['Total Milestones:', '9', 'Critical Path Milestones:', ', '.join(critical_path_milestones)],
        ['Total Duration:', f'{total_weeks} weeks ({total_days} days)', 'Total Payment Points:', f'{total_payment}%'],
        ['Phase Count:', '6 phases', 'Estimated Team:', '8-10 resources (peak)'],
        ['Total Deliverables:', f'{sum(len(m["deliverables"]) for m in MILESTONES)} items', 'High Risk Milestones:', f'{sum(1 for m in MILESTONES if m["risk_level"] == "High")}']
    ]

    for data_row in summary_data:
        ws.cell(row, 1, data_row[0]).font = Font(bold=True)
        ws.cell(row, 2, data_row[1])
        ws.cell(row, 3, data_row[2]).font = Font(bold=True)
        ws.merge_cells(f'D{row}:F{row}')
        ws.cell(row, 4, data_row[3])
        row += 1

    # Freeze panes
    ws.freeze_panes = 'A4'

    # Save workbook
    wb.save('Karishma Proposal Documents/Project_Plan_Phase2_1_9 SMEdits.xlsx')

    print('\n[SUCCESS] Milestones sheet updated with industry-grade project plan!')
    print(f'\nProject Timeline Summary:')
    print(f'  Total Duration: {total_weeks} weeks ({total_days} days)')
    print(f'  Total Milestones: 9')
    print(f'  Critical Path Milestones: {len(critical_path_milestones)}')
    print(f'  Payment Milestones: {sum(1 for m in MILESTONES if m["payment_percent"] > 0)}')
    print(f'\nPhase Breakdown:')
    for phase_name, phase_info in PHASES.items():
        phase_milestones = [m for m in MILESTONES if m['phase'] == phase_name]
        phase_days = sum(m['duration_days'] for m in phase_milestones)
        phase_weeks = (phase_days + 4) // 5
        print(f'  {phase_name}: {phase_weeks} weeks ({phase_days} days) - {len(phase_milestones)} milestone(s)')

    print(f'\nCritical Path (Must-Complete Milestones):')
    for m_id in critical_path_milestones:
        m = next(m for m in MILESTONES if m['id'] == m_id)
        print(f'  {m_id}: {m["name"]}')

    print(f'\nPayment Schedule:')
    for m in MILESTONES:
        if m['payment_percent'] > 0:
            print(f'  {m["id"]}: {m["payment_percent"]}% - {m["name"]}')

if __name__ == '__main__':
    update_milestones_sheet()
