"""
Compare FIS line items:
- WORKBOOK2.xlsx (expected output)
- Our JSON extraction (echelon_bank_fis_extraction_ai.json)
"""
import openpyxl
import json
from collections import defaultdict

def read_workbook2_fis_items():
    """Read all FIS line items from WORKBOOK2.xlsx"""

    wb = openpyxl.load_workbook('WORKBOOK2.xlsx')
    ws = wb['Line Items']

    items = []
    current_section = None

    # Read from row 6 onwards (after headers)
    for row in range(6, 160):
        col_b = ws.cell(row, 2).value  # Fee Type
        col_o = ws.cell(row, 15).value  # Solution Name
        col_p = ws.cell(row, 16).value  # Category
        col_q = ws.cell(row, 17).value  # Per Unit Rate
        col_c = ws.cell(row, 3).value  # Proposal (quantity or indicator)

        # Check for section headers
        if col_b and isinstance(col_b, str) and 'Recurring' in col_b:
            if 'Bundle FIS' in col_b:
                current_section = 'Bundle FIS Products'
            elif 'Non-Bundle REQUIRED FIS' in col_b:
                current_section = 'Non-Bundle REQUIRED FIS Products'
            elif 'Non-Bundle Required Third' in col_b:
                current_section = 'Non-Bundle REQUIRED Third Parties'
            continue

        if col_b and isinstance(col_b, str) and 'Implementation' in col_b and 'FIS' in col_b:
            current_section = 'Implementation Credits and One-Time Fees'
            continue

        if col_b and isinstance(col_b, str) and 'Non-Bundle Optional FIS' in col_b:
            current_section = 'Non-Bundle OPTIONAL FIS Solutions'
            continue

        # Extract line items
        if col_b in ['Monthly F', 'Monthly V', 'Annual', 'One-Time'] and col_o:
            item = {
                'row': row,
                'section': current_section or 'Unknown',
                'solution_name': col_o,
                'category': col_p or '',
                'fee_type': col_b,
                'per_unit_rate': col_q or 0,
                'quantity': col_c if isinstance(col_c, (int, float)) else 0
            }
            items.append(item)

    wb.close()
    return items


def read_our_extraction():
    """Read our JSON extraction"""

    with open('Extracted JSON/echelon_bank_fis_extraction_ai.json', 'r') as f:
        data = json.load(f)

    return data.get('line_items', [])


def normalize_name(name):
    """Normalize solution name for comparison"""
    if not name:
        return ''
    return name.lower().strip().replace('-', ' ').replace('_', ' ')


def compare_extractions():
    """Compare WORKBOOK2 vs our extraction"""

    print(f"\n{'='*120}")
    print("FIS LINE ITEMS COMPARISON")
    print(f"{'='*120}\n")

    # Read both
    workbook2_items = read_workbook2_fis_items()
    our_items = read_our_extraction()

    print(f"WORKBOOK2 FIS items: {len(workbook2_items)}")
    print(f"Our extraction items: {len(our_items)}\n")

    # Group by section
    print(f"\n{'='*120}")
    print("WORKBOOK2 ITEMS BY SECTION:")
    print(f"{'='*120}\n")

    sections = defaultdict(list)
    for item in workbook2_items:
        sections[item['section']].append(item)

    for section, items in sections.items():
        print(f"\n{section}: ({len(items)} items)")
        for item in items:
            rate = item['per_unit_rate']
            rate_str = f"${rate:,.4f}" if isinstance(rate, (int, float)) else str(rate)
            print(f"  - {item['solution_name']:60s} | {item['fee_type']:10s} | {rate_str:>12s} | {item['category']}")

    # Group our items by category
    print(f"\n{'='*120}")
    print("OUR EXTRACTION ITEMS BY CATEGORY:")
    print(f"{'='*120}\n")

    our_categories = defaultdict(list)
    for item in our_items:
        our_categories[item.get('category', 'Unknown')].append(item)

    for category, items in sorted(our_categories.items()):
        print(f"\n{category}: ({len(items)} items)")
        for item in items:
            monthly = item.get('monthly_fee', 0) or 0
            per_unit = item.get('per_unit_rate', 0) or 0
            one_time = item.get('one_time_fee', 0) or 0
            print(f"  - {item['solution_name']:60s} | {item['fee_type']:10s} | Monthly:${monthly:,.2f} Per Unit:${per_unit:,.4f}")

    # Find matches
    print(f"\n{'='*120}")
    print("MATCHING ANALYSIS:")
    print(f"{'='*120}\n")

    workbook2_names = {normalize_name(item['solution_name']): item for item in workbook2_items}
    our_names = {normalize_name(item['solution_name']): item for item in our_items}

    # Find matches
    matched = []
    workbook2_only = []
    our_only = []

    for name, item in workbook2_names.items():
        if name in our_names:
            matched.append((item, our_names[name]))
        else:
            workbook2_only.append(item)

    for name, item in our_names.items():
        if name not in workbook2_names:
            our_only.append(item)

    print(f"MATCHED items: {len(matched)}")
    print(f"WORKBOOK2 ONLY: {len(workbook2_only)}")
    print(f"OUR EXTRACTION ONLY: {len(our_only)}")

    # Show WORKBOOK2 items missing from our extraction
    if workbook2_only:
        print(f"\n{'='*120}")
        print("MISSING FROM OUR EXTRACTION (in WORKBOOK2 but not extracted):")
        print(f"{'='*120}\n")

        for item in workbook2_only:
            print(f"  [-] {item['solution_name']:60s} | {item['section']:40s} | {item['fee_type']}")

    # Show items we extracted but not in WORKBOOK2
    if our_only:
        print(f"\n{'='*120}")
        print("EXTRA IN OUR EXTRACTION (extracted but not in WORKBOOK2):")
        print(f"{'='*120}\n")

        for item in our_only:
            print(f"  [+] {item['solution_name']:60s} | {item.get('category', 'N/A'):40s} | {item['fee_type']}")

    # Show matched items
    print(f"\n{'='*120}")
    print("MATCHED ITEMS (partial name matches):")
    print(f"{'='*120}\n")

    for wb_item, our_item in matched[:10]:  # Show first 10
        print(f"  [OK] {wb_item['solution_name']}")
        print(f"    WB2: {wb_item['fee_type']} | Rate: ${wb_item['per_unit_rate'] if isinstance(wb_item['per_unit_rate'], (int, float)) else wb_item['per_unit_rate']}")
        print(f"    Our: {our_item['fee_type']} | Monthly: ${our_item.get('monthly_fee', 0):,.2f} | Per Unit: ${our_item.get('per_unit_rate', 0):,.4f}")
        print()

    print(f"  ... ({len(matched)} total matches)\n")

    print(f"\n{'='*120}")
    print("SUMMARY:")
    print(f"{'='*120}\n")
    print(f"  Total in WORKBOOK2: {len(workbook2_items)}")
    print(f"  Total in our extraction: {len(our_items)}")
    print(f"  Matched (by name): {len(matched)}")
    print(f"  Missing from extraction: {len(workbook2_only)}")
    print(f"  Extra in extraction: {len(our_only)}")
    print(f"  Match rate: {len(matched)/len(workbook2_items)*100:.1f}%")
    print()


if __name__ == '__main__':
    compare_extractions()
