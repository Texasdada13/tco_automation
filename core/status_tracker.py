"""
Status Tracker Module

Provides pipeline status tracking, logging, and reporting.
Maintains a status.json file with extraction history and metrics.

Features:
- Real-time status tracking
- Extraction history
- Success/failure metrics
- Daily summary reports
"""

import json
import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Any, Optional
from dataclasses import dataclass, field, asdict
from enum import Enum

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

logger = logging.getLogger(__name__)


class ExtractionStatus(str, Enum):
    """Status of an extraction."""
    PENDING = "pending"
    IN_PROGRESS = "in_progress"
    SUCCESS = "success"
    REVIEW_NEEDED = "review_needed"
    FAILED = "failed"


@dataclass
class ExtractionRecord:
    """Record of a single extraction."""
    id: str
    file_name: str
    vendor: str
    client: str
    status: str
    confidence: float
    items_extracted: int
    items_need_review: int
    started_at: str
    completed_at: Optional[str] = None
    error_message: Optional[str] = None
    output_path: Optional[str] = None
    qa_report_path: Optional[str] = None

    def to_dict(self) -> Dict:
        return asdict(self)


@dataclass
class PipelineStatus:
    """Overall pipeline status."""
    last_run: str
    total_processed: int
    success_count: int
    review_needed_count: int
    failed_count: int
    success_rate: float
    average_confidence: float
    recent_extractions: List[Dict]
    pending_count: int = 0

    def to_dict(self) -> Dict:
        return asdict(self)


class StatusTracker:
    """
    Tracks extraction pipeline status and history.

    Usage:
        tracker = StatusTracker()
        record_id = tracker.start_extraction("proposal.docx", "FIS", "First Bank")
        tracker.update_extraction(record_id, status="success", confidence=0.95)
        tracker.save()
    """

    def __init__(self, status_file: str = None, output_dir: str = None):
        """
        Initialize the status tracker.

        Args:
            status_file: Path to status.json file
            output_dir: Base output directory
        """
        if output_dir is None:
            project_root = Path(__file__).parent.parent
            output_dir = str(project_root / "output")

        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

        if status_file is None:
            status_file = str(self.output_dir / "status.json")

        self.status_file = Path(status_file)
        self.records: Dict[str, ExtractionRecord] = {}
        self._load()

        logger.info(f"StatusTracker initialized: {self.status_file}")

    def _load(self) -> None:
        """Load existing status data."""
        if self.status_file.exists():
            try:
                with open(self.status_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)

                for record_data in data.get('extractions', []):
                    record = ExtractionRecord(**record_data)
                    self.records[record.id] = record

                logger.info(f"Loaded {len(self.records)} extraction records")
            except Exception as e:
                logger.error(f"Failed to load status file: {e}")

    def save(self) -> None:
        """Save status data to file."""
        status = self.get_status()

        data = {
            'status': status.to_dict(),
            'extractions': [r.to_dict() for r in self.records.values()],
            'updated_at': datetime.now().isoformat()
        }

        with open(self.status_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, default=str)

        logger.info(f"Status saved to: {self.status_file}")

    def start_extraction(
        self,
        file_name: str,
        vendor: str,
        client: str
    ) -> str:
        """
        Record the start of a new extraction.

        Args:
            file_name: Name of source file
            vendor: Vendor name
            client: Client name

        Returns:
            Unique record ID
        """
        record_id = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{vendor}_{client}".replace(' ', '_')

        record = ExtractionRecord(
            id=record_id,
            file_name=file_name,
            vendor=vendor,
            client=client,
            status=ExtractionStatus.IN_PROGRESS.value,
            confidence=0.0,
            items_extracted=0,
            items_need_review=0,
            started_at=datetime.now().isoformat()
        )

        self.records[record_id] = record
        self.save()

        logger.info(f"Started extraction: {record_id}")
        return record_id

    def update_extraction(
        self,
        record_id: str,
        status: str = None,
        confidence: float = None,
        items_extracted: int = None,
        items_need_review: int = None,
        error_message: str = None,
        output_path: str = None,
        qa_report_path: str = None
    ) -> None:
        """
        Update an extraction record.

        Args:
            record_id: ID of the record to update
            status: New status
            confidence: Confidence score
            items_extracted: Number of items extracted
            items_need_review: Number of items needing review
            error_message: Error message if failed
            output_path: Path to output files
            qa_report_path: Path to QA report
        """
        if record_id not in self.records:
            logger.warning(f"Record not found: {record_id}")
            return

        record = self.records[record_id]

        if status:
            record.status = status
        if confidence is not None:
            record.confidence = confidence
        if items_extracted is not None:
            record.items_extracted = items_extracted
        if items_need_review is not None:
            record.items_need_review = items_need_review
        if error_message:
            record.error_message = error_message
        if output_path:
            record.output_path = output_path
        if qa_report_path:
            record.qa_report_path = qa_report_path

        # Set completion time if status is terminal
        if status in [ExtractionStatus.SUCCESS.value, ExtractionStatus.FAILED.value, ExtractionStatus.REVIEW_NEEDED.value]:
            record.completed_at = datetime.now().isoformat()

        self.save()
        logger.info(f"Updated extraction {record_id}: status={status}, confidence={confidence}")

    def complete_extraction(
        self,
        record_id: str,
        success: bool,
        confidence: float,
        items_extracted: int,
        items_need_review: int,
        output_path: str = None,
        qa_report_path: str = None,
        error_message: str = None
    ) -> None:
        """
        Mark an extraction as complete.

        Args:
            record_id: ID of the record
            success: Whether extraction succeeded
            confidence: Overall confidence score
            items_extracted: Number of items extracted
            items_need_review: Number needing review
            output_path: Path to outputs
            qa_report_path: Path to QA report
            error_message: Error message if failed
        """
        if success:
            if items_need_review > 0 or confidence < 0.90:
                status = ExtractionStatus.REVIEW_NEEDED.value
            else:
                status = ExtractionStatus.SUCCESS.value
        else:
            status = ExtractionStatus.FAILED.value

        self.update_extraction(
            record_id,
            status=status,
            confidence=confidence,
            items_extracted=items_extracted,
            items_need_review=items_need_review,
            output_path=output_path,
            qa_report_path=qa_report_path,
            error_message=error_message
        )

    def get_status(self) -> PipelineStatus:
        """
        Get overall pipeline status.

        Returns:
            PipelineStatus object
        """
        total = len(self.records)
        success = sum(1 for r in self.records.values() if r.status == ExtractionStatus.SUCCESS.value)
        review_needed = sum(1 for r in self.records.values() if r.status == ExtractionStatus.REVIEW_NEEDED.value)
        failed = sum(1 for r in self.records.values() if r.status == ExtractionStatus.FAILED.value)
        pending = sum(1 for r in self.records.values() if r.status in [ExtractionStatus.PENDING.value, ExtractionStatus.IN_PROGRESS.value])

        completed = success + review_needed + failed
        success_rate = (success + review_needed) / completed * 100 if completed > 0 else 0

        confidences = [r.confidence for r in self.records.values() if r.confidence > 0]
        avg_confidence = sum(confidences) / len(confidences) if confidences else 0

        # Get recent extractions (last 10)
        sorted_records = sorted(
            self.records.values(),
            key=lambda r: r.started_at,
            reverse=True
        )[:10]

        recent = [{
            'file': r.file_name,
            'vendor': r.vendor,
            'status': r.status,
            'confidence': r.confidence,
            'completed_at': r.completed_at
        } for r in sorted_records]

        last_run = sorted_records[0].started_at if sorted_records else datetime.now().isoformat()

        return PipelineStatus(
            last_run=last_run,
            total_processed=total,
            success_count=success,
            review_needed_count=review_needed,
            failed_count=failed,
            success_rate=success_rate,
            average_confidence=avg_confidence,
            recent_extractions=recent,
            pending_count=pending
        )

    def get_records_by_date(self, date: str = None) -> List[ExtractionRecord]:
        """
        Get extraction records for a specific date.

        Args:
            date: Date string (YYYY-MM-DD), defaults to today

        Returns:
            List of ExtractionRecord objects
        """
        if date is None:
            date = datetime.now().strftime('%Y-%m-%d')

        return [
            r for r in self.records.values()
            if r.started_at.startswith(date)
        ]

    def generate_daily_summary(self, date: str = None, output_path: str = None) -> str:
        """
        Generate a daily summary Excel report.

        Args:
            date: Date string (YYYY-MM-DD), defaults to today
            output_path: Output path for Excel file

        Returns:
            Path to generated report
        """
        if not EXCEL_AVAILABLE:
            logger.warning("openpyxl not available, cannot generate Excel summary")
            return ""

        if date is None:
            date = datetime.now().strftime('%Y-%m-%d')

        if output_path is None:
            output_path = str(self.output_dir / date / "summary" / f"daily_summary_{date}.xlsx")

        # Ensure directory exists
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        records = self.get_records_by_date(date)

        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Summary"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # Header
        ws['A1'] = f"TCO Extraction Daily Summary - {date}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:H1')

        # Summary metrics
        status = self.get_status()
        ws['A3'] = "Total Processed:"
        ws['B3'] = len(records)
        ws['C3'] = "Success Rate:"
        ws['D3'] = f"{status.success_rate:.1f}%"
        ws['E3'] = "Avg Confidence:"
        ws['F3'] = f"{status.average_confidence:.1%}"

        # Detail table
        headers = ['File', 'Vendor', 'Client', 'Status', 'Confidence', 'Items', 'Review Items', 'Completed']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, record in enumerate(records, 6):
            ws.cell(row=row_idx, column=1, value=record.file_name)
            ws.cell(row=row_idx, column=2, value=record.vendor)
            ws.cell(row=row_idx, column=3, value=record.client)
            ws.cell(row=row_idx, column=4, value=record.status)
            ws.cell(row=row_idx, column=5, value=f"{record.confidence:.1%}")
            ws.cell(row=row_idx, column=6, value=record.items_extracted)
            ws.cell(row=row_idx, column=7, value=record.items_need_review)
            ws.cell(row=row_idx, column=8, value=record.completed_at or "In Progress")

        # Adjust column widths
        for col in range(1, 9):
            ws.column_dimensions[chr(64 + col)].width = 15

        ws.column_dimensions['A'].width = 30  # File name
        ws.column_dimensions['H'].width = 20  # Completed

        wb.save(output_path)
        logger.info(f"Daily summary saved to: {output_path}")

        return output_path

    def print_status(self) -> None:
        """Print status to console."""
        status = self.get_status()

        print("\n" + "=" * 60)
        print("TCO EXTRACTION PIPELINE STATUS")
        print("=" * 60)
        print(f"Last Run: {status.last_run}")
        print(f"Total Processed: {status.total_processed}")
        print(f"Success Rate: {status.success_rate:.1f}%")
        print(f"Average Confidence: {status.average_confidence:.1%}")
        print()
        print(f"  Success: {status.success_count}")
        print(f"  Review Needed: {status.review_needed_count}")
        print(f"  Failed: {status.failed_count}")
        print(f"  Pending: {status.pending_count}")
        print()
        print("Recent Extractions:")
        for ext in status.recent_extractions[:5]:
            conf = f"{ext['confidence']:.0%}" if ext['confidence'] else "N/A"
            print(f"  - {ext['file']}: {ext['status']} ({conf})")
        print("=" * 60)


# Global instance
_status_tracker: Optional[StatusTracker] = None


def get_status_tracker(output_dir: str = None) -> StatusTracker:
    """Get or create the global status tracker."""
    global _status_tracker
    if _status_tracker is None:
        _status_tracker = StatusTracker(output_dir=output_dir)
    return _status_tracker


def print_pipeline_status():
    """Print current pipeline status to console."""
    tracker = get_status_tracker()
    tracker.print_status()
