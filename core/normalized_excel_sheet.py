"""
Normalized Comparison Excel Sheet Generator

Adds a 6-bucket "Normalized Comparison" sheet to TCO workbooks for
apples-to-apples vendor comparison.

This module creates a summary sheet that shows costs organized by the
Universal Cost Breakdown Structure (6 buckets) regardless of how each
vendor labels their costs.
"""

import json
import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, Optional, List
from dataclasses import dataclass

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from .cost_taxonomy import CostBucket, get_bucket_display_order

logger = logging.getLogger(__name__)


# =============================================================================
# STYLING CONSTANTS
# =============================================================================

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BUCKET_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
BUCKET_FONT = Font(bold=True, size=11)
TOTAL_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
TOTAL_FONT = Font(bold=True, color="FFFFFF", size=12)
CURRENCY_FORMAT = '_("$"* #,##0.00_);_("$"* (#,##0.00);_("$"* "-"??_);_(@_)'
PERCENT_FORMAT = '0.0%'

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


@dataclass
class NormalizedSheetConfig:
    """Configuration for the normalized comparison sheet"""
    sheet_name: str = "Normalized Comparison"
    include_line_items: bool = True
    include_metrics: bool = True
    include_bucket_summary: bool = True


def add_normalized_sheet(
    workbook_path: str,
    normalized_data: dict,
    config: NormalizedSheetConfig = None
) -> str:
    """
    Add a normalized comparison sheet to an existing TCO workbook.

    Args:
        workbook_path: Path to the existing Excel workbook
        normalized_data: Normalized extraction data (from normalized_extraction.json)
        config: Optional configuration for sheet layout

    Returns:
        Path to the updated workbook
    """
    if config is None:
        config = NormalizedSheetConfig()

    # Load existing workbook
    wb = load_workbook(workbook_path)

    # Remove existing sheet if present
    if config.sheet_name in wb.sheetnames:
        del wb[config.sheet_name]

    # Create new sheet
    ws = wb.create_sheet(config.sheet_name)

    # Extract proposal data
    proposal = normalized_data.get("proposal", normalized_data)
    vendor = proposal.get("vendor", "Unknown")
    client = proposal.get("client", "Unknown")
    bucket_totals = proposal.get("bucket_totals", {})
    normalized_metrics = proposal.get("normalized_metrics", {})
    line_items = proposal.get("line_items", [])

    current_row = 1

    # === HEADER SECTION ===
    current_row = _write_header_section(ws, vendor, client, proposal, current_row)
    current_row += 1

    # === BUCKET SUMMARY SECTION ===
    if config.include_bucket_summary:
        current_row = _write_bucket_summary(ws, bucket_totals, current_row)
        current_row += 1

    # === COST-PER-UNIT METRICS ===
    if config.include_metrics:
        current_row = _write_metrics_section(ws, normalized_metrics, proposal, current_row)
        current_row += 1

    # === LINE ITEM DETAILS BY BUCKET ===
    if config.include_line_items:
        current_row = _write_line_items_by_bucket(ws, line_items, current_row)

    # Auto-fit column widths
    _auto_fit_columns(ws)

    # Save workbook
    wb.save(workbook_path)
    logger.info(f"Added '{config.sheet_name}' sheet to: {workbook_path}")

    return workbook_path


def _write_header_section(ws, vendor: str, client: str, proposal: dict, start_row: int) -> int:
    """Write the header section with proposal info"""
    row = start_row

    # Title
    ws.cell(row=row, column=1, value="NORMALIZED COST COMPARISON")
    ws.cell(row=row, column=1).font = Font(bold=True, size=16)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1

    # Subtitle
    ws.cell(row=row, column=1, value="Universal 6-Bucket Cost Breakdown Structure")
    ws.cell(row=row, column=1).font = Font(italic=True, size=11, color="666666")
    row += 2

    # Proposal details
    details = [
        ("Vendor:", vendor),
        ("Client:", client),
        ("Contract Term:", f"{proposal.get('contract_term_years', 7)} years"),
        ("Annual CPI:", f"{proposal.get('annual_cpi_rate', 0.03):.1%}"),
        ("Normalized:", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]

    for label, value in details:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    return row + 1


def _write_bucket_summary(ws, bucket_totals: dict, start_row: int) -> int:
    """Write the 6-bucket cost summary section"""
    row = start_row

    # Section header
    ws.cell(row=row, column=1, value="COST SUMMARY BY BUCKET")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

    # Column headers
    headers = ["Cost Bucket", "Items", "Annual Cost", "7-Year TCO", "Required", "Optional"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER
    row += 1

    # Bucket data rows
    total_annual = 0
    total_7year = 0
    total_items = 0

    for bucket in get_bucket_display_order():
        data = bucket_totals.get(bucket.value, {})
        item_count = data.get("item_count", 0)
        annual = data.get("annual_cost", 0)
        total_7yr = data.get("total_7_year", 0)
        required = data.get("required_7_year", 0)
        optional = data.get("optional_7_year", 0)

        # Bucket name
        cell = ws.cell(row=row, column=1, value=bucket.value)
        cell.fill = BUCKET_FILL
        cell.font = BUCKET_FONT
        cell.border = THIN_BORDER

        # Item count
        ws.cell(row=row, column=2, value=item_count).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')

        # Annual cost
        cell = ws.cell(row=row, column=3, value=annual)
        cell.number_format = CURRENCY_FORMAT
        cell.border = THIN_BORDER

        # 7-Year TCO
        cell = ws.cell(row=row, column=4, value=total_7yr)
        cell.number_format = CURRENCY_FORMAT
        cell.border = THIN_BORDER

        # Required
        cell = ws.cell(row=row, column=5, value=required)
        cell.number_format = CURRENCY_FORMAT
        cell.border = THIN_BORDER

        # Optional
        cell = ws.cell(row=row, column=6, value=optional)
        cell.number_format = CURRENCY_FORMAT
        cell.border = THIN_BORDER

        total_annual += annual
        total_7year += total_7yr
        total_items += item_count
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=row, column=1).fill = TOTAL_FILL
    ws.cell(row=row, column=1).border = THIN_BORDER

    ws.cell(row=row, column=2, value=total_items).font = TOTAL_FONT
    ws.cell(row=row, column=2).fill = TOTAL_FILL
    ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=2).border = THIN_BORDER

    cell = ws.cell(row=row, column=3, value=total_annual)
    cell.number_format = CURRENCY_FORMAT
    cell.font = TOTAL_FONT
    cell.fill = TOTAL_FILL
    cell.border = THIN_BORDER

    cell = ws.cell(row=row, column=4, value=total_7year)
    cell.number_format = CURRENCY_FORMAT
    cell.font = TOTAL_FONT
    cell.fill = TOTAL_FILL
    cell.border = THIN_BORDER

    for col in [5, 6]:
        ws.cell(row=row, column=col).fill = TOTAL_FILL
        ws.cell(row=row, column=col).border = THIN_BORDER

    return row + 2


def _write_metrics_section(ws, metrics: dict, proposal: dict, start_row: int) -> int:
    """Write the cost-per-unit metrics section"""
    row = start_row

    # Section header
    ws.cell(row=row, column=1, value="COST-PER-UNIT METRICS")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1

    # Column headers
    headers = ["Metric", "Value", "Formula", "Use Case"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER
    row += 1

    # Institution params
    params = proposal.get("institution_params", {})

    metric_details = [
        ("Cost per Account", metrics.get("Cost per Account", 0),
         "7-Year TCO / Accounts", f"Based on {params.get('total_accounts', 50000):,} accounts"),
        ("Cost per User", metrics.get("Cost per User", 0),
         "7-Year TCO / Users", f"Based on {params.get('total_users', 500):,} users"),
        ("Cost per Transaction", metrics.get("Cost per Transaction", 0),
         "7-Year TCO / Annual Txns", f"Based on {params.get('annual_transactions', 1000000):,} txns/year"),
        ("Cost per Month", metrics.get("Cost per Month", 0),
         "7-Year TCO / Contract Months", f"{params.get('contract_months', 84)} months"),
        ("Cost per Asset ($M)", metrics.get("Cost per Asset", 0),
         "7-Year TCO / Total Assets", f"Based on ${params.get('total_assets_millions', 500)}M assets"),
    ]

    for name, value, formula, use_case in metric_details:
        ws.cell(row=row, column=1, value=name).border = THIN_BORDER

        cell = ws.cell(row=row, column=2, value=value)
        cell.number_format = CURRENCY_FORMAT
        cell.border = THIN_BORDER

        ws.cell(row=row, column=3, value=formula).border = THIN_BORDER
        ws.cell(row=row, column=4, value=use_case).border = THIN_BORDER
        row += 1

    return row + 1


def _write_line_items_by_bucket(ws, line_items: list, start_row: int) -> int:
    """Write detailed line items organized by bucket"""
    row = start_row

    # Section header
    ws.cell(row=row, column=1, value="LINE ITEM DETAILS BY BUCKET")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1

    # Column headers
    headers = ["Solution", "Category", "Fee Type", "Annual Cost", "7-Year TCO",
               "Required", "Variable", "Confidence"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER
    row += 1

    # Group items by bucket
    for bucket in get_bucket_display_order():
        bucket_items = [i for i in line_items if i.get("level_1_bucket") == bucket.value]

        if not bucket_items:
            continue

        # Bucket header row
        ws.cell(row=row, column=1, value=bucket.value)
        ws.cell(row=row, column=1).font = BUCKET_FONT
        ws.cell(row=row, column=1).fill = BUCKET_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 1

        # Items
        for item in bucket_items:
            ws.cell(row=row, column=1, value=item.get("solution_name", "")).border = THIN_BORDER
            ws.cell(row=row, column=2, value=item.get("level_2_category", "")).border = THIN_BORDER
            ws.cell(row=row, column=3, value=item.get("original_fee_type", "")).border = THIN_BORDER

            cell = ws.cell(row=row, column=4, value=item.get("annual_cost", 0))
            cell.number_format = CURRENCY_FORMAT
            cell.border = THIN_BORDER

            cell = ws.cell(row=row, column=5, value=item.get("total_7_year_cost", 0))
            cell.number_format = CURRENCY_FORMAT
            cell.border = THIN_BORDER

            ws.cell(row=row, column=6, value="Yes" if not item.get("is_optional") else "No").border = THIN_BORDER
            ws.cell(row=row, column=7, value="Yes" if item.get("is_variable") else "No").border = THIN_BORDER

            cell = ws.cell(row=row, column=8, value=item.get("confidence_score", 0))
            cell.number_format = PERCENT_FORMAT
            cell.border = THIN_BORDER

            row += 1

        row += 1  # Space between buckets

    return row


def _auto_fit_columns(ws):
    """Auto-fit column widths based on content"""
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0

        for row in range(1, ws.max_row + 1):
            try:
                cell = ws.cell(row=row, column=col_idx)
                if cell.value and not isinstance(cell, type(None)):
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass

        # Set minimum and maximum widths
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def create_comparison_workbook(
    normalized_proposals: List[dict],
    output_path: str
) -> str:
    """
    Create a new Excel workbook comparing multiple vendors using 6-bucket structure.

    Args:
        normalized_proposals: List of normalized proposal data
        output_path: Path for output Excel file

    Returns:
        Path to created workbook
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendor Comparison"

    row = 1

    # Title
    ws.cell(row=row, column=1, value="VENDOR COMPARISON - 6-BUCKET ANALYSIS")
    ws.cell(row=row, column=1).font = Font(bold=True, size=16)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(normalized_proposals) + 2)
    row += 2

    # Get vendor names
    vendors = [p.get("proposal", p).get("vendor", f"Vendor {i+1}")
               for i, p in enumerate(normalized_proposals)]

    # Header row
    ws.cell(row=row, column=1, value="Cost Bucket").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=1).border = THIN_BORDER

    for col, vendor in enumerate(vendors, 2):
        cell = ws.cell(row=row, column=col, value=vendor)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    # Winner column
    cell = ws.cell(row=row, column=len(vendors) + 2, value="Lowest Cost")
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = THIN_BORDER
    row += 1

    # Bucket rows
    for bucket in get_bucket_display_order():
        ws.cell(row=row, column=1, value=bucket.value).border = THIN_BORDER
        ws.cell(row=row, column=1).fill = BUCKET_FILL
        ws.cell(row=row, column=1).font = BUCKET_FONT

        costs = []
        for col, proposal in enumerate(normalized_proposals, 2):
            prop_data = proposal.get("proposal", proposal)
            bucket_data = prop_data.get("bucket_totals", {}).get(bucket.value, {})
            cost = bucket_data.get("total_7_year", 0)
            costs.append((vendors[col-2], cost))

            cell = ws.cell(row=row, column=col, value=cost)
            cell.number_format = CURRENCY_FORMAT
            cell.border = THIN_BORDER

        # Winner
        if costs:
            winner = min(costs, key=lambda x: x[1] if x[1] > 0 else float('inf'))
            ws.cell(row=row, column=len(vendors) + 2, value=winner[0]).border = THIN_BORDER

        row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL 7-YEAR TCO").font = TOTAL_FONT
    ws.cell(row=row, column=1).fill = TOTAL_FILL
    ws.cell(row=row, column=1).border = THIN_BORDER

    totals = []
    for col, proposal in enumerate(normalized_proposals, 2):
        prop_data = proposal.get("proposal", proposal)
        total = sum(
            prop_data.get("bucket_totals", {}).get(b.value, {}).get("total_7_year", 0)
            for b in get_bucket_display_order()
        )
        totals.append((vendors[col-2], total))

        cell = ws.cell(row=row, column=col, value=total)
        cell.number_format = CURRENCY_FORMAT
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER

    # Overall winner
    if totals:
        winner = min(totals, key=lambda x: x[1])
        cell = ws.cell(row=row, column=len(vendors) + 2, value=winner[0])
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER

    _auto_fit_columns(ws)

    # Save
    wb.save(output_path)
    logger.info(f"Created comparison workbook: {output_path}")

    return output_path
