#!/usr/bin/env python3
"""
TCO Comparison Tool

Compares source proposals with populated TCO output to verify accuracy.
Creates a side-by-side comparison showing:
- What was in the source
- What ended up in the TCO
- Any discrepancies
"""

from openpyxl import load_workbook
from docx import Document
import pandas as pd
from extractors import extract_fis_proposal, extract_jack_henry_proposal
from config import TCO_COLUMNS


def compare_fis_to_tco(fis_proposal: str, tco_file: str, term: str = '7_year'):
    """
    Compare FIS proposal with populated TCO template
    
    Returns:
        DataFrame with comparison results
    """
    print("\n" + "="*70)
    print("COMPARING FIS PROPOSAL TO TCO OUTPUT")
    print("="*70 + "\n")
    
    # Extract FIS data
    print("Extracting FIS proposal data...")
    fis_data = extract_fis_proposal(fis_proposal)
    
    # Load TCO file
    print("Loading TCO output...")
    wb = load_workbook(tco_file, data_only=True)
    sheet = wb['Line Items']
    col_map = TCO_COLUMNS['FIS']
    
    comparisons = []
    
    # Compare bundle pricing
    print("\nComparing Bundle Pricing...")
    for year, prices in fis_data['bundle_pricing'].items():
        source_value = prices.get(term, 0)
        
        # Find this in TCO (look for solution name matching the year)
        year_num = year.split()[-1]  # Extract "1" from "Year 1"
        
        # Search for bundle item in TCO
        found = False
        for row in range(7, 50):
            solution_name = sheet[f"{col_map['solution_name']}{row}"].value
            if solution_name and 'CORE PROCESSING' in str(solution_name) and year in str(solution_name):
                tco_value = sheet[f"{col_map['proposal']}{row}"].value or 0
                
                match = abs(source_value - tco_value) < 0.01
                comparisons.append({
                    'Category': 'Bundle',
                    'Item': year,
                    'Source': f"${source_value:,.2f}",
                    'TCO': f"${tco_value:,.2f}",
                    'Match': 'Y' if match else 'N',
                    'Difference': f"${abs(source_value - tco_value):,.2f}"
                })
                found = True
                break
        
        if not found and source_value > 0:
            comparisons.append({
                'Category': 'Bundle',
                'Item': year,
                'Source': f"${source_value:,.2f}",
                'TCO': 'NOT FOUND',
                'Match': 'N',
                'Difference': 'N/A'
            })
    
    # Compare monthly fees (sample first 5)
    print("\nComparing Monthly Fees (sample)...")
    for idx, fee in enumerate(fis_data['monthly_fees'][:5]):
        comparisons.append({
            'Category': 'Monthly Fee',
            'Item': fee['solution_name'][:40],  # Truncate long names
            'Source': f"${fee['monthly_fee']:,.2f}",
            'TCO': 'Check manually',
            'Match': '?',
            'Difference': 'N/A'
        })
    
    # Compare one-time credits
    print("\nComparing One-Time Credits...")
    for desc, values in fis_data['one_time_credits'].items():
        source_value = values.get(term, 0)
        if source_value != 0:
            comparisons.append({
                'Category': 'One-Time',
                'Item': desc[:40],
                'Source': f"${source_value:,.2f}",
                'TCO': 'Check manually',
                'Match': '?',
                'Difference': 'N/A'
            })
    
    df = pd.DataFrame(comparisons)
    
    # Print summary
    matches = df[df['Match'] == 'Y'].shape[0]
    mismatches = df[df['Match'] == 'N'].shape[0]

    print("\n" + "="*70)
    print("COMPARISON SUMMARY")
    print("="*70)
    print(f"[OK] Matches: {matches}")
    print(f"[X] Mismatches: {mismatches}")
    print(f"? Manual check needed: {df[df['Match'] == '?'].shape[0]}")
    
    return df


def compare_jh_to_tco(jh_proposal: str, tco_file: str, scenario: str = 'Proposal_1'):
    """
    Compare Jack Henry proposal with populated TCO template
    
    Returns:
        DataFrame with comparison results
    """
    print("\n" + "="*70)
    print("COMPARING JACK HENRY PROPOSAL TO TCO OUTPUT")
    print("="*70 + "\n")
    
    # Extract JH data
    print("Extracting Jack Henry proposal data...")
    jh_data = extract_jack_henry_proposal(jh_proposal)
    
    # Load TCO file
    print("Loading TCO output...")
    wb = load_workbook(tco_file, data_only=True)
    sheet = wb['Line Items']
    col_map = TCO_COLUMNS['JACK_HENRY']
    
    # Find the scenario
    scenario_data = None
    for s in jh_data['scenarios']:
        if s['scenario_name'] == scenario:
            scenario_data = s
            break
    
    if not scenario_data:
        print(f"Error: Scenario {scenario} not found")
        return pd.DataFrame()
    
    comparisons = []
    
    # Compare sample products (first 10)
    print(f"\nComparing Products from {scenario} (sample)...")
    
    for product in scenario_data['products'][:10]:
        # Calculate expected monthly from source
        source_monthly = product['monthly_net']
        
        comparisons.append({
            'Product': product['product_description'][:40],
            'Source Monthly': f"${source_monthly:,.2f}",
            'Source Install': f"${product['install_net']:,.2f}",
            'Source License': f"${product['license_net']:,.2f}",
            'TCO Status': 'Check manually',
            'Notes': f"Family: {product.get('product_family', 'N/A')}"
        })
    
    df = pd.DataFrame(comparisons)
    
    print("\n" + "="*70)
    print(f"Showing sample of {len(comparisons)} products from {scenario}")
    print(f"Total products in scenario: {len(scenario_data['products'])}")
    print("="*70)
    
    return df


def create_comparison_report(fis_proposal: str = None, jh_proposal: str = None, 
                            tco_file: str = None, output_excel: str = "comparison_report.xlsx"):
    """
    Create comprehensive comparison report in Excel
    """
    print("\nCreating Comparison Report...\n")
    
    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        
        if fis_proposal and tco_file:
            df_fis = compare_fis_to_tco(fis_proposal, tco_file)
            df_fis.to_excel(writer, sheet_name='FIS Comparison', index=False)
        
        if jh_proposal and tco_file:
            df_jh = compare_jh_to_tco(jh_proposal, tco_file)
            df_jh.to_excel(writer, sheet_name='JH Comparison', index=False)
    
    print(f"\n[DONE] Comparison report saved to: {output_excel}")
    return output_excel


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Compare source proposals with TCO output')
    parser.add_argument('--fis', type=str, help='Path to FIS proposal')
    parser.add_argument('--jh', type=str, help='Path to Jack Henry proposal')
    parser.add_argument('--tco', type=str, required=True, help='Path to populated TCO file')
    parser.add_argument('--output', type=str, default='comparison_report.xlsx', 
                       help='Output Excel report path')
    
    args = parser.parse_args()
    
    create_comparison_report(args.fis, args.jh, args.tco, args.output)
