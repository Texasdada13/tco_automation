"""
Multi-Vendor Comparison Generator

Generates a standalone Excel file comparing N unique vendors for a specific client.
Automatically finds all vendor extractions for the client and creates a side-by-side comparison.

Features (Phase 4 Enhanced):
- Bucket-level comparison (original)
- Product-by-product comparison across vendors (NEW)
- Gap detection (product in one vendor but not others) (NEW)
- Unmatched products tab (NEW)
- Match confidence indicators (NEW)

Usage:
    python generate_comparison.py <client_name>

Examples:
    python generate_comparison.py "echelon_bank"
    python generate_comparison.py "liberty_capital_bank"
    python generate_comparison.py "heritage"
"""

import sys
import json
import re
from pathlib import Path
from datetime import datetime
from glob import glob
from collections import defaultdict

from dotenv import load_dotenv
load_dotenv()  # Load environment variables from .env file

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Product matching imports
from core.product_matcher import ProductMatcher, MatchResult, MatchMethod


def parse_vendor_from_filename(filename: str) -> tuple:
    """
    Parse filename to extract client name and vendor type.

    Examples:
        'echelon_bank_fis_1_extraction_ai' -> ('echelon_bank', 'FIS')
        'liberty_capital_bank_csi_extraction_ai' -> ('liberty_capital_bank', 'CSI')
    """
    known_vendors = ['FIS', 'CSI', 'JH', 'Fiserv', 'Finastra', 'Q2', 'Temenos']

    # Remove extraction suffix
    name = filename.replace('_extraction_ai', '').replace('_raw_extraction', '')

    # Strip numeric suffix (e.g., _1, _2, _retry, _v2, _final)
    cleaned_name = re.sub(r'_(\d+|retry|v\d+|final)$', '', name, flags=re.IGNORECASE)

    # Check if name ends with a known vendor
    for vendor in known_vendors:
        if cleaned_name.upper().endswith(f"_{vendor.upper()}"):
            client = cleaned_name[:-(len(vendor) + 1)]
            return (client.lower(), vendor.upper())

    # Fallback: assume last part after underscore is vendor
    parts = cleaned_name.rsplit('_', 1)
    if len(parts) == 2:
        return (parts[0].lower(), parts[1].upper())

    return (cleaned_name.lower(), 'Unknown')


def find_client_vendors(client_name: str) -> dict:
    """
    Find all unique vendors for a client, keeping most recent file per vendor.

    Args:
        client_name: Client name pattern to search for

    Returns:
        Dict mapping vendor name -> json file path (most recent)
    """
    # Search for extraction files matching the client pattern
    patterns = [
        f"Extracted JSON/{client_name}*_extraction_ai.json",
        f"Extracted JSON/{client_name.lower()}*_extraction_ai.json",
        f"Extracted JSON/{client_name.upper()}*_extraction_ai.json",
    ]

    matching_files = []
    for pattern in patterns:
        matching_files.extend(glob(pattern))

    # Remove duplicates
    matching_files = list(set(matching_files))

    # Group by vendor, keeping track of file modification times
    vendor_files = defaultdict(list)

    for json_path in matching_files:
        filename = Path(json_path).stem
        parsed_client, vendor = parse_vendor_from_filename(filename)

        # Check if this file belongs to the requested client
        if client_name.lower() in parsed_client.lower():
            mtime = Path(json_path).stat().st_mtime
            vendor_files[vendor].append((json_path, mtime))

    # Keep only the most recent file per vendor
    unique_vendors = {}
    for vendor, files in vendor_files.items():
        # Sort by modification time (newest first)
        files.sort(key=lambda x: x[1], reverse=True)
        unique_vendors[vendor] = files[0][0]  # Keep newest

    return unique_vendors


def normalize_proposal(json_path: str) -> dict:
    """
    Load and normalize a proposal extraction into 6-bucket structure.
    """
    from core import CostNormalizer

    with open(json_path, 'r', encoding='utf-8') as f:
        extraction_data = json.load(f)

    normalizer = CostNormalizer()
    proposal = normalizer.normalize_proposal(
        extraction_data=extraction_data,
        source_file=json_path
    )

    return proposal


def match_products_across_vendors(vendors: dict, matcher: ProductMatcher) -> dict:
    """
    Match all line items to canonical categories across all vendors.

    Args:
        vendors: Dict mapping vendor name -> NormalizedProposal
        matcher: ProductMatcher instance

    Returns:
        Dict with structure:
        {
            'by_category': {category_key: {vendor: [matched_items]}},
            'unmatched': {vendor: [unmatched_items]},
            'match_results': {vendor: [MatchResult objects]},
            'statistics': {vendor: stats_dict}
        }
    """
    result = {
        'by_category': defaultdict(lambda: defaultdict(list)),
        'unmatched': defaultdict(list),
        'match_results': {},
        'statistics': {}
    }

    for vendor_name, proposal in vendors.items():
        vendor_results = []

        for item in proposal.line_items:
            # Create a dict for the matcher
            item_dict = {
                'solution_name': item.solution_name,
                'vendor': vendor_name
            }

            match_result = matcher.match(item.solution_name, vendor_name)
            vendor_results.append(match_result)

            # Store matched items by category
            if match_result.canonical_category:
                result['by_category'][match_result.canonical_category][vendor_name].append({
                    'item': item,
                    'match_result': match_result
                })
            else:
                result['unmatched'][vendor_name].append({
                    'item': item,
                    'match_result': match_result
                })

        result['match_results'][vendor_name] = vendor_results
        result['statistics'][vendor_name] = matcher.get_match_statistics(vendor_results)

    return result


def create_product_comparison_sheet(wb: Workbook, matched_data: dict, vendor_names: list, matcher: ProductMatcher):
    """
    Create the Product-by-Product comparison sheet.

    Shows equivalent products side-by-side across vendors, grouped by canonical category.
    """
    ws = wb.create_sheet("Product Comparison")

    # Styling
    DARK_BLUE = "1F4E79"
    MEDIUM_BLUE = "2E75B6"
    LIGHT_BLUE = "BDD7EE"
    LIGHT_GRAY = "F2F2F2"
    GREEN = "C6EFCE"
    DARK_GREEN = "006100"
    YELLOW = "FFF2CC"
    ORANGE = "FCE4D6"

    TITLE_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    HEADER_FILL = PatternFill(start_color=MEDIUM_BLUE, end_color=MEDIUM_BLUE, fill_type="solid")
    CATEGORY_FILL = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    ALT_ROW_FILL = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
    WINNER_FILL = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
    GAP_FILL = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")

    TITLE_FONT = Font(bold=True, color="FFFFFF", size=16)
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    CATEGORY_FONT = Font(bold=True, color=DARK_BLUE, size=12)
    NORMAL_FONT = Font(size=10)
    BOLD_FONT = Font(bold=True, size=10)
    GAP_FONT = Font(italic=True, color="996600", size=10)

    CURRENCY_FORMAT = '_("$"* #,##0_);_("$"* (#,##0);_("$"* "-"_);_(@_)'

    THIN_BORDER = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    num_vendors = len(vendor_names)
    total_cols = 1 + (num_vendors * 2) + 1  # Category + (Product, Cost) per vendor + Winner

    row = 1

    # Title
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    cell = ws.cell(row=row, column=1, value="PRODUCT-BY-PRODUCT COMPARISON")
    cell.fill = TITLE_FILL
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 28
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    cell = ws.cell(row=row, column=1, value="Equivalent products matched across vendors using canonical product categories")
    cell.font = Font(italic=True, color="666666", size=11)
    row += 2

    # Header row
    headers = ["Product Category"]
    for vendor in vendor_names:
        headers.extend([f"{vendor} Product", f"{vendor} 7-Yr Cost"])
    headers.append("Lowest Cost")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 30
    row += 1

    # Sort categories for display
    by_category = matched_data['by_category']
    sorted_categories = sorted(by_category.keys())

    category_totals = {vendor: 0 for vendor in vendor_names}
    gaps_detected = []

    for idx, category_key in enumerate(sorted_categories):
        vendor_items = by_category[category_key]

        # Get canonical name
        cat_info = matcher.get_category_info(category_key)
        category_name = cat_info.get('canonical_name', category_key) if cat_info else category_key

        # Collect costs per vendor for this category
        vendor_costs = {}
        vendor_products = {}

        for vendor in vendor_names:
            items = vendor_items.get(vendor, [])
            if items:
                # Sum all items in this category for this vendor
                total_cost = sum(item['item'].total_7_year_cost for item in items)
                product_names = [item['item'].solution_name for item in items]
                vendor_costs[vendor] = total_cost
                vendor_products[vendor] = product_names
                category_totals[vendor] += total_cost
            else:
                vendor_costs[vendor] = None
                vendor_products[vendor] = None

        # Detect gaps
        has_gap = any(v is None for v in vendor_costs.values()) and any(v is not None for v in vendor_costs.values())
        if has_gap:
            gaps_detected.append({
                'category': category_name,
                'present_in': [v for v in vendor_names if vendor_costs[v] is not None],
                'missing_from': [v for v in vendor_names if vendor_costs[v] is None]
            })

        row_fill = ALT_ROW_FILL if idx % 2 == 0 else None

        # Category name cell
        cell = ws.cell(row=row, column=1, value=category_name)
        cell.font = BOLD_FONT
        cell.border = THIN_BORDER
        if row_fill:
            cell.fill = row_fill

        # Find winner (lowest cost among those that have it)
        valid_costs = {v: c for v, c in vendor_costs.items() if c is not None and c > 0}
        winner = min(valid_costs.keys(), key=lambda v: valid_costs[v]) if valid_costs else None

        col = 2
        for vendor in vendor_names:
            cost = vendor_costs[vendor]
            products = vendor_products[vendor]

            # Product name cell
            if products:
                # Show first product, truncate if multiple
                if len(products) == 1:
                    product_display = products[0][:40]
                else:
                    product_display = f"{products[0][:30]}... (+{len(products)-1})"
                cell = ws.cell(row=row, column=col, value=product_display)
                cell.font = NORMAL_FONT
            else:
                cell = ws.cell(row=row, column=col, value="[Not offered]")
                cell.font = GAP_FONT
                cell.fill = GAP_FILL
            cell.border = THIN_BORDER
            if row_fill and not (products is None):
                cell.fill = row_fill

            # Cost cell
            cell = ws.cell(row=row, column=col + 1, value=cost if cost else 0)
            cell.number_format = CURRENCY_FORMAT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='right')

            if cost is None:
                cell.fill = GAP_FILL
            elif vendor == winner:
                cell.fill = WINNER_FILL
                cell.font = Font(bold=True, color=DARK_GREEN)
            elif row_fill:
                cell.fill = row_fill

            col += 2

        # Winner cell
        cell = ws.cell(row=row, column=col, value=winner if winner else "N/A")
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        cell.font = BOLD_FONT
        if row_fill:
            cell.fill = row_fill

        row += 1

    # Total row
    row += 1
    cell = ws.cell(row=row, column=1, value="TOTAL (Matched Products)")
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.border = THIN_BORDER

    col = 2
    category_winner = min(category_totals.keys(), key=lambda v: category_totals[v]) if category_totals else None

    for vendor in vendor_names:
        cell = ws.cell(row=row, column=col, value="")
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

        cell = ws.cell(row=row, column=col + 1, value=category_totals[vendor])
        cell.number_format = CURRENCY_FORMAT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='right')

        if vendor == category_winner:
            cell.fill = WINNER_FILL
            cell.font = Font(bold=True, color=DARK_GREEN, size=11)
        else:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

        col += 2

    cell = ws.cell(row=row, column=col, value=category_winner if category_winner else "")
    cell.fill = WINNER_FILL
    cell.font = Font(bold=True, color=DARK_GREEN)
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='center')

    # Column widths
    ws.column_dimensions['A'].width = 35
    col_idx = 2
    for _ in vendor_names:
        ws.column_dimensions[get_column_letter(col_idx)].width = 30
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = 14
        col_idx += 2
    ws.column_dimensions[get_column_letter(col_idx)].width = 12

    ws.freeze_panes = 'A5'

    return gaps_detected


def create_gaps_sheet(wb: Workbook, gaps_detected: list, matched_data: dict, vendor_names: list, matcher: ProductMatcher):
    """
    Create the Gap Detection sheet.

    Shows products/categories present in some vendors but missing from others.
    """
    ws = wb.create_sheet("Gap Analysis")

    # Styling
    DARK_BLUE = "1F4E79"
    MEDIUM_BLUE = "2E75B6"
    YELLOW = "FFF2CC"
    LIGHT_GRAY = "F2F2F2"

    TITLE_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    HEADER_FILL = PatternFill(start_color=MEDIUM_BLUE, end_color=MEDIUM_BLUE, fill_type="solid")
    GAP_FILL = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
    ALT_ROW_FILL = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")

    TITLE_FONT = Font(bold=True, color="FFFFFF", size=16)
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    NORMAL_FONT = Font(size=10)
    BOLD_FONT = Font(bold=True, size=10)

    THIN_BORDER = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    row = 1

    # Title
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1, value="GAP ANALYSIS")
    cell.fill = TITLE_FILL
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 28
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1, value="Products/categories offered by some vendors but not others")
    cell.font = Font(italic=True, color="666666", size=11)
    row += 2

    if not gaps_detected:
        ws.cell(row=row, column=1, value="No gaps detected - all vendors offer equivalent products in all categories.")
        ws.cell(row=row, column=1).font = Font(italic=True, size=11)
        return

    # Summary
    cell = ws.cell(row=row, column=1, value=f"Found {len(gaps_detected)} categories with coverage gaps:")
    cell.font = BOLD_FONT
    row += 2

    # Headers
    headers = ["Category", "Present In", "Missing From", "Cost Impact"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
    row += 1

    by_category = matched_data['by_category']

    for idx, gap in enumerate(gaps_detected):
        row_fill = ALT_ROW_FILL if idx % 2 == 0 else None

        # Category
        cell = ws.cell(row=row, column=1, value=gap['category'])
        cell.font = BOLD_FONT
        cell.border = THIN_BORDER
        if row_fill:
            cell.fill = row_fill

        # Present in
        cell = ws.cell(row=row, column=2, value=", ".join(gap['present_in']))
        cell.border = THIN_BORDER
        if row_fill:
            cell.fill = row_fill

        # Missing from
        cell = ws.cell(row=row, column=3, value=", ".join(gap['missing_from']))
        cell.fill = GAP_FILL
        cell.border = THIN_BORDER

        # Cost impact - show what the present vendors are charging
        # Find category key from name
        cat_key = None
        for key in by_category.keys():
            cat_info = matcher.get_category_info(key)
            if cat_info and cat_info.get('canonical_name') == gap['category']:
                cat_key = key
                break

        if cat_key:
            costs = []
            for vendor in gap['present_in']:
                items = by_category[cat_key].get(vendor, [])
                total = sum(item['item'].total_7_year_cost for item in items)
                costs.append(f"{vendor}: ${total:,.0f}")
            cost_str = "; ".join(costs)
        else:
            cost_str = "N/A"

        cell = ws.cell(row=row, column=4, value=cost_str)
        cell.border = THIN_BORDER
        if row_fill:
            cell.fill = row_fill

        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 40


def create_unmatched_sheet(wb: Workbook, matched_data: dict, vendor_names: list):
    """
    Create the Unmatched Products sheet.

    Shows products that couldn't be matched to any canonical category.
    """
    ws = wb.create_sheet("Unmatched Products")

    # Styling
    DARK_BLUE = "1F4E79"
    MEDIUM_BLUE = "2E75B6"
    ORANGE = "FCE4D6"
    LIGHT_GRAY = "F2F2F2"

    TITLE_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    HEADER_FILL = PatternFill(start_color=MEDIUM_BLUE, end_color=MEDIUM_BLUE, fill_type="solid")
    UNMATCHED_FILL = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
    ALT_ROW_FILL = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")

    TITLE_FONT = Font(bold=True, color="FFFFFF", size=16)
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    NORMAL_FONT = Font(size=10)
    BOLD_FONT = Font(bold=True, size=10)

    CURRENCY_FORMAT = '_("$"* #,##0_);_("$"* (#,##0);_("$"* "-"_);_(@_)'

    THIN_BORDER = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    row = 1

    # Title
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    cell = ws.cell(row=row, column=1, value="UNMATCHED PRODUCTS")
    cell.fill = TITLE_FILL
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 28
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    cell = ws.cell(row=row, column=1, value="Products that could not be matched to a canonical category (may need manual review)")
    cell.font = Font(italic=True, color="666666", size=11)
    row += 2

    unmatched = matched_data['unmatched']
    total_unmatched = sum(len(items) for items in unmatched.values())

    if total_unmatched == 0:
        ws.cell(row=row, column=1, value="All products were successfully matched to canonical categories.")
        ws.cell(row=row, column=1).font = Font(italic=True, size=11)
        return

    # Summary by vendor
    cell = ws.cell(row=row, column=1, value=f"Total unmatched products: {total_unmatched}")
    cell.font = BOLD_FONT
    row += 1

    for vendor in vendor_names:
        count = len(unmatched.get(vendor, []))
        ws.cell(row=row, column=1, value=f"  {vendor}: {count} unmatched")
    row += 2

    # Headers
    headers = ["Vendor", "Product Name", "7-Year Cost", "Cost Bucket", "Fuzzy Suggestion"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
    row += 1

    idx = 0
    for vendor in vendor_names:
        items = unmatched.get(vendor, [])
        for item_data in items:
            item = item_data['item']
            match_result = item_data['match_result']

            row_fill = ALT_ROW_FILL if idx % 2 == 0 else None

            # Vendor
            cell = ws.cell(row=row, column=1, value=vendor)
            cell.font = BOLD_FONT
            cell.border = THIN_BORDER
            if row_fill:
                cell.fill = row_fill

            # Product name
            cell = ws.cell(row=row, column=2, value=item.solution_name)
            cell.border = THIN_BORDER
            cell.fill = UNMATCHED_FILL

            # Cost
            cell = ws.cell(row=row, column=3, value=item.total_7_year_cost)
            cell.number_format = CURRENCY_FORMAT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='right')
            if row_fill:
                cell.fill = row_fill

            # Bucket
            cell = ws.cell(row=row, column=4, value=item.level_1_bucket)
            cell.border = THIN_BORDER
            if row_fill:
                cell.fill = row_fill

            # Fuzzy suggestion
            if match_result.match_method == MatchMethod.FUZZY:
                suggestion = f"{match_result.canonical_name} ({match_result.confidence:.0f}%)"
            else:
                suggestion = "No suggestion"
            cell = ws.cell(row=row, column=5, value=suggestion)
            cell.border = THIN_BORDER
            cell.font = Font(italic=True, size=10)
            if row_fill:
                cell.fill = row_fill

            row += 1
            idx += 1

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 35


def create_match_statistics_sheet(wb: Workbook, matched_data: dict, vendor_names: list):
    """
    Create the Match Statistics sheet.

    Shows match quality metrics for transparency.
    """
    ws = wb.create_sheet("Match Statistics")

    # Styling
    DARK_BLUE = "1F4E79"
    MEDIUM_BLUE = "2E75B6"
    GREEN = "C6EFCE"
    YELLOW = "FFF2CC"
    LIGHT_GRAY = "F2F2F2"

    TITLE_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    HEADER_FILL = PatternFill(start_color=MEDIUM_BLUE, end_color=MEDIUM_BLUE, fill_type="solid")
    GOOD_FILL = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
    WARN_FILL = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")

    TITLE_FONT = Font(bold=True, color="FFFFFF", size=16)
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    NORMAL_FONT = Font(size=10)
    BOLD_FONT = Font(bold=True, size=10)

    PERCENT_FORMAT = '0.0%'

    THIN_BORDER = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    row = 1

    # Title
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell = ws.cell(row=row, column=1, value="PRODUCT MATCHING STATISTICS")
    cell.fill = TITLE_FILL
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 28
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell = ws.cell(row=row, column=1, value="Quality metrics for product matching - higher match rates indicate better cross-vendor comparability")
    cell.font = Font(italic=True, color="666666", size=11)
    row += 2

    # Headers
    headers = ["Vendor", "Total Items", "Exact Match", "Fuzzy Match", "Unmatched", "Match Rate", "Auto-Approved"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
    row += 1

    stats = matched_data['statistics']

    totals = {'total': 0, 'exact': 0, 'fuzzy': 0, 'unmatched': 0, 'auto_approved': 0}

    for idx, vendor in enumerate(vendor_names):
        vendor_stats = stats.get(vendor, {})
        row_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid") if idx % 2 == 0 else None

        # Vendor
        cell = ws.cell(row=row, column=1, value=vendor)
        cell.font = BOLD_FONT
        cell.border = THIN_BORDER
        if row_fill:
            cell.fill = row_fill

        # Total
        total = vendor_stats.get('total', 0)
        cell = ws.cell(row=row, column=2, value=total)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        if row_fill:
            cell.fill = row_fill
        totals['total'] += total

        # Exact
        exact = vendor_stats.get('exact_matches', 0)
        cell = ws.cell(row=row, column=3, value=exact)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        if row_fill:
            cell.fill = row_fill
        totals['exact'] += exact

        # Fuzzy
        fuzzy = vendor_stats.get('fuzzy_matches', 0)
        cell = ws.cell(row=row, column=4, value=fuzzy)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        if row_fill:
            cell.fill = row_fill
        totals['fuzzy'] += fuzzy

        # Unmatched
        unmatched = vendor_stats.get('unmatched', 0)
        cell = ws.cell(row=row, column=5, value=unmatched)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        if row_fill:
            cell.fill = row_fill
        totals['unmatched'] += unmatched

        # Match rate
        match_rate = vendor_stats.get('match_rate', 0) / 100
        cell = ws.cell(row=row, column=6, value=match_rate)
        cell.number_format = PERCENT_FORMAT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        cell.font = BOLD_FONT
        if match_rate >= 0.8:
            cell.fill = GOOD_FILL
        elif match_rate >= 0.5:
            cell.fill = WARN_FILL
        elif row_fill:
            cell.fill = row_fill

        # Auto-approved
        auto = vendor_stats.get('auto_approved', 0)
        cell = ws.cell(row=row, column=7, value=auto)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        if row_fill:
            cell.fill = row_fill
        totals['auto_approved'] += auto

        row += 1

    # Total row
    row += 1
    cell = ws.cell(row=row, column=1, value="TOTAL")
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.border = THIN_BORDER

    for col, key in enumerate(['total', 'exact', 'fuzzy', 'unmatched'], 2):
        cell = ws.cell(row=row, column=col, value=totals[key])
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    overall_rate = (totals['exact'] + totals['fuzzy']) / totals['total'] if totals['total'] > 0 else 0
    cell = ws.cell(row=row, column=6, value=overall_rate)
    cell.number_format = PERCENT_FORMAT
    cell.fill = GOOD_FILL if overall_rate >= 0.8 else WARN_FILL
    cell.font = Font(bold=True, size=11)
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='center')

    cell = ws.cell(row=row, column=7, value=totals['auto_approved'])
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='center')

    # Column widths
    ws.column_dimensions['A'].width = 18
    for col in range(2, 8):
        ws.column_dimensions[get_column_letter(col)].width = 14


def create_comparison_excel(client_name: str, vendors: dict, output_path: str, matcher: ProductMatcher = None):
    """
    Create a standalone Excel file with the multi-vendor comparison sheet.

    Args:
        client_name: Client name for the title
        vendors: Dict mapping vendor name -> NormalizedProposal
        output_path: Path for the output Excel file
        matcher: ProductMatcher instance for product-level comparison (optional)

    Returns:
        Tuple of (output_path, matched_data) where matched_data contains product matching results
    """
    from core import get_bucket_display_order

    # Initialize matcher if not provided
    if matcher is None:
        matcher = ProductMatcher()

    # Match products across vendors for the new comparison features
    matched_data = match_products_across_vendors(vendors, matcher)
    vendor_names = list(vendors.keys())

    # ==========================================================================
    # STYLING DEFINITIONS
    # ==========================================================================

    DARK_BLUE = "1F4E79"
    MEDIUM_BLUE = "2E75B6"
    LIGHT_BLUE = "BDD7EE"
    LIGHT_GRAY = "F2F2F2"
    GREEN = "C6EFCE"
    DARK_GREEN = "006100"
    RED = "FFC7CE"
    DARK_RED = "9C0006"
    ORANGE = "FCE4D6"

    TITLE_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    HEADER_FILL = PatternFill(start_color=MEDIUM_BLUE, end_color=MEDIUM_BLUE, fill_type="solid")
    SUBHEADER_FILL = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    ALT_ROW_FILL = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
    WINNER_FILL = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
    LOSER_FILL = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
    SUMMARY_BOX_FILL = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")

    TITLE_FONT = Font(bold=True, color="FFFFFF", size=18)
    SUBTITLE_FONT = Font(italic=True, color="666666", size=11)
    SECTION_FONT = Font(bold=True, color=DARK_BLUE, size=14)
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    SUBHEADER_FONT = Font(bold=True, size=11)
    NORMAL_FONT = Font(size=10)
    BOLD_FONT = Font(bold=True, size=10)
    WINNER_FONT = Font(bold=True, color=DARK_GREEN, size=11)
    LOSER_FONT = Font(bold=True, color=DARK_RED, size=11)
    TOTAL_FONT = Font(bold=True, color="FFFFFF", size=12)
    RECOMMENDATION_FONT = Font(bold=True, color=DARK_GREEN, size=12)

    CURRENCY_FORMAT = '_("$"* #,##0_);_("$"* (#,##0);_("$"* "-"_);_(@_)'
    PERCENT_FORMAT = '0.0%'

    THIN_BORDER = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    # ==========================================================================
    # CREATE WORKBOOK
    # ==========================================================================

    wb = Workbook()
    ws = wb.active
    ws.title = "Vendor Comparison"

    # ==========================================================================
    # CALCULATE VENDOR DATA & RANKINGS
    # ==========================================================================

    vendor_names = list(vendors.keys())
    num_vendors = len(vendor_names)

    # Pre-calculate totals for each vendor
    totals = {}
    for vendor_name, proposal in vendors.items():
        total = 0
        for bucket in get_bucket_display_order():
            total += proposal.bucket_totals.get(bucket.value, {}).get('total_7_year', 0)
        totals[vendor_name] = total

    # Create ranked list of vendors by total TCO (lowest first)
    ranked_vendors = sorted(vendor_names, key=lambda v: totals[v])

    winner = ranked_vendors[0]
    loser = ranked_vendors[-1]
    winner_tco = totals[winner]
    loser_tco = totals[loser]
    cost_range = loser_tco - winner_tco
    savings_pct = cost_range / loser_tco if loser_tco > 0 else 0

    row = 1

    # ==========================================================================
    # SECTION 1: TITLE HEADER
    # ==========================================================================

    title_merge_cols = max(num_vendors + 2, 6)

    for col in range(1, title_merge_cols + 1):
        ws.cell(row=row, column=col).fill = TITLE_FILL

    display_client = client_name.upper().replace('_', ' ')
    ws.cell(row=row, column=1, value=f"VENDOR COMPARISON: {display_client}")
    ws.cell(row=row, column=1).font = TITLE_FONT
    ws.cell(row=row, column=1).alignment = Alignment(vertical='center')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=title_merge_cols)
    ws.row_dimensions[row].height = 30
    row += 1

    ws.cell(row=row, column=1, value=f"Normalized Cost Analysis  |  {num_vendors} Vendors  |  Generated: {datetime.now().strftime('%B %d, %Y')}")
    ws.cell(row=row, column=1).font = SUBTITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=title_merge_cols)
    row += 2

    # ==========================================================================
    # SECTION 2: EXECUTIVE SUMMARY - VENDOR RANKINGS
    # ==========================================================================

    ws.cell(row=row, column=1, value="EXECUTIVE SUMMARY")
    ws.cell(row=row, column=1).font = SECTION_FONT
    row += 1

    summary_headers = ["Vendors Analyzed", "Cost Range (High - Low)", "Potential Savings", "Recommended Vendor"]
    summary_values = [num_vendors, cost_range, savings_pct, winner]

    for col, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = BOLD_FONT
        cell.fill = SUMMARY_BOX_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
    row += 1

    for col, value in enumerate(summary_values, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

        if col == 2:
            cell.number_format = CURRENCY_FORMAT
            cell.font = BOLD_FONT
        elif col == 3:
            cell.number_format = PERCENT_FORMAT
            cell.font = WINNER_FONT
        elif col == 4:
            cell.fill = WINNER_FILL
            cell.font = RECOMMENDATION_FONT
    row += 2

    # Vendor Rankings Table
    ws.cell(row=row, column=1, value="VENDOR RANKINGS BY 7-YEAR TCO")
    ws.cell(row=row, column=1).font = SECTION_FONT
    row += 1

    rank_headers = ["Rank", "Vendor", "7-Year TCO", "vs Lowest", "Status"]
    for col, header in enumerate(rank_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
    row += 1

    for idx, vendor in enumerate(ranked_vendors):
        rank = idx + 1
        tco = totals[vendor]
        vs_lowest = tco - winner_tco

        if rank == 1:
            row_fill = WINNER_FILL
            status = "LOWEST COST"
            status_font = WINNER_FONT
        elif rank == num_vendors:
            row_fill = LOSER_FILL
            status = "HIGHEST COST"
            status_font = LOSER_FONT
        else:
            row_fill = ALT_ROW_FILL if rank % 2 == 0 else None
            status = ""
            status_font = NORMAL_FONT

        cell = ws.cell(row=row, column=1, value=rank)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        cell.font = BOLD_FONT
        if row_fill:
            cell.fill = row_fill

        cell = ws.cell(row=row, column=2, value=vendor)
        cell.border = THIN_BORDER
        cell.font = BOLD_FONT
        if row_fill:
            cell.fill = row_fill

        cell = ws.cell(row=row, column=3, value=tco)
        cell.number_format = CURRENCY_FORMAT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='right')
        if row_fill:
            cell.fill = row_fill

        cell = ws.cell(row=row, column=4, value=vs_lowest if rank > 1 else 0)
        cell.number_format = '_("$"* +#,##0_);_("$"* -#,##0_);_("$"* "-"_);_(@_)'
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='right')
        if row_fill:
            cell.fill = row_fill

        cell = ws.cell(row=row, column=5, value=status)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        cell.font = status_font
        if row_fill:
            cell.fill = row_fill

        row += 1

    row += 1

    # ==========================================================================
    # SECTION 3: COST BREAKDOWN BY BUCKET
    # ==========================================================================

    ws.cell(row=row, column=1, value="7-YEAR TCO BY COST CATEGORY")
    ws.cell(row=row, column=1).font = SECTION_FONT
    row += 1

    headers = ["Cost Category"] + vendor_names + ["Lowest Cost"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 25
    row += 1

    for idx, bucket in enumerate(get_bucket_display_order()):
        bucket_name = bucket.value
        row_fill = ALT_ROW_FILL if idx % 2 == 0 else None

        cell = ws.cell(row=row, column=1, value=bucket_name)
        cell.font = SUBHEADER_FONT
        cell.border = THIN_BORDER
        if row_fill:
            cell.fill = row_fill

        bucket_values = {}
        for vendor_name, proposal in vendors.items():
            tco = proposal.bucket_totals.get(bucket_name, {}).get('total_7_year', 0)
            bucket_values[vendor_name] = tco

        min_vendor = min(bucket_values.keys(), key=lambda v: bucket_values[v])
        min_value = bucket_values[min_vendor]

        for col, vendor in enumerate(vendor_names, 2):
            tco = bucket_values[vendor]
            cell = ws.cell(row=row, column=col, value=tco)
            cell.number_format = CURRENCY_FORMAT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='right')

            if vendor == min_vendor and tco > 0:
                cell.fill = WINNER_FILL
            elif row_fill:
                cell.fill = row_fill

        cell = ws.cell(row=row, column=num_vendors + 2, value=min_vendor if min_value > 0 else "N/A")
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        cell.font = BOLD_FONT
        if row_fill:
            cell.fill = row_fill

        row += 1

    # Total row
    cell = ws.cell(row=row, column=1, value="TOTAL 7-YEAR TCO")
    cell.fill = HEADER_FILL
    cell.font = TOTAL_FONT
    cell.border = THIN_BORDER

    for col, vendor in enumerate(vendor_names, 2):
        cell = ws.cell(row=row, column=col, value=totals[vendor])
        cell.number_format = CURRENCY_FORMAT
        cell.fill = HEADER_FILL
        cell.font = TOTAL_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='right')

        if vendor == winner:
            cell.fill = WINNER_FILL
            cell.font = Font(bold=True, color=DARK_GREEN, size=12)

    cell = ws.cell(row=row, column=num_vendors + 2, value=winner)
    cell.fill = WINNER_FILL
    cell.font = RECOMMENDATION_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='center')

    row += 2

    # ==========================================================================
    # SECTION 4: RECOMMENDATION
    # ==========================================================================

    rec_cols = num_vendors + 2

    for col in range(1, rec_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = WINNER_FILL
        cell.border = THIN_BORDER

    ws.cell(row=row, column=1, value="RECOMMENDATION")
    ws.cell(row=row, column=1).font = Font(bold=True, color=DARK_GREEN, size=14)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=rec_cols)
    row += 1

    for col in range(1, rec_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = WINNER_FILL
        cell.border = THIN_BORDER

    if num_vendors == 2:
        recommendation = f"{winner} offers the lowest 7-year TCO at ${winner_tco:,.0f}, saving ${cost_range:,.0f} ({savings_pct:.1%}) compared to {loser}."
    else:
        recommendation = f"{winner} offers the lowest 7-year TCO at ${winner_tco:,.0f}, saving ${cost_range:,.0f} ({savings_pct:.1%}) compared to the highest-cost option ({loser})."

    ws.cell(row=row, column=1, value=recommendation)
    ws.cell(row=row, column=1).font = Font(size=12, color=DARK_GREEN)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=rec_cols)
    ws.row_dimensions[row].height = 25
    row += 2

    # ==========================================================================
    # SECTION 5: LINE ITEM DETAILS
    # ==========================================================================

    ws.cell(row=row, column=1, value="DETAILED LINE ITEM COMPARISON")
    ws.cell(row=row, column=1).font = SECTION_FONT
    row += 1

    ws.cell(row=row, column=1, value="Line items organized by cost category for detailed analysis")
    ws.cell(row=row, column=1).font = SUBTITLE_FONT
    row += 2

    for bucket in get_bucket_display_order():
        bucket_name = bucket.value

        vendor_items = {}
        has_items = False
        for vendor_name, proposal in vendors.items():
            items = [i for i in proposal.line_items if i.level_1_bucket == bucket_name]
            vendor_items[vendor_name] = sorted(items, key=lambda x: x.total_7_year_cost, reverse=True)
            if items:
                has_items = True

        if not has_items:
            continue

        total_cols = num_vendors * 2
        for col in range(1, total_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = SUBHEADER_FILL
            cell.border = THIN_BORDER

        ws.cell(row=row, column=1, value=bucket_name)
        ws.cell(row=row, column=1).font = SUBHEADER_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
        ws.row_dimensions[row].height = 22
        row += 1

        col = 1
        for vendor in vendor_names:
            cell = ws.cell(row=row, column=col, value=f"{vendor} Solution")
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')

            cell = ws.cell(row=row, column=col + 1, value=f"{vendor} Cost")
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')

            col += 2
        row += 1

        max_items = max(len(items) for items in vendor_items.values()) if vendor_items else 0

        for item_idx in range(max_items):
            row_fill = ALT_ROW_FILL if item_idx % 2 == 0 else None

            col = 1
            for vendor in vendor_names:
                items = vendor_items[vendor]

                if item_idx < len(items):
                    item = items[item_idx]

                    cell = ws.cell(row=row, column=col, value=item.solution_name)
                    cell.font = NORMAL_FONT
                    cell.border = THIN_BORDER
                    if row_fill:
                        cell.fill = row_fill

                    cell = ws.cell(row=row, column=col + 1, value=item.total_7_year_cost)
                    cell.number_format = CURRENCY_FORMAT
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(horizontal='right')
                    if row_fill:
                        cell.fill = row_fill
                else:
                    for c in [col, col + 1]:
                        cell = ws.cell(row=row, column=c, value="")
                        cell.border = THIN_BORDER
                        if row_fill:
                            cell.fill = row_fill

                col += 2

            row += 1

        subtotals = {vendor: sum(i.total_7_year_cost for i in vendor_items[vendor]) for vendor in vendor_names}
        min_subtotal_vendor = min(subtotals.keys(), key=lambda v: subtotals[v]) if subtotals else None

        col = 1
        for vendor in vendor_names:
            cell = ws.cell(row=row, column=col, value=f"{vendor} Subtotal")
            cell.font = BOLD_FONT
            cell.fill = SUBHEADER_FILL
            cell.border = THIN_BORDER

            cell = ws.cell(row=row, column=col + 1, value=subtotals[vendor])
            cell.number_format = CURRENCY_FORMAT
            cell.font = BOLD_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='right')

            if vendor == min_subtotal_vendor and subtotals[vendor] > 0:
                cell.fill = WINNER_FILL
            else:
                cell.fill = SUBHEADER_FILL

            col += 2

        row += 2

    # ==========================================================================
    # COLUMN WIDTHS
    # ==========================================================================

    ws.column_dimensions['A'].width = 30

    col_idx = 2
    for _ in vendor_names:
        ws.column_dimensions[get_column_letter(col_idx)].width = 28
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = 15
        col_idx += 2

    ws.column_dimensions[get_column_letter(num_vendors + 2)].width = 14

    ws.freeze_panes = 'A5'

    # ==========================================================================
    # SECTION 6: NEW PRODUCT-LEVEL COMPARISON SHEETS (Phase 4)
    # ==========================================================================

    # Create Product-by-Product Comparison sheet
    gaps_detected = create_product_comparison_sheet(wb, matched_data, vendor_names, matcher)

    # Create Gap Analysis sheet
    create_gaps_sheet(wb, gaps_detected, matched_data, vendor_names, matcher)

    # Create Unmatched Products sheet
    create_unmatched_sheet(wb, matched_data, vendor_names)

    # Create Match Statistics sheet
    create_match_statistics_sheet(wb, matched_data, vendor_names)

    # Reorder sheets - put main comparison first
    sheet_order = ['Vendor Comparison', 'Product Comparison', 'Gap Analysis', 'Unmatched Products', 'Match Statistics']
    for i, sheet_name in enumerate(sheet_order):
        if sheet_name in wb.sheetnames:
            wb.move_sheet(wb[sheet_name], offset=i - wb.sheetnames.index(sheet_name))

    # Save workbook
    wb.save(output_path)

    return output_path, matched_data


def main():
    if len(sys.argv) < 2:
        print("Usage: python generate_comparison.py <client_name>")
        print()
        print("Examples:")
        print("  python generate_comparison.py echelon_bank")
        print("  python generate_comparison.py liberty_capital_bank")
        print("  python generate_comparison.py heritage")
        print()
        print("The script will find all vendor extractions for the client and compare them.")
        sys.exit(1)

    client_name = sys.argv[1]

    print("=" * 70)
    print("MULTI-VENDOR COMPARISON GENERATOR")
    print("=" * 70)
    print(f"Client: {client_name}")
    print()

    # Find all unique vendors for this client
    print("Searching for vendor extractions...")
    unique_vendors = find_client_vendors(client_name)

    if not unique_vendors:
        print(f"ERROR: No vendor extractions found for client '{client_name}'")
        print()
        print("Available extraction files:")
        for f in glob("Extracted JSON/*_extraction_ai.json"):
            print(f"  - {Path(f).stem}")
        sys.exit(1)

    print(f"Found {len(unique_vendors)} unique vendor(s):")
    for vendor, json_path in unique_vendors.items():
        print(f"  - {vendor}: {json_path}")
    print()

    if len(unique_vendors) < 2:
        print("WARNING: Only 1 vendor found. Comparison requires at least 2 vendors.")
        print("         The file will be generated but no comparison is possible.")
        print()

    # Normalize all proposals
    print("Normalizing vendor proposals...")
    normalized_vendors = {}
    for vendor, json_path in unique_vendors.items():
        print(f"  Processing {vendor}...")
        proposal = normalize_proposal(json_path)
        normalized_vendors[vendor] = proposal
    print()

    # Generate output filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_filename = f"TCO Output/{client_name.upper()}_COMPARISON_{timestamp}.xlsx"

    # Initialize product matcher
    print("Initializing product matcher...")
    matcher = ProductMatcher()

    # Create the comparison Excel file
    print("Generating comparison Excel file...")
    output_path, matched_data = create_comparison_excel(client_name, normalized_vendors, output_filename, matcher)

    print()
    print("=" * 70)
    print("COMPARISON COMPLETE")
    print("=" * 70)
    print(f"Output file: {output_path}")
    print()

    # Print product matching statistics
    print("PRODUCT MATCHING STATISTICS:")
    print("-" * 70)
    stats = matched_data['statistics']
    total_items = sum(s.get('total', 0) for s in stats.values())
    total_matched = sum(s.get('exact_matches', 0) + s.get('fuzzy_matches', 0) for s in stats.values())
    total_unmatched = sum(s.get('unmatched', 0) for s in stats.values())
    overall_rate = (total_matched / total_items * 100) if total_items > 0 else 0

    for vendor in normalized_vendors.keys():
        vendor_stats = stats.get(vendor, {})
        rate = vendor_stats.get('match_rate', 0)
        print(f"  {vendor}: {vendor_stats.get('total', 0)} items, {rate:.1f}% matched")

    print()
    print(f"  OVERALL: {total_items} items, {overall_rate:.1f}% matched ({total_unmatched} unmatched)")
    print()

    # Print summary
    print("SUMMARY:")
    print("-" * 70)

    from core import get_bucket_display_order

    totals = {}
    for vendor, proposal in normalized_vendors.items():
        total = sum(proposal.bucket_totals.get(b.value, {}).get('total_7_year', 0)
                    for b in get_bucket_display_order())
        totals[vendor] = total

    ranked = sorted(totals.items(), key=lambda x: x[1])

    for rank, (vendor, tco) in enumerate(ranked, 1):
        status = " (LOWEST)" if rank == 1 else " (HIGHEST)" if rank == len(ranked) else ""
        print(f"  {rank}. {vendor}: ${tco:,.0f}{status}")

    if len(ranked) >= 2:
        savings = ranked[-1][1] - ranked[0][1]
        print()
        print(f"  >> Potential savings: ${savings:,.0f} by choosing {ranked[0][0]}")

    print()


if __name__ == "__main__":
    main()
