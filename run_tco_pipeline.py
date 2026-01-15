"""
Single-Command TCO Pipeline
Runs complete extraction → Excel generation → Normalized Comparison pipeline in one command

Usage:
    python run_tco_pipeline.py <input_file> <vendor_name>

Example:
    python run_tco_pipeline.py "proposal.pdf" "csi"
    python run_tco_pipeline.py "WORKBOOK1.xlsx" "liberty"
"""

import sys
import subprocess
import json
from pathlib import Path
from datetime import datetime
from glob import glob


def parse_vendor_name(vendor_name: str) -> tuple:
    """
    Parse vendor name to extract client name and vendor type.

    Examples:
        'Liberty_Bank_FIS' -> ('Liberty_Bank', 'FIS')
        'Liberty_Bank_CSI' -> ('Liberty_Bank', 'CSI')
        'Echelon_Bank_FIS' -> ('Echelon_Bank', 'FIS')
        'echelon_bank_fis_1' -> ('echelon_bank', 'FIS')  # handles numbered suffixes
        'echelon_bank_fis_2' -> ('echelon_bank', 'FIS')  # handles numbered suffixes

    Known vendor suffixes: FIS, CSI, JH (Jack Henry), Fiserv, Finastra
    """
    import re

    known_vendors = ['FIS', 'CSI', 'JH', 'Fiserv', 'Finastra', 'Q2', 'Temenos']

    # Strip numeric suffix (e.g., _1, _2, _retry, _v2, _final) from the end
    cleaned_name = re.sub(r'_(\d+|retry|v\d+|final)$', '', vendor_name, flags=re.IGNORECASE)

    # Check if name ends with a known vendor
    for vendor in known_vendors:
        if cleaned_name.upper().endswith(f"_{vendor.upper()}"):
            # Extract client name (everything before the vendor suffix)
            client = cleaned_name[:-(len(vendor) + 1)]  # +1 for underscore
            return (client, vendor.upper())

    # Fallback: assume last part after underscore is vendor
    parts = cleaned_name.rsplit('_', 1)
    if len(parts) == 2:
        return (parts[0], parts[1].upper())

    return (vendor_name, 'Unknown')


def find_other_vendor_extractions(client_name: str, current_vendor: str) -> list:
    """
    Find other vendor extraction files for the same client.

    Args:
        client_name: Client name (e.g., 'Liberty_Bank')
        current_vendor: Current vendor to exclude (e.g., 'FIS')

    Returns:
        List of tuples: [(vendor_name, json_path), ...]
    """
    other_vendors = []

    # Search for extraction files matching the client pattern
    pattern = f"Extracted JSON/{client_name}_*_extraction_ai.json"
    matching_files = glob(pattern)

    # Also try lowercase pattern
    pattern_lower = f"Extracted JSON/{client_name.lower()}_*_extraction_ai.json"
    matching_files.extend(glob(pattern_lower))

    for json_path in matching_files:
        # Extract vendor from filename
        filename = Path(json_path).stem  # e.g., 'Liberty_Bank_CSI_extraction_ai'
        vendor_part = filename.replace('_extraction_ai', '').replace('_raw_extraction', '')

        # Parse to get vendor name
        _, vendor = parse_vendor_name(vendor_part)

        # Skip if it's the current vendor
        if vendor.upper() != current_vendor.upper():
            other_vendors.append((vendor, json_path))

    return other_vendors


def add_normalized_comparison_sheet(json_file: str, vendor_name: str) -> str:
    """
    Add Normalized Comparison sheet to the generated Excel file.

    This step normalizes the extracted data into the 6-bucket cost structure
    and automatically detects other vendor proposals for the same client
    to create a side-by-side comparison.

    Args:
        json_file: Path to the AI extraction JSON file
        vendor_name: Vendor name for finding the Excel file

    Returns:
        Path to the Excel file with normalized sheet, or None on failure
    """
    try:
        # Import core modules
        from core import CostNormalizer, add_normalized_sheet, NormalizedSheetConfig
        from dataclasses import asdict

        # Parse vendor name to get client and vendor type
        client_name, vendor_type = parse_vendor_name(vendor_name)
        print(f"  Client: {client_name}")
        print(f"  Vendor: {vendor_type}")

        # Find the generated Excel file (most recent one for this vendor)
        vendor_upper = vendor_name.upper().replace(' ', '_')
        today = datetime.now().strftime('%Y%m%d')
        excel_pattern = f"TCO Output/{vendor_upper}_TCO_New_{today}.xlsx"

        excel_files = glob(excel_pattern)
        if not excel_files:
            # Try without date constraint
            excel_pattern = f"TCO Output/{vendor_upper}_TCO_New_*.xlsx"
            excel_files = sorted(glob(excel_pattern), key=lambda x: Path(x).stat().st_mtime, reverse=True)

        if not excel_files:
            print(f"  ERROR: No Excel file found matching pattern: {excel_pattern}")
            return None

        excel_file = excel_files[0]
        print(f"  Excel file: {excel_file}")

        # Load the primary JSON extraction data
        with open(json_file, 'r', encoding='utf-8') as f:
            extraction_data = json.load(f)

        print(f"  Line items to normalize: {len(extraction_data.get('line_items', []))}")

        # Normalize the primary extraction
        normalizer = CostNormalizer()
        primary_proposal = normalizer.normalize_proposal(
            extraction_data=extraction_data,
            source_file=json_file
        )

        # Find other vendors for the same client
        other_vendors = find_other_vendor_extractions(client_name, vendor_type)

        all_proposals = [(vendor_type, primary_proposal)]

        if other_vendors:
            print(f"\n  MULTI-VENDOR COMPARISON DETECTED")
            print(f"  Found {len(other_vendors)} other vendor(s) for {client_name}:")

            for other_vendor, other_json_path in other_vendors:
                print(f"    - {other_vendor}: {other_json_path}")

                # Load and normalize the other vendor's extraction
                with open(other_json_path, 'r', encoding='utf-8') as f:
                    other_extraction = json.load(f)

                other_proposal = normalizer.normalize_proposal(
                    extraction_data=other_extraction,
                    source_file=other_json_path
                )
                all_proposals.append((other_vendor, other_proposal))

        # Create the comparison sheet
        if len(all_proposals) > 1:
            # Multi-vendor comparison
            result_path = _create_multi_vendor_comparison_sheet(
                excel_file, client_name, all_proposals
            )
        else:
            # Single vendor - use original sheet format
            normalized_data = {"proposal": asdict(primary_proposal)}
            config = NormalizedSheetConfig(
                sheet_name="Normalized Comparison",
                include_line_items=True,
                include_metrics=True,
                include_bucket_summary=True
            )
            result_path = add_normalized_sheet(
                workbook_path=excel_file,
                normalized_data=normalized_data,
                config=config
            )

        # Print summary
        print(f"\n  NORMALIZED COMPARISON SUMMARY (6-BUCKET STRUCTURE)")
        print(f"  " + "=" * 70)

        if len(all_proposals) > 1:
            # Print header row
            header = f"  {'Cost Bucket':<25}"
            for vendor, _ in all_proposals:
                header += f" {vendor:>18}"
            header += f" {'Difference':>15}"
            print(header)
            print(f"  " + "-" * 70)

            # Get bucket display order
            from core import get_bucket_display_order

            totals = {vendor: 0 for vendor, _ in all_proposals}

            for bucket in get_bucket_display_order():
                bucket_name = bucket.value
                row = f"  {bucket_name:<25}"

                values = []
                for vendor, proposal in all_proposals:
                    tco = proposal.bucket_totals.get(bucket_name, {}).get('total_7_year', 0)
                    values.append(tco)
                    totals[vendor] += tco
                    row += f" ${tco:>15,.0f}"

                # Calculate difference (first vendor - second vendor)
                if len(values) >= 2:
                    diff = values[0] - values[1]
                    diff_str = f"${diff:>+13,.0f}"
                    row += f" {diff_str}"

                print(row)

            print(f"  " + "-" * 70)

            # Total row
            total_row = f"  {'TOTAL 7-YEAR TCO':<25}"
            total_values = []
            for vendor, _ in all_proposals:
                total_row += f" ${totals[vendor]:>15,.0f}"
                total_values.append(totals[vendor])

            if len(total_values) >= 2:
                total_diff = total_values[0] - total_values[1]
                total_row += f" ${total_diff:>+13,.0f}"

            print(total_row)
            print(f"  " + "=" * 70)

            # Winner announcement
            if len(total_values) >= 2:
                if total_values[0] < total_values[1]:
                    winner = all_proposals[0][0]
                    savings = total_values[1] - total_values[0]
                else:
                    winner = all_proposals[1][0]
                    savings = total_values[0] - total_values[1]
                print(f"\n  >> RECOMMENDATION: {winner} is ${savings:,.0f} less expensive over 7 years")
        else:
            # Single vendor summary
            bucket_totals = primary_proposal.bucket_totals
            total_tco = 0
            for bucket_name, bucket_data in bucket_totals.items():
                tco = bucket_data.get('total_7_year', 0)
                total_tco += tco
                item_count = bucket_data.get('item_count', 0)
                if item_count > 0:
                    print(f"  {bucket_name:<30} ${tco:>12,.2f}  ({item_count} items)")

            print(f"  " + "-" * 50)
            print(f"  {'TOTAL 7-YEAR TCO':<30} ${total_tco:>12,.2f}")

        return result_path

    except ImportError as e:
        print(f"  ERROR: Could not import core modules: {e}")
        print(f"  Make sure the 'core' package is available")
        return None
    except Exception as e:
        print(f"  ERROR: Failed to add normalized sheet: {e}")
        import traceback
        traceback.print_exc()
        return None


def _create_multi_vendor_comparison_sheet(excel_file: str, client_name: str,
                                           all_proposals: list) -> str:
    """
    Create a professionally formatted side-by-side multi-vendor comparison sheet.
    Dynamically supports N vendors (2, 3, 5, 10, or any number).

    Args:
        excel_file: Path to the Excel workbook
        client_name: Client name for the header
        all_proposals: List of (vendor_name, NormalizedProposal) tuples

    Returns:
        Path to the updated Excel file
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from core import get_bucket_display_order

    # ==========================================================================
    # STYLING DEFINITIONS
    # ==========================================================================

    # Colors
    DARK_BLUE = "1F4E79"
    MEDIUM_BLUE = "2E75B6"
    LIGHT_BLUE = "BDD7EE"
    LIGHT_GRAY = "F2F2F2"
    GREEN = "C6EFCE"
    DARK_GREEN = "006100"
    RED = "FFC7CE"
    DARK_RED = "9C0006"
    ORANGE = "FCE4D6"
    YELLOW = "FFEB9C"

    # Fills
    TITLE_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    HEADER_FILL = PatternFill(start_color=MEDIUM_BLUE, end_color=MEDIUM_BLUE, fill_type="solid")
    SUBHEADER_FILL = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    ALT_ROW_FILL = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
    WINNER_FILL = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
    LOSER_FILL = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
    SUMMARY_BOX_FILL = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
    RANK_2_FILL = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")

    # Fonts
    TITLE_FONT = Font(bold=True, color="FFFFFF", size=18)
    SUBTITLE_FONT = Font(italic=True, color="666666", size=11)
    SECTION_FONT = Font(bold=True, color=DARK_BLUE, size=14)
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    SUBHEADER_FONT = Font(bold=True, size=11)
    NORMAL_FONT = Font(size=10)
    BOLD_FONT = Font(bold=True, size=10)
    WINNER_FONT = Font(bold=True, color=DARK_GREEN, size=11)
    LOSER_FONT = Font(bold=True, color=DARK_RED, size=11)
    TOTAL_FONT = Font(bold=True, color="FFFFFF", size=12)
    RECOMMENDATION_FONT = Font(bold=True, color=DARK_GREEN, size=12)

    # Formats
    CURRENCY_FORMAT = '_("$"* #,##0_);_("$"* (#,##0);_("$"* "-"_);_(@_)'
    PERCENT_FORMAT = '0.0%'

    # Borders
    THIN_BORDER = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    # ==========================================================================
    # SETUP WORKBOOK
    # ==========================================================================

    wb = load_workbook(excel_file)

    if "Normalized Comparison" in wb.sheetnames:
        del wb["Normalized Comparison"]

    ws = wb.create_sheet("Normalized Comparison")

    # ==========================================================================
    # CALCULATE VENDOR DATA & RANKINGS
    # ==========================================================================

    vendors = [v[0] for v in all_proposals]
    num_vendors = len(vendors)

    # Pre-calculate totals for each vendor
    totals = {v: 0 for v in vendors}
    for vendor, proposal in all_proposals:
        for bucket in get_bucket_display_order():
            totals[vendor] += proposal.bucket_totals.get(bucket.value, {}).get('total_7_year', 0)

    # Create ranked list of vendors by total TCO (lowest first)
    ranked_vendors = sorted(vendors, key=lambda v: totals[v])
    vendor_ranks = {v: idx + 1 for idx, v in enumerate(ranked_vendors)}

    # Winner and loser
    winner = ranked_vendors[0]  # Lowest cost
    loser = ranked_vendors[-1]  # Highest cost
    winner_tco = totals[winner]
    loser_tco = totals[loser]
    cost_range = loser_tco - winner_tco
    savings_pct = cost_range / loser_tco if loser_tco > 0 else 0

    # Calculate dynamic column counts
    # Cost breakdown: Category + N vendors + "Lowest Cost" column
    cost_table_cols = 1 + num_vendors + 1
    # Line items: N vendors × 2 columns (Solution, Cost)
    line_item_cols = num_vendors * 2

    row = 1

    # ==========================================================================
    # SECTION 1: TITLE HEADER
    # ==========================================================================

    title_merge_cols = max(cost_table_cols, 6)

    for col in range(1, title_merge_cols + 1):
        ws.cell(row=row, column=col).fill = TITLE_FILL

    ws.cell(row=row, column=1, value=f"VENDOR COMPARISON: {client_name.upper().replace('_', ' ')}")
    ws.cell(row=row, column=1).font = TITLE_FONT
    ws.cell(row=row, column=1).alignment = Alignment(vertical='center')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=title_merge_cols)
    ws.row_dimensions[row].height = 30
    row += 1

    # Subtitle
    ws.cell(row=row, column=1, value=f"Normalized Cost Analysis  |  {num_vendors} Vendors  |  Generated: {datetime.now().strftime('%B %d, %Y')}")
    ws.cell(row=row, column=1).font = SUBTITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=title_merge_cols)
    row += 2

    # ==========================================================================
    # SECTION 2: EXECUTIVE SUMMARY - VENDOR RANKINGS
    # ==========================================================================

    ws.cell(row=row, column=1, value="EXECUTIVE SUMMARY")
    ws.cell(row=row, column=1).font = SECTION_FONT
    row += 1

    # Summary metrics row
    summary_headers = ["Vendors Analyzed", "Cost Range (High - Low)", "Potential Savings", "Recommended Vendor"]
    summary_values = [num_vendors, cost_range, savings_pct, winner]

    for col, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = BOLD_FONT
        cell.fill = SUMMARY_BOX_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
    row += 1

    for col, value in enumerate(summary_values, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

        if col == 2:  # Cost Range
            cell.number_format = CURRENCY_FORMAT
            cell.font = BOLD_FONT
        elif col == 3:  # Savings %
            cell.number_format = PERCENT_FORMAT
            cell.font = WINNER_FONT
        elif col == 4:  # Recommended
            cell.fill = WINNER_FILL
            cell.font = RECOMMENDATION_FONT
    row += 2

    # Vendor Rankings Table
    ws.cell(row=row, column=1, value="VENDOR RANKINGS BY 7-YEAR TCO")
    ws.cell(row=row, column=1).font = SECTION_FONT
    row += 1

    # Rankings header
    rank_headers = ["Rank", "Vendor", "7-Year TCO", "vs Lowest", "Status"]
    for col, header in enumerate(rank_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
    row += 1

    # Vendor ranking rows
    for idx, vendor in enumerate(ranked_vendors):
        rank = idx + 1
        tco = totals[vendor]
        vs_lowest = tco - winner_tco

        # Determine row styling based on rank
        if rank == 1:
            row_fill = WINNER_FILL
            status = "LOWEST COST"
            status_font = WINNER_FONT
        elif rank == num_vendors:
            row_fill = LOSER_FILL
            status = "HIGHEST COST"
            status_font = LOSER_FONT
        else:
            row_fill = ALT_ROW_FILL if rank % 2 == 0 else None
            status = ""
            status_font = NORMAL_FONT

        # Rank
        cell = ws.cell(row=row, column=1, value=rank)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        cell.font = BOLD_FONT
        if row_fill:
            cell.fill = row_fill

        # Vendor name
        cell = ws.cell(row=row, column=2, value=vendor)
        cell.border = THIN_BORDER
        cell.font = BOLD_FONT
        if row_fill:
            cell.fill = row_fill

        # TCO
        cell = ws.cell(row=row, column=3, value=tco)
        cell.number_format = CURRENCY_FORMAT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='right')
        if row_fill:
            cell.fill = row_fill

        # vs Lowest
        cell = ws.cell(row=row, column=4, value=vs_lowest if rank > 1 else 0)
        cell.number_format = '_("$"* +#,##0_);_("$"* -#,##0_);_("$"* "-"_);_(@_)'
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='right')
        if row_fill:
            cell.fill = row_fill

        # Status
        cell = ws.cell(row=row, column=5, value=status)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        cell.font = status_font
        if row_fill:
            cell.fill = row_fill

        row += 1

    row += 1

    # ==========================================================================
    # SECTION 3: COST BREAKDOWN BY BUCKET (N VENDORS)
    # ==========================================================================

    ws.cell(row=row, column=1, value="7-YEAR TCO BY COST CATEGORY")
    ws.cell(row=row, column=1).font = SECTION_FONT
    row += 1

    # Header row: Category + N vendor columns + Lowest Cost
    headers = ["Cost Category"] + vendors + ["Lowest Cost"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 25
    row += 1

    # Bucket rows
    for idx, bucket in enumerate(get_bucket_display_order()):
        bucket_name = bucket.value
        row_fill = ALT_ROW_FILL if idx % 2 == 0 else None

        # Category name
        cell = ws.cell(row=row, column=1, value=bucket_name)
        cell.font = SUBHEADER_FONT
        cell.border = THIN_BORDER
        if row_fill:
            cell.fill = row_fill

        # Collect values for all vendors
        bucket_values = {}
        for vendor, proposal in all_proposals:
            tco = proposal.bucket_totals.get(bucket_name, {}).get('total_7_year', 0)
            bucket_values[vendor] = tco

        # Find lowest cost vendor for this bucket
        min_vendor = min(bucket_values.keys(), key=lambda v: bucket_values[v])
        min_value = bucket_values[min_vendor]

        # Write vendor values
        for col, vendor in enumerate(vendors, 2):
            tco = bucket_values[vendor]
            cell = ws.cell(row=row, column=col, value=tco)
            cell.number_format = CURRENCY_FORMAT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='right')

            # Highlight lowest cost cell
            if vendor == min_vendor and tco > 0:
                cell.fill = WINNER_FILL
            elif row_fill:
                cell.fill = row_fill

        # Lowest Cost column (vendor name)
        cell = ws.cell(row=row, column=num_vendors + 2, value=min_vendor if min_value > 0 else "N/A")
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        cell.font = BOLD_FONT
        if row_fill:
            cell.fill = row_fill

        row += 1

    # Total row
    cell = ws.cell(row=row, column=1, value="TOTAL 7-YEAR TCO")
    cell.fill = HEADER_FILL
    cell.font = TOTAL_FONT
    cell.border = THIN_BORDER

    for col, vendor in enumerate(vendors, 2):
        cell = ws.cell(row=row, column=col, value=totals[vendor])
        cell.number_format = CURRENCY_FORMAT
        cell.fill = HEADER_FILL
        cell.font = TOTAL_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='right')

        # Highlight overall winner
        if vendor == winner:
            cell.fill = WINNER_FILL
            cell.font = Font(bold=True, color=DARK_GREEN, size=12)

    # Overall winner in last column
    cell = ws.cell(row=row, column=num_vendors + 2, value=winner)
    cell.fill = WINNER_FILL
    cell.font = RECOMMENDATION_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='center')

    row += 2

    # ==========================================================================
    # SECTION 4: RECOMMENDATION
    # ==========================================================================

    rec_cols = num_vendors + 2

    # Recommendation header
    for col in range(1, rec_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = WINNER_FILL
        cell.border = THIN_BORDER

    ws.cell(row=row, column=1, value="RECOMMENDATION")
    ws.cell(row=row, column=1).font = Font(bold=True, color=DARK_GREEN, size=14)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=rec_cols)
    row += 1

    # Recommendation text
    for col in range(1, rec_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = WINNER_FILL
        cell.border = THIN_BORDER

    if num_vendors == 2:
        recommendation = f"{winner} offers the lowest 7-year TCO at ${winner_tco:,.0f}, saving ${cost_range:,.0f} ({savings_pct:.1%}) compared to {loser}."
    else:
        recommendation = f"{winner} offers the lowest 7-year TCO at ${winner_tco:,.0f}, saving ${cost_range:,.0f} ({savings_pct:.1%}) compared to the highest-cost option ({loser})."

    ws.cell(row=row, column=1, value=recommendation)
    ws.cell(row=row, column=1).font = Font(size=12, color=DARK_GREEN)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=rec_cols)
    ws.row_dimensions[row].height = 25
    row += 2

    # ==========================================================================
    # SECTION 5: LINE ITEM DETAILS (N VENDORS)
    # ==========================================================================

    ws.cell(row=row, column=1, value="DETAILED LINE ITEM COMPARISON")
    ws.cell(row=row, column=1).font = SECTION_FONT
    row += 1

    ws.cell(row=row, column=1, value="Line items organized by cost category for detailed analysis")
    ws.cell(row=row, column=1).font = SUBTITLE_FONT
    row += 2

    # Process each bucket
    for bucket in get_bucket_display_order():
        bucket_name = bucket.value

        # Get items for each vendor
        vendor_items = {}
        has_items = False
        for vendor, proposal in all_proposals:
            items = [i for i in proposal.line_items if i.level_1_bucket == bucket_name]
            vendor_items[vendor] = sorted(items, key=lambda x: x.total_7_year_cost, reverse=True)
            if items:
                has_items = True

        if not has_items:
            continue

        # Bucket header - spans all vendor columns
        total_cols = num_vendors * 2
        for col in range(1, total_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = SUBHEADER_FILL
            cell.border = THIN_BORDER

        ws.cell(row=row, column=1, value=bucket_name)
        ws.cell(row=row, column=1).font = SUBHEADER_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
        ws.row_dimensions[row].height = 22
        row += 1

        # Column headers for line items - dynamic for N vendors
        col = 1
        for vendor in vendors:
            cell = ws.cell(row=row, column=col, value=f"{vendor} Solution")
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')

            cell = ws.cell(row=row, column=col + 1, value=f"{vendor} Cost")
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')

            col += 2
        row += 1

        # Get max items across all vendors
        max_items = max(len(items) for items in vendor_items.values()) if vendor_items else 0

        # Write items side by side for all vendors
        for item_idx in range(max_items):
            row_fill = ALT_ROW_FILL if item_idx % 2 == 0 else None

            col = 1
            for vendor in vendors:
                items = vendor_items[vendor]

                if item_idx < len(items):
                    item = items[item_idx]

                    # Solution name
                    cell = ws.cell(row=row, column=col, value=item.solution_name)
                    cell.font = NORMAL_FONT
                    cell.border = THIN_BORDER
                    if row_fill:
                        cell.fill = row_fill

                    # Cost
                    cell = ws.cell(row=row, column=col + 1, value=item.total_7_year_cost)
                    cell.number_format = CURRENCY_FORMAT
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(horizontal='right')
                    if row_fill:
                        cell.fill = row_fill
                else:
                    # Empty cells
                    for c in [col, col + 1]:
                        cell = ws.cell(row=row, column=c, value="")
                        cell.border = THIN_BORDER
                        if row_fill:
                            cell.fill = row_fill

                col += 2

            row += 1

        # Bucket subtotals for all vendors
        subtotals = {vendor: sum(i.total_7_year_cost for i in vendor_items[vendor]) for vendor in vendors}
        min_subtotal_vendor = min(subtotals.keys(), key=lambda v: subtotals[v]) if subtotals else None

        col = 1
        for vendor in vendors:
            cell = ws.cell(row=row, column=col, value=f"{vendor} Subtotal")
            cell.font = BOLD_FONT
            cell.fill = SUBHEADER_FILL
            cell.border = THIN_BORDER

            cell = ws.cell(row=row, column=col + 1, value=subtotals[vendor])
            cell.number_format = CURRENCY_FORMAT
            cell.font = BOLD_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='right')

            # Highlight lowest subtotal
            if vendor == min_subtotal_vendor and subtotals[vendor] > 0:
                cell.fill = WINNER_FILL
            else:
                cell.fill = SUBHEADER_FILL

            col += 2

        row += 2

    # ==========================================================================
    # DYNAMIC COLUMN WIDTHS
    # ==========================================================================

    # Column A is always category/solution names
    ws.column_dimensions['A'].width = 30

    # Vendor columns alternate: Solution (28), Cost (15)
    col_idx = 2
    for vendor in vendors:
        ws.column_dimensions[get_column_letter(col_idx)].width = 28  # Solution
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = 15  # Cost
        col_idx += 2

    # "Lowest Cost" column in breakdown table
    ws.column_dimensions[get_column_letter(num_vendors + 2)].width = 14

    # Freeze panes at row 5 (after title/subtitle)
    ws.freeze_panes = 'A5'

    # Save workbook
    wb.save(excel_file)

    return excel_file


def run_pipeline(input_file: str, vendor_name: str):
    """
    Execute complete TCO pipeline:
    1. Extract proposal data to JSON
    2. Generate Excel TCO report from JSON
    3. Add Normalized Comparison sheet (6-bucket structure for vendor comparison)
    """
    print("=" * 80)
    print("TCO AUTOMATION PIPELINE - SINGLE COMMAND MODE")
    print("=" * 80)
    print(f"Input file: {input_file}")
    print(f"Vendor: {vendor_name}")
    print()

    # Validate input file exists
    if not Path(input_file).exists():
        print(f"ERROR: Input file not found: {input_file}")
        sys.exit(1)

    # Step 1: Extract proposal data
    print("STEP 1: Extracting proposal data...")
    print("-" * 80)

    extraction_cmd = [sys.executable, "extract_proposal.py", input_file, vendor_name]
    result = subprocess.run(extraction_cmd, capture_output=False)

    if result.returncode != 0:
        print(f"\nERROR: Extraction failed with exit code {result.returncode}")
        sys.exit(1)

    print()
    print("[OK] Extraction completed successfully")
    print()

    # Step 2: Generate Excel TCO report
    print("STEP 2: Generating Excel TCO report...")
    print("-" * 80)

    # Determine JSON file path (using AI-enhanced version)
    json_file = f"Extracted JSON/{vendor_name}_extraction_ai.json"

    if not Path(json_file).exists():
        # Try alternative naming convention
        json_file = f"Extracted JSON/{vendor_name.lower().replace(' ', '_')}_extraction_ai.json"

    if not Path(json_file).exists():
        print(f"\nERROR: Expected JSON file not found: {json_file}")
        print("Extraction may have failed or used different naming.")
        sys.exit(1)

    excel_cmd = [sys.executable, "scripts/json_to_excel_mapper.py", json_file]
    result = subprocess.run(excel_cmd, capture_output=False)

    if result.returncode != 0:
        print(f"\nERROR: Excel generation failed with exit code {result.returncode}")
        sys.exit(1)

    print()
    print("[OK] Excel generation completed successfully")
    print()

    # Step 3: Add Normalized Comparison sheet
    print("STEP 3: Adding Normalized Comparison sheet...")
    print("-" * 80)

    excel_file = add_normalized_comparison_sheet(json_file, vendor_name)

    if excel_file:
        print()
        print("[OK] Normalized Comparison sheet added successfully")
        print()
    else:
        print()
        print("[WARNING] Could not add Normalized Comparison sheet - continuing")
        print()
        excel_file = f"TCO Output/{vendor_name.upper()}_TCO_New_*.xlsx"

    # Summary
    print("=" * 80)
    print("PIPELINE COMPLETED SUCCESSFULLY")
    print("=" * 80)
    print(f"Extracted JSON: {json_file}")
    print(f"TCO Excel: {excel_file}")
    print(f"Normalized Comparison: Included in Excel (6-bucket structure)")
    print()


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python run_tco_pipeline.py <input_file> <vendor_name>")
        print()
        print("Examples:")
        print("  python run_tco_pipeline.py 'proposal.pdf' 'csi'")
        print("  python run_tco_pipeline.py 'WORKBOOK1.xlsx' 'liberty'")
        print("  python run_tco_pipeline.py 'deal_sheet.xlsx' 'jh'")
        print()
        print("Vendor names: fis, jh, csi, liberty")
        sys.exit(1)

    input_file = sys.argv[1]
    vendor_name = sys.argv[2]

    run_pipeline(input_file, vendor_name)
