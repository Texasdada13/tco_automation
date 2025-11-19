"""
Jack Henry Excel Extractor
Extracts pricing data from Jack Henry deal sheet Excel files
"""

from openpyxl import load_workbook
import pandas as pd
from typing import Dict, List, Any, Tuple


class JackHenryExtractor:
    """Extract structured data from Jack Henry Excel proposals"""

    def __init__(self, filepath: str):
        self.filepath = filepath
        self.wb = load_workbook(filepath, data_only=True)
        self.wb_formulas = load_workbook(filepath, data_only=False)  # For formula extraction
        self.extracted_data = {
            'vendor': 'Jack Henry',
            'summary': {},
            'products': [],
            'term_options': None,
            'scenarios': [],
            'cell_validation': {
                'total_cells_checked': 0,
                'cells_with_data': 0,
                'cells_extracted': 0,
                'cells_with_formulas': 0,
                'cells_with_comments': 0,
                'hidden_rows': [],
                'hidden_columns': []
            },
            'comments': {},
            'formulas': {}
        }
        self.column_schemas = {}  # Will store expected data types per column
    
    def extract(self) -> Dict[str, Any]:
        """Main extraction method"""
        print(f"Extracting data from Jack Henry proposal: {self.filepath}")
        print(f"Available sheets: {self.wb.sheetnames}")
        
        # Extract from Summary sheet
        self._extract_summary()
        
        # Extract from Proposal sheets
        for sheet_name in ['Proposal_1', 'Proposal_2', 'Proposal_3']:
            if sheet_name in self.wb.sheetnames:
                self._extract_proposal(sheet_name)
        
        return self.extracted_data
    
    def _extract_summary(self):
        """Extract high-level summary data"""
        if 'Summary' not in self.wb.sheetnames:
            print("Warning: Summary sheet not found")
            return
        
        print("\nExtracting Summary sheet...")
        sheet = self.wb['Summary']
        
        # Extract key summary metrics
        summary = {}
        
        # Look for bank name, assets, etc. in first ~20 rows
        for row_idx in range(1, 20):
            row = sheet[row_idx]
            for cell in row[:5]:
                if cell.value and isinstance(cell.value, str):
                    if ': Bank Name' in cell.value or 'Bank Name' in cell.value:
                        summary['bank_name'] = row[0].value if row[0].value else 'Unknown'
                    elif ': Assets' in cell.value or 'Asset Size' in cell.value:
                        summary['assets'] = row[1].value if len(row) > 1 else 0
                    elif ': JHA Core Product' in cell.value:
                        summary['core_product'] = row[1].value if len(row) > 1 else 'Unknown'
        
        # Extract monthly fees summary (around rows 28-32)
        for row_idx in range(25, 35):
            row = sheet[row_idx]
            cell_value = str(row[0].value) if row[0].value else ''
            
            if 'Monthly Fees' in cell_value:
                summary['total_monthly_list'] = row[1].value if len(row) > 1 else 0
            elif 'Net Monthly Fees' in cell_value:
                summary['total_monthly_net'] = row[1].value if len(row) > 1 else 0
            elif 'Annualized List Price' in cell_value:
                summary['annualized_list'] = row[1].value if len(row) > 1 else 0
            elif 'Annualized Net Price' in cell_value:
                summary['annualized_net'] = row[1].value if len(row) > 1 else 0
        
        self.extracted_data['summary'] = summary
        print(f"Summary extracted: {summary}")
    
    def _extract_proposal(self, sheet_name: str):
        """Extract detailed product pricing from proposal sheet"""
        print(f"\nExtracting {sheet_name}...")
        sheet = self.wb[sheet_name]
        sheet_formulas = self.wb_formulas[sheet_name]

        # Check for hidden rows and columns
        self._check_hidden_data(sheet, sheet_name)

        # Extract cell comments
        self._extract_sheet_comments(sheet, sheet_name)

        # Find the header row (usually around row 11-13)
        header_row_idx = None
        for row_idx in range(1, 20):
            row = sheet[row_idx]
            if any('Product Description' in str(cell.value) for cell in row[:10] if cell.value):
                header_row_idx = row_idx
                break

        if not header_row_idx:
            print(f"Warning: Could not find header row in {sheet_name}")
            return

        print(f"Found header row at index {header_row_idx}")

        # Get column indices
        header_row = sheet[header_row_idx]
        col_indices = self._get_column_indices(header_row)

        # Build column schema for validation
        self._build_column_schema(col_indices)

        # Dynamically find data boundaries
        start_row, end_row = self._find_data_boundaries(sheet, header_row_idx)
        print(f"Data range: rows {start_row} to {end_row}")

        # Extract products from data range
        products = []
        for row_idx in range(start_row, end_row + 1):
            row = sheet[row_idx]
            row_formulas = sheet_formulas[row_idx]

            # Get product description from column B (index 1)
            product_desc = row[1].value if len(row) > 1 else None

            # Skip empty rows or section headers
            if not product_desc or product_desc == 'Product Description':
                continue

            # Skip if it looks like a header or summary row
            if isinstance(product_desc, str) and any(x in product_desc.lower() for x in ['total', 'summary', 'subtotal']):
                continue

            # Validate all cells in this row
            self._validate_row_cells(row, row_formulas, col_indices, row_idx)

            # Extract product data
            product_data = self._extract_product_row(row, col_indices, row_idx)
            if product_data:
                products.append(product_data)

        print(f"Extracted {len(products)} products from {sheet_name}")

        # Store in extracted_data
        self.extracted_data['scenarios'].append({
            'scenario_name': sheet_name,
            'products': products,
            'product_count': len(products)
        })
    
    def _get_column_indices(self, header_row) -> Dict[str, int]:
        """Map column names to indices"""
        col_map = {}
        
        for idx, cell in enumerate(header_row):
            if cell.value:
                value = str(cell.value).strip()
                
                if 'Product Description' in value:
                    col_map['product_description'] = idx
                elif 'Order Type' in value:
                    col_map['order_type'] = idx
                elif 'Delivery' in value:
                    col_map['delivery'] = idx
                elif 'Optional' in value:
                    col_map['optional'] = idx
                elif 'Category' in value:
                    col_map['category'] = idx
                elif 'Product Family' in value:
                    col_map['product_family'] = idx
                elif 'Quantity' in value:
                    col_map['quantity'] = idx
                elif 'License List' in value:
                    col_map['license_list'] = idx
                elif 'License Net' in value:
                    col_map['license_net'] = idx
                elif 'Install List' in value:
                    col_map['install_list'] = idx
                elif 'Install Net' in value:
                    col_map['install_net'] = idx
                elif 'Maintenance List' in value:
                    col_map['maintenance_list'] = idx
                elif 'Maintenance Net' in value:
                    col_map['maintenance_net'] = idx
                elif 'New Monthly List' in value:
                    col_map['monthly_list'] = idx
                elif 'New Monthly Net' in value:
                    col_map['monthly_net'] = idx
        
        return col_map
    
    def _extract_product_row(self, row, col_indices: Dict[str, int], row_idx: int) -> Dict[str, Any]:
        """Extract data from a single product row"""
        try:
            product_desc = row[col_indices.get('product_description', 1)].value
            if not product_desc:
                return None

            product = {
                'product_description': str(product_desc),
                'order_type': self._get_cell_value(row, col_indices, 'order_type', ''),
                'delivery': self._get_cell_value(row, col_indices, 'delivery', ''),
                'optional': self._get_cell_value(row, col_indices, 'optional', 0),
                'category': self._get_cell_value(row, col_indices, 'category', ''),
                'product_family': self._get_cell_value(row, col_indices, 'product_family', ''),
                'quantity': self._get_cell_value(row, col_indices, 'quantity', 0, numeric=True),
                'license_list': self._get_cell_value(row, col_indices, 'license_list', 0, numeric=True),
                'license_net': self._get_cell_value(row, col_indices, 'license_net', 0, numeric=True),
                'install_list': self._get_cell_value(row, col_indices, 'install_list', 0, numeric=True),
                'install_net': self._get_cell_value(row, col_indices, 'install_net', 0, numeric=True),
                'maintenance_list': self._get_cell_value(row, col_indices, 'maintenance_list', 0, numeric=True),
                'maintenance_net': self._get_cell_value(row, col_indices, 'maintenance_net', 0, numeric=True),
                'monthly_list': self._get_cell_value(row, col_indices, 'monthly_list', 0, numeric=True),
                'monthly_net': self._get_cell_value(row, col_indices, 'monthly_net', 0, numeric=True),
                'row_number': row_idx  # Track source row for validation
            }

            return product

        except Exception as e:
            print(f"Error extracting product row {row_idx}: {e}")
            print(f"  Product: {product_desc if 'product_desc' in locals() else 'Unknown'}")
            print(f"  Cell values (first 10): {[c.value for c in row[:10]]}")
            return None
    
    def _get_cell_value(self, row, col_indices: Dict[str, int], key: str, default, numeric: bool = False):
        """Safely get cell value with default"""
        try:
            idx = col_indices.get(key)
            if idx is None or idx >= len(row):
                return default
            
            value = row[idx].value
            
            if value is None:
                return default
            
            if numeric:
                return float(value) if value else 0.0
            
            return value
        except:
            return default
    
    def _check_hidden_data(self, sheet, sheet_name: str):
        """Check for hidden rows and columns that might contain data"""
        hidden_rows = []
        hidden_cols = []

        # Check hidden rows
        for row_idx in range(1, sheet.max_row + 1):
            if row_idx in sheet.row_dimensions and sheet.row_dimensions[row_idx].hidden:
                # Check if hidden row has data
                row_has_data = any(sheet.cell(row_idx, col).value for col in range(1, 20))
                if row_has_data:
                    hidden_rows.append(row_idx)

        # Check hidden columns
        for col_idx in range(1, sheet.max_column + 1):
            col_letter = sheet.cell(1, col_idx).column_letter
            if col_letter in sheet.column_dimensions and sheet.column_dimensions[col_letter].hidden:
                # Check if hidden column has data
                col_has_data = any(sheet.cell(row, col_idx).value for row in range(1, 20))
                if col_has_data:
                    hidden_cols.append(col_letter)

        if hidden_rows:
            print(f"  WARNING: {len(hidden_rows)} hidden rows with data in {sheet_name}: {hidden_rows[:5]}")
            self.extracted_data['cell_validation']['hidden_rows'].extend(
                [f"{sheet_name}:{r}" for r in hidden_rows]
            )

        if hidden_cols:
            print(f"  WARNING: {len(hidden_cols)} hidden columns with data in {sheet_name}: {hidden_cols[:5]}")
            self.extracted_data['cell_validation']['hidden_columns'].extend(
                [f"{sheet_name}:{c}" for c in hidden_cols]
            )

    def _extract_sheet_comments(self, sheet, sheet_name: str):
        """Extract all cell comments which may contain important pricing notes"""
        comment_count = 0

        for row in sheet.iter_rows():
            for cell in row:
                if cell.comment:
                    comment_key = f"{sheet_name}!{cell.coordinate}"
                    self.extracted_data['comments'][comment_key] = {
                        'cell': cell.coordinate,
                        'value': cell.value,
                        'comment': cell.comment.text
                    }
                    comment_count += 1

        if comment_count > 0:
            print(f"  Extracted {comment_count} cell comments from {sheet_name}")
            self.extracted_data['cell_validation']['cells_with_comments'] += comment_count

    def _build_column_schema(self, col_indices: Dict[str, int]):
        """Build expected data type schema for columns"""
        numeric_fields = [
            'quantity', 'license_list', 'license_net', 'install_list',
            'install_net', 'maintenance_list', 'maintenance_net',
            'monthly_list', 'monthly_net'
        ]

        string_fields = [
            'product_description', 'order_type', 'delivery',
            'category', 'product_family'
        ]

        for field in numeric_fields:
            if field in col_indices:
                self.column_schemas[col_indices[field]] = (int, float)

        for field in string_fields:
            if field in col_indices:
                self.column_schemas[col_indices[field]] = str

    def _find_data_boundaries(self, sheet, header_row_idx: int) -> Tuple[int, int]:
        """Dynamically find where actual data starts and ends"""
        start_row = header_row_idx + 1
        end_row = start_row
        empty_count = 0
        max_empty_threshold = 5  # Stop after 5 consecutive empty rows

        for row_idx in range(start_row, sheet.max_row + 1):
            row = sheet[row_idx]
            product_desc = row[1].value if len(row) > 1 else None

            # Check if this is a data row
            if product_desc and not isinstance(product_desc, str):
                # Convert to string if it's a number
                product_desc = str(product_desc)

            is_data_row = (
                product_desc
                and product_desc != 'Product Description'
                and not (isinstance(product_desc, str) and any(
                    x in product_desc.lower() for x in ['total', 'summary', 'subtotal']
                ))
            )

            if is_data_row:
                end_row = row_idx
                empty_count = 0
            else:
                empty_count += 1
                if empty_count >= max_empty_threshold:
                    break

        return start_row, end_row

    def _validate_row_cells(self, row, row_formulas, col_indices: Dict[str, int], row_idx: int):
        """Validate all cells in a row for data types and extract formulas"""
        for field_name, col_idx in col_indices.items():
            if col_idx >= len(row):
                continue

            cell = row[col_idx]
            cell_formula = row_formulas[col_idx]

            # Track total cells checked
            self.extracted_data['cell_validation']['total_cells_checked'] += 1

            # Check if cell has data
            if cell.value is not None:
                self.extracted_data['cell_validation']['cells_with_data'] += 1
                self.extracted_data['cell_validation']['cells_extracted'] += 1

                # Validate data type against schema
                if col_idx in self.column_schemas:
                    expected_types = self.column_schemas[col_idx]
                    if not isinstance(expected_types, tuple):
                        expected_types = (expected_types,)

                    if cell.value and not isinstance(cell.value, expected_types):
                        # Allow for type conversions (e.g., string numbers to float)
                        if (int, float) in expected_types or expected_types == (int, float):
                            try:
                                float(cell.value)
                            except (ValueError, TypeError):
                                print(f"  WARNING: Type mismatch at row {row_idx}, {field_name}: "
                                      f"expected {expected_types}, got {type(cell.value).__name__}")

            # Extract formula if present
            if cell_formula.value and isinstance(cell_formula.value, str) and cell_formula.value.startswith('='):
                formula_key = f"Row{row_idx}_{field_name}"
                self.extracted_data['formulas'][formula_key] = {
                    'cell': cell.coordinate,
                    'formula': cell_formula.value,
                    'calculated_value': cell.value
                }
                self.extracted_data['cell_validation']['cells_with_formulas'] += 1

    def get_summary(self) -> Dict[str, Any]:
        """Get summary of extracted data"""
        total_products = sum(scenario['product_count'] for scenario in self.extracted_data['scenarios'])

        return {
            'vendor': self.extracted_data['vendor'],
            'bank_name': self.extracted_data['summary'].get('bank_name', 'Unknown'),
            'core_product': self.extracted_data['summary'].get('core_product', 'Unknown'),
            'scenarios_count': len(self.extracted_data['scenarios']),
            'total_products': total_products,
            'monthly_net': self.extracted_data['summary'].get('total_monthly_net', 0)
        }

    def get_validation_report(self) -> Dict[str, Any]:
        """Get detailed validation report"""
        val = self.extracted_data['cell_validation']

        coverage_pct = 0
        if val['cells_with_data'] > 0:
            coverage_pct = (val['cells_extracted'] / val['cells_with_data']) * 100

        return {
            'coverage_percentage': round(coverage_pct, 2),
            'total_cells_checked': val['total_cells_checked'],
            'cells_with_data': val['cells_with_data'],
            'cells_extracted': val['cells_extracted'],
            'cells_with_formulas': val['cells_with_formulas'],
            'cells_with_comments': val['cells_with_comments'],
            'hidden_rows_count': len(val['hidden_rows']),
            'hidden_columns_count': len(val['hidden_columns']),
            'total_comments': len(self.extracted_data['comments']),
            'total_formulas': len(self.extracted_data['formulas'])
        }


def extract_jack_henry_proposal(filepath: str) -> Dict[str, Any]:
    """Convenience function to extract Jack Henry proposal data"""
    extractor = JackHenryExtractor(filepath)
    data = extractor.extract()
    summary = extractor.get_summary()
    validation = extractor.get_validation_report()

    print("\n" + "="*70)
    print("EXTRACTION SUMMARY")
    print("="*70)
    for key, value in summary.items():
        print(f"{key}: {value}")
    print("="*70)

    print("\n" + "="*70)
    print("VALIDATION REPORT")
    print("="*70)
    for key, value in validation.items():
        print(f"{key}: {value}")
    print("="*70)

    # Print sample comments if any
    if data['comments']:
        print(f"\nSample Cell Comments (showing first 3):")
        for idx, (cell_ref, comment_data) in enumerate(list(data['comments'].items())[:3]):
            print(f"  {cell_ref}: {comment_data['comment'][:100]}...")

    # Print sample formulas if any
    if data['formulas']:
        print(f"\nSample Formulas (showing first 3):")
        for idx, (key, formula_data) in enumerate(list(data['formulas'].items())[:3]):
            print(f"  {formula_data['cell']}: {formula_data['formula']} = {formula_data['calculated_value']}")

    return data


if __name__ == "__main__":
    # Test extraction
    import sys
    if len(sys.argv) > 1:
        filepath = sys.argv[1]
    else:
        filepath = "../data/Updated_With_all_products__Deal_Sheet_Clearwater__FL_-_Echelon_Bank__InOrg__-_New_Core_SilverLake_OL_PAP_08_27_25.xlsx"
    
    data = extract_jack_henry_proposal(filepath)
    
    # Print sample products
    if data['scenarios']:
        print(f"\nSample products from {data['scenarios'][0]['scenario_name']}:")
        for product in data['scenarios'][0]['products'][:5]:
            print(f"  - {product['product_description']}: ${product['monthly_net']}/month")
