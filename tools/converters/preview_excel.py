"""Quick preview of Excel vendor extract"""
import openpyxl
from pathlib import Path

excel_file = "Vendor Extracts/Echelon Bank Fis_Extract_20260113.xlsx"

wb = openpyxl.load_workbook(excel_file)
ws = wb.active

print(f"\n{'='*80}")
print(f"PREVIEW: {excel_file}")
print(f"{'='*80}\n")

# Show header info (rows 1-4)
print("HEADER INFORMATION:")
print(f"  {ws['A2'].value} {ws['B2'].value}")
print(f"  {ws['A3'].value} {ws['B3'].value}")
print(f"  {ws['A4'].value} {ws['B4'].value}")
print(f"  {ws['E2'].value} {ws['F2'].value}")
print(f"  {ws['E3'].value} {ws['F3'].value}")
print()

# Show column headers (row 6)
print("COLUMNS:")
headers = []
for col in range(1, ws.max_column + 1):
    header = ws.cell(6, col).value
    if header:
        headers.append(header)
        if col <= 7:
            print(f"  {col}. {header}")

print(f"  ... Year columns (Year 1 through Year 10)")
print(f"  {len(headers)-1}. {headers[-2]}")  # Total
print(f"  {len(headers)}. {headers[-1]}")    # Notes
print()

# Show first 5 line items
print("FIRST 5 LINE ITEMS:")
for row in range(7, min(12, ws.max_row)):
    solution = ws.cell(row, 1).value
    category = ws.cell(row, 2).value
    fee_type = ws.cell(row, 3).value
    monthly = ws.cell(row, 4).value
    year_1 = ws.cell(row, 8).value
    total = ws.cell(row, ws.max_column - 1).value

    print(f"\n  {row-6}. {solution}")
    print(f"     Category: {category} | Type: {fee_type}")
    print(f"     Monthly: ${monthly:,.2f} | Year 1 Annual: ${year_1:,.2f}")
    if isinstance(total, str) and total.startswith('='):
        print(f"     Total (10 years): [Formula: {total}]")

print(f"\n  ... (showing 5 of {ws.max_row - 7} line items)")

# Show summary row
summary_row = ws.max_row
print(f"\n{'='*80}")
print("SUMMARY ROW:")
print(f"  {ws.cell(summary_row, 1).value}")
year_1_total = ws.cell(summary_row, 8).value
print(f"  Year 1 Total: {year_1_total}")
print(f"  Grand Total: {ws.cell(summary_row, ws.max_column - 1).value}")

print(f"\n{'='*80}\n")

wb.close()
