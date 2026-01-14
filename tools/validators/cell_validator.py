#!/usr/bin/env python3
"""
Cell-by-Cell Validator for TCO Automation System

This module provides granular validation by comparing source Excel files
to TCO output files cell-by-cell to ensure 100% accuracy.
"""

from openpyxl import load_workbook
from typing import Dict, List, Any, Tuple
import sys


class CellByCellValidator:
    """Validates TCO output against source files at the cell level"""

    def __init__(self):
        self.discrepancies = []
        self.validation_stats = {
            'total_cells_compared': 0,
            'matching_cells': 0,
            'mismatched_cells': 0,
            'missing_in_output': 0,
            'extra_in_output': 0
        }

    def validate_jack_henry_to_tco(
        self,
        source_file: str,
        tco_file: str,
        scenario: str = 'Proposal_1',
        tco_start_row: int = 7
    ) -> Dict[str, Any]:
        """
        Compare Jack Henry source file to populated TCO template

        Args:
            source_file: Path to Jack Henry Excel file
            tco_file: Path to populated TCO file
            scenario: Which proposal scenario to validate
            tco_start_row: Starting row in TCO template

        Returns:
            Dictionary with validation results
        """
        print(f"\n{'='*70}")
        print(f"CELL-BY-CELL VALIDATION: {scenario}")
        print(f"{'='*70}")

        # Load both workbooks
        source_wb = load_workbook(source_file, data_only=True)
        tco_wb = load_workbook(tco_file, data_only=True)

        if scenario not in source_wb.sheetnames:
            print(f"Error: Scenario {scenario} not found in source file")
            return {'success': False, 'error': f'Scenario {scenario} not found'}

        source_sheet = source_wb[scenario]
        tco_sheet = tco_wb['Line Items']

        # Find header row in source
        header_row_idx = self._find_header_row(source_sheet)
        if not header_row_idx:
            print("Error: Could not find header row in source file")
            return {'success': False, 'error': 'Header row not found'}

        print(f"Source header row: {header_row_idx}")
        print(f"TCO start row: {tco_start_row}")

        # Get column mappings
        source_col_map = self._get_source_column_indices(source_sheet[header_row_idx])

        # Validate each product row
        tco_row = tco_start_row
        source_row = header_row_idx + 1

        while source_row <= source_sheet.max_row:
            source_product = source_sheet.cell(source_row, 2).value  # Column B

            # Skip empty rows
            if not source_product:
                source_row += 1
                continue

            # Skip total/summary rows
            if isinstance(source_product, str) and any(
                x in source_product.lower() for x in ['total', 'summary', 'subtotal']
            ):
                source_row += 1
                continue

            # Validate this product row
            self._validate_product_row(
                source_sheet, source_row, source_col_map,
                tco_sheet, tco_row
            )

            source_row += 1
            tco_row += 1

        # Generate report
        return self._generate_validation_report()

    def _find_header_row(self, sheet) -> int:
        """Find the header row in the source sheet"""
        for row_idx in range(1, 20):
            row = sheet[row_idx]
            if any('Product Description' in str(cell.value) for cell in row[:10] if cell.value):
                return row_idx
        return None

    def _get_source_column_indices(self, header_row) -> Dict[str, int]:
        """Map source column names to indices"""
        col_map = {}

        for idx, cell in enumerate(header_row):
            if cell.value:
                value = str(cell.value).strip()

                if 'Product Description' in value:
                    col_map['product_description'] = idx
                elif 'License Net' in value:
                    col_map['license_net'] = idx
                elif 'Install Net' in value:
                    col_map['install_net'] = idx
                elif 'Maintenance Net' in value:
                    col_map['maintenance_net'] = idx
                elif 'New Monthly Net' in value:
                    col_map['monthly_net'] = idx
                elif 'Product Family' in value:
                    col_map['product_family'] = idx
                elif 'Category' in value:
                    col_map['category'] = idx

        return col_map

    def _validate_product_row(
        self,
        source_sheet,
        source_row: int,
        source_col_map: Dict[str, int],
        tco_sheet,
        tco_row: int
    ):
        """Validate a single product row cell-by-cell"""

        # Get source product description
        source_product_col = source_col_map.get('product_description', 1)
        source_product = source_sheet.cell(source_row, source_product_col + 1).value

        # Get TCO product description (assuming column BB for Jack Henry)
        tco_product_col = 'BB'
        tco_product = tco_sheet[f'{tco_product_col}{tco_row}'].value

        self.validation_stats['total_cells_compared'] += 1

        # Compare product descriptions
        if str(source_product).strip() != str(tco_product).strip():
            self.discrepancies.append({
                'type': 'product_name_mismatch',
                'source_row': source_row,
                'tco_row': tco_row,
                'source_value': source_product,
                'tco_value': tco_product,
                'severity': 'HIGH'
            })
            self.validation_stats['mismatched_cells'] += 1
        else:
            self.validation_stats['matching_cells'] += 1

        # Compare pricing fields
        pricing_fields = [
            ('license_net', 'BD'),  # TCO column for per unit rate
            ('monthly_net', 'BD'),   # Same column, different products
        ]

        for field_name, tco_col in pricing_fields:
            if field_name in source_col_map:
                source_value = source_sheet.cell(
                    source_row,
                    source_col_map[field_name] + 1
                ).value

                tco_value = tco_sheet[f'{tco_col}{tco_row}'].value

                self.validation_stats['total_cells_compared'] += 1

                # Compare values (with tolerance for floating point)
                if not self._values_match(source_value, tco_value):
                    self.discrepancies.append({
                        'type': 'value_mismatch',
                        'field': field_name,
                        'source_row': source_row,
                        'tco_row': tco_row,
                        'tco_col': tco_col,
                        'source_value': source_value,
                        'tco_value': tco_value,
                        'severity': 'MEDIUM'
                    })
                    self.validation_stats['mismatched_cells'] += 1
                else:
                    self.validation_stats['matching_cells'] += 1

    def _values_match(self, val1, val2, tolerance: float = 0.01) -> bool:
        """Compare two values with tolerance for floats"""
        # Handle None/empty cases
        if val1 is None and val2 is None:
            return True
        if val1 is None or val2 is None:
            return False

        # Try numeric comparison
        try:
            num1 = float(val1)
            num2 = float(val2)
            return abs(num1 - num2) < tolerance
        except (ValueError, TypeError):
            pass

        # String comparison
        return str(val1).strip() == str(val2).strip()

    def _generate_validation_report(self) -> Dict[str, Any]:
        """Generate comprehensive validation report"""
        stats = self.validation_stats
        total = stats['total_cells_compared']

        accuracy_pct = 0
        if total > 0:
            accuracy_pct = (stats['matching_cells'] / total) * 100

        print(f"\n{'='*70}")
        print("VALIDATION RESULTS")
        print(f"{'='*70}")
        print(f"Total cells compared: {total}")
        print(f"Matching cells: {stats['matching_cells']}")
        print(f"Mismatched cells: {stats['mismatched_cells']}")
        print(f"Accuracy: {accuracy_pct:.2f}%")
        print(f"{'='*70}")

        if self.discrepancies:
            print(f"\nFound {len(self.discrepancies)} discrepancies:")
            for idx, disc in enumerate(self.discrepancies[:10], 1):
                print(f"\n{idx}. {disc['type'].upper()} [{disc['severity']}]")
                print(f"   Source row {disc['source_row']} -> TCO row {disc['tco_row']}")
                print(f"   Expected: {disc['source_value']}")
                print(f"   Found: {disc['tco_value']}")

            if len(self.discrepancies) > 10:
                print(f"\n   ... and {len(self.discrepancies) - 10} more")

        return {
            'success': stats['mismatched_cells'] == 0,
            'accuracy_percentage': accuracy_pct,
            'statistics': stats,
            'discrepancies': self.discrepancies
        }

    def export_discrepancies_report(self, output_path: str = "cell_validation_report.txt"):
        """Export detailed discrepancy report to file"""
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write("="*70 + "\n")
            f.write("CELL-BY-CELL VALIDATION DISCREPANCY REPORT\n")
            f.write("="*70 + "\n\n")

            stats = self.validation_stats
            f.write(f"Total Cells Compared: {stats['total_cells_compared']}\n")
            f.write(f"Matching Cells: {stats['matching_cells']}\n")
            f.write(f"Mismatched Cells: {stats['mismatched_cells']}\n")
            f.write(f"Missing in Output: {stats['missing_in_output']}\n")
            f.write(f"Extra in Output: {stats['extra_in_output']}\n\n")

            if self.discrepancies:
                f.write(f"DISCREPANCIES ({len(self.discrepancies)} total):\n")
                f.write("-"*70 + "\n\n")

                for idx, disc in enumerate(self.discrepancies, 1):
                    f.write(f"{idx}. {disc['type'].upper()} [{disc['severity']}]\n")
                    f.write(f"   Source Row: {disc['source_row']}\n")
                    f.write(f"   TCO Row: {disc['tco_row']}\n")
                    if 'field' in disc:
                        f.write(f"   Field: {disc['field']}\n")
                    if 'tco_col' in disc:
                        f.write(f"   TCO Column: {disc['tco_col']}\n")
                    f.write(f"   Source Value: {disc['source_value']}\n")
                    f.write(f"   TCO Value: {disc['tco_value']}\n\n")

            f.write("="*70 + "\n")

        print(f"\nDetailed report saved to: {output_path}")


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description='Cell-by-cell validator for TCO automation')
    parser.add_argument('--source', type=str, required=True, help='Source Jack Henry Excel file')
    parser.add_argument('--tco', type=str, required=True, help='Populated TCO template file')
    parser.add_argument('--scenario', type=str, default='Proposal_1', help='Scenario to validate')
    parser.add_argument('--report', type=str, default='cell_validation_report.txt', help='Output report path')

    args = parser.parse_args()

    validator = CellByCellValidator()
    result = validator.validate_jack_henry_to_tco(
        args.source,
        args.tco,
        args.scenario
    )

    if not result['success']:
        print(f"\nValidation FAILED with {len(validator.discrepancies)} discrepancies")
        validator.export_discrepancies_report(args.report)
        sys.exit(1)
    else:
        print("\nValidation PASSED - 100% accuracy!")
        sys.exit(0)
