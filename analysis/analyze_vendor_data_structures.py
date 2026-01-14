"""
Analyze data structures from multiple vendors to create universal schema
"""

import json
import openpyxl
from openpyxl.utils import get_column_letter

def analyze_workbook2_structure():
    """Analyze WORKBOOK2 to extract column structure"""
    wb = openpyxl.load_workbook('WORKBOOK2.xlsx')
    ws = wb['Line Items']

    print("="*120)
    print("WORKBOOK2 COLUMN STRUCTURE ANALYSIS")
    print("="*120)

    # Get headers from row 4 and 5
    columns = {}
    for col in range(1, 41):  # Check first 40 columns
        col_letter = get_column_letter(col)
        row4_val = ws.cell(4, col).value
        row5_val = ws.cell(5, col).value

        if row4_val or row5_val:
            columns[col] = {
                'letter': col_letter,
                'index': col,
                'header_row4': row4_val,
                'header_row5': row5_val
            }

    # Print key columns
    print("\nKEY DATA COLUMNS:")
    print("-" * 120)

    key_cols = [
        (2, 'B', 'Fee Type'),
        (3, 'C', 'Proposal Quantity'),
        (4, 'D', 'Average Monthly QTY'),
        (15, 'O', 'Solution Name/Description'),
        (16, 'P', 'Category'),
        (17, 'Q', 'Per Unit Rate'),
        (18, 'R', 'Year 1 Monthly Cost'),
        (19, 'S', 'Year 1 Cost (Annual)')
    ]

    for idx, letter, desc in key_cols:
        row4 = ws.cell(4, idx).value
        row5 = ws.cell(5, idx).value
        print(f"Column {letter} (idx {idx:2d}): {desc:30s}")
        print(f"  Row 4: {row4}")
        print(f"  Row 5: {row5}")
        print()

    # Sample a few line items
    print("="*120)
    print("SAMPLE LINE ITEMS (Rows 7-12)")
    print("="*120)

    for row in range(7, 13):
        fee_type = ws.cell(row, 2).value
        solution = ws.cell(row, 15).value
        category = ws.cell(row, 16).value
        rate = ws.cell(row, 17).value

        if solution:
            print(f"Row {row}:")
            print(f"  Solution: {solution}")
            print(f"  Fee Type: {fee_type}")
            print(f"  Category: {category}")
            print(f"  Rate: {rate}")
            print()

    wb.close()
    return columns


def analyze_vendor_extractions():
    """Analyze all vendor JSON extractions to understand data variations"""

    print("\n" + "="*120)
    print("VENDOR EXTRACTION ANALYSIS")
    print("="*120)

    vendor_files = {
        'Echelon FIS': 'Extracted JSON/echelon_bank_fis_extraction_ai.json',
        'Liberty Capital FIS': 'Extracted JSON/liberty_capital_bank_fis_extraction_ai.json',
        'Liberty Capital CSI': 'Extracted JSON/liberty_capital_bank_csi_extraction_ai.json'
    }

    for vendor_name, file_path in vendor_files.items():
        try:
            with open(file_path, 'r') as f:
                data = json.load(f)

            print(f"\n{vendor_name}")
            print("-" * 120)
            print(f"Contract Term: {data.get('contract_term')} years")
            print(f"Total Items: {len(data.get('line_items', []))}")

            # Analyze first line item structure
            if data.get('line_items'):
                item = data['line_items'][0]
                print(f"\nSample Item Fields:")
                for key, value in item.items():
                    print(f"  {key:25s}: {type(value).__name__:10s} = {value}")

            # Summary stats
            summary = data.get('summary', {})
            print(f"\nSummary:")
            print(f"  Monthly Required:  ${summary.get('total_monthly_required', 0):,.2f}")
            print(f"  Monthly Optional:  ${summary.get('total_monthly_optional', 0):,.2f}")
            print(f"  One-Time Fees:     ${summary.get('total_one_time_fees', 0):,.2f}")
            print(f"  One-Time Credits:  ${summary.get('total_one_time_credits', 0):,.2f}")

        except Exception as e:
            print(f"  ERROR: {e}")


def create_universal_schema():
    """Define universal column schema for all vendors"""

    print("\n" + "="*120)
    print("PROPOSED UNIVERSAL EXCEL SCHEMA")
    print("="*120)

    schema = [
        {
            'column': 'A',
            'name': 'Item #',
            'type': 'integer',
            'description': 'Sequential line item number',
            'required': True
        },
        {
            'column': 'B',
            'name': 'Solution Name',
            'type': 'string',
            'description': 'Full product/service name',
            'required': True
        },
        {
            'column': 'C',
            'name': 'Category',
            'type': 'string',
            'description': 'Product category (standardized)',
            'required': True
        },
        {
            'column': 'D',
            'name': 'Fee Type',
            'type': 'enum',
            'description': 'Monthly F, Monthly V, Annual, One-Time',
            'required': True
        },
        {
            'column': 'E',
            'name': 'Monthly Fee',
            'type': 'currency',
            'description': 'Fixed monthly fee (or estimated for variable)',
            'required': False
        },
        {
            'column': 'F',
            'name': 'Per Unit Rate',
            'type': 'currency',
            'description': 'Cost per unit (for variable fees)',
            'required': False
        },
        {
            'column': 'G',
            'name': 'Unit Description',
            'type': 'string',
            'description': 'What the unit represents (e.g., per transaction, per user)',
            'required': False
        },
        {
            'column': 'H',
            'name': 'Estimated Volume',
            'type': 'integer',
            'description': 'Estimated monthly volume/quantity',
            'required': False
        },
        {
            'column': 'I',
            'name': 'One-Time Fee',
            'type': 'currency',
            'description': 'Implementation/setup fee (can be negative for credits)',
            'required': False
        },
        {
            'column': 'J',
            'name': 'Year 1 Annual Cost',
            'type': 'currency',
            'description': 'Total cost for Year 1 (calculated)',
            'required': True
        },
        {
            'column': 'K',
            'name': 'Year 2 Annual Cost',
            'type': 'currency',
            'description': 'Year 2 with growth applied',
            'required': False
        },
        {
            'column': 'L',
            'name': 'Year 3 Annual Cost',
            'type': 'currency',
            'description': 'Year 3 with growth applied',
            'required': False
        },
        {
            'column': 'M',
            'name': 'Year 5 Annual Cost',
            'type': 'currency',
            'description': 'Year 5 with growth applied',
            'required': False
        },
        {
            'column': 'N',
            'name': 'Year 7 Annual Cost',
            'type': 'currency',
            'description': 'Year 7 with growth applied',
            'required': False
        },
        {
            'column': 'O',
            'name': 'Optional',
            'type': 'boolean',
            'description': 'Yes/No - is this an optional product',
            'required': True
        },
        {
            'column': 'P',
            'name': 'Third Party',
            'type': 'boolean',
            'description': 'Yes/No - is this a third-party solution',
            'required': True
        },
        {
            'column': 'Q',
            'name': 'Notes',
            'type': 'string',
            'description': 'Additional notes/extraction notes',
            'required': False
        }
    ]

    print("\nCOLUMN DEFINITIONS:")
    print("-" * 120)
    for col in schema:
        print(f"{col['column']:3s} | {col['name']:25s} | {col['type']:10s} | {col['description']}")

    print("\n" + "="*120)
    print("KEY DESIGN DECISIONS:")
    print("="*120)
    print("""
1. SIMPLIFIED COLUMNS
   - Reduced from 40+ columns to 17 essential columns
   - Focus on data needed for TCO comparison

2. VENDOR-AGNOSTIC
   - Works for FIS bundle pricing, CSI organic growth, Jack Henry tiered
   - Monthly Fee + Per Unit Rate accommodates all pricing models

3. YEAR PROJECTIONS
   - Include Year 1, 2, 3, 5, 7 (not all 10 years to reduce clutter)
   - Can expand to 10 years if needed

4. NEGATIVE FEES FOR CREDITS
   - One-Time Fee can be negative for credits/incentives
   - This captures FIS credits, CSI credits, etc.

5. CONSISTENT FEE TYPES
   - Monthly F: Fixed monthly fee
   - Monthly V: Variable fee based on volume
   - Annual: Annual fee (billed yearly)
   - One-Time: Implementation, setup, or credit

6. CATEGORIZATION
   - Standardized categories across all vendors
   - Maps to Data_Dictionary/enum_mappings.json
""")

    return schema


if __name__ == '__main__':
    # Step 1: Analyze WORKBOOK2
    workbook2_cols = analyze_workbook2_structure()

    # Step 2: Analyze vendor extractions
    analyze_vendor_extractions()

    # Step 3: Propose universal schema
    schema = create_universal_schema()

    # Save schema to JSON
    with open('universal_schema.json', 'w') as f:
        json.dump(schema, f, indent=2)

    print("\n[OK] Schema saved to universal_schema.json")
