"""
Analyze WORKBOOK2.xlsx to understand expected structure for FIS line items
"""
import openpyxl
from openpyxl.utils import get_column_letter

def analyze_workbook2():
    """Read and analyze WORKBOOK2.xlsx"""

    wb = openpyxl.load_workbook('WORKBOOK2.xlsx', data_only=False)

    print(f"\n{'='*80}")
    print("WORKBOOK2.xlsx ANALYSIS")
    print(f"{'='*80}\n")

    # List all sheets
    print("SHEETS IN WORKBOOK:")
    for idx, sheet_name in enumerate(wb.sheetnames, 1):
        print(f"  {idx}. {sheet_name}")
    print()

    # Analyze each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'='*80}")
        print(f"SHEET: {sheet_name}")
        print(f"{'='*80}")
        print(f"Dimensions: {ws.max_row} rows x {ws.max_column} columns\n")

        # Show first 20 rows with actual content
        print("FIRST 20 ROWS WITH CONTENT:")
        for row in range(1, min(21, ws.max_row + 1)):
            row_data = []
            has_content = False
            for col in range(1, min(15, ws.max_column + 1)):  # First 15 columns
                cell = ws.cell(row, col)
                value = cell.value
                if value is not None:
                    has_content = True
                    # Truncate long values
                    if isinstance(value, str) and len(value) > 40:
                        value = value[:37] + "..."
                    row_data.append(f"{get_column_letter(col)}: {value}")

            if has_content:
                print(f"\n  Row {row}:")
                for item in row_data[:5]:  # Show first 5 columns with content
                    print(f"    {item}")

        print(f"\n  ... (showing first 20 rows)")

        # Look for FIS-specific content
        print(f"\n  SEARCHING FOR 'FIS' IN SHEET...")
        fis_rows = []
        for row in range(1, min(100, ws.max_row + 1)):
            for col in range(1, min(20, ws.max_column + 1)):
                cell_value = ws.cell(row, col).value
                if cell_value and 'FIS' in str(cell_value).upper():
                    fis_rows.append((row, col, cell_value))

        if fis_rows:
            print(f"  Found {len(fis_rows)} cells containing 'FIS':")
            for row, col, value in fis_rows[:10]:  # Show first 10
                print(f"    Row {row}, Col {get_column_letter(col)}: {value}")
        else:
            print(f"  No 'FIS' references found in first 100 rows")

    wb.close()
    print(f"\n{'='*80}\n")

if __name__ == '__main__':
    analyze_workbook2()
