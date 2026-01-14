"""
Read WORKBOOK2.xlsx raw structure - focus on FIS section
"""
import openpyxl

def read_workbook2_fis():
    """Read FIS section in detail"""

    wb = openpyxl.load_workbook('WORKBOOK2.xlsx')
    ws = wb['Line Items']

    print(f"\n{'='*120}")
    print("WORKBOOK2 - Line Items Sheet - FIS SECTION DETAILED VIEW")
    print(f"{'='*120}\n")

    # Focus on rows 6-30 (Bundle FIS section) and columns A-U
    print("Rows 6-30 (Bundle FIS Products Section):")
    print(f"{'='*120}\n")

    # Column headers from row 5
    headers = {}
    for col in range(1, 22):  # A through U
        header = ws.cell(5, col).value
        if header:
            headers[col] = str(header).strip()

    print("HEADERS (Row 5):")
    for col, header in headers.items():
        col_letter = chr(64 + col)
        print(f"  {col_letter}: {header}")
    print()

    # Now read rows 7-30
    print(f"\n{'='*120}")
    print("LINE ITEMS DATA:")
    print(f"{'='*120}\n")

    for row in range(7, 31):
        # Get key columns
        col_a = ws.cell(row, 1).value  # A
        col_b = ws.cell(row, 2).value  # B - Fee Type
        col_c = ws.cell(row, 3).value  # C - Proposal value (maybe?)
        col_o = ws.cell(row, 15).value  # O - FIS Solution Name
        col_p = ws.cell(row, 16).value  # P - Category
        col_q = ws.cell(row, 17).value  # Q - Per Unit Rate

        # Check if row has data
        if col_b or col_o:
            print(f"Row {row}:")
            if col_a:
                print(f"  A: {col_a}")
            if col_b:
                print(f"  B (Fee Type): {col_b}")
            if col_c:
                val = col_c if not isinstance(col_c, str) or not col_c.startswith('=') else f"[Formula: {col_c[:50]}...]"
                print(f"  C (Proposal): {val}")
            if col_o:
                print(f"  O (Solution Name): {col_o}")
            if col_p:
                print(f"  P (Category): {col_p}")
            if col_q:
                print(f"  Q (Per Unit Rate): {col_q}")
            print()

    # Check other sections
    print(f"\n{'='*120}")
    print("Rows 22-30 (Non-Bundle REQUIRED FIS Products Section):")
    print(f"{'='*120}\n")

    for row in range(22, 31):
        col_b = ws.cell(row, 2).value
        col_o = ws.cell(row, 15).value
        col_p = ws.cell(row, 16).value

        if col_b or col_o:
            print(f"Row {row}: Fee={col_b} | Solution={col_o} | Category={col_p}")

    wb.close()

if __name__ == '__main__':
    read_workbook2_fis()
