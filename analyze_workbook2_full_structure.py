"""
Complete structural analysis of WORKBOOK2.xlsx to understand exact format requirements
"""
import openpyxl
from openpyxl.utils import get_column_letter
import json

def analyze_full_structure():
    """Complete analysis of WORKBOOK2 Line Items sheet"""

    wb = openpyxl.load_workbook('WORKBOOK2.xlsx')
    ws = wb['Line Items']

    print(f"\n{'='*120}")
    print("WORKBOOK2.xlsx - COMPLETE STRUCTURAL ANALYSIS")
    print(f"{'='*120}\n")

    # Step 1: Map all column headers
    print("STEP 1: COLUMN MAPPING")
    print("="*120)

    headers_row1 = {}
    headers_row4 = {}
    headers_row5 = {}

    for col in range(1, 30):  # Check first 30 columns
        val1 = ws.cell(1, col).value
        val4 = ws.cell(4, col).value
        val5 = ws.cell(5, col).value

        if val1 or val4 or val5:
            col_letter = get_column_letter(col)
            if val1:
                headers_row1[col] = val1
            if val4:
                headers_row4[col] = val4
            if val5:
                headers_row5[col] = val5

            print(f"\nColumn {col_letter} (index {col}):")
            if val1:
                print(f"  Row 1: {val1}")
            if val4:
                print(f"  Row 4: {val4}")
            if val5:
                print(f"  Row 5: {val5}")

    # Step 2: Extract complete FIS line item structure
    print(f"\n{'='*120}")
    print("STEP 2: COMPLETE FIS LINE ITEMS WITH ALL COLUMNS")
    print("="*120)

    fis_items = []
    current_section = None

    for row in range(6, 160):
        col_a = ws.cell(row, 1).value  # A
        col_b = ws.cell(row, 2).value  # B - Fee Type
        col_c = ws.cell(row, 3).value  # C - Proposal
        col_d = ws.cell(row, 4).value  # D - Avg Monthly Qty
        col_o = ws.cell(row, 15).value  # O - Solution Name
        col_p = ws.cell(row, 16).value  # P - Category
        col_q = ws.cell(row, 17).value  # Q - Per Unit Rate
        col_r = ws.cell(row, 18).value  # R - Year 1 Monthly Cost
        col_s = ws.cell(row, 19).value  # S - Year 1 Cost

        # Check for section headers
        if col_b and isinstance(col_b, str):
            if 'Bundle FIS' in col_b:
                current_section = 'Bundle FIS Products'
                print(f"\n{'='*120}")
                print(f"SECTION: {current_section} (Row {row})")
                print(f"{'='*120}\n")
                continue
            elif 'Non-Bundle REQUIRED FIS' in col_b:
                current_section = 'Non-Bundle REQUIRED FIS Products'
                print(f"\n{'='*120}")
                print(f"SECTION: {current_section} (Row {row})")
                print(f"{'='*120}\n")
                continue
            elif 'Non-Bundle Required Third' in col_b:
                current_section = 'Non-Bundle REQUIRED Third Parties'
                print(f"\n{'='*120}")
                print(f"SECTION: {current_section} (Row {row})")
                print(f"{'='*120}\n")
                continue
            elif 'Implementation' in col_b and 'FIS' in col_b:
                current_section = 'Implementation Credits and One-Time Fees'
                print(f"\n{'='*120}")
                print(f"SECTION: {current_section} (Row {row})")
                print(f"{'='*120}\n")
                continue
            elif 'Non-Bundle Optional FIS' in col_b:
                current_section = 'Non-Bundle OPTIONAL FIS Solutions'
                print(f"\n{'='*120}")
                print(f"SECTION: {current_section} (Row {row})")
                print(f"{'='*120}\n")
                continue

        # Extract line items
        if col_b in ['Monthly F', 'Monthly V', 'Annual', 'One-Time'] and col_o:
            item = {
                'row': row,
                'section': current_section or 'Unknown',
                'fee_type': col_b,
                'proposal_qty': col_c if isinstance(col_c, (int, float)) else 0,
                'avg_monthly_qty': col_d,
                'solution_name': col_o,
                'category': col_p or '',
                'per_unit_rate': col_q or 0,
                'year_1_monthly_cost': col_r,
                'year_1_annual_cost': col_s
            }

            fis_items.append(item)

            # Print in readable format
            qty_str = f"{item['proposal_qty']:.0f}" if isinstance(item['proposal_qty'], (int, float)) else str(item['proposal_qty'])
            rate_str = f"${item['per_unit_rate']:,.4f}" if isinstance(item['per_unit_rate'], (int, float)) else str(item['per_unit_rate'])

            print(f"Row {row:3d}: {item['solution_name'][:50]:50s}")
            print(f"         Fee Type: {item['fee_type']:10s} | Category: {item['category'][:30]:30s}")
            print(f"         Qty: {qty_str:>8s} | Rate: {rate_str:>12s}")
            if isinstance(item['year_1_annual_cost'], str):
                print(f"         Year 1 Annual: [Formula]")
            elif item['year_1_annual_cost']:
                print(f"         Year 1 Annual: ${item['year_1_annual_cost']:,.2f}")
            print()

    # Step 3: Save complete structure
    print(f"\n{'='*120}")
    print("STEP 3: SAVING COMPLETE STRUCTURE TO JSON")
    print("="*120)

    output = {
        'total_items': len(fis_items),
        'column_mapping': {
            'A': 'Row ID/Notes',
            'B': 'Fee Type',
            'C': 'Proposal (11/2025)',
            'D': 'Average Monthly QTY',
            'E-N': 'Quantities by Year (1-10)',
            'O': 'Solution Name/Description',
            'P': 'Category',
            'Q': 'Per Unit Rate',
            'R': 'Year 1 Monthly Cost',
            'S': 'Year 1 Cost',
            'T-Z, AA-AN': 'Year 2-10 Costs + Total'
        },
        'sections': {},
        'line_items': fis_items
    }

    # Group by section
    for item in fis_items:
        section = item['section']
        if section not in output['sections']:
            output['sections'][section] = []
        output['sections'][section].append(item)

    with open('workbook2_complete_structure.json', 'w') as f:
        json.dump(output, f, indent=2, default=str)

    print(f"\nSaved to: workbook2_complete_structure.json")
    print(f"Total FIS line items: {len(fis_items)}")

    # Summary by section
    print(f"\n{'='*120}")
    print("SUMMARY BY SECTION:")
    print("="*120)

    for section, items in output['sections'].items():
        print(f"\n{section}: {len(items)} items")

    wb.close()

if __name__ == '__main__':
    analyze_full_structure()
