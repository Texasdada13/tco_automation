#!/usr/bin/env python3
"""
Reporting Module for TCO Automation System

Provides:
1. ProcessingMetrics - Surfaces processing metrics to users
2. TraceabilityReport - Generates source-to-output traceability reports
"""

import os
import json
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional
from dataclasses import dataclass, field
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =============================================================================
# PROCESSING METRICS
# =============================================================================

@dataclass
class StageMetrics:
    """Metrics for a single processing stage."""
    name: str
    start_time: Optional[float] = None
    end_time: Optional[float] = None
    items_processed: int = 0
    items_successful: int = 0
    items_failed: int = 0
    warnings: List[str] = field(default_factory=list)

    @property
    def duration_seconds(self) -> float:
        if self.start_time and self.end_time:
            return self.end_time - self.start_time
        return 0.0

    @property
    def success_rate(self) -> float:
        if self.items_processed == 0:
            return 0.0
        return (self.items_successful / self.items_processed) * 100


@dataclass
class ExtractionMetrics:
    """Metrics specific to extraction stage."""
    total_fields_expected: int = 0
    total_fields_extracted: int = 0
    high_confidence_count: int = 0  # >= 0.85
    medium_confidence_count: int = 0  # 0.7 - 0.85
    low_confidence_count: int = 0  # < 0.7
    extraction_method_counts: Dict[str, int] = field(default_factory=dict)

    @property
    def accuracy_percentage(self) -> float:
        if self.total_fields_expected == 0:
            return 0.0
        return (self.total_fields_extracted / self.total_fields_expected) * 100

    @property
    def high_confidence_percentage(self) -> float:
        total = self.high_confidence_count + self.medium_confidence_count + self.low_confidence_count
        if total == 0:
            return 0.0
        return (self.high_confidence_count / total) * 100


@dataclass
class CoverageMetrics:
    """Metrics for data coverage."""
    total_cells_in_source: int = 0
    cells_with_data: int = 0
    cells_extracted: int = 0
    cells_mapped: int = 0
    cells_written_to_output: int = 0
    formulas_found: int = 0
    comments_found: int = 0
    hidden_rows: int = 0
    hidden_columns: int = 0

    @property
    def extraction_coverage(self) -> float:
        if self.cells_with_data == 0:
            return 0.0
        return (self.cells_extracted / self.cells_with_data) * 100

    @property
    def output_coverage(self) -> float:
        if self.cells_extracted == 0:
            return 0.0
        return (self.cells_written_to_output / self.cells_extracted) * 100


class ProcessingMetrics:
    """
    Collects and displays processing metrics to users.

    Usage:
        metrics = ProcessingMetrics()
        metrics.start_stage("ingestion")
        # ... do work ...
        metrics.end_stage("ingestion", items_processed=5, items_successful=5)
        metrics.display_summary()
    """

    def __init__(self, vendor: str = "Unknown", document_name: str = "Unknown"):
        self.vendor = vendor
        self.document_name = document_name
        self.pipeline_start_time: Optional[float] = None
        self.pipeline_end_time: Optional[float] = None

        self.stages: Dict[str, StageMetrics] = {}
        self.extraction: ExtractionMetrics = ExtractionMetrics()
        self.coverage: CoverageMetrics = CoverageMetrics()

        self._stage_order = [
            "ingestion", "preprocessing", "extraction",
            "validation", "mapping", "output"
        ]

    def start_pipeline(self):
        """Mark pipeline start time."""
        self.pipeline_start_time = time.time()

    def end_pipeline(self):
        """Mark pipeline end time."""
        self.pipeline_end_time = time.time()

    @property
    def total_duration_seconds(self) -> float:
        if self.pipeline_start_time and self.pipeline_end_time:
            return self.pipeline_end_time - self.pipeline_start_time
        return 0.0

    def start_stage(self, stage_name: str):
        """Start tracking a processing stage."""
        if stage_name not in self.stages:
            self.stages[stage_name] = StageMetrics(name=stage_name)
        self.stages[stage_name].start_time = time.time()

    def end_stage(
        self,
        stage_name: str,
        items_processed: int = 0,
        items_successful: int = 0,
        items_failed: int = 0,
        warnings: List[str] = None
    ):
        """End tracking a processing stage with results."""
        if stage_name not in self.stages:
            self.stages[stage_name] = StageMetrics(name=stage_name)

        stage = self.stages[stage_name]
        stage.end_time = time.time()
        stage.items_processed = items_processed
        stage.items_successful = items_successful
        stage.items_failed = items_failed
        if warnings:
            stage.warnings = warnings

    def record_extraction_metrics(
        self,
        fields_expected: int,
        fields_extracted: int,
        confidences: List[float],
        methods: Dict[str, int]
    ):
        """Record extraction-specific metrics."""
        self.extraction.total_fields_expected = fields_expected
        self.extraction.total_fields_extracted = fields_extracted
        self.extraction.extraction_method_counts = methods

        for conf in confidences:
            if conf >= 0.85:
                self.extraction.high_confidence_count += 1
            elif conf >= 0.7:
                self.extraction.medium_confidence_count += 1
            else:
                self.extraction.low_confidence_count += 1

    def record_coverage_metrics(
        self,
        total_cells: int = 0,
        cells_with_data: int = 0,
        cells_extracted: int = 0,
        cells_mapped: int = 0,
        cells_written: int = 0,
        formulas: int = 0,
        comments: int = 0,
        hidden_rows: int = 0,
        hidden_cols: int = 0
    ):
        """Record coverage metrics."""
        self.coverage.total_cells_in_source = total_cells
        self.coverage.cells_with_data = cells_with_data
        self.coverage.cells_extracted = cells_extracted
        self.coverage.cells_mapped = cells_mapped
        self.coverage.cells_written_to_output = cells_written
        self.coverage.formulas_found = formulas
        self.coverage.comments_found = comments
        self.coverage.hidden_rows = hidden_rows
        self.coverage.hidden_columns = hidden_cols

    def display_summary(self, output_to_console: bool = True) -> str:
        """
        Display a formatted summary of all metrics.

        Returns:
            Formatted string summary
        """
        lines = []

        # Header
        lines.append("")
        lines.append("=" * 70)
        lines.append("PROCESSING METRICS SUMMARY")
        lines.append("=" * 70)
        lines.append(f"Vendor: {self.vendor}")
        lines.append(f"Document: {self.document_name}")
        lines.append(f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("")

        # Pipeline timing
        lines.append("-" * 70)
        lines.append("TIMING")
        lines.append("-" * 70)
        lines.append(f"Total Processing Time: {self.total_duration_seconds:.2f} seconds")
        lines.append("")

        # Stage breakdown
        lines.append("Stage Breakdown:")
        for stage_name in self._stage_order:
            if stage_name in self.stages:
                stage = self.stages[stage_name]
                status = "PASS" if stage.items_failed == 0 else "FAIL"
                lines.append(
                    f"  {stage_name.capitalize():15} | "
                    f"{stage.duration_seconds:6.2f}s | "
                    f"{stage.items_successful}/{stage.items_processed} items | "
                    f"[{status}]"
                )
        lines.append("")

        # Extraction metrics
        lines.append("-" * 70)
        lines.append("EXTRACTION ACCURACY")
        lines.append("-" * 70)
        lines.append(f"Fields Expected:    {self.extraction.total_fields_expected}")
        lines.append(f"Fields Extracted:   {self.extraction.total_fields_extracted}")
        lines.append(f"Accuracy:           {self.extraction.accuracy_percentage:.1f}%")
        lines.append("")
        lines.append("Confidence Distribution:")
        lines.append(f"  High (>=85%):     {self.extraction.high_confidence_count} ({self.extraction.high_confidence_percentage:.1f}%)")
        lines.append(f"  Medium (70-85%):  {self.extraction.medium_confidence_count}")
        lines.append(f"  Low (<70%):       {self.extraction.low_confidence_count}")
        lines.append("")

        if self.extraction.extraction_method_counts:
            lines.append("Extraction Methods Used:")
            for method, count in self.extraction.extraction_method_counts.items():
                lines.append(f"  {method}: {count}")
        lines.append("")

        # Coverage metrics
        lines.append("-" * 70)
        lines.append("DATA COVERAGE")
        lines.append("-" * 70)
        lines.append(f"Source Cells (with data): {self.coverage.cells_with_data}")
        lines.append(f"Cells Extracted:          {self.coverage.cells_extracted}")
        lines.append(f"Cells Written to Output:  {self.coverage.cells_written_to_output}")
        lines.append(f"Extraction Coverage:      {self.coverage.extraction_coverage:.1f}%")
        lines.append(f"Output Coverage:          {self.coverage.output_coverage:.1f}%")
        lines.append("")
        lines.append(f"Formulas Found:           {self.coverage.formulas_found}")
        lines.append(f"Comments Found:           {self.coverage.comments_found}")

        if self.coverage.hidden_rows > 0 or self.coverage.hidden_columns > 0:
            lines.append("")
            lines.append("WARNINGS:")
            if self.coverage.hidden_rows > 0:
                lines.append(f"  ! {self.coverage.hidden_rows} hidden rows detected")
            if self.coverage.hidden_columns > 0:
                lines.append(f"  ! {self.coverage.hidden_columns} hidden columns detected")

        lines.append("")
        lines.append("=" * 70)

        # Collect any warnings
        all_warnings = []
        for stage in self.stages.values():
            all_warnings.extend(stage.warnings)

        if all_warnings:
            lines.append("WARNINGS:")
            for warning in all_warnings[:10]:  # Limit to first 10
                lines.append(f"  - {warning}")
            if len(all_warnings) > 10:
                lines.append(f"  ... and {len(all_warnings) - 10} more")
            lines.append("=" * 70)

        summary = "\n".join(lines)

        if output_to_console:
            print(summary)

        return summary

    def to_dict(self) -> Dict[str, Any]:
        """Export metrics as dictionary."""
        return {
            "vendor": self.vendor,
            "document": self.document_name,
            "timestamp": datetime.now().isoformat(),
            "total_duration_seconds": self.total_duration_seconds,
            "stages": {
                name: {
                    "duration_seconds": s.duration_seconds,
                    "items_processed": s.items_processed,
                    "items_successful": s.items_successful,
                    "items_failed": s.items_failed,
                    "success_rate": s.success_rate
                }
                for name, s in self.stages.items()
            },
            "extraction": {
                "fields_expected": self.extraction.total_fields_expected,
                "fields_extracted": self.extraction.total_fields_extracted,
                "accuracy_percentage": self.extraction.accuracy_percentage,
                "high_confidence_count": self.extraction.high_confidence_count,
                "medium_confidence_count": self.extraction.medium_confidence_count,
                "low_confidence_count": self.extraction.low_confidence_count,
                "methods": self.extraction.extraction_method_counts
            },
            "coverage": {
                "cells_with_data": self.coverage.cells_with_data,
                "cells_extracted": self.coverage.cells_extracted,
                "cells_written": self.coverage.cells_written_to_output,
                "extraction_coverage": self.coverage.extraction_coverage,
                "output_coverage": self.coverage.output_coverage,
                "formulas_found": self.coverage.formulas_found,
                "comments_found": self.coverage.comments_found
            }
        }

    def save_to_json(self, output_path: str):
        """Save metrics to JSON file."""
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(self.to_dict(), f, indent=2)
        print(f"Metrics saved to: {output_path}")


# =============================================================================
# TRACEABILITY REPORT
# =============================================================================

@dataclass
class TraceabilityRecord:
    """Single record tracking data from source to output."""
    source_file: str
    source_location: str  # e.g., "Sheet1!B15" or "Table 3, Row 5"
    source_value: Any

    mapped_field: str
    mapped_value: Any
    mapping_confidence: float
    mapping_method: str

    output_file: str
    output_location: str  # e.g., "Line Items!BB7"
    output_value: Any

    status: str = "matched"  # matched, transformed, missing, error
    notes: str = ""


class TraceabilityReport:
    """
    Generates source-to-output traceability reports.

    Tracks data flow: Source Document -> Extraction -> Mapping -> TCO Output

    Usage:
        report = TraceabilityReport("FIS", "proposal.docx", "output.xlsx")
        report.add_record(source_loc="Table1!B5", source_val=1500.00, ...)
        report.generate_excel_report("traceability_report.xlsx")
    """

    def __init__(
        self,
        vendor: str,
        source_file: str,
        output_file: str
    ):
        self.vendor = vendor
        self.source_file = source_file
        self.output_file = output_file
        self.created_at = datetime.now()

        self.records: List[TraceabilityRecord] = []
        self.summary_stats = {
            "total_records": 0,
            "matched": 0,
            "transformed": 0,
            "missing": 0,
            "errors": 0
        }

    def add_record(
        self,
        source_location: str,
        source_value: Any,
        mapped_field: str,
        mapped_value: Any,
        output_location: str,
        output_value: Any,
        confidence: float = 1.0,
        method: str = "direct",
        status: str = "matched",
        notes: str = ""
    ):
        """Add a traceability record."""
        record = TraceabilityRecord(
            source_file=self.source_file,
            source_location=source_location,
            source_value=source_value,
            mapped_field=mapped_field,
            mapped_value=mapped_value,
            mapping_confidence=confidence,
            mapping_method=method,
            output_file=self.output_file,
            output_location=output_location,
            output_value=output_value,
            status=status,
            notes=notes
        )
        self.records.append(record)

        self.summary_stats["total_records"] += 1
        if status in self.summary_stats:
            self.summary_stats[status] += 1

    def add_batch_records(
        self,
        source_data: List[Dict[str, Any]],
        mapped_data: List[Dict[str, Any]],
        output_data: List[Dict[str, Any]],
        field_mappings: Dict[str, str]
    ):
        """
        Add multiple records by matching source, mapped, and output data.

        Args:
            source_data: List of extracted source records
            mapped_data: List of mapped/normalized records
            output_data: List of records written to output
            field_mappings: Dict mapping source fields to output fields
        """
        for idx, (src, mapped, out) in enumerate(zip(source_data, mapped_data, output_data)):
            for src_field, out_field in field_mappings.items():
                src_val = src.get(src_field)
                mapped_val = mapped.get(src_field) or mapped.get(out_field)
                out_val = out.get(out_field)

                # Determine status
                if src_val is None:
                    status = "missing"
                elif src_val == out_val:
                    status = "matched"
                elif out_val is not None:
                    status = "transformed"
                else:
                    status = "error"

                self.add_record(
                    source_location=f"Row {idx + 1}, {src_field}",
                    source_value=src_val,
                    mapped_field=out_field,
                    mapped_value=mapped_val,
                    output_location=f"Row {idx + 1}, {out_field}",
                    output_value=out_val,
                    confidence=mapped.get('confidence', 1.0) if isinstance(mapped, dict) else 1.0,
                    method=mapped.get('method', 'direct') if isinstance(mapped, dict) else 'direct',
                    status=status
                )

    def generate_text_report(self, output_path: str = None) -> str:
        """Generate a text-based traceability report."""
        lines = []

        # Header
        lines.append("=" * 80)
        lines.append("SOURCE-TO-OUTPUT TRACEABILITY REPORT")
        lines.append("=" * 80)
        lines.append(f"Generated: {self.created_at.strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"Vendor: {self.vendor}")
        lines.append(f"Source: {self.source_file}")
        lines.append(f"Output: {self.output_file}")
        lines.append("")

        # Summary
        lines.append("-" * 80)
        lines.append("SUMMARY")
        lines.append("-" * 80)
        lines.append(f"Total Records Traced:  {self.summary_stats['total_records']}")
        lines.append(f"Matched (exact):       {self.summary_stats['matched']}")
        lines.append(f"Transformed:           {self.summary_stats['transformed']}")
        lines.append(f"Missing in Output:     {self.summary_stats['missing']}")
        lines.append(f"Errors:                {self.summary_stats['errors']}")

        if self.summary_stats['total_records'] > 0:
            match_rate = (self.summary_stats['matched'] / self.summary_stats['total_records']) * 100
            lines.append(f"Match Rate:            {match_rate:.1f}%")
        lines.append("")

        # Detail section
        lines.append("-" * 80)
        lines.append("DETAILED TRACEABILITY")
        lines.append("-" * 80)
        lines.append("")

        # Group by status for better readability
        for status in ["matched", "transformed", "missing", "error"]:
            status_records = [r for r in self.records if r.status == status]
            if not status_records:
                continue

            status_label = {
                "matched": "MATCHED RECORDS (Source = Output)",
                "transformed": "TRANSFORMED RECORDS (Value Changed)",
                "missing": "MISSING RECORDS (Not in Output)",
                "error": "ERROR RECORDS (Processing Failed)"
            }.get(status, status.upper())

            lines.append(f"\n{status_label} ({len(status_records)} records)")
            lines.append("-" * 40)

            # Show first 20 of each type
            for record in status_records[:20]:
                lines.append(f"  Source:  {record.source_location}")
                lines.append(f"    Value: {record.source_value}")
                lines.append(f"  Mapped:  {record.mapped_field} (conf: {record.mapping_confidence:.0%})")
                lines.append(f"    Value: {record.mapped_value}")
                lines.append(f"  Output:  {record.output_location}")
                lines.append(f"    Value: {record.output_value}")
                if record.notes:
                    lines.append(f"  Notes:   {record.notes}")
                lines.append("")

            if len(status_records) > 20:
                lines.append(f"  ... and {len(status_records) - 20} more {status} records")
                lines.append("")

        lines.append("=" * 80)

        report_text = "\n".join(lines)

        if output_path:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(report_text)
            print(f"Traceability report saved to: {output_path}")

        return report_text

    def generate_excel_report(self, output_path: str):
        """Generate an Excel traceability report with formatting."""
        wb = Workbook()

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        matched_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        transformed_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        missing_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        error_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # === Summary Sheet ===
        ws_summary = wb.active
        ws_summary.title = "Summary"

        ws_summary['A1'] = "SOURCE-TO-OUTPUT TRACEABILITY REPORT"
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary.merge_cells('A1:D1')

        summary_data = [
            ("Generated", self.created_at.strftime('%Y-%m-%d %H:%M:%S')),
            ("Vendor", self.vendor),
            ("Source File", self.source_file),
            ("Output File", self.output_file),
            ("", ""),
            ("STATISTICS", ""),
            ("Total Records", self.summary_stats['total_records']),
            ("Matched", self.summary_stats['matched']),
            ("Transformed", self.summary_stats['transformed']),
            ("Missing", self.summary_stats['missing']),
            ("Errors", self.summary_stats['errors']),
        ]

        if self.summary_stats['total_records'] > 0:
            match_rate = (self.summary_stats['matched'] / self.summary_stats['total_records']) * 100
            summary_data.append(("Match Rate", f"{match_rate:.1f}%"))

        for row_idx, (label, value) in enumerate(summary_data, start=3):
            ws_summary[f'A{row_idx}'] = label
            ws_summary[f'B{row_idx}'] = value
            if label == "STATISTICS":
                ws_summary[f'A{row_idx}'].font = Font(bold=True)

        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 50

        # === Detail Sheet ===
        ws_detail = wb.create_sheet("Traceability Detail")

        headers = [
            "Status", "Source Location", "Source Value",
            "Mapped Field", "Mapped Value", "Confidence",
            "Output Location", "Output Value", "Method", "Notes"
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws_detail.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Add records
        for row_idx, record in enumerate(self.records, start=2):
            row_data = [
                record.status.upper(),
                record.source_location,
                str(record.source_value) if record.source_value is not None else "",
                record.mapped_field,
                str(record.mapped_value) if record.mapped_value is not None else "",
                f"{record.mapping_confidence:.0%}",
                record.output_location,
                str(record.output_value) if record.output_value is not None else "",
                record.mapping_method,
                record.notes
            ]

            # Determine row fill based on status
            row_fill = {
                "matched": matched_fill,
                "transformed": transformed_fill,
                "missing": missing_fill,
                "error": error_fill
            }.get(record.status, None)

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_detail.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if row_fill and col_idx == 1:  # Only color the status column
                    cell.fill = row_fill

        # Adjust column widths
        col_widths = [12, 25, 20, 20, 20, 12, 25, 20, 12, 30]
        for col_idx, width in enumerate(col_widths, start=1):
            ws_detail.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze header row
        ws_detail.freeze_panes = 'A2'

        # Add auto-filter
        ws_detail.auto_filter.ref = f"A1:J{len(self.records) + 1}"

        # Save
        wb.save(output_path)
        print(f"Excel traceability report saved to: {output_path}")

    def to_dict(self) -> Dict[str, Any]:
        """Export report as dictionary."""
        return {
            "vendor": self.vendor,
            "source_file": self.source_file,
            "output_file": self.output_file,
            "created_at": self.created_at.isoformat(),
            "summary": self.summary_stats,
            "records": [
                {
                    "source_location": r.source_location,
                    "source_value": r.source_value,
                    "mapped_field": r.mapped_field,
                    "mapped_value": r.mapped_value,
                    "confidence": r.mapping_confidence,
                    "method": r.mapping_method,
                    "output_location": r.output_location,
                    "output_value": r.output_value,
                    "status": r.status,
                    "notes": r.notes
                }
                for r in self.records
            ]
        }

    def save_to_json(self, output_path: str):
        """Save report to JSON file."""
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(self.to_dict(), f, indent=2, default=str)
        print(f"Traceability JSON saved to: {output_path}")


# =============================================================================
# CONVENIENCE FUNCTIONS
# =============================================================================

def create_metrics_from_pipeline_result(result, vendor: str, document: str) -> ProcessingMetrics:
    """
    Create ProcessingMetrics from a PipelineResult object.

    Args:
        result: PipelineResult from pipeline.py
        vendor: Vendor name
        document: Document name

    Returns:
        Populated ProcessingMetrics instance
    """
    metrics = ProcessingMetrics(vendor=vendor, document_name=document)

    # Set pipeline times
    metrics.pipeline_start_time = result.start_time.timestamp() if hasattr(result.start_time, 'timestamp') else result.start_time
    metrics.pipeline_end_time = result.end_time.timestamp() if hasattr(result.end_time, 'timestamp') else result.end_time

    # Process stages
    for stage in result.stages:
        metrics.stages[stage.name] = StageMetrics(
            name=stage.name,
            start_time=stage.start_time,
            end_time=stage.end_time,
            items_processed=result.total_documents if stage.name == "ingestion" else 0,
            items_successful=1 if stage.status.value == "completed" else 0,
            items_failed=1 if stage.status.value == "failed" else 0,
            warnings=[]
        )

    return metrics


def generate_full_report(
    vendor: str,
    source_file: str,
    output_file: str,
    extracted_data: List[Dict],
    mapped_data: List[Dict],
    output_data: List[Dict],
    metrics: ProcessingMetrics = None,
    report_dir: str = "./reports"
) -> Dict[str, str]:
    """
    Generate complete metrics and traceability reports.

    Args:
        vendor: Vendor name
        source_file: Path to source file
        output_file: Path to output file
        extracted_data: Raw extracted data
        mapped_data: Normalized/mapped data
        output_data: Data written to output
        metrics: Optional pre-populated metrics
        report_dir: Directory for report outputs

    Returns:
        Dict with paths to generated reports
    """
    os.makedirs(report_dir, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    report_paths = {}

    # Generate traceability report
    trace_report = TraceabilityReport(vendor, source_file, output_file)

    # Define field mappings based on vendor
    if vendor == "FIS":
        field_mappings = {
            "solution_name": "solution_name",
            "monthly_fee": "per_unit_rate",
            "one_time_fee": "one_time_fee",
            "fee_type": "fee_type",
            "category": "category"
        }
    else:  # Jack Henry
        field_mappings = {
            "product_description": "solution_name",
            "monthly_net": "per_unit_rate",
            "license_net": "one_time_fee",
            "product_family": "category"
        }

    # Add records (handle length mismatches)
    min_len = min(len(extracted_data), len(mapped_data), len(output_data))
    for idx in range(min_len):
        src = extracted_data[idx] if idx < len(extracted_data) else {}
        mapped = mapped_data[idx] if idx < len(mapped_data) else {}
        out = output_data[idx] if idx < len(output_data) else {}

        for src_field, out_field in field_mappings.items():
            src_val = src.get(src_field)
            mapped_val = mapped.get(out_field, mapped.get(src_field))
            out_val = out.get(out_field)

            if src_val is None and mapped_val is None and out_val is None:
                continue

            status = "matched" if src_val == out_val else (
                "transformed" if out_val is not None else "missing"
            )

            trace_report.add_record(
                source_location=f"Row {idx + 1}",
                source_value=src_val,
                mapped_field=out_field,
                mapped_value=mapped_val,
                output_location=f"Row {idx + 1}",
                output_value=out_val,
                confidence=mapped.get('confidence', 1.0) if isinstance(mapped, dict) else 1.0,
                status=status
            )

    # Save traceability reports
    trace_text_path = os.path.join(report_dir, f"traceability_{vendor}_{timestamp}.txt")
    trace_excel_path = os.path.join(report_dir, f"traceability_{vendor}_{timestamp}.xlsx")
    trace_json_path = os.path.join(report_dir, f"traceability_{vendor}_{timestamp}.json")

    trace_report.generate_text_report(trace_text_path)
    trace_report.generate_excel_report(trace_excel_path)
    trace_report.save_to_json(trace_json_path)

    report_paths["traceability_text"] = trace_text_path
    report_paths["traceability_excel"] = trace_excel_path
    report_paths["traceability_json"] = trace_json_path

    # Save metrics if provided
    if metrics:
        metrics_path = os.path.join(report_dir, f"metrics_{vendor}_{timestamp}.json")
        metrics.save_to_json(metrics_path)
        report_paths["metrics"] = metrics_path

        # Also display to console
        metrics.display_summary()

    return report_paths


# =============================================================================
# MAIN - Demo/Test
# =============================================================================

if __name__ == "__main__":
    print("=" * 70)
    print("REPORTING MODULE DEMO")
    print("=" * 70)

    # Demo ProcessingMetrics
    print("\n--- ProcessingMetrics Demo ---\n")

    metrics = ProcessingMetrics(vendor="Jack Henry", document_name="test_proposal.xlsx")
    metrics.start_pipeline()

    # Simulate stages
    import time

    metrics.start_stage("ingestion")
    time.sleep(0.1)
    metrics.end_stage("ingestion", items_processed=1, items_successful=1)

    metrics.start_stage("preprocessing")
    time.sleep(0.05)
    metrics.end_stage("preprocessing", items_processed=150, items_successful=148, items_failed=2)

    metrics.start_stage("extraction")
    time.sleep(0.2)
    metrics.end_stage("extraction", items_processed=148, items_successful=145, items_failed=3)

    metrics.start_stage("validation")
    time.sleep(0.05)
    metrics.end_stage("validation", items_processed=145, items_successful=145)

    metrics.start_stage("mapping")
    time.sleep(0.1)
    metrics.end_stage("mapping", items_processed=145, items_successful=145)

    metrics.start_stage("output")
    time.sleep(0.1)
    metrics.end_stage("output", items_processed=145, items_successful=145)

    metrics.end_pipeline()

    # Record extraction metrics
    metrics.record_extraction_metrics(
        fields_expected=150,
        fields_extracted=145,
        confidences=[0.95, 0.88, 0.72, 0.91, 0.65, 0.99] * 24,
        methods={"llm": 100, "regex": 30, "ner": 15}
    )

    # Record coverage metrics
    metrics.record_coverage_metrics(
        total_cells=5000,
        cells_with_data=1500,
        cells_extracted=1480,
        cells_mapped=1480,
        cells_written=1475,
        formulas=250,
        comments=15,
        hidden_rows=2,
        hidden_cols=0
    )

    # Display summary
    metrics.display_summary()

    # Demo TraceabilityReport
    print("\n--- TraceabilityReport Demo ---\n")

    trace = TraceabilityReport(
        vendor="Jack Henry",
        source_file="proposal.xlsx",
        output_file="tco_output.xlsx"
    )

    # Add sample records
    trace.add_record(
        source_location="Proposal_1!B15",
        source_value=1500.00,
        mapped_field="per_unit_rate",
        mapped_value=1500.00,
        output_location="Line Items!BD7",
        output_value=1500.00,
        confidence=0.95,
        method="direct",
        status="matched"
    )

    trace.add_record(
        source_location="Proposal_1!B16",
        source_value="Digital Banking",
        mapped_field="solution_name",
        mapped_value="Digital Banking Platform",
        output_location="Line Items!BB8",
        output_value="Digital Banking Platform",
        confidence=0.85,
        method="fuzzy_match",
        status="transformed",
        notes="Name normalized via fuzzy matching"
    )

    trace.add_record(
        source_location="Proposal_1!B17",
        source_value=250.00,
        mapped_field="monthly_fee",
        mapped_value=250.00,
        output_location="Line Items!BD9",
        output_value=None,
        confidence=0.90,
        method="direct",
        status="missing",
        notes="Output cell was empty"
    )

    # Generate text report
    print(trace.generate_text_report())

    print("\nDemo complete. In production, use:")
    print("  metrics.save_to_json('metrics.json')")
    print("  trace.generate_excel_report('traceability.xlsx')")
