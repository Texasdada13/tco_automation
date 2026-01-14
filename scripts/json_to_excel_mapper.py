"""
JSON to Excel Mapper - Transform JSON extractions to New TCO Excel format

This script loads JSON extraction files and populates the New_TCO_Excel_v1.xlsx
template with all data, applying transformations, validations, and calculations.

Usage:
    python json_to_excel_mapper.py <json_file> [output_file]

Example:
    python json_to_excel_mapper.py "Extracted JSON/liberty_extraction_ai.json"
"""

import json
import logging
import sys
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('logs/mapping_errors.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class EnumNormalizer:
    """Handles normalization of enum values"""

    def __init__(self, enum_mappings_file='Data_Dictionary/enum_mappings.json'):
        with open(enum_mappings_file, 'r') as f:
            self.mappings = json.load(f)

    def normalize_fee_type(self, value: Any) -> str:
        """Normalize fee type to standard values"""
        if not value:
            return "Monthly F"

        value_str = str(value).strip()
        mappings = self.mappings['fee_type_mappings']['mappings']

        for standard, variants in mappings.items():
            if value_str in variants:
                return standard

        logger.warning(f"Could not normalize fee_type '{value}', using default 'Monthly F'")
        return "Monthly F"

    def normalize_category(self, value: Any) -> str:
        """Normalize category to standard values"""
        if not value:
            return "Other"

        value_str = str(value).strip()
        mappings = self.mappings['category_mappings']['mappings']

        # Exact match first
        for standard, variants in mappings.items():
            if value_str in variants:
                return standard

        # Keyword match
        for standard, variants in mappings.items():
            for variant in variants:
                if variant.lower() in value_str.lower():
                    return standard

        logger.warning(f"Unknown category '{value}', using 'Other'")
        return "Other"

    def normalize_boolean(self, value: Any) -> bool:
        """Normalize various boolean representations"""
        true_values = self.mappings['boolean_mappings']['mappings']['TRUE']
        return value in true_values

    def normalize_vendor(self, value: Any) -> str:
        """Normalize vendor name"""
        if not value:
            return "Other"

        value_str = str(value).strip()
        mappings = self.mappings['vendor_mappings']['mappings']

        for standard, variants in mappings.items():
            if value_str in variants:
                return standard

        return value_str  # Return as-is if not in mappings


class LineItemProcessor:
    """Processes and transforms JSON line items"""

    def __init__(self, normalizer: EnumNormalizer):
        self.normalizer = normalizer
        self.data_quality_issues = []

    def split_item_if_needed(self, item: Dict) -> List[Dict]:
        """Split items with both monthly and one-time fees"""
        # Handle both old format (monthly_fee) and new direct extraction format (proposed_monthly_fee)
        monthly_fee = item.get('proposed_monthly_fee') or item.get('monthly_fee', 0) or 0
        one_time_fee = item.get('one_time_fee', 0) or 0

        rows = []

        # Create monthly fee row if present
        if monthly_fee and monthly_fee != 0:
            monthly_row = item.copy()
            monthly_row['_per_unit_rate'] = monthly_fee
            monthly_row['_is_split_monthly'] = True
            rows.append(monthly_row)

        # Create one-time fee row if present
        if one_time_fee and one_time_fee != 0:
            onetime_row = item.copy()
            onetime_row['fee_type'] = 'One-Time'
            onetime_row['_per_unit_rate'] = one_time_fee
            onetime_row['solution_name'] = item.get('solution_name', '') + ' - Implementation Fee'
            onetime_row['_is_split_onetime'] = True
            rows.append(onetime_row)

        # If no split occurred, return original item
        if not rows:
            item['_per_unit_rate'] = item.get('per_unit_rate', 0) or monthly_fee or one_time_fee
            rows.append(item)

        return rows

    def transform_item(self, item: Dict, row_id: int) -> Dict:
        """Transform a single JSON item to Excel row format"""
        try:
            # Normalize enums
            fee_type = self.normalizer.normalize_fee_type(item.get('fee_type', 'Monthly F'))
            category = self.normalizer.normalize_category(item.get('category', 'Other'))
            third_party = self.normalizer.normalize_boolean(item.get('third_party', False))
            optional = self.normalizer.normalize_boolean(item.get('optional', False))

            # Get per_unit_rate (prioritize _per_unit_rate from splitting)
            per_unit_rate = item.get('_per_unit_rate') or item.get('per_unit_rate', 0) or 0

            # Solution name
            solution_name = item.get('solution_name', '').strip()
            if not solution_name:
                self.data_quality_issues.append({
                    'row_id': row_id,
                    'issue_type': 'Missing Required Field',
                    'description': 'Missing solution_name'
                })
                solution_name = "UNNAMED SOLUTION"

            # Unit description
            unit_desc = item.get('unit_description', '')
            if not unit_desc:
                if fee_type == "Monthly F":
                    unit_desc = "per month"
                elif fee_type == "Monthly V":
                    unit_desc = "per transaction"
                elif fee_type == "Annual":
                    unit_desc = "per year"
                elif fee_type == "One-Time":
                    unit_desc = "one-time"

            # Average monthly quantity (handle both formats)
            # Direct extraction uses 'volume', old format uses 'average_monthly_qty'
            avg_qty = item.get('average_monthly_qty') or item.get('volume') or (1 if fee_type == "Monthly V" else None)

            # Validation: Monthly V requires quantity
            if fee_type == "Monthly V" and not avg_qty:
                self.data_quality_issues.append({
                    'row_id': row_id,
                    'solution_name': solution_name,
                    'issue_type': 'Missing Quantity for Variable Fee',
                    'description': 'Monthly V requires average_monthly_qty'
                })
                avg_qty = 1  # Default to prevent division errors

            # Confidence score
            confidence = item.get('overall_confidence')
            if confidence is not None:
                confidence = max(0.0, min(1.0, float(confidence)))  # Clamp to 0-1
                if confidence < 0.80:
                    self.data_quality_issues.append({
                        'row_id': row_id,
                        'solution_name': solution_name,
                        'issue_type': 'Low Confidence Score',
                        'confidence': confidence,
                        'description': f'Confidence score {confidence:.2%} below 80% threshold'
                    })

            # Zero cost validation
            if per_unit_rate == 0 and fee_type != "One-Time":
                self.data_quality_issues.append({
                    'row_id': row_id,
                    'solution_name': solution_name,
                    'issue_type': 'Zero Cost Item',
                    'description': 'Item has zero per_unit_rate'
                })

            # Negative cost flag (may be credit)
            if per_unit_rate < 0 and fee_type != "One-Time":
                self.data_quality_issues.append({
                    'row_id': row_id,
                    'solution_name': solution_name,
                    'issue_type': 'Negative Cost (Review Required)',
                    'description': f'Negative per_unit_rate: ${per_unit_rate:,.2f}'
                })

            return {
                'row_id': row_id,
                'fee_type': fee_type,
                'solution_name': solution_name,
                'category': category,
                'third_party': third_party,
                'optional': optional,
                'per_unit_rate': per_unit_rate,
                'unit_description': unit_desc,
                'average_monthly_qty': avg_qty,
                'confidence_score': confidence,
                'extraction_notes': item.get('extraction_notes', '')
            }

        except Exception as e:
            logger.error(f"Error transforming item: {e}")
            logger.error(f"Item data: {item}")
            self.data_quality_issues.append({
                'row_id': row_id,
                'issue_type': 'Transformation Error',
                'description': str(e)
            })
            return None


class ExcelWriter:
    """Writes transformed data to Excel template"""

    def __init__(self, template_path='Templates/New_TCO_Excel_v1.xlsx'):
        self.wb = openpyxl.load_workbook(template_path)

        # Get worksheet references (using old names from template)
        self.ws_meta = self.wb['Metadata']
        self.ws_items = self.wb['Line_Items']
        self.ws_summary = self.wb['Summary']
        self.ws_years = self.wb['Year_Summary']

        # === CANONICAL STRUCTURE ENFORCEMENT ===

        # 1. Remove unwanted sheets
        if 'Data_Quality' in self.wb.sheetnames:
            del self.wb['Data_Quality']

        if 'Enums' in self.wb.sheetnames:
            del self.wb['Enums']

        # 2. Rename sheets to canonical names
        self.wb['Line_Items'].title = 'Line Items'
        self.wb['Year_Summary'].title = 'Years 1-7'
        # Summary and Metadata keep their original names

        # 3. Reorder sheets: Line Items, Years 1-7, Summary, Metadata
        self.wb._sheets = [
            self.wb['Line Items'],
            self.wb['Years 1-7'],
            self.wb['Summary'],
            self.wb['Metadata']
        ]

    def write_metadata(self, json_data: Dict):
        """Write metadata to Metadata sheet"""
        self.ws_meta['B2'] = json_data.get('vendor', 'Unknown')
        self.ws_meta['B3'] = json_data.get('client', 'Unknown')
        self.ws_meta['B4'] = json_data.get('proposal_type', 'N/A')
        # Handle both formats: document_date (old) and proposal_date (direct extraction)
        self.ws_meta['B5'] = json_data.get('proposal_date') or json_data.get('document_date', 'N/A')
        self.ws_meta['B6'] = json_data.get('contract_term', 7)

        # Extraction metadata
        self.ws_meta['B10'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        # AI Model Used row (B11) - REMOVED per canonical structure rules
        self.ws_meta['A11'] = None  # Clear label
        self.ws_meta['B11'] = None  # Clear value
        self.ws_meta['B12'] = json_data.get('source_file', 'N/A')

    def write_line_items(self, items: List[Dict]):
        """Write line items to Line_Items sheet"""
        # Sort items: optional (F, T), third_party (F, T), fee_type, category, solution_name
        items_sorted = sorted(items, key=lambda x: (
            x['optional'],
            x['third_party'],
            x['fee_type'],
            x['category'],
            x['solution_name']
        ))

        row_num = 2
        for item in items_sorted:
            # Column A: Row ID
            self.ws_items[f'A{row_num}'] = item['row_id']

            # Column B: Fee Type
            self.ws_items[f'B{row_num}'] = item['fee_type']

            # Column C: Solution Name
            self.ws_items[f'C{row_num}'] = item['solution_name']

            # Column D: Category
            self.ws_items[f'D{row_num}'] = item['category']

            # Column E: Third Party
            self.ws_items[f'E{row_num}'] = item['third_party']

            # Column F: Optional
            self.ws_items[f'F{row_num}'] = item['optional']

            # Column G: Per Unit Rate
            self.ws_items[f'G{row_num}'] = item['per_unit_rate']
            self.ws_items[f'G{row_num}'].number_format = '$#,##0.00'

            # Column H: Unit Description
            self.ws_items[f'H{row_num}'] = item['unit_description']

            # Column I: Average Monthly Qty
            if item['average_monthly_qty'] is not None:
                self.ws_items[f'I{row_num}'] = item['average_monthly_qty']
                self.ws_items[f'I{row_num}'].number_format = '#,##0'

            # Column J: Year 1 Monthly Cost (Formula)
            formula_j = f'=IF(B{row_num}="Monthly F",G{row_num},IF(B{row_num}="Monthly V",G{row_num}*I{row_num},IF(B{row_num}="Annual",G{row_num}/12,0)))'
            self.ws_items[f'J{row_num}'] = formula_j
            self.ws_items[f'J{row_num}'].number_format = '$#,##0.00'

            # Column K: Year 1 Annual Cost (Formula)
            formula_k = f'=IF(B{row_num}="One-Time",G{row_num},J{row_num}*12)'
            self.ws_items[f'K{row_num}'] = formula_k
            self.ws_items[f'K{row_num}'].number_format = '$#,##0.00'

            # Columns L-Q: Years 2-7 (Formulas)
            prev_col = 'K'
            for col_letter in ['L', 'M', 'N', 'O', 'P', 'Q']:
                formula = f'=IF(B{row_num}="One-Time",0,{prev_col}{row_num}*(1+Metadata!$B$8))'
                self.ws_items[f'{col_letter}{row_num}'] = formula
                self.ws_items[f'{col_letter}{row_num}'].number_format = '$#,##0.00'
                prev_col = col_letter

            # Column R: Total 7-Year Cost (Formula)
            formula_r = f'=SUM(K{row_num}:Q{row_num})'
            self.ws_items[f'R{row_num}'] = formula_r
            self.ws_items[f'R{row_num}'].number_format = '$#,##0.00'

            # Column S: Confidence Score
            if item['confidence_score'] is not None:
                self.ws_items[f'S{row_num}'] = item['confidence_score']
                self.ws_items[f'S{row_num}'].number_format = '0.00%'

            # Column T: Notes
            self.ws_items[f'T{row_num}'] = item['extraction_notes']

            row_num += 1

    def write_data_quality_issues(self, issues: List[Dict]):
        """Write data quality issues to Data_Quality sheet"""
        row_num = 3
        for issue in issues:
            self.ws_quality[f'A{row_num}'] = issue.get('row_id', 'N/A')
            self.ws_quality[f'B{row_num}'] = issue.get('solution_name', 'N/A')
            self.ws_quality[f'C{row_num}'] = issue.get('issue_type', 'Unknown')

            if 'confidence' in issue:
                self.ws_quality[f'D{row_num}'] = issue['confidence']
                self.ws_quality[f'D{row_num}'].number_format = '0.00%'

            self.ws_quality[f'E{row_num}'] = issue.get('description', '')
            row_num += 1

    def populate_summary_sheet(self, items: List[Dict]):
        """
        Populate Summary sheet with dynamic Excel formulas.

        Automatically calculates 7-year totals by category and monthly averages.
        All calculations use Excel formulas, not static values.
        """
        logger.info("Populating Summary sheet with formulas...")

        # Extract unique categories from items, sorted alphabetically
        categories = sorted(set(item['category'] for item in items))

        # Summary sheet starts at row 5 (after header rows)
        summary_row = 5

        # Column mapping: A=Category, B=7-Year Total, C=Monthly Average
        for category in categories:
            # Column A: Category name
            self.ws_summary[f'A{summary_row}'] = category

            # Column B: 7-Year Total using SUMIFS formula
            # SUMIFS(sum_range, criteria_range, criteria)
            # Sum Line Items column R (Total 7-Year Cost) where column D (Category) equals this category
            formula_total = f"=SUMIFS('Line Items'!$R:$R,'Line Items'!$D:$D,A{summary_row})"
            self.ws_summary[f'B{summary_row}'] = formula_total
            self.ws_summary[f'B{summary_row}'].number_format = '$#,##0.00'

            # Column C: Monthly Average = 7-Year Total / 84 months
            # This gives the true average cost per month across the 7-year period (7 years × 12 months = 84 months)
            formula_monthly = f'=B{summary_row}/84'
            self.ws_summary[f'C{summary_row}'] = formula_monthly
            self.ws_summary[f'C{summary_row}'].number_format = '$#,##0.00'

            summary_row += 1

        # Add separator row
        summary_row += 1

        # Add "Total Required TCO" row
        total_row = summary_row
        self.ws_summary[f'A{total_row}'] = 'Total Required TCO'
        self.ws_summary[f'A{total_row}'].font = Font(bold=True)

        # Sum all category totals
        first_category_row = 5
        last_category_row = total_row - 2  # -2 to skip the separator row

        if last_category_row >= first_category_row:
            formula_grand_total = f'=SUM(B{first_category_row}:B{last_category_row})'
            self.ws_summary[f'B{total_row}'] = formula_grand_total
            self.ws_summary[f'B{total_row}'].number_format = '$#,##0.00'
            self.ws_summary[f'B{total_row}'].font = Font(bold=True)

            # Monthly average for total (7 years × 12 months = 84 months)
            formula_total_monthly = f'=B{total_row}/84'
            self.ws_summary[f'C{total_row}'] = formula_total_monthly
            self.ws_summary[f'C{total_row}'].number_format = '$#,##0.00'
            self.ws_summary[f'C{total_row}'].font = Font(bold=True)

        # Optional: Add optional items summary (if any)
        optional_items = [item for item in items if item.get('optional', False)]
        if optional_items:
            optional_row = total_row + 2
            self.ws_summary[f'A{optional_row}'] = 'Optional Items Total'
            self.ws_summary[f'A{optional_row}'].font = Font(italic=True)

            # Sum all items where Optional = TRUE
            formula_optional = "=SUMIFS('Line Items'!$R:$R,'Line Items'!$F:$F,TRUE)"
            self.ws_summary[f'B{optional_row}'] = formula_optional
            self.ws_summary[f'B{optional_row}'].number_format = '$#,##0.00'

            formula_optional_monthly = f'=B{optional_row}/84'
            self.ws_summary[f'C{optional_row}'] = formula_optional_monthly
            self.ws_summary[f'C{optional_row}'].number_format = '$#,##0.00'

        logger.info(f"Summary sheet populated with {len(categories)} categories")

    def populate_year_summary_sheet(self):
        """
        Populate Years 1-7 sheet with dynamic Excel formulas.

        Creates clean 3-row structure:
        - Required Annual Fees (Optional = FALSE)
        - Optional Annual Fees (Optional = TRUE)
        - Total Annual Cost (sum of above)

        All calculations use Excel formulas for automatic recalculation.
        """
        logger.info("Populating Years 1-7 sheet with formulas...")

        # Year_Summary sheet structure:
        # Row 1: Title
        # Row 2: Headers (Category, Year 1, Year 2, ..., Year 7)
        # Row 3: Required Annual Fees
        # Row 4: Optional Annual Fees
        # Row 5: Total Annual Cost (moved from row 7, removing One-Time Fees row)

        # Clear the One-Time Fees row (row 5) - we don't need it
        self.ws_years['A5'] = None
        for col in range(2, 9):  # Columns B-H (Year 1-7)
            col_letter = get_column_letter(col)
            self.ws_years[f'{col_letter}5'] = None

        # Move "Total Annual Cost" from row 7 to row 5
        self.ws_years['A5'] = 'Total Annual Cost'
        self.ws_years['A5'].font = Font(bold=True)

        # Clear old row 7
        self.ws_years['A7'] = None

        # Line Items columns:
        # F = Optional (TRUE/FALSE)
        # K = Year 1, L = Year 2, M = Year 3, N = Year 4, O = Year 5, P = Year 6, Q = Year 7
        year_columns = ['K', 'L', 'M', 'N', 'O', 'P', 'Q']  # Years 1-7 in Line Items

        # Populate formulas for each year (columns B through H in Years 1-7)
        for idx, year_col in enumerate(year_columns, start=2):  # Start at column B (index 2)
            col_letter = get_column_letter(idx)

            # Row 3: Required Annual Fees (Optional = FALSE)
            # SUMIFS(sum_range, criteria_range, criteria)
            formula_required = f"=SUMIFS('Line Items'!${year_col}:${year_col},'Line Items'!$F:$F,FALSE)"
            self.ws_years[f'{col_letter}3'] = formula_required
            self.ws_years[f'{col_letter}3'].number_format = '$#,##0.00'

            # Row 4: Optional Annual Fees (Optional = TRUE)
            formula_optional = f"=SUMIFS('Line Items'!${year_col}:${year_col},'Line Items'!$F:$F,TRUE)"
            self.ws_years[f'{col_letter}4'] = formula_optional
            self.ws_years[f'{col_letter}4'].number_format = '$#,##0.00'

            # Row 5: Total Annual Cost (Required + Optional)
            formula_total = f'={col_letter}3+{col_letter}4'
            self.ws_years[f'{col_letter}5'] = formula_total
            self.ws_years[f'{col_letter}5'].number_format = '$#,##0.00'
            self.ws_years[f'{col_letter}5'].font = Font(bold=True)

        logger.info("Years 1-7 sheet populated with formulas for all 7 years")

    def save(self, output_path: Path):
        """Save the workbook"""
        self.wb.save(output_path)
        logger.info(f"Saved Excel file to: {output_path}")


def transform_json_to_excel(json_file: Path, output_file: Path = None):
    """Main transformation pipeline"""
    logger.info("=" * 80)
    logger.info("JSON TO EXCEL TRANSFORMATION")
    logger.info("=" * 80)
    logger.info(f"Input: {json_file}")

    # Step 1: Load JSON
    try:
        with open(json_file, 'r') as f:
            json_data = json.load(f)
        logger.info(f"[OK] Loaded JSON file")
    except Exception as e:
        logger.error(f"Failed to load JSON: {e}")
        return None

    # Step 2: Validate JSON structure
    if 'line_items' not in json_data or not json_data['line_items']:
        logger.error("Invalid JSON: missing or empty 'line_items'")
        return None

    logger.info(f"Line items in JSON: {len(json_data['line_items'])}")

    # Step 3: Initialize processors
    normalizer = EnumNormalizer()
    processor = LineItemProcessor(normalizer)

    # Step 4: Process line items
    transformed_items = []
    row_id = 1

    for item in json_data['line_items']:
        # Split item if needed
        split_items = processor.split_item_if_needed(item)

        for split_item in split_items:
            transformed = processor.transform_item(split_item, row_id)
            if transformed:
                transformed_items.append(transformed)
                row_id += 1

    logger.info(f"Transformed items: {len(transformed_items)}")
    logger.info(f"Items split: {len(transformed_items) - len(json_data['line_items'])}")
    logger.info(f"Data quality issues: {len(processor.data_quality_issues)}")

    # Step 5: Write to Excel
    writer = ExcelWriter()
    writer.write_metadata(json_data)
    writer.write_line_items(transformed_items)
    # Data_Quality and Enums sheets removed - not included in output
    writer.populate_summary_sheet(transformed_items)  # Auto-populate Summary sheet
    writer.populate_year_summary_sheet()  # Auto-populate Years 1-7 sheet

    # Step 6: Determine output filename
    if not output_file:
        vendor = normalizer.normalize_vendor(json_data.get('vendor', 'Unknown'))
        timestamp = datetime.now().strftime('%Y%m%d')
        output_file = Path(f"TCO Output/{vendor}_TCO_New_{timestamp}.xlsx")

    output_file.parent.mkdir(parents=True, exist_ok=True)

    # Step 7: Save
    writer.save(output_file)

    logger.info("=" * 80)
    logger.info("TRANSFORMATION COMPLETE")
    logger.info(f"Output: {output_file}")
    logger.info(f"Total rows written: {len(transformed_items)}")
    logger.info(f"Data quality issues logged: {len(processor.data_quality_issues)}")
    logger.info("=" * 80)

    return output_file


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python json_to_excel_mapper.py <json_file> [output_file]")
        print()
        print("Example:")
        print("  python json_to_excel_mapper.py 'Extracted JSON/liberty_extraction_ai.json'")
        sys.exit(1)

    json_file = Path(sys.argv[1])
    output_file = Path(sys.argv[2]) if len(sys.argv) > 2 else None

    if not json_file.exists():
        print(f"ERROR: File not found: {json_file}")
        sys.exit(1)

    transform_json_to_excel(json_file, output_file)
