"""
Multi-Vendor Comparison Excel Generator
Creates a comprehensive Excel workbook comparing multiple vendors for a single client

Features:
- Separate sheet for each vendor with all line items
- Summary comparison sheet
- Year-by-year cost projections
- Visual formatting for easy comparison
"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime


class MultiVendorComparison:
    """Generate multi-vendor comparison Excel workbook"""

    def __init__(self, client_name, growth_rate=0.20):
        """
        Initialize comparison generator

        Args:
            client_name: Name of the client
            growth_rate: Year-over-year growth rate (default 20%)
        """
        self.client_name = client_name
        self.growth_rate = growth_rate
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet

        # Styling
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=11)
        self.subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        self.subheader_font = Font(bold=True, size=10)
        self.section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.section_font = Font(bold=True, italic=True)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        self.vendors_data = {}

    def load_vendor_data(self, vendor_name, json_file):
        """
        Load vendor data from JSON file

        Args:
            vendor_name: Display name for vendor (e.g., "FIS", "CSI")
            json_file: Path to JSON extraction file
        """
        with open(json_file, 'r') as f:
            data = json.load(f)

        self.vendors_data[vendor_name] = data
        print(f"[OK] Loaded {vendor_name}: {len(data['line_items'])} line items")

    def create_vendor_sheet(self, vendor_name):
        """
        Create detailed sheet for a specific vendor

        Args:
            vendor_name: Name of vendor (must be in vendors_data)
        """
        if vendor_name not in self.vendors_data:
            print(f"WARNING: No data loaded for {vendor_name}")
            return

        data = self.vendors_data[vendor_name]
        ws = self.wb.create_sheet(f"{vendor_name} Line Items")

        # Title section
        ws.merge_cells('A1:M1')
        ws['A1'] = f"{self.client_name} - {vendor_name} Pricing"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Metadata
        ws['A2'] = "Client:"
        ws['B2'] = self.client_name
        ws['A3'] = "Vendor:"
        ws['B3'] = vendor_name
        ws['A4'] = "Proposal Date:"
        ws['B4'] = data.get('document_date', 'N/A')
        ws['A5'] = "Contract Term:"
        ws['B5'] = f"{data.get('contract_term', 0)} years"
        ws['A6'] = "Growth Rate:"
        ws['B6'] = f"{self.growth_rate*100:.0f}%"

        # Summary totals
        summary = data.get('summary', {})
        ws['D2'] = "Total Monthly (Required):"
        ws['E2'] = summary.get('total_monthly_required', 0)
        ws['E2'].number_format = '$#,##0.00'

        ws['D3'] = "Total Monthly (Optional):"
        ws['E3'] = summary.get('total_monthly_optional', 0)
        ws['E3'].number_format = '$#,##0.00'

        ws['D4'] = "Total One-Time Fees:"
        ws['E4'] = summary.get('total_one_time_fees', 0)
        ws['E4'].number_format = '$#,##0.00'

        ws['D5'] = "Total One-Time Credits:"
        ws['E5'] = -summary.get('total_one_time_credits', 0)
        ws['E5'].number_format = '$#,##0.00'
        ws['E5'].font = Font(color="00AA00")  # Green for credits

        ws['D6'] = "Items Extracted:"
        ws['E6'] = summary.get('items_extracted', 0)

        # Column headers (row 8)
        headers = [
            'Solution Name',
            'Category',
            'Fee Type',
            'Monthly Fee',
            'Per Unit Rate',
            'Unit Description',
            'One-Time Fee',
            'Year 1 Annual',
            'Year 3 Annual',
            'Year 5 Annual',
            'Year 7 Annual',
            'Optional',
            'Third Party'
        ]

        row = 8
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row, col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border

        # Freeze panes
        ws.freeze_panes = 'A9'

        # Process line items
        row = 9
        current_category = None

        # Sort items by category, then required/optional
        line_items = sorted(
            data.get('line_items', []),
            key=lambda x: (x.get('optional', False), x.get('category', 'Unknown'))
        )

        for item in line_items:
            # Add category header if changed
            category = item.get('category', 'Unknown')
            if category != current_category:
                ws.merge_cells(f'A{row}:M{row}')
                cell = ws.cell(row, 1)
                cell.value = f"CATEGORY: {category}"
                cell.fill = self.section_fill
                cell.font = self.section_font
                cell.border = self.border
                row += 1
                current_category = category

            # Item data
            solution_name = item.get('solution_name', '')
            fee_type = item.get('fee_type', '')
            monthly_fee = item.get('monthly_fee', 0)
            per_unit_rate = item.get('per_unit_rate', 0)
            unit_desc = item.get('unit_description', '')
            one_time_fee = item.get('one_time_fee', 0)
            optional = item.get('optional', False)
            third_party = item.get('third_party', False)

            # Calculate annual costs
            year_1_annual = self._calculate_annual_cost(fee_type, monthly_fee, one_time_fee, 1)
            year_3_annual = self._calculate_annual_cost(fee_type, monthly_fee, one_time_fee, 3)
            year_5_annual = self._calculate_annual_cost(fee_type, monthly_fee, one_time_fee, 5)
            year_7_annual = self._calculate_annual_cost(fee_type, monthly_fee, one_time_fee, 7)

            # Write row
            ws.cell(row, 1).value = solution_name
            ws.cell(row, 2).value = category
            ws.cell(row, 3).value = fee_type
            ws.cell(row, 4).value = monthly_fee
            ws.cell(row, 5).value = per_unit_rate
            ws.cell(row, 6).value = unit_desc or ''
            ws.cell(row, 7).value = one_time_fee
            ws.cell(row, 8).value = year_1_annual
            ws.cell(row, 9).value = year_3_annual
            ws.cell(row, 10).value = year_5_annual
            ws.cell(row, 11).value = year_7_annual
            ws.cell(row, 12).value = 'Yes' if optional else 'No'
            ws.cell(row, 13).value = 'Yes' if third_party else 'No'

            # Format currency columns
            for col in [4, 5, 7, 8, 9, 10, 11]:
                ws.cell(row, col).number_format = '$#,##0.00'

            # Highlight optional items
            if optional:
                for col in range(1, 14):
                    ws.cell(row, col).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

            row += 1

        # Auto-size columns
        for col in range(1, 14):
            ws.column_dimensions[get_column_letter(col)].width = 18

        ws.column_dimensions['A'].width = 50  # Solution name wider
        ws.column_dimensions['F'].width = 25  # Unit description wider

        print(f"[OK] Created sheet for {vendor_name}: {len(line_items)} items")

    def create_summary_sheet(self):
        """
        Create summary comparison sheet across all vendors
        """
        ws = self.wb.create_sheet("Summary Comparison", 0)  # Insert at beginning

        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = f"{self.client_name} - Vendor Comparison Summary"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws['A2'] = f"Comparison Date: {datetime.now().strftime('%Y-%m-%d')}"

        # Vendor comparison table
        row = 4
        ws.merge_cells(f'A{row}:H{row}')
        ws.cell(row, 1).value = "MONTHLY RECURRING COSTS"
        ws.cell(row, 1).fill = self.header_fill
        ws.cell(row, 1).font = self.header_font

        row += 1
        ws.cell(row, 1).value = "Vendor"
        ws.cell(row, 2).value = "Required Monthly"
        ws.cell(row, 3).value = "Optional Monthly"
        ws.cell(row, 4).value = "Total Monthly"

        for col in range(1, 5):
            ws.cell(row, col).fill = self.subheader_fill
            ws.cell(row, col).font = self.subheader_font

        row += 1
        for vendor_name, data in self.vendors_data.items():
            summary = data.get('summary', {})
            req_monthly = summary.get('total_monthly_required', 0)
            opt_monthly = summary.get('total_monthly_optional', 0)
            total_monthly = req_monthly + opt_monthly

            ws.cell(row, 1).value = vendor_name
            ws.cell(row, 2).value = req_monthly
            ws.cell(row, 3).value = opt_monthly
            ws.cell(row, 4).value = total_monthly

            for col in [2, 3, 4]:
                ws.cell(row, col).number_format = '$#,##0.00'

            row += 1

        # One-time costs
        row += 2
        ws.merge_cells(f'A{row}:H{row}')
        ws.cell(row, 1).value = "ONE-TIME COSTS"
        ws.cell(row, 1).fill = self.header_fill
        ws.cell(row, 1).font = self.header_font

        row += 1
        ws.cell(row, 1).value = "Vendor"
        ws.cell(row, 2).value = "Implementation Fees"
        ws.cell(row, 3).value = "Implementation Credits"
        ws.cell(row, 4).value = "Net One-Time Cost"

        for col in range(1, 5):
            ws.cell(row, col).fill = self.subheader_fill
            ws.cell(row, col).font = self.subheader_font

        row += 1
        for vendor_name, data in self.vendors_data.items():
            summary = data.get('summary', {})
            one_time_fees = summary.get('total_one_time_fees', 0)
            one_time_credits = summary.get('total_one_time_credits', 0)
            net_one_time = one_time_fees - one_time_credits

            ws.cell(row, 1).value = vendor_name
            ws.cell(row, 2).value = one_time_fees
            ws.cell(row, 3).value = -one_time_credits  # Show as negative
            ws.cell(row, 4).value = net_one_time

            for col in [2, 3, 4]:
                ws.cell(row, col).number_format = '$#,##0.00'

            ws.cell(row, 3).font = Font(color="00AA00")  # Green for credits

            row += 1

        # 7-Year TCO Projection
        row += 2
        ws.merge_cells(f'A{row}:H{row}')
        ws.cell(row, 1).value = "7-YEAR TOTAL COST OF OWNERSHIP (TCO)"
        ws.cell(row, 1).fill = self.header_fill
        ws.cell(row, 1).font = self.header_font

        row += 1
        ws.cell(row, 1).value = "Vendor"
        ws.cell(row, 2).value = "Year 1"
        ws.cell(row, 3).value = "Year 3"
        ws.cell(row, 4).value = "Year 5"
        ws.cell(row, 5).value = "Year 7"
        ws.cell(row, 6).value = "7-Year Total"

        for col in range(1, 7):
            ws.cell(row, col).fill = self.subheader_fill
            ws.cell(row, col).font = self.subheader_font

        row += 1
        for vendor_name, data in self.vendors_data.items():
            summary = data.get('summary', {})
            base_monthly = summary.get('total_monthly_required', 0) + summary.get('total_monthly_optional', 0)
            one_time = summary.get('total_one_time_fees', 0) - summary.get('total_one_time_credits', 0)

            # Calculate TCO for each year
            year_1 = (base_monthly * 12) + one_time
            year_3 = base_monthly * 12 * ((1 + self.growth_rate) ** 2)
            year_5 = base_monthly * 12 * ((1 + self.growth_rate) ** 4)
            year_7 = base_monthly * 12 * ((1 + self.growth_rate) ** 6)

            # 7-year total with compounding growth
            seven_year_total = one_time
            for year in range(1, 8):
                seven_year_total += base_monthly * 12 * ((1 + self.growth_rate) ** (year - 1))

            ws.cell(row, 1).value = vendor_name
            ws.cell(row, 2).value = year_1
            ws.cell(row, 3).value = year_3
            ws.cell(row, 4).value = year_5
            ws.cell(row, 5).value = year_7
            ws.cell(row, 6).value = seven_year_total

            for col in [2, 3, 4, 5, 6]:
                ws.cell(row, col).number_format = '$#,##0.00'
                ws.cell(row, col).font = Font(bold=True)

            row += 1

        # Key insights
        row += 2
        ws.merge_cells(f'A{row}:H{row}')
        ws.cell(row, 1).value = "KEY INSIGHTS"
        ws.cell(row, 1).fill = self.header_fill
        ws.cell(row, 1).font = self.header_font

        row += 1
        ws.cell(row, 1).value = "Item Count Comparison:"
        ws.cell(row, 1).font = Font(bold=True)

        row += 1
        for vendor_name, data in self.vendors_data.items():
            summary = data.get('summary', {})
            items = summary.get('items_extracted', 0)
            ws.cell(row, 1).value = f"  {vendor_name}:"
            ws.cell(row, 2).value = f"{items} line items extracted"
            row += 1

        # Auto-size columns
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 20

        print("[OK] Created summary comparison sheet")

    def _calculate_annual_cost(self, fee_type, monthly_fee, one_time_fee, year_num):
        """
        Calculate annual cost for a given year

        Args:
            fee_type: Type of fee (Monthly F, Monthly V, Annual, One-Time)
            monthly_fee: Monthly fee amount
            one_time_fee: One-time fee amount
            year_num: Year number (1-7)

        Returns:
            Annual cost for that year
        """
        if fee_type == 'One-Time':
            return one_time_fee if year_num == 1 else 0
        elif fee_type == 'Annual':
            # Annual fee grows each year
            return monthly_fee * 12 * ((1 + self.growth_rate) ** (year_num - 1))
        else:  # Monthly F or Monthly V
            # Monthly fees grow each year
            return monthly_fee * 12 * ((1 + self.growth_rate) ** (year_num - 1))

    def save(self, output_file=None):
        """
        Save the workbook

        Args:
            output_file: Output file path (optional)

        Returns:
            Path to saved file
        """
        if not output_file:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = f"TCO Output/{self.client_name}_Multi_Vendor_Comparison_{timestamp}.xlsx"

        Path(output_file).parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(output_file)

        print(f"\n{'='*80}")
        print(f"SAVED: {output_file}")
        print(f"{'='*80}\n")

        return output_file


def create_multi_vendor_comparison(
    client_name,
    vendor_files,
    output_file=None,
    growth_rate=0.20
):
    """
    Create multi-vendor comparison Excel workbook

    Args:
        client_name: Name of the client
        vendor_files: Dict mapping vendor names to JSON file paths
                     e.g., {'FIS': 'path/to/fis.json', 'CSI': 'path/to/csi.json'}
        output_file: Output file path (optional)
        growth_rate: Year-over-year growth rate (default 20%)

    Returns:
        Path to generated file
    """
    print(f"\n{'='*80}")
    print(f"MULTI-VENDOR COMPARISON GENERATOR")
    print(f"{'='*80}")
    print(f"Client: {client_name}")
    print(f"Vendors: {', '.join(vendor_files.keys())}")
    print(f"Growth Rate: {growth_rate*100:.0f}%")
    print(f"{'='*80}\n")

    comparison = MultiVendorComparison(client_name, growth_rate)

    # Load vendor data
    for vendor_name, json_file in vendor_files.items():
        comparison.load_vendor_data(vendor_name, json_file)

    # Create sheets
    for vendor_name in vendor_files.keys():
        comparison.create_vendor_sheet(vendor_name)

    comparison.create_summary_sheet()

    # Save
    output = comparison.save(output_file)

    print("\n{'='*80}")
    print("COMPARISON COMPLETE")
    print(f"{'='*80}\n")

    return output


if __name__ == '__main__':
    import sys

    # Default: Liberty Capital Bank with FIS and CSI
    client_name = "Liberty Capital Bank"
    vendor_files = {
        'FIS': 'Extracted JSON/liberty_capital_bank_fis_extraction_ai.json',
        'CSI': 'Extracted JSON/liberty_capital_bank_csi_extraction_ai.json'
    }

    # Check for command line arguments
    if len(sys.argv) > 1:
        client_name = sys.argv[1]

    if len(sys.argv) > 3:
        vendor_files = {
            sys.argv[2]: sys.argv[3]  # First vendor
        }
        if len(sys.argv) > 5:
            vendor_files[sys.argv[4]] = sys.argv[5]  # Second vendor

    create_multi_vendor_comparison(client_name, vendor_files)
