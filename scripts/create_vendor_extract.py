"""
Create Vendor Extract Excel - Simple, clean format for Arriba Advisors
Takes JSON extraction and creates Excel with line items and yearly charges
"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime

def create_vendor_extract_excel(json_file, output_file=None, growth_rate=0.20):
    """
    Create a clean vendor extract Excel from JSON

    Args:
        json_file: Path to JSON extraction file
        output_file: Output Excel file path (optional)
        growth_rate: Annual growth rate for costs (default 20%)
    """

    # Load JSON data
    with open(json_file, 'r') as f:
        data = json.load(f)

    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create Line Items sheet
    ws = wb.create_sheet("Line Items")

    # Get contract details
    vendor = data.get('vendor', 'Unknown Vendor')
    client = data.get('client', 'Unknown Client')
    contract_term = data.get('contract_term', 7)
    proposal_date = data.get('document_date', data.get('proposal_date', 'N/A'))

    # Add header section
    add_header_section(ws, vendor, client, proposal_date, contract_term, growth_rate)

    # Add column headers (starting at row 6)
    add_column_headers(ws, contract_term)

    # Process and add line items
    line_items = data.get('line_items', [])
    current_row = 7  # Start after headers

    for item in line_items:
        current_row = add_line_item(ws, item, current_row, contract_term, growth_rate)

    # Add summary section
    add_summary_section(ws, current_row + 1, contract_term)

    # Format the sheet
    format_sheet(ws, current_row)

    # Auto-size columns
    auto_size_columns(ws)

    # Determine output filename
    if not output_file:
        vendor_clean = vendor.replace('_', ' ').title()
        timestamp = datetime.now().strftime('%Y%m%d')
        output_file = f"Vendor Extracts/{vendor_clean}_Extract_{timestamp}.xlsx"

    # Ensure output directory exists
    Path(output_file).parent.mkdir(parents=True, exist_ok=True)

    # Save workbook
    wb.save(output_file)
    print(f"\n[OK] Created vendor extract: {output_file}")
    print(f"  - Line items: {len(line_items)}")
    print(f"  - Contract term: {contract_term} years")
    print(f"  - Growth rate: {growth_rate*100}%")

    return output_file


def add_header_section(ws, vendor, client, proposal_date, contract_term, growth_rate):
    """Add header section with metadata"""

    # Title
    ws['A1'] = 'VENDOR PRICING EXTRACT'
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    ws.merge_cells('A1:D1')

    # Metadata
    ws['A2'] = 'Vendor:'
    ws['B2'] = vendor.replace('_', ' ').title()
    ws['A3'] = 'Client:'
    ws['B3'] = client if client != 'Not specified' else 'Echelon Bank'
    ws['A4'] = 'Proposal Date:'
    ws['B4'] = proposal_date

    ws['E2'] = 'Contract Term:'
    ws['F2'] = f'{contract_term} years'
    ws['E3'] = 'Growth Rate:'
    ws['F3'] = f'{growth_rate*100}%'
    ws['E4'] = 'Extracted:'
    ws['F4'] = datetime.now().strftime('%Y-%m-%d')

    # Bold labels
    for cell in ['A2', 'A3', 'A4', 'E2', 'E3', 'E4']:
        ws[cell].font = Font(bold=True)


def add_column_headers(ws, contract_term):
    """Add column headers for line items table"""

    headers = [
        'Solution Name',
        'Category',
        'Fee Type',
        'Monthly Fee',
        'Per Unit Rate',
        'Unit',
        'One-Time Fee',
    ]

    # Add year columns
    for year in range(1, contract_term + 1):
        headers.append(f'Year {year} Annual')

    # Add total column
    headers.append('Total (All Years)')
    headers.append('Notes')

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def add_line_item(ws, item, row, contract_term, growth_rate):
    """Add a single line item to the sheet"""

    # Basic info
    ws.cell(row=row, column=1).value = item.get('solution_name', '')
    ws.cell(row=row, column=2).value = item.get('category', '')
    ws.cell(row=row, column=3).value = item.get('fee_type', '')

    # Fees
    monthly_fee = item.get('monthly_fee', 0) or 0
    per_unit_rate = item.get('per_unit_rate', 0) or 0
    unit_description = item.get('unit_description', '')
    one_time_fee = item.get('one_time_fee', 0) or 0
    fee_type = item.get('fee_type', '')

    ws.cell(row=row, column=4).value = monthly_fee
    ws.cell(row=row, column=4).number_format = '$#,##0.00'

    ws.cell(row=row, column=5).value = per_unit_rate
    ws.cell(row=row, column=5).number_format = '$#,##0.0000'

    ws.cell(row=row, column=6).value = unit_description or ''

    ws.cell(row=row, column=7).value = one_time_fee
    ws.cell(row=row, column=7).number_format = '$#,##0.00'

    # Calculate yearly costs
    year_col_start = 8

    for year_num in range(1, contract_term + 1):
        col_idx = year_col_start + year_num - 1

        # Calculate annual cost for this year
        if fee_type == 'One-Time':
            # One-time fees only appear in Year 1
            annual_cost = one_time_fee if year_num == 1 else 0
        elif fee_type == 'Monthly F':
            # Fixed monthly fee
            year_1_annual = monthly_fee * 12
            annual_cost = year_1_annual * ((1 + growth_rate) ** (year_num - 1))
        elif fee_type == 'Monthly V':
            # Variable monthly fee (we'll use the monthly_fee as baseline)
            year_1_annual = monthly_fee * 12
            annual_cost = year_1_annual * ((1 + growth_rate) ** (year_num - 1))
        elif fee_type == 'Annual':
            # Annual fee
            year_1_fee = monthly_fee  # Sometimes stored in monthly_fee field
            annual_cost = year_1_fee * ((1 + growth_rate) ** (year_num - 1))
        else:
            annual_cost = 0

        cell = ws.cell(row=row, column=col_idx)
        cell.value = annual_cost
        cell.number_format = '$#,##0.00'

    # Total column
    total_col = year_col_start + contract_term
    total_formula = f"=SUM({get_column_letter(year_col_start)}{row}:{get_column_letter(year_col_start + contract_term - 1)}{row})"
    ws.cell(row=row, column=total_col).value = total_formula
    ws.cell(row=row, column=total_col).number_format = '$#,##0.00'
    ws.cell(row=row, column=total_col).font = Font(bold=True)

    # Notes column
    notes_col = total_col + 1
    notes = item.get('extraction_notes', '')
    ws.cell(row=row, column=notes_col).value = notes

    return row + 1


def add_summary_section(ws, start_row, contract_term):
    """Add summary totals at the bottom"""

    # Add blank row
    start_row += 1

    # Summary label
    ws.cell(row=start_row, column=1).value = 'TOTAL COSTS'
    ws.cell(row=start_row, column=1).font = Font(bold=True, size=12)

    # Calculate totals for each year
    year_col_start = 8
    for year_num in range(1, contract_term + 1):
        col_idx = year_col_start + year_num - 1
        col_letter = get_column_letter(col_idx)

        # Sum all costs in this year column (from row 7 to start_row - 2)
        formula = f"=SUM({col_letter}7:{col_letter}{start_row-2})"
        cell = ws.cell(row=start_row, column=col_idx)
        cell.value = formula
        cell.number_format = '$#,##0.00'
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

    # Grand total
    total_col = year_col_start + contract_term
    total_col_letter = get_column_letter(total_col)
    formula = f"=SUM({total_col_letter}7:{total_col_letter}{start_row-2})"
    cell = ws.cell(row=start_row, column=total_col)
    cell.value = formula
    cell.number_format = '$#,##0.00'
    cell.font = Font(bold=True, size=12)
    cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')


def format_sheet(ws, last_data_row):
    """Apply formatting to the sheet"""

    # Freeze panes (freeze first 6 rows and first column)
    ws.freeze_panes = 'B7'

    # Add borders to data table
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )

    # Apply borders to all data cells
    for row in range(6, last_data_row + 3):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).border = thin_border

            # Align text
            if col <= 3:  # Text columns
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            else:  # Number columns
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='right', vertical='top')


def auto_size_columns(ws):
    """Auto-size columns based on content"""

    column_widths = {
        1: 35,  # Solution Name
        2: 20,  # Category
        3: 12,  # Fee Type
        4: 12,  # Monthly Fee
        5: 12,  # Per Unit Rate
        6: 18,  # Unit
        7: 12,  # One-Time Fee
    }

    # Year columns
    for col in range(8, ws.max_column - 1):
        column_widths[col] = 14

    # Total column
    column_widths[ws.max_column - 1] = 16

    # Notes column
    column_widths[ws.max_column] = 30

    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


if __name__ == '__main__':
    import sys

    # Default to Echelon FIS file
    default_file = 'Extracted JSON/echelon_bank_fis_extraction_ai.json'

    if len(sys.argv) > 1:
        json_file = sys.argv[1]
    else:
        json_file = default_file

    if len(sys.argv) > 2:
        output_file = sys.argv[2]
    else:
        output_file = None

    # Growth rate (default 20%)
    growth_rate = 0.20 if len(sys.argv) <= 3 else float(sys.argv[3])

    print(f"\n{'='*70}")
    print("CREATING VENDOR EXTRACT EXCEL")
    print(f"{'='*70}")
    print(f"Input: {json_file}")

    create_vendor_extract_excel(json_file, output_file, growth_rate)

    print(f"{'='*70}")
    print("[COMPLETE]")
    print(f"{'='*70}\n")
