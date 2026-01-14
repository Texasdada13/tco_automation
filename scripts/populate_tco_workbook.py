"""
TCO Workbook Population Script

Maps extracted JSON data to WORKBOOK2.xlsx TCO template.

HARDCODED RULE: All TCO Excel outputs are saved to "TCO Output/" folder
"""

import json
import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
from pathlib import Path

from extraction_config import get_tco_output_path, TCO_OUTPUT_DIR


# =============================================================================
# CONFIGURATION
# =============================================================================

FIS_COLUMN_MAP = {
    'fee_type': 'B',
    'proposal_qty': 'C',
    'solution_name': 'O',
    'category': 'P',
    'per_unit_rate': 'Q',
}

FIS_ROW_SECTIONS = {
    'Bundle': {'start_row': 7, 'end_row': 20},
    'Non-Bundle Required': {'start_row': 22, 'end_row': 54},
    'Non-Bundle Required Third-Party': {'start_row': 55, 'end_row': 85},
    'One-Time': {'start_row': 86, 'end_row': 109},
    'One-Time Third-Party': {'start_row': 110, 'end_row': 120},
    'Non-Bundle Optional': {'start_row': 121, 'end_row': 129},
    'Non-Bundle Optional Third-Party': {'start_row': 130, 'end_row': 164},
}


def get_excel_section(json_item, is_one_time_row=False):
    """
    Determine which Excel row section a JSON item belongs to.

    Args:
        json_item: The JSON item dictionary
        is_one_time_row: If True, this is a split one-time fee row
    """
    category = json_item.get('category', '')
    is_optional = json_item.get('optional', False)
    is_third_party = json_item.get('third_party', False)
    fee_type = json_item.get('fee_type', '')

    # If this is explicitly a one-time row (split from monthly)
    if is_one_time_row:
        if is_third_party:
            return 'One-Time Third-Party', FIS_ROW_SECTIONS['One-Time Third-Party']
        return 'One-Time', FIS_ROW_SECTIONS['One-Time']

    # One-Time items (explicit fee_type)
    if fee_type == 'One-Time' or 'One-Time' in category:
        if is_third_party:
            return 'One-Time Third-Party', FIS_ROW_SECTIONS['One-Time Third-Party']
        return 'One-Time', FIS_ROW_SECTIONS['One-Time']

    # Bundle items - check for various bundle indicators
    if category == 'Bundle' or 'Bundle' in category:
        return 'Bundle', FIS_ROW_SECTIONS['Bundle']
    # Liberty format: "Existing Service - Core" could be bundle
    if 'Core' in category and 'HORIZON' in json_item.get('solution_name', '').upper():
        return 'Bundle', FIS_ROW_SECTIONS['Bundle']

    # Optional items
    if is_optional or 'Optional' in category:
        if is_third_party:
            return 'Non-Bundle Optional Third-Party', FIS_ROW_SECTIONS['Non-Bundle Optional Third-Party']
        return 'Non-Bundle Optional', FIS_ROW_SECTIONS['Non-Bundle Optional']

    # Required items (default)
    if is_third_party:
        return 'Non-Bundle Required Third-Party', FIS_ROW_SECTIONS['Non-Bundle Required Third-Party']
    return 'Non-Bundle Required', FIS_ROW_SECTIONS['Non-Bundle Required']


def split_item_if_needed(json_item):
    """
    Split items that have both monthly AND one-time fees into separate rows.

    Returns list of (item_dict, is_one_time_row) tuples
    """
    monthly_fee = json_item.get('monthly_fee', 0) or json_item.get('per_unit_rate', 0)
    one_time_fee = json_item.get('one_time_fee', 0)

    rows = []

    # Add monthly row if there's a monthly fee
    if monthly_fee and monthly_fee != 0:
        rows.append((json_item, False))

    # Add one-time row if there's a one-time fee
    if one_time_fee and one_time_fee != 0:
        one_time_item = json_item.copy()
        one_time_item['fee_type'] = 'One-Time'
        one_time_item['_rate_override'] = one_time_fee
        one_time_item['solution_name'] = json_item.get('solution_name', '') + ' - Implementation Fee'
        rows.append((one_time_item, True))

    # If no fees at all, still include the item
    if not rows:
        rows.append((json_item, False))

    return rows


def safe_write(ws, cell_ref, value):
    """Safely write to a cell, skipping merged cells."""
    from openpyxl.cell.cell import MergedCell
    try:
        cell = ws[cell_ref]
        if isinstance(cell, MergedCell):
            return False
        cell.value = value
        return True
    except Exception as e:
        print(f"  Warning: Could not write to {cell_ref}: {e}")
        return False


def populate_workbook(json_file, template_file='WORKBOOK2.xlsx', output_file=None, version='v1'):
    """
    Populate TCO workbook with JSON extracted data.

    Args:
        json_file: Path to extracted JSON file
        template_file: Path to WORKBOOK2.xlsx template (default: 'WORKBOOK2.xlsx')
        output_file: Path for output file (if None, uses auto-generated path in TCO Output/)
        version: Version identifier for auto-generated filename (default: 'v1')

    Returns:
        Path to the saved output file
    """
    # Load JSON data
    with open(json_file, 'r') as f:
        data = json.load(f)

    # Auto-generate output filename if not provided
    if output_file is None:
        vendor_name = data.get('vendor', 'Unknown')
        output_file = get_tco_output_path(vendor_name, version)
        print(f"Auto-generating output filename: {output_file}")

    # Ensure output file is in TCO Output directory
    output_path = Path(output_file)
    if output_path.parent.name != "TCO Output":
        # Move to TCO Output directory
        output_file = TCO_OUTPUT_DIR / output_path.name
        print(f"Redirecting output to: {output_file}")

    # Load workbook
    wb = openpyxl.load_workbook(template_file)
    ws = wb['Line Items']

    # Track row usage per section
    section_counters = {section: config['start_row'] for section, config in FIS_ROW_SECTIONS.items()}

    print("=" * 80)
    print(f"POPULATING TCO WORKBOOK")
    print("=" * 80)
    print(f"Source JSON: {json_file}")
    print(f"Template: {template_file}")
    print(f"Output: {output_file}")
    print(f"Total items: {len(data.get('line_items', []))}")
    print()

    # Process each line item
    mapping_report = []

    for item in data.get('line_items', []):
        # Split items that have both monthly AND one-time fees
        split_rows = split_item_if_needed(item)

        for row_item, is_one_time_row in split_rows:
            section_name, section_config = get_excel_section(row_item, is_one_time_row)
            current_row = section_counters[section_name]

            # Check if we've exceeded the section
            if current_row > section_config['end_row']:
                print(f"  WARNING: Section '{section_name}' is full! Skipping: {row_item['solution_name']}")
                continue

            # Get the rate to use
            if '_rate_override' in row_item:
                rate = row_item['_rate_override']
            else:
                rate = row_item.get('monthly_fee', 0) or row_item.get('per_unit_rate', 0) or row_item.get('one_time_fee', 0)

            # Handle negative rates (credits)
            if row_item.get('one_time_fee', 0) < 0 and not is_one_time_row:
                rate = row_item['one_time_fee']

            # Write to cells using safe_write
            safe_write(ws, f"B{current_row}", row_item.get('fee_type', 'Monthly F'))
            safe_write(ws, f"C{current_row}", 1)  # Proposal quantity
            safe_write(ws, f"O{current_row}", row_item.get('solution_name', ''))
            safe_write(ws, f"P{current_row}", row_item.get('category', ''))
            safe_write(ws, f"Q{current_row}", rate)

            # Record mapping
            mapping_report.append({
                'solution_name': row_item.get('solution_name', ''),
                'section': section_name,
                'row': current_row,
                'fee_type': row_item.get('fee_type', ''),
                'rate': rate,
                'cells': f"B{current_row}, C{current_row}, O{current_row}, P{current_row}, Q{current_row}"
            })

            # Increment counter
            section_counters[section_name] += 1

    # Save workbook
    wb.save(output_file)

    # Print mapping report
    print("-" * 80)
    print("MAPPING REPORT")
    print("-" * 80)
    print(f"{'Row':<6} {'Section':<35} {'Solution Name':<40} {'Rate':>12}")
    print("-" * 100)

    for m in mapping_report:
        print(f"{m['row']:<6} {m['section']:<35} {m['solution_name'][:38]:<40} ${m['rate']:>10,.2f}")

    # Print summary by section
    print()
    print("-" * 80)
    print("SECTION SUMMARY")
    print("-" * 80)

    for section, config in FIS_ROW_SECTIONS.items():
        items_written = section_counters[section] - config['start_row']
        available = config['end_row'] - config['start_row'] + 1
        print(f"  {section:<40}: {items_written:>3} items (rows {config['start_row']}-{section_counters[section]-1 if items_written > 0 else 'N/A'})")

    print()
    print(f"Output saved to: {output_file}")
    print("=" * 80)

    return output_file, mapping_report


def show_sample_mapping(json_file):
    """Show sample mapping without writing to Excel."""
    with open(json_file, 'r') as f:
        data = json.load(f)

    print("=" * 100)
    print("SAMPLE MAPPING PREVIEW")
    print("=" * 100)
    print(f"Source: {json_file}")
    print(f"Client: {data.get('client', 'N/A')}")
    print(f"Vendor: {data.get('vendor', 'N/A')}")
    print()

    # Group items by section
    sections = {}
    for item in data.get('line_items', []):
        section_name, _ = get_excel_section(item)
        if section_name not in sections:
            sections[section_name] = []
        sections[section_name].append(item)

    # Print by section
    for section_name, items in sections.items():
        config = FIS_ROW_SECTIONS[section_name]
        print(f"\n{'=' * 100}")
        print(f"SECTION: {section_name}")
        print(f"Excel Rows: {config['start_row']} - {config['end_row']} ({config['end_row'] - config['start_row'] + 1} available)")
        print(f"Items to map: {len(items)}")
        print("-" * 100)
        print(f"{'Row':<6} | {'Col B (Type)':<12} | {'Col O (Solution Name)':<45} | {'Col Q (Rate)':>15}")
        print("-" * 100)

        for idx, item in enumerate(items[:10]):  # Show first 10 per section
            row = config['start_row'] + idx
            rate = item.get('monthly_fee', 0) or item.get('per_unit_rate', 0) or item.get('one_time_fee', 0)
            name = item.get('solution_name', '')[:43]

            print(f"{row:<6} | {item.get('fee_type', ''):<12} | {name:<45} | ${rate:>13,.2f}")

        if len(items) > 10:
            print(f"  ... and {len(items) - 10} more items")


if __name__ == "__main__":
    import sys

    # Default files
    json_file = "Extracted JSON/liberty_extraction_ai.json"

    if len(sys.argv) > 1:
        if sys.argv[1] == "--preview":
            show_sample_mapping(json_file)
        elif sys.argv[1] == "--help":
            print("TCO Workbook Population Script")
            print("=" * 60)
            print()
            print("Usage:")
            print("  python populate_tco_workbook.py <json_file> [version]")
            print()
            print("Examples:")
            print("  python populate_tco_workbook.py 'Extracted JSON/csi_extraction_ai.json'")
            print("  python populate_tco_workbook.py 'Extracted JSON/csi_extraction_ai.json' v2")
            print()
            print("Output:")
            print("  Auto-saves to 'TCO Output/{Vendor}_TCO_Output_{version}.xlsx'")
            print()
            print("Options:")
            print("  --preview   Show mapping preview without writing")
            print("  --help      Show this help message")
        else:
            json_file = sys.argv[1]
            version = sys.argv[2] if len(sys.argv) > 2 else 'v1'
            output_path, report = populate_workbook(json_file, version=version)
            print(f"\nSuccess! File saved to: {output_path}")
    else:
        # Run preview by default
        show_sample_mapping(json_file)
        print("\n" + "=" * 100)
        print("To actually populate the workbook, run:")
        print(f"  python populate_tco_workbook.py '{json_file}'")
        print()
        print("Output will be auto-saved to: TCO Output/")
        print("=" * 100)
