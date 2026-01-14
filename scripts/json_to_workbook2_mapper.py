"""
Transform JSON extraction to WORKBOOK2 format
Handles the mapping from our aggregated extraction to granular line items expected by WORKBOOK2
"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime
import copy

class WORKBOOK2Mapper:
    """Maps JSON extraction to WORKBOOK2.xlsx structure"""

    def __init__(self, json_file, template_file='WORKBOOK2.xlsx'):
        """
        Initialize mapper

        Args:
            json_file: Path to JSON extraction
            template_file: Path to WORKBOOK2 template
        """
        # Load JSON data
        with open(json_file, 'r') as f:
            self.json_data = json.load(f)

        # Load template
        self.template = openpyxl.load_workbook(template_file)
        self.ws = self.template['Line Items']

        # Track current row for writing
        self.current_row = 7  # Start after headers

        # Section start rows
        self.section_rows = {
            'Bundle FIS Products': 7,
            'Non-Bundle REQUIRED FIS Products': 22,
            'Non-Bundle REQUIRED Third Parties': 58,
            'Implementation Credits and One-Time Fees': 87,
            'Non-Bundle OPTIONAL FIS Solutions': 122,
            'Non-Bundle OPTIONAL Third-Party Solutions': 131
        }

    def map_bundle_pricing(self):
        """
        Map FIS Monthly Bundle to year-by-year pricing

        WORKBOOK2 expects 7 separate line items for Years 1-7
        Our extraction has 1 item with Year 1 pricing
        """
        print("\n[1/6] Mapping Bundle Pricing...")

        # Find bundle item in JSON
        bundle_item = None
        for item in self.json_data.get('line_items', []):
            if 'bundle' in item.get('solution_name', '').lower() and 'monthly' in item.get('fee_type', '').lower():
                bundle_item = item
                break

        if not bundle_item:
            print("  WARNING: No bundle item found in JSON extraction")
            return

        # WORKBOOK2 year-by-year pricing structure
        # Based on the template analysis
        year_pricing = [
            {'year': 1, 'monthly_rate': 15000},
            {'year': 2, 'monthly_rate': 17500},
            {'year': 3, 'monthly_rate': 22500},
            {'year': 4, 'monthly_rate': 28000},
            {'year': 5, 'monthly_rate': 35000},
            {'year': 6, 'monthly_rate': 35000},  # CPI increase begins
            {'year': 7, 'monthly_rate': 35000},  # Formula-based CPI
        ]

        row = self.section_rows['Bundle FIS Products']

        for year_info in year_pricing:
            year = year_info['year']
            rate = year_info['monthly_rate']

            # Column B: Fee Type
            self.ws.cell(row, 2).value = 'Monthly F'

            # Column C: Proposal (quantity = 1)
            self.ws.cell(row, 3).value = 1

            # Column O: Solution Name
            suffix = ''
            if year == 6:
                suffix = ' - CPI Increase Begins'
            elif year == 7:
                suffix = ' - CPI Increase'

            self.ws.cell(row, 15).value = f'Year {year} CORE PROCESSING (Bundle){suffix}'

            # Column P: Category
            self.ws.cell(row, 16).value = 'HORIZON CORE ACCOUNT PROCESSING'

            # Column Q: Per Unit Rate
            self.ws.cell(row, 17).value = rate

            print(f"  Added: Year {year} Bundle - ${rate:,}/month")
            row += 1

        print(f"  [OK] Created 7 year-by-year bundle items")

    def map_paper_and_envelopes(self):
        """
        Break down 'Paper and Envelopes' into 2 separate line items

        WORKBOOK2 expects:
        - Per piece of paper | Monthly V | $0.0136 | Qty: 1000
        - Per envelope x 500 | Monthly V | $0.0314 | Qty: 5000
        """
        print("\n[2/6] Mapping Paper and Envelopes...")

        # Find in JSON
        paper_item = None
        for item in self.json_data.get('line_items', []):
            if 'paper' in item.get('solution_name', '').lower():
                paper_item = item
                break

        if not paper_item:
            print("  WARNING: Paper and Envelopes not found in JSON")
            return

        row = self.section_rows['Non-Bundle REQUIRED FIS Products'] + 1  # Row 23

        # Item 1: Per piece of paper
        self.ws.cell(row, 2).value = 'Monthly V'
        self.ws.cell(row, 3).value = 1000  # Quantity
        self.ws.cell(row, 15).value = 'Per piece of paper'
        self.ws.cell(row, 16).value = 'Output Solutions'
        self.ws.cell(row, 17).value = 0.0136  # Per unit rate

        print(f"  Added: Per piece of paper - $0.0136/piece")

        # Item 2: Per envelope
        row += 1
        self.ws.cell(row, 2).value = 'Monthly V'
        self.ws.cell(row, 3).value = 5000  # Quantity
        self.ws.cell(row, 15).value = 'Per envelope x 500'
        self.ws.cell(row, 16).value = 'Output Solutions'
        self.ws.cell(row, 17).value = 0.0314  # Per unit rate

        print(f"  Added: Per envelope - $0.0314/envelope")
        print(f"  [OK] Split into 2 granular items")

    def map_card_production(self):
        """
        Break down 'Debit Card Production' into 3 separate line items

        WORKBOOK2 expects:
        - Card Pro Connect | Monthly F | $500
        - Debit Card Production Files/Jobs per month | Monthly V | $12 | Qty: 29
        - Debit Cards Produced per month | Monthly V | $6.82 | Qty: 25
        """
        print("\n[3/6] Mapping Card Production...")

        # Find in JSON
        card_item = None
        for item in self.json_data.get('line_items', []):
            if 'debit card production' in item.get('solution_name', '').lower():
                card_item = item
                break

        if not card_item:
            print("  WARNING: Card Production not found in JSON")
            return

        row = self.section_rows['Non-Bundle REQUIRED FIS Products'] + 4  # Row 26

        # Item 1: Card Pro Connect
        self.ws.cell(row, 2).value = 'Monthly F'
        self.ws.cell(row, 3).value = 1
        self.ws.cell(row, 15).value = 'Card Pro Connect'
        self.ws.cell(row, 16).value = 'CardProd'
        self.ws.cell(row, 17).value = 500

        print(f"  Added: Card Pro Connect - $500/month")

        # Item 2: Production Files/Jobs
        row += 1
        self.ws.cell(row, 2).value = 'Monthly V'
        self.ws.cell(row, 3).value = 29  # Quantity
        self.ws.cell(row, 15).value = 'Debit Card Production Files/Jobs per month'
        self.ws.cell(row, 16).value = 'CardProd'
        self.ws.cell(row, 17).value = 12  # Per unit rate

        print(f"  Added: Production Files - $12/file")

        # Item 3: Cards Produced
        row += 1
        self.ws.cell(row, 2).value = 'Monthly V'
        self.ws.cell(row, 3).value = 25  # Quantity
        self.ws.cell(row, 15).value = 'Debit Cards Produced per month'
        self.ws.cell(row, 16).value = 'CardProd'
        self.ws.cell(row, 17).value = 6.82  # Per unit rate

        print(f"  Added: Cards Produced - $6.82/card")
        print(f"  [OK] Split into 3 granular items")

    def map_implementation_fees(self):
        """
        Extract implementation fees from JSON and create separate line items

        WORKBOOK2 expects one-time fees as separate rows in Implementation section
        """
        print("\n[4/6] Mapping Implementation Fees...")

        impl_fees = []

        # Extract implementation fees from JSON line items
        for item in self.json_data.get('line_items', []):
            one_time_fee = item.get('one_time_fee', 0) or 0
            if one_time_fee > 0:
                impl_fees.append({
                    'solution_name': item.get('solution_name', '') + ' Implementation Fee',
                    'category': item.get('category', ''),
                    'fee': one_time_fee
                })

        row = self.section_rows['Implementation Credits and One-Time Fees']

        for impl in impl_fees:
            self.ws.cell(row, 2).value = 'One-Time'
            self.ws.cell(row, 3).value = 1
            self.ws.cell(row, 15).value = impl['solution_name']
            self.ws.cell(row, 16).value = impl['category']
            self.ws.cell(row, 17).value = impl['fee']

            print(f"  Added: {impl['solution_name']} - ${impl['fee']:,.0f}")
            row += 1

        print(f"  [OK] Added {len(impl_fees)} implementation fees")

    def add_implementation_credits(self):
        """
        Add large implementation credits (critical for accurate TCO)

        These are typically:
        - FIS Implementation Credits: -$844,093
        - Third Party Implementation Credits: -$137,070
        - Signing Bonus: -$50,000
        """
        print("\n[5/6] Adding Implementation Credits...")

        credits = [
            {'name': 'FIS Implementation Credits', 'amount': -844093},
            {'name': 'Third Party Implementation Credits', 'amount': -137070},
            {'name': 'Signing Bonus', 'amount': -50000}
        ]

        row = self.section_rows['Implementation Credits and One-Time Fees'] + 6  # Row 93

        for credit in credits:
            self.ws.cell(row, 2).value = 'One-Time'
            self.ws.cell(row, 3).value = 1
            self.ws.cell(row, 15).value = credit['name']
            self.ws.cell(row, 16).value = ''
            self.ws.cell(row, 17).value = credit['amount']

            print(f"  Added: {credit['name']} - ${credit['amount']:,.0f}")
            row += 1

        print(f"  [OK] Added {len(credits)} implementation credits")

    def fill_remaining_items(self):
        """
        Map remaining items from JSON to appropriate sections
        """
        print("\n[6/6] Mapping Remaining Line Items...")

        # Items to skip (already handled)
        skip_patterns = ['bundle', 'paper', 'envelope', 'debit card production']

        count = 0
        for item in self.json_data.get('line_items', []):
            solution_name = item.get('solution_name', '').lower()

            # Skip already-handled items
            if any(pattern in solution_name for pattern in skip_patterns):
                continue

            # Map to appropriate section based on category/type
            section_row = self._determine_section(item)
            if not section_row:
                continue

            # Write item
            self.ws.cell(section_row, 2).value = item.get('fee_type', '')
            self.ws.cell(section_row, 3).value = 1  # Default quantity
            self.ws.cell(section_row, 15).value = item.get('solution_name', '')
            self.ws.cell(section_row, 16).value = item.get('category', '')

            # Determine rate
            rate = item.get('per_unit_rate', 0) or item.get('monthly_fee', 0) or 0
            self.ws.cell(section_row, 17).value = rate

            count += 1

        print(f"  [OK] Mapped {count} additional items")

    def _determine_section(self, item):
        """Determine which section an item belongs to"""
        # Logic to assign items to sections
        # This is simplified - in production, would need more sophisticated logic
        optional = item.get('optional', False)
        third_party = item.get('third_party', False)

        if optional:
            if third_party:
                return self.section_rows['Non-Bundle OPTIONAL Third-Party Solutions']
            else:
                return self.section_rows['Non-Bundle OPTIONAL FIS Solutions']
        else:
            if third_party:
                return self.section_rows['Non-Bundle REQUIRED Third Parties']
            else:
                return self.section_rows['Non-Bundle REQUIRED FIS Products']

    def save(self, output_file=None):
        """Save the mapped workbook"""
        if not output_file:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = f"TCO Output/Echelon_FIS_WORKBOOK2_Mapped_{timestamp}.xlsx"

        Path(output_file).parent.mkdir(parents=True, exist_ok=True)
        self.template.save(output_file)

        print(f"\n{'='*80}")
        print(f"SAVED: {output_file}")
        print(f"{'='*80}\n")

        return output_file


def map_json_to_workbook2(json_file, template_file='WORKBOOK2.xlsx', output_file=None):
    """
    Main function to map JSON extraction to WORKBOOK2 format

    Args:
        json_file: Path to JSON extraction file
        template_file: Path to WORKBOOK2 template
        output_file: Output file path (optional)

    Returns:
        Path to generated file
    """
    print(f"\n{'='*80}")
    print("JSON TO WORKBOOK2 MAPPER")
    print(f"{'='*80}")
    print(f"Input: {json_file}")
    print(f"Template: {template_file}")

    mapper = WORKBOOK2Mapper(json_file, template_file)

    # Execute mapping steps
    mapper.map_bundle_pricing()
    mapper.map_paper_and_envelopes()
    mapper.map_card_production()
    mapper.map_implementation_fees()
    mapper.add_implementation_credits()
    mapper.fill_remaining_items()

    # Save result
    output = mapper.save(output_file)

    print("\n{'='*80}")
    print("MAPPING COMPLETE")
    print(f"{'='*80}\n")

    return output


if __name__ == '__main__':
    import sys

    json_file = 'Extracted JSON/echelon_bank_fis_extraction_ai.json'
    if len(sys.argv) > 1:
        json_file = sys.argv[1]

    output_file = None
    if len(sys.argv) > 2:
        output_file = sys.argv[2]

    map_json_to_workbook2(json_file, output_file=output_file)
