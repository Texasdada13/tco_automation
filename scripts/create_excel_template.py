"""
Creates the New_TCO_Excel_v1.xlsx template file

This script generates a professional, finance-ready Excel template
based on the derived_schema.json specification.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import json
from pathlib import Path

def create_new_tco_template():
    """Create the complete New TCO Excel template"""

    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Define colors
    HEADER_BLUE = "4472C4"
    LIGHT_GRAY = "F2F2F2"
    WHITE = "FFFFFF"

    # Define fonts
    header_font = Font(name='Calibri', size=11, bold=True, color=WHITE)
    data_font = Font(name='Calibri', size=10)
    title_font = Font(name='Calibri', size=14, bold=True)

    # Define borders
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    # ========== SHEET 1: Metadata ==========
    ws_meta = wb.create_sheet("Metadata", 0)

    # Set column widths
    ws_meta.column_dimensions['A'].width = 30
    ws_meta.column_dimensions['B'].width = 40

    # Title
    ws_meta['A1'] = "TCO Analysis - Proposal Metadata"
    ws_meta['A1'].font = title_font

    # Metadata fields
    metadata_fields = [
        ("A2", "Primary Vendor:", "B2"),
        ("A3", "Client Name:", "B3"),
        ("A4", "Proposal Type:", "B4"),
        ("A5", "Proposal Date:", "B5"),
        ("A6", "Contract Term (Years):", "B6"),
        ("A7", "Annual Growth Rate:", "B7"),
        ("A8", "Annual CPI Rate:", "B8"),
        ("A10", "Data Extraction Date:", "B10"),
        ("A11", "AI Model Used:", "B11"),
        ("A12", "Source Proposal File:", "B12")
    ]

    for label_cell, label_text, value_cell in metadata_fields:
        ws_meta[label_cell] = label_text
        ws_meta[label_cell].font = Font(name='Calibri', size=10, bold=True)
        ws_meta[label_cell].alignment = Alignment(horizontal='right')

    # Default values
    ws_meta['B6'] = 7
    ws_meta['B7'] = 0.02
    ws_meta['B7'].number_format = '0.00%'
    ws_meta['B8'] = 0.02
    ws_meta['B8'].number_format = '0.00%'

    # Freeze panes
    ws_meta.freeze_panes = 'A2'

    # ========== SHEET 2: Enums ==========
    ws_enums = wb.create_sheet("Enums", 1)

    # Fee Types
    ws_enums['A1'] = "Fee Types"
    ws_enums['A1'].font = header_font
    ws_enums['A1'].fill = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type='solid')

    fee_types = ["Monthly F", "Monthly V", "Annual", "One-Time"]
    for i, ft in enumerate(fee_types, start=2):
        ws_enums[f'A{i}'] = ft

    # Categories
    ws_enums['B1'] = "Categories"
    ws_enums['B1'].font = header_font
    ws_enums['B1'].fill = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type='solid')

    categories = [
        "Core", "Digital", "EFT", "Risk, Fraud & Compliance",
        "Treasury", "Image Solutions", "Item Processing", "FOS",
        "Lending", "ACH", "Accounts Payable", "Security Plus",
        "Network", "Other"
    ]
    for i, cat in enumerate(categories, start=2):
        ws_enums[f'B{i}'] = cat

    # Boolean values
    ws_enums['C1'] = "Boolean"
    ws_enums['C1'].font = header_font
    ws_enums['C1'].fill = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type='solid')
    ws_enums['C2'] = "TRUE"
    ws_enums['C3'] = "FALSE"

    # Set column widths
    ws_enums.column_dimensions['A'].width = 15
    ws_enums.column_dimensions['B'].width = 25
    ws_enums.column_dimensions['C'].width = 10

    # ========== SHEET 3: Line_Items ==========
    ws_items = wb.create_sheet("Line_Items", 2)

    # Define columns
    columns = [
        ("A", "Row ID", 8),
        ("B", "Fee Type", 12),
        ("C", "Solution Name", 40),
        ("D", "Category", 20),
        ("E", "Third Party", 12),
        ("F", "Optional", 10),
        ("G", "Per Unit Rate", 15),
        ("H", "Unit Description", 18),
        ("I", "Avg Monthly Qty", 15),
        ("J", "Year 1 Monthly", 15),
        ("K", "Year 1 Annual", 15),
        ("L", "Year 2", 15),
        ("M", "Year 3", 15),
        ("N", "Year 4", 15),
        ("O", "Year 5", 15),
        ("P", "Year 6", 15),
        ("Q", "Year 7", 15),
        ("R", "Total 7-Year Cost", 18),
        ("S", "Confidence", 12),
        ("T", "Notes", 30)
    ]

    # Write headers
    for col_letter, col_name, col_width in columns:
        cell = ws_items[f'{col_letter}1']
        cell.value = col_name
        cell.font = header_font
        cell.fill = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        ws_items.column_dimensions[col_letter].width = col_width

    # Add data validation for Fee Type (B column)
    dv_fee_type = DataValidation(type="list", formula1='=Enums!$A$2:$A$5', allow_blank=False)
    dv_fee_type.error = 'Invalid Fee Type'
    dv_fee_type.errorTitle = 'Invalid Entry'
    ws_items.add_data_validation(dv_fee_type)
    dv_fee_type.add(f'B2:B1000')

    # Add data validation for Category (D column)
    dv_category = DataValidation(type="list", formula1='=Enums!$B$2:$B$15', allow_blank=False)
    dv_category.error = 'Invalid Category'
    dv_category.errorTitle = 'Invalid Entry'
    ws_items.add_data_validation(dv_category)
    dv_category.add(f'D2:D1000')

    # Add data validation for Boolean columns (E, F)
    dv_boolean = DataValidation(type="list", formula1='=Enums!$C$2:$C$3', allow_blank=False)
    ws_items.add_data_validation(dv_boolean)
    dv_boolean.add(f'E2:E1000')
    dv_boolean.add(f'F2:F1000')

    # Freeze panes
    ws_items.freeze_panes = 'A2'

    # ========== SHEET 4: Summary ==========
    ws_summary = wb.create_sheet("Summary", 3)

    # Title
    ws_summary['A1'] = "Total Cost of Ownership (TCO) Summary"
    ws_summary['A1'].font = title_font
    ws_summary.merge_cells('A1:D1')

    # Vendor name
    ws_summary['A2'] = "Vendor:"
    ws_summary['A2'].font = Font(name='Calibri', size=11, bold=True)
    ws_summary['B2'] = "=Metadata!B2"
    ws_summary['B2'].font = Font(name='Calibri', size=11)

    # Headers
    ws_summary['A4'] = "Category"
    ws_summary['B4'] = "7-Year Total"
    ws_summary['C4'] = "Monthly Average"

    for cell in ['A4', 'B4', 'C4']:
        ws_summary[cell].font = header_font
        ws_summary[cell].fill = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type='solid')
        ws_summary[cell].border = thin_border

    # Summary rows
    summary_rows = [
        ("A5", "Bundle Products"),
        ("A6", "Non-Bundle Required - Vendor"),
        ("A7", "Non-Bundle Required - Third Party"),
        ("A8", "One-Time Fees"),
        ("A10", "Total Required TCO", True),
        ("A12", "Optional Solutions"),
        ("A14", "Grand Total (Required + Optional)", True)
    ]

    row_num = 5
    for cell, label, *is_bold in summary_rows:
        ws_summary[cell] = label
        if is_bold:
            ws_summary[cell].font = Font(name='Calibri', size=11, bold=True)
            ws_summary[cell].fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')

    # Format currency columns
    for row in range(5, 15):
        for col in ['B', 'C']:
            cell = ws_summary[f'{col}{row}']
            cell.number_format = '$#,##0.00'
            cell.border = thin_border

    # Set column widths
    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 20
    ws_summary.column_dimensions['C'].width = 20

    # Freeze panes
    ws_summary.freeze_panes = 'A5'

    # ========== SHEET 5: Year_Summary ==========
    ws_years = wb.create_sheet("Year_Summary", 4)

    # Title
    ws_years['A1'] = "Annual Cost Breakdown (7-Year Projection)"
    ws_years['A1'].font = title_font

    # Headers
    year_headers = ["Category"] + [f"Year {i}" for i in range(1, 8)]
    for i, header in enumerate(year_headers, start=1):
        cell = ws_years.cell(row=2, column=i)
        cell.value = header
        cell.font = header_font
        cell.fill = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Category rows
    year_categories = [
        "Required Annual Fees",
        "Optional Annual Fees",
        "One-Time Fees",
        "",
        "Total Annual Cost"
    ]

    for i, category in enumerate(year_categories, start=3):
        ws_years.cell(row=i, column=1).value = category
        if category == "Total Annual Cost":
            ws_years.cell(row=i, column=1).font = Font(name='Calibri', size=11, bold=True)
            ws_years.cell(row=i, column=1).fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')

    # Format currency cells
    for row in range(3, 8):
        for col in range(2, 9):
            cell = ws_years.cell(row=row, column=col)
            cell.number_format = '$#,##0.00'
            cell.border = thin_border

    # Set column widths
    ws_years.column_dimensions['A'].width = 30
    for col in range(2, 9):
        ws_years.column_dimensions[get_column_letter(col)].width = 15

    # Freeze panes
    ws_years.freeze_panes = 'A3'

    # ========== SHEET 6: Data_Quality ==========
    ws_quality = wb.create_sheet("Data_Quality", 5)

    # Title
    ws_quality['A1'] = "Data Quality & Validation Issues"
    ws_quality['A1'].font = title_font

    # Headers
    quality_headers = ["Row ID", "Solution Name", "Issue Type", "Confidence", "Description"]
    for i, header in enumerate(quality_headers, start=1):
        cell = ws_quality.cell(row=2, column=i)
        cell.value = header
        cell.font = header_font
        cell.fill = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type='solid')
        cell.border = thin_border

    # Set column widths
    ws_quality.column_dimensions['A'].width = 10
    ws_quality.column_dimensions['B'].width = 40
    ws_quality.column_dimensions['C'].width = 25
    ws_quality.column_dimensions['D'].width = 12
    ws_quality.column_dimensions['E'].width = 50

    # Freeze panes
    ws_quality.freeze_panes = 'A3'

    # Save workbook
    output_path = Path("Templates/New_TCO_Excel_v1.xlsx")
    wb.save(output_path)
    print(f"[OK] Created: {output_path}")
    print(f"  Sheets: {', '.join(wb.sheetnames)}")
    print(f"  Line Items columns: 20 (A-T)")
    print(f"  Data validation: Fee Type, Category, Boolean fields")
    print(f"  Professional formatting: Headers, borders, currency formats")

    return output_path


if __name__ == "__main__":
    create_new_tco_template()
